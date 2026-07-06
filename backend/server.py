from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Request, Form, Body
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse, PlainTextResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import hmac
import hashlib
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta, date
from passlib.context import CryptContext
from jose import JWTError, jwt
import httpx
from enum import Enum
import io
import csv
import base64
import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import json
import qrcode

# Facebook Business SDK
try:
    from facebook_business.api import FacebookAdsApi
    from facebook_business.adobjects.lead import Lead
    from facebook_business.adobjects.adaccount import AdAccount
    from facebook_business.adobjects.page import Page
    FACEBOOK_SDK_AVAILABLE = True
except ImportError:
    FACEBOOK_SDK_AVAILABLE = False
    logger.warning("Facebook Business SDK not available")

# LLM Integration for AI Insights
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    LLM_AVAILABLE = True
except ImportError:
    LLM_AVAILABLE = False
    logging.warning("emergentintegrations not available - AI insights will use rule-based fallback")

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440
MSG91_AUTH_KEY = os.environ.get('MSG91_AUTH_KEY', '')

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

class UserRole(str, Enum):
    PLACEMENT_MANAGER = "Placement Manager"
    ADMIN = "Admin"  # Super Admin - full access to all branches
    BRANCH_ADMIN = "Branch Admin"  # Branch Admin - full access to their branch only
    COUNSELLOR = "Counsellor"
    FRONT_DESK = "Front Desk Executive"
    CERTIFICATE_MANAGER = "Certificate Manager"  # Manages certificate requests
    TRAINER = "Trainer"  # Trainer for batch management
    ACADEMIC_CONTROLLER = "Academic Controller"  # Creates quiz exams and curriculum
    WIZBANG = "Wizbang"  # Finance role - manages global income/expenses with locked opening bank balance. UI label: "Wizbang Admin"
    WIZBANG_BRAND_MANAGER = "Wizbang Brand Manager"  # Manages brand clients - content plans + monthly reports

class LeadStatus(str, Enum):
    NEW = "New"
    CONTACT_ATTEMPTED = "Contact Attempted"
    CONNECTED = "Connected"
    CONTACTED = "Contacted"
    INTERESTED = "Interested"
    COUNSELLING_SCHEDULED = "Counselling Scheduled"
    COUNSELLING_COMPLETED = "Counselling Completed"
    PARENT_DISCUSSION_PENDING = "Parent Discussion Pending"
    DEMO_BOOKED = "Demo Booked"  # legacy alias for "Demo Scheduled"
    DEMO_SCHEDULED = "Demo Scheduled"
    DEMO_ATTENDED = "Demo Attended"
    FOLLOW_UP = "Follow-up"
    ADMISSION_LIKELY = "Admission Likely"
    CONVERTED = "Converted"  # legacy alias for "Admitted"
    ADMITTED = "Admitted"
    FUTURE_PROSPECT = "Future Prospect"
    NOT_INTERESTED = "Not Interested"
    LOST = "Lost"


class LeadPriority(str, Enum):
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"


class ParentStatus(str, Enum):
    NOT_CONTACTED = "Not Contacted"
    CONTACTED = "Contacted"
    INTERESTED = "Interested"
    APPROVAL_PENDING = "Approval Pending"
    APPROVED = "Approved"

class FollowUpStatus(str, Enum):
    PENDING = "Pending"
    COMPLETED = "Completed"
    CANCELLED = "Cancelled"

class PaymentMode(str, Enum):
    CASH = "Cash"
    CARD = "Card"
    UPI = "UPI"
    NET_BANKING = "Net Banking"
    CHEQUE = "Cheque"

class PaymentPlanType(str, Enum):
    ONE_TIME = "One-time"
    INSTALLMENTS = "Installments"

# Models
class Branch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    location: str
    address: str
    city: str
    state: str
    pincode: str
    state_code: Optional[str] = None  # Auto-generated: first 2 chars of state
    city_code: Optional[str] = None   # Auto-generated: first 3 chars of city
    owner_name: str
    owner_email: EmailStr
    owner_phone: str
    owner_designation: str
    branch_phone: str
    branch_email: EmailStr
    lead_counter: int = 0      # For custom ID generation
    enrollment_counter: int = 0
    receipt_counter: int = 0
    webhook_key: Optional[str] = None  # Unique key for external lead capture (Google Ads, Meta)
    royalty_percentage: float = 0.0  # Royalty % charged to branch on enrollment payments
    meta_ads_sheet_url: Optional[str] = None  # Legacy single sheet (kept for backwards-compat)
    meta_ads_sheet_last_sync: Optional[datetime] = None
    meta_ads_sheet_last_row: int = 0  # Track last imported row to avoid duplicates
    meta_ads_sheets: List[dict] = Field(default_factory=list)  # Multi-sheet: [{id,url,label,last_row,last_sync}]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BranchCreate(BaseModel):
    name: str
    location: str
    address: str
    city: str
    state: str
    pincode: str
    owner_name: str
    owner_email: EmailStr
    owner_phone: str
    owner_designation: str
    branch_phone: str
    branch_email: EmailStr
    royalty_percentage: float = 0.0

class BranchUpdate(BaseModel):
    name: Optional[str] = None
    location: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    owner_name: Optional[str] = None
    owner_email: Optional[EmailStr] = None
    owner_phone: Optional[str] = None
    owner_designation: Optional[str] = None
    branch_phone: Optional[str] = None
    branch_email: Optional[EmailStr] = None
    royalty_percentage: Optional[float] = None
    meta_ads_sheet_url: Optional[str] = None

# Audit Log Model for tracking changes
class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    user_name: str
    user_role: str
    branch_id: Optional[str] = None
    action: str  # create, update, delete, login, logout, etc.
    entity_type: str  # lead, enrollment, payment, student, batch, etc.
    entity_id: Optional[str] = None
    entity_name: Optional[str] = None  # Human readable identifier
    changes: Optional[dict] = None  # What was changed (old -> new values)
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Program(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    duration: str
    fee: float
    max_discount_percent: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProgramCreate(BaseModel):
    name: str
    duration: str
    fee: float
    max_discount_percent: float

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Optional[UserRole] = UserRole.COUNSELLOR
    branch_id: Optional[str] = None
    phone: Optional[str] = None
    alternate_phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    date_of_birth: Optional[str] = None
    designation: Optional[str] = None
    photo_url: Optional[str] = None
    opening_balance: Optional[float] = None  # Required only for Wizbang role; locked once set

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    name: str
    role: UserRole
    branch_id: Optional[str] = None
    hashed_password: str
    phone: Optional[str] = None
    alternate_phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    date_of_birth: Optional[str] = None
    designation: Optional[str] = None
    photo_url: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: UserRole
    branch_id: Optional[str] = None
    phone: Optional[str] = None
    alternate_phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    date_of_birth: Optional[str] = None
    designation: Optional[str] = None
    photo_url: Optional[str] = None
    is_active: Optional[bool] = True
    monthly_target: Optional[int] = None              # ₹ target (Counsellor)
    monthly_target_leads: Optional[int] = None        # count target (FDE)
    monthly_target_placements: Optional[int] = None   # count target (Placement Manager)

class PasswordChange(BaseModel):
    current_password: Optional[str] = None  # Required for self-change
    new_password: str

class UserStatusUpdate(BaseModel):
    is_active: bool

# Task Management
class TaskStatus(str, Enum):
    PENDING = "Pending"
    IN_PROGRESS = "In Progress"
    COMPLETED = "Completed"

class Task(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    assigned_by: str
    assigned_by_name: str
    assigned_to: str
    assigned_to_name: str
    branch_id: str
    status: TaskStatus = TaskStatus.PENDING
    priority: str = "Normal"  # Low, Normal, High, Urgent
    due_date: Optional[str] = None
    completed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    assigned_to: str
    priority: str = "Normal"
    due_date: Optional[str] = None

class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    due_date: Optional[str] = None

# International Exams
class InternationalExam(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    price: float
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExamBooking(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    booking_id: Optional[str] = None  # Custom ID like PBPTKX0001
    student_name: str
    student_phone: str
    student_email: Optional[str] = None
    exam_id: str
    exam_name: str
    exam_price: float
    branch_id: str
    status: str = "Pending"  # Pending, Confirmed, Completed, Cancelled
    exam_date: Optional[str] = None
    notes: Optional[str] = None
    created_by: str  # Counsellor who booked the exam (for incentive tracking)
    counsellor_incentive: float = 0.0  # 10% of exam_price when completed
    incentive_status: str = "Pending"  # Pending, Earned, Cancelled
    refund_status: Optional[str] = None  # None, Pending, Processed (when cancelled)
    refund_amount: Optional[float] = None  # Amount to refund when cancelled
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Quiz-Based Exams
class QuizQuestion(BaseModel):
    question_number: int
    question_text: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct_answer: str  # A, B, C, or D

class QuizExam(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    duration_minutes: int = 30  # Time limit in minutes
    pass_percentage: int = 60  # Percentage needed to pass
    questions_per_attempt: int = 100  # # of random questions shown to each student
    questions: List[QuizQuestion] = []  # Question bank (any size)
    is_active: bool = True
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuizExamCreate(BaseModel):
    name: str
    description: Optional[str] = None
    duration_minutes: int = 30
    pass_percentage: int = 60
    questions_per_attempt: int = 100
    questions: List[dict] = []

class QuizAttempt(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    exam_id: str
    exam_name: str
    enrollment_number: str  # Student enters this to take the exam
    student_name: Optional[str] = None
    question_numbers: List[int] = []  # The question_numbers randomly picked for this attempt
    answers: dict = {}  # {question_number: "A" or "B" or "C" or "D"}
    score: int = 0  # Number of correct answers
    total_questions: int = 0
    percentage: float = 0.0
    passed: bool = False
    started_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None
    time_taken_seconds: Optional[int] = None

class QuizAttemptCreate(BaseModel):
    exam_id: str
    enrollment_number: str

class QuizAttemptSubmit(BaseModel):
    answers: dict  # {question_number: "A" or "B" or "C" or "D"}

# Add-on Course for existing enrollments
class AddOnCourse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enrollment_id: str
    program_id: str
    program_name: str
    fee_quoted: float
    discount_percent: Optional[float] = 0
    discount_amount: Optional[float] = 0
    final_fee: float
    added_by: str
    added_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AddOnCourseCreate(BaseModel):
    enrollment_id: str
    program_id: str
    fee_quoted: float
    discount_percent: Optional[float] = 0
    discount_amount: Optional[float] = 0

# Schools/Colleges Outreach Management
class OrganizationType(str, Enum):
    SCHOOL = "School"
    COLLEGE = "College"

class Organization(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_type: OrganizationType
    name: str
    city: str
    address: Optional[str] = None
    contact_person_name: str
    contact_number: str
    email: Optional[str] = None
    alternate_number: Optional[str] = None
    alternate_email: Optional[str] = None
    notes: Optional[str] = None
    created_by: str
    branch_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class OrganizationCreate(BaseModel):
    organization_type: str
    name: str
    city: str
    address: Optional[str] = None
    contact_person_name: str
    contact_number: str
    email: Optional[str] = None
    alternate_number: Optional[str] = None
    alternate_email: Optional[str] = None
    notes: Optional[str] = None

class OrganizationUpdate(BaseModel):
    organization_type: Optional[str] = None
    name: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    contact_person_name: Optional[str] = None
    contact_number: Optional[str] = None
    email: Optional[str] = None
    alternate_number: Optional[str] = None
    alternate_email: Optional[str] = None
    notes: Optional[str] = None

class OrganizationFollowUp(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    follow_up_date: str
    follow_up_time: Optional[str] = None
    notes: str
    outcome: Optional[str] = None  # Interested, Not Interested, Call Back, Meeting Scheduled
    created_by: str
    created_by_name: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OrganizationFollowUpCreate(BaseModel):
    organization_id: str
    follow_up_date: str
    follow_up_time: Optional[str] = None
    notes: str
    outcome: Optional[str] = None

# Batch Management Models
class Batch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    program_id: str
    program_name: Optional[str] = None
    trainer_id: str
    trainer_name: Optional[str] = None
    branch_id: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    timing: Optional[str] = None  # e.g., "10:00 AM - 12:00 PM"
    max_students: Optional[int] = 30
    status: str = "Active"  # Active, Completed, Cancelled
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BatchCreate(BaseModel):
    name: str
    program_id: str
    trainer_id: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    timing: Optional[str] = None
    max_students: Optional[int] = 30

class BatchUpdate(BaseModel):
    name: Optional[str] = None
    trainer_id: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    timing: Optional[str] = None
    max_students: Optional[int] = None
    status: Optional[str] = None

class StudentBatchAssignment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enrollment_id: str
    student_name: Optional[str] = None
    batch_id: str
    batch_name: Optional[str] = None
    trainer_id: str
    trainer_name: Optional[str] = None
    assigned_by: str
    assigned_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StudentBatchAssignmentCreate(BaseModel):
    enrollment_id: str

# Attendance Model
class Attendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    batch_id: str
    enrollment_id: str
    student_name: Optional[str] = None
    date: str  # YYYY-MM-DD
    status: str  # Present, Absent, Late
    marked_by: str
    marked_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    remarks: Optional[str] = None

class AttendanceCreate(BaseModel):
    batch_id: str
    enrollment_id: str
    date: str
    status: str
    remarks: Optional[str] = None

class AttendanceBulkCreate(BaseModel):
    batch_id: str
    date: str
    attendance_records: List[dict]  # [{enrollment_id, status, remarks}]

# Curriculum Model
class Curriculum(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    program_id: str
    program_name: Optional[str] = None
    title: str
    description: Optional[str] = None
    topics: List[str] = []  # List of topic names
    duration_weeks: Optional[int] = None
    created_by: str
    branch_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CurriculumCreate(BaseModel):
    program_id: str
    title: str
    description: Optional[str] = None
    topics: List[str] = []
    duration_weeks: Optional[int] = None

# Course Completion Model
class CourseCompletion(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enrollment_id: str
    student_name: Optional[str] = None
    program_id: str
    program_name: Optional[str] = None
    batch_id: Optional[str] = None
    completion_date: str
    exam_status: str = "Pending"  # Pending, Passed, Failed
    exam_score: Optional[float] = None
    marked_by: str
    marked_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    certificate_issued: bool = False
    remarks: Optional[str] = None

# Student Feedback Model
class StudentFeedback(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enrollment_id: str
    student_name: str
    student_phone: Optional[str] = None
    program_name: Optional[str] = None
    branch_id: str
    month: str  # YYYY-MM format
    doubt_clearance: int = 0  # Rating 1-5
    teacher_behavior: int = 0  # Rating 1-5
    facilities: int = 0  # Rating 1-5
    overall_rating: int = 0  # Rating 1-5
    remarks: Optional[str] = None
    collected_by: str  # Counsellor ID
    collected_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StudentFeedbackCreate(BaseModel):
    enrollment_id: str
    doubt_clearance: int
    teacher_behavior: int
    facilities: int
    overall_rating: int
    remarks: Optional[str] = None

# Feedback List Model (auto-generated monthly)
class FeedbackList(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    month: str  # YYYY-MM format
    branch_id: str
    students: List[dict] = []  # List of {enrollment_id, student_name, program_name, feedback_status}
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Fixed Batch Timings (auto-created for trainers)
FIXED_BATCH_TIMINGS = [
    {"slot": 1, "name": "Batch 1", "timing": "9:00 AM to 10:30 AM"},
    {"slot": 2, "name": "Batch 2", "timing": "10:30 AM to 12:00 PM"},
    {"slot": 3, "name": "Batch 3", "timing": "12:00 PM to 1:30 PM"},
    {"slot": 4, "name": "Batch 4", "timing": "2:00 PM to 3:30 PM"},
    {"slot": 5, "name": "Batch 5", "timing": "3:30 PM to 5:00 PM"},
    {"slot": 6, "name": "Batch 6", "timing": "5:00 PM to 6:30 PM"},
]


# User Responsibility Model
class Responsibility(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: Optional[str] = None  # If None, applies to role
    role: str  # Target role (Admin, Branch Admin, Counsellor, etc.)
    branch_id: Optional[str] = None  # If None, applies to all branches
    title: str
    description: str
    priority: str = "medium"  # high, medium, low
    category: str = "general"  # general, daily, weekly, monthly
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True

class ResponsibilityCreate(BaseModel):
    user_id: Optional[str] = None
    role: str
    branch_id: Optional[str] = None
    title: str
    description: str
    priority: str = "medium"
    category: str = "general"

class ResponsibilityUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    priority: Optional[str] = None
    category: Optional[str] = None
    is_active: Optional[bool] = None


# Payment Plan Edit Model
class PaymentPlanEdit(BaseModel):
    plan_type: Optional[str] = None
    total_installments: Optional[int] = None
    installments: Optional[List[dict]] = None  # [{amount, due_date}]

# Cash Handling Model
class CashHandling(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # YYYY-MM-DD format
    branch_id: str
    total_cash: float = 0
    deposit_receipt_url: Optional[str] = None
    remarks: Optional[str] = None
    status: str = "Pending"  # Pending, Deposited
    submitted_by: Optional[str] = None  # FDE user ID
    submitted_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CashHandlingUpdate(BaseModel):
    deposit_receipt_url: Optional[str] = None
    remarks: Optional[str] = None
    status: str = "Deposited"

# Meta Integration Models
class MetaConfig(BaseModel):
    """Meta/Facebook integration configuration per branch"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch_id: str
    app_id: str
    app_secret: str
    page_id: str
    page_ids: List[str] = []  # Multiple pages support
    ad_account_id: Optional[str] = None
    instagram_account_id: Optional[str] = None
    access_token: Optional[str] = None
    webhook_verify_token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_sync_at: Optional[datetime] = None

class MetaConfigCreate(BaseModel):
    branch_id: str
    app_id: str
    app_secret: str
    page_id: str
    page_ids: List[str] = []
    ad_account_id: Optional[str] = None
    instagram_account_id: Optional[str] = None
    access_token: Optional[str] = None

class MetaConfigUpdate(BaseModel):
    app_id: Optional[str] = None
    app_secret: Optional[str] = None
    page_id: Optional[str] = None
    page_ids: Optional[List[str]] = None
    ad_account_id: Optional[str] = None
    instagram_account_id: Optional[str] = None
    access_token: Optional[str] = None
    is_active: Optional[bool] = None

class MetaLead(BaseModel):
    """Lead imported from Facebook Lead Ads"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch_id: str
    leadgen_id: str  # Facebook's lead ID
    page_id: str
    form_id: Optional[str] = None
    ad_id: Optional[str] = None
    campaign_id: Optional[str] = None
    campaign_name: Optional[str] = None
    ad_name: Optional[str] = None
    # Lead data
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    field_data: Dict[str, Any] = {}  # All form fields
    # CRM linkage
    crm_lead_id: Optional[str] = None  # ID of lead created in CRM
    is_synced_to_crm: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    fb_created_time: Optional[str] = None

class MetaAdInsight(BaseModel):
    """Cached ad performance data"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch_id: str
    date: str  # YYYY-MM-DD
    account_id: str
    campaign_id: Optional[str] = None
    campaign_name: Optional[str] = None
    adset_id: Optional[str] = None
    adset_name: Optional[str] = None
    ad_id: Optional[str] = None
    ad_name: Optional[str] = None
    # Metrics
    impressions: int = 0
    reach: int = 0
    clicks: int = 0
    spend: float = 0
    cpc: float = 0
    cpm: float = 0
    ctr: float = 0
    conversions: int = 0
    cost_per_conversion: float = 0
    # Breakdown
    level: str = "account"  # account, campaign, adset, ad
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse
    session: Optional[str] = None  # Academic session (e.g., "2024" for April 2024 - March 2025)

# Academic Session Helper Functions
def get_current_academic_session() -> str:
    """Get current academic session based on today's date.
    Session runs from April 1 to March 31.
    E.g., if today is Jan 2025, session is "2024" (April 2024 - March 2025)
    """
    today = datetime.now()
    if today.month >= 4:  # April onwards
        return str(today.year)
    else:  # Jan-March
        return str(today.year - 1)

def get_session_date_range(session: str) -> tuple:
    """Get start and end dates for an academic session.
    Session "2024" = April 1, 2024 to March 31, 2025
    Returns (start_date, end_date) as datetime objects
    """
    if session == "all":
        return None, None
    # Handle both "2026" and "2026-2027" formats
    if "-" in str(session):
        year = int(session.split("-")[0])
    else:
        year = int(session)
    start_date = datetime(year, 4, 1, 0, 0, 0)
    end_date = datetime(year + 1, 3, 31, 23, 59, 59)
    return start_date, end_date

def get_available_sessions() -> list:
    """Get list of available academic sessions."""
    sessions = [
        {"value": "2024", "label": "2024-2025"},
        {"value": "2025", "label": "2025-2026"},
        {"value": "2026", "label": "2026-2027"}
    ]
    return sessions

# Session model for custom sessions
class AcademicSessionCreate(BaseModel):
    year: int  # Start year (e.g., 2025 for 2025-26)
    label: Optional[str] = None  # Custom label, defaults to "2025-26"
    is_active: bool = True

def get_session_filter(session: Optional[str], date_field: str = "created_at") -> dict:
    """Generate MongoDB date filter for academic session.
    Returns a dict that can be merged into a query.
    """
    if not session or session == "all":
        return {}
    
    start_date, end_date = get_session_date_range(session)
    if start_date and end_date:
        return {
            date_field: {
                "$gte": start_date.isoformat(),
                "$lte": end_date.isoformat()
            }
        }
    return {}

async def get_session_from_request(request: Request) -> Optional[str]:
    """Extract session from request header or query param."""
    # Try header first
    session = request.headers.get('X-Academic-Session')
    if session:
        return session
    # Try query param
    return request.query_params.get('session')

class LeadCreate(BaseModel):
    name: str
    number: str
    alternate_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    email: EmailStr
    program_id: str
    fee_quoted: Optional[float] = None
    discount_percent: Optional[float] = None
    discount_amount: Optional[float] = None  # Direct amount discount
    payment_plan: Optional[str] = None
    lead_source: str
    lead_date: Optional[str] = None  # Custom lead date
    branch_id: Optional[str] = None  # Admin must specify; counsellor/branch admin inherits from user
    counsellor_id: Optional[str] = None  # Lead Owner — Admin/Branch Admin/Front Desk can assign
    # Phase C — Parent & priority
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_email: Optional[str] = None
    parent_status: Optional[ParentStatus] = None
    parent_notes: Optional[str] = None

class LeadUpdate(BaseModel):
    name: Optional[str] = None
    number: Optional[str] = None
    alternate_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    email: Optional[EmailStr] = None
    program_id: Optional[str] = None
    fee_quoted: Optional[float] = None
    discount_percent: Optional[float] = None
    discount_amount: Optional[float] = None  # Direct amount discount
    payment_plan: Optional[str] = None
    lead_source: Optional[str] = None
    status: Optional[LeadStatus] = None
    lead_date: Optional[str] = None  # Custom lead date
    # Demo booking fields
    demo_date: Optional[str] = None
    demo_time: Optional[str] = None
    trainer_name: Optional[str] = None
    # Reassignment
    counsellor_id: Optional[str] = None  # Lead Owner reassign (Admin/Branch Admin/Front Desk only)
    # Phase C — Parent & priority
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_email: Optional[str] = None
    parent_status: Optional[ParentStatus] = None
    parent_notes: Optional[str] = None

class Lead(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: Optional[str] = None  # Custom ID: PBPTKL0001
    name: str
    number: str
    alternate_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    email: str
    program_id: str
    program_name: Optional[str] = None
    fee_quoted: Optional[float] = None
    discount_percent: Optional[float] = None
    discount_amount: Optional[float] = None  # Direct amount discount
    payment_plan: Optional[str] = None
    lead_source: str
    lead_date: Optional[str] = None  # Custom lead date
    status: LeadStatus = LeadStatus.NEW
    branch_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    counsellor_id: Optional[str] = None
    counsellor_name: Optional[str] = None  # Cached counsellor name for quick display
    # Demo booking fields
    demo_date: Optional[str] = None
    demo_time: Optional[str] = None
    trainer_name: Optional[str] = None
    # Soft delete fields
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None
    deleted_by_name: Optional[str] = None
    deletion_reason: Optional[str] = None
    # Meta (Facebook) integration fields
    meta_lead_id: Optional[str] = None  # Link to MetaLead
    meta_campaign: Optional[str] = None  # Campaign name from Facebook
    meta_ad: Optional[str] = None  # Ad name from Facebook
    # Phase C — Parent info + Priority
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_email: Optional[str] = None
    parent_status: Optional[ParentStatus] = None
    parent_notes: Optional[str] = None
    priority: Optional[LeadPriority] = None  # Auto-computed on create/update

class FollowUpCreate(BaseModel):
    lead_id: str
    note: str
    followup_date: datetime
    reminder_time: Optional[str] = None

class FollowUpOutcome(str, Enum):
    CONNECTED = "Connected"
    NOT_CONNECTED = "Not Connected"
    BUSY = "Busy"
    NO_ANSWER = "No Answer"
    SWITCHED_OFF = "Switched Off"
    WRONG_NUMBER = "Wrong Number"
    CALLBACK_REQUESTED = "Callback Requested"

class FollowUpLogCreate(BaseModel):
    """Model for logging follow-up call attempts"""
    followup_id: str
    outcome: FollowUpOutcome
    notes: str
    next_action: Optional[str] = None
    next_followup_date: Optional[datetime] = None

class FollowUpLog(BaseModel):
    """Trail/history entry for a follow-up"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    followup_id: str
    lead_id: str
    branch_id: Optional[str] = None
    outcome: FollowUpOutcome
    notes: str
    next_action: Optional[str] = None
    next_followup_date: Optional[datetime] = None
    logged_by: str
    logged_by_name: Optional[str] = None
    logged_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FollowUp(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str
    branch_id: Optional[str] = None  # Branch ID from the lead
    note: str
    followup_date: datetime
    reminder_time: Optional[str] = None
    status: FollowUpStatus = FollowUpStatus.PENDING
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    created_by_name: Optional[str] = None
    lead_name: Optional[str] = None
    lead_number: Optional[str] = None
    # Enhanced fields
    call_attempts: int = 0
    last_outcome: Optional[str] = None
    last_call_at: Optional[datetime] = None

# Notification Model
class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    branch_id: Optional[str] = None
    type: str  # "followup_reminder", "lead_converted", "cash_pending"
    title: str
    message: str
    data: Optional[dict] = None  # Additional data like lead_id, followup_id
    is_read: bool = False
    play_audio: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Expense Management
class ExpenseCategory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExpenseCategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None

# Lead Source Management (Admin configurable)
class LeadSource(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeadSourceCreate(BaseModel):
    name: str
    description: Optional[str] = None
    name: str
    description: Optional[str] = None

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch_id: str
    category_id: str
    category_name: Optional[str] = None
    name: str
    amount: float
    payment_mode: PaymentMode
    expense_date: date
    remarks: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExpenseCreate(BaseModel):
    category_id: str
    name: str
    amount: float
    payment_mode: PaymentMode
    expense_date: str
    remarks: Optional[str] = None

# Campaign Management for Branch Admin
class Campaign(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch_id: str
    campaign_name: str
    platform: str  # Google, Meta, Instagram, etc.
    campaign_link: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None  # Optional - filled when campaign completes
    total_spend: float = 0.0
    total_leads: int = 0
    total_messages: int = 0
    status: str = "Active"  # Active, Completed, Paused
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CampaignCreate(BaseModel):
    campaign_name: str
    platform: str
    campaign_link: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None  # Optional - filled when campaign completes
    total_spend: float = 0.0
    total_leads: int = 0
    total_messages: int = 0
    status: str = "Active"
    notes: Optional[str] = None

# WhatsApp Notification Settings
class WhatsAppTemplate(BaseModel):
    template_name: str
    template_namespace: str
    variables: List[str] = []  # List of variable names like ["name", "course", "amount"]

class WhatsAppEventConfig(BaseModel):
    enabled: bool = True
    template_name: str = ""
    namespace: str = ""
    variables: List[str] = []  # Variable names for this event

class WhatsAppSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enabled: bool = True
    integrated_number: str = "918728054145"  # Your MSG91 registered WhatsApp number
    # Per-event template configuration with specific variables
    events: dict = Field(default_factory=lambda: {
        "enquiry_saved": {
            "enabled": True,
            "template_name": "eti_enquiry_confirmation",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name", "course"],
            "description": "When a new enquiry/lead is saved"
        },
        "demo_booked": {
            "enabled": True,
            "template_name": "eti_demo_confirmation",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name", "demo_date", "demo_time", "trainer"],
            "description": "When demo is scheduled for a lead"
        },
        "enrollment_confirmed": {
            "enabled": True,
            "template_name": "eti_enrollment_confirmation",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name", "enrollment_number", "course"],
            "description": "Thank you message when enrollment is confirmed"
        },
        "payment_received": {
            "enabled": True,
            "template_name": "eti_payment",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name", "amount", "total_fee", "paid_fee", "pending_fee", "receipt_number"],
            "description": "Fee payment confirmation with details"
        },
        "fee_reminder": {
            "enabled": True,
            "template_name": "eti_fee_reminder",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name", "amount_due", "due_date"],
            "description": "Automatic pending fee reminders"
        },
        "birthday_wishes": {
            "enabled": True,
            "template_name": "eti_birthday_wishes",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name"],
            "description": "Birthday wishes sent on student's DOB"
        },
        "certificate_ready": {
            "enabled": True,
            "template_name": "eti_certificate",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            # MSG91 `eti_certificate` template body variable order (per user):
            #   {{1}} = student name   {{2}} = course name   {{3}} = certificate ID
            "variables": ["name", "course", "certificate_id"],
            "description": "Sent when certificate is marked as Printed and ready to hand over"
        },
        "demo_reminder": {
            "enabled": True,
            "template_name": "eti_demo_reminder",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name", "trainer", "demo_time"],
            "description": "30-minute reminder before scheduled demo"
        },
        "placement_confirmed": {
            "enabled": True,
            "template_name": "eti_placement_confirmed",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name", "company", "designation", "salary", "city"],
            "description": "Congratulations message when student is marked placed"
        },
        "fee_late": {
            "enabled": True,
            "template_name": "eti_fee_late",
            "namespace": "73fda5e9_77e9_445f_82ac_9c2e532b32f4",
            "variables": ["name"],
            "description": "Sent once when a fee installment is overdue (past due date and unpaid)"
        }
    })
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None

class WhatsAppSettingsUpdate(BaseModel):
    enabled: Optional[bool] = None
    integrated_number: Optional[str] = None
    events: Optional[dict] = None

# Push Notifications
class PushNotification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sender_id: str
    sender_name: str
    sender_role: str
    recipient_ids: List[str]  # List of user IDs
    recipient_role: Optional[str] = None  # Target role if sending to all of a role
    branch_id: Optional[str] = None
    title: str
    message: str
    notification_type: str = "general"  # general, task, reminder
    is_read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PushNotificationCreate(BaseModel):
    recipient_ids: Optional[List[str]] = None
    recipient_role: Optional[str] = None  # Send to all users of this role
    title: str
    message: str
    notification_type: str = "general"

# Browser Push Subscription for Web Push Notifications
class PushSubscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    endpoint: str
    keys: dict  # Contains p256dh and auth keys
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PushSubscriptionCreate(BaseModel):
    endpoint: str
    keys: dict

# Webhook Lead from external sources (Google Ads, Meta)
class WebhookLeadCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    source: Optional[str] = None  # e.g., "Google Ads", "Meta", "Facebook"
    campaign: Optional[str] = None  # Campaign name
    ad_name: Optional[str] = None  # Ad name
    form_name: Optional[str] = None  # Form name
    program_name: Optional[str] = None  # Program interest if captured
    city: Optional[str] = None
    state: Optional[str] = None
    additional_data: Optional[dict] = None  # Any extra fields from the platform

# Certificate Management
class CertificateStatus(str, Enum):
    PENDING = "Pending"
    APPROVED = "Approved"
    REJECTED = "Rejected"
    READY = "Ready"  # Downloaded/Issued
    PRINTED = "Printed"  # Physically printed and handed over to the student

class TrainingMode(str, Enum):
    OFFLINE = "Offline"
    ONLINE = "Online"
    HYBRID = "Hybrid"

class CertificateRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    certificate_id: Optional[str] = None  # Custom ID: ETI-2025-00001
    enrollment_id: str
    enrollment_number: str  # The visible enrollment number like PBPTKE0001
    
    # Student Info (auto-fetched from enrollment)
    student_name: str
    program_name: str
    program_duration: str
    branch_name: str
    branch_id: str
    
    # Student-provided info
    email: str
    phone: str
    program_start_date: str
    program_end_date: str
    training_mode: TrainingMode = TrainingMode.OFFLINE
    training_hours: Optional[int] = None
    
    # Certificate details
    registration_number: Optional[str] = None  # ETI-STU-XXXX
    verification_id: Optional[str] = None  # Unique QR verification code
    
    # Status and workflow
    status: CertificateStatus = CertificateStatus.PENDING
    rejection_reason: Optional[str] = None
    
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    approved_at: Optional[datetime] = None
    approved_by: Optional[str] = None
    approved_by_name: Optional[str] = None
    issued_at: Optional[datetime] = None  # When certificate was downloaded
    issued_by: Optional[str] = None

    # Manual overrides — when True, live-enrichment will not overwrite the
    # student/program/branch fields on this certificate. Set automatically on
    # any edit through the PUT endpoint or when created via /manual.
    is_manually_edited: bool = False
    is_manually_created: bool = False

class CertificateRequestCreate(BaseModel):
    enrollment_number: str  # Student enters this
    email: str
    phone: str
    program_start_date: str
    program_end_date: str
    training_mode: TrainingMode = TrainingMode.OFFLINE
    training_hours: Optional[int] = None

class CertificateRequestUpdate(BaseModel):
    # Any field on the certificate can be edited by Admin / Certificate Manager.
    student_name: Optional[str] = None
    program_name: Optional[str] = None
    program_duration: Optional[str] = None
    branch_id: Optional[str] = None
    branch_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    program_start_date: Optional[str] = None
    program_end_date: Optional[str] = None
    training_mode: Optional[TrainingMode] = None
    training_hours: Optional[int] = None
    registration_number: Optional[str] = None
    certificate_id: Optional[str] = None
    verification_id: Optional[str] = None


class CertificateManualCreate(BaseModel):
    """Payload for a Cert Manager / Admin to create a certificate from scratch,
    bypassing all eligibility gates (fees, course completion, exam pass, duplicates).
    Enables issuing multiple certificates to the same student even against a single
    enrollment (reissue, correction, multi-module recognitions, etc.).
    """
    student_name: str
    program_name: str
    program_duration: Optional[str] = ""
    branch_id: Optional[str] = None
    branch_name: Optional[str] = None
    email: Optional[str] = ""
    phone: Optional[str] = ""
    program_start_date: str
    program_end_date: str
    training_mode: TrainingMode = TrainingMode.OFFLINE
    training_hours: Optional[int] = None
    enrollment_number: Optional[str] = None  # Optional link to an existing enrollment
    registration_number: Optional[str] = None  # If omitted, auto-generated
    certificate_id: Optional[str] = None  # If omitted, auto-generated
    initial_status: Optional[str] = "Approved"  # "Pending" or "Approved" (default: Approved)


# Enrollment Management
class EnrollmentStatus(str, Enum):
    ACTIVE = "Active"
    COMPLETED = "Completed"
    CANCELLED = "Cancelled"
    DROPPED = "Dropped"

class Enrollment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enrollment_id: Optional[str] = None  # Custom ID: PBPTKE0001
    lead_id: str
    branch_id: str
    student_name: str
    email: str
    phone: str
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    
    # Photo uploads
    student_photo_url: Optional[str] = None
    aadhar_photo_url: Optional[str] = None
    aadhar_documents: Optional[List[str]] = None  # Multiple aadhar images/PDFs
    
    # Academic Info
    highest_qualification: Optional[str] = None
    institution_name: Optional[str] = None
    passing_year: Optional[str] = None
    percentage: Optional[float] = None
    
    # Program Info
    program_id: str
    program_name: str
    fee_quoted: float
    discount_percent: Optional[float] = None
    discount_amount: Optional[float] = None  # Direct amount discount
    final_fee: float
    
    # Status
    status: EnrollmentStatus = EnrollmentStatus.ACTIVE
    payment_status: str = "Pending"  # Pending, Partial, Paid
    total_paid: float = 0
    
    enrollment_date: date
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    cancelled_at: Optional[datetime] = None
    cancelled_by: Optional[str] = None
    cancellation_reason: Optional[str] = None

class EnrollmentCreate(BaseModel):
    lead_id: str
    student_name: str
    email: EmailStr
    phone: str
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    highest_qualification: Optional[str] = None
    institution_name: Optional[str] = None
    passing_year: Optional[str] = None
    percentage: Optional[float] = None
    program_id: str
    fee_quoted: float
    discount_percent: Optional[float] = None
    discount_amount: Optional[float] = None  # Direct amount discount
    enrollment_date: str
    student_photo_url: Optional[str] = None
    aadhar_photo_url: Optional[str] = None
    aadhar_documents: Optional[List[str]] = None  # Multiple aadhar images/PDFs

# Payment Management
class PaymentPlan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enrollment_id: str
    plan_type: PaymentPlanType
    total_amount: float
    installments_count: Optional[int] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enrollment_id: str
    payment_plan_id: Optional[str] = None
    branch_id: str
    amount: float
    payment_mode: PaymentMode
    payment_date: date
    installment_number: Optional[int] = None
    remarks: Optional[str] = None
    receipt_number: str = Field(default_factory=lambda: f"RCP-{str(uuid.uuid4())[:8].upper()}")
    student_name: Optional[str] = None  # For reports
    program_name: Optional[str] = None  # For reports
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Marketing Resources Model
class ResourceType(str, Enum):
    BROCHURE = "Brochure"
    CREATIVE = "Creative"
    VIDEO = "Video"
    DOCUMENT = "Document"

class MarketingResource(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    resource_type: ResourceType
    file_url: Optional[str] = None
    video_link: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MarketingResourceCreate(BaseModel):
    title: str
    description: Optional[str] = None
    resource_type: ResourceType
    file_url: Optional[str] = None
    video_link: Optional[str] = None

class PaymentCreate(BaseModel):
    enrollment_id: str
    payment_plan_id: Optional[str] = None
    amount: float
    payment_mode: PaymentMode
    payment_date: str
    installment_number: Optional[int] = None
    remarks: Optional[str] = None

class PaymentPlanCreate(BaseModel):
    enrollment_id: str
    plan_type: PaymentPlanType
    total_amount: float
    installments_count: Optional[int] = None
    installments: Optional[List[dict]] = None

# ============== STUDENT PORTAL MODELS ==============
class StudentTopicProgress(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str
    enrollment_id: str
    program_id: str
    curriculum_id: str
    topic_index: int
    topic_name: str
    done: bool = True
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StudentTopicProgressUpdate(BaseModel):
    enrollment_id: str
    program_id: str
    curriculum_id: str
    topic_index: int
    topic_name: str
    done: bool

class StudentFeedback(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str
    enrollment_id: str
    student_name: str
    student_email: Optional[str] = None
    student_phone: Optional[str] = None
    branch_id: str
    program_id: Optional[str] = None
    program_name: Optional[str] = None
    message: str
    rating: Optional[int] = None  # 1-5
    is_read: bool = False
    read_by: Optional[str] = None
    read_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StudentFeedbackCreate(BaseModel):
    enrollment_id: str
    message: str
    program_id: Optional[str] = None
    rating: Optional[int] = None

class StudentLoginRequest(BaseModel):
    enrollment_number: str
    password: str

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if user is None:
        raise credentials_exception
    user_obj = User(**user)
    if not user_obj.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Your account has been deactivated. Please contact your administrator.",
        )
    return user_obj

def require_role(allowed_roles: List[UserRole]):
    async def role_checker(current_user: User = Depends(get_current_user)):
        if current_user.role not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return current_user
    return role_checker

async def get_current_student(token: str = Depends(oauth2_scheme)):
    """Resolve the current student from the JWT.
    
    Student JWTs use role='Student' and sub=enrollment_id (any one of the student's enrollments).
    Returns a dict with student identity + branch + lead_id so downstream queries can fetch
    all enrollments belonging to the same person.
    """
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate student credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("role") != "Student":
            raise credentials_exception
        enrollment_id = payload.get("sub")
        if not enrollment_id:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    enrollment = await db.enrollments.find_one(
        {"$or": [{"enrollment_id": enrollment_id}, {"id": enrollment_id}]},
        {"_id": 0}
    )
    if not enrollment:
        raise credentials_exception
    
    return {
        "enrollment_id": enrollment.get("enrollment_id") or enrollment.get("id"),
        "lead_id": enrollment.get("lead_id"),
        "name": enrollment.get("student_name"),
        "email": enrollment.get("email"),
        "phone": enrollment.get("phone"),
        "branch_id": enrollment.get("branch_id"),
    }

# Helper function to generate state/city codes
def generate_state_code(state: str) -> str:
    """Generate 2-letter state code"""
    state_map = {
        "punjab": "PB", "haryana": "HR", "delhi": "DL", "uttar pradesh": "UP",
        "maharashtra": "MH", "karnataka": "KA", "tamil nadu": "TN", "kerala": "KL",
        "west bengal": "WB", "gujarat": "GJ", "rajasthan": "RJ", "madhya pradesh": "MP",
        "andhra pradesh": "AP", "telangana": "TG", "bihar": "BR", "odisha": "OR",
        "jharkhand": "JH", "chhattisgarh": "CG", "assam": "AS", "himachal pradesh": "HP",
        "uttarakhand": "UK", "goa": "GA", "jammu and kashmir": "JK", "chandigarh": "CH"
    }
    state_lower = state.lower().strip()
    return state_map.get(state_lower, state[:2].upper())

def generate_city_code(city: str) -> str:
    """Generate 3-letter city code"""
    city_clean = city.strip().upper()
    # Remove spaces and get first 3 consonants or characters
    consonants = ''.join([c for c in city_clean if c not in 'AEIOU '])
    if len(consonants) >= 3:
        return consonants[:3]
    return city_clean[:3]

async def generate_custom_id(branch_id: str, id_type: str) -> str:
    """Generate custom ID based on branch state/city codes
    id_type: 'L' for Lead, 'E' for Enrollment, 'R' for Receipt
    Returns format: PBPTKL0001
    """
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        return None
    
    state_code = branch.get('state_code') or generate_state_code(branch.get('state', 'XX'))
    city_code = branch.get('city_code') or generate_city_code(branch.get('city', 'XXX'))
    
    # Get and increment counter
    counter_field = {
        'L': 'lead_counter',
        'E': 'enrollment_counter', 
        'R': 'receipt_counter'
    }.get(id_type, 'lead_counter')
    
    result = await db.branches.find_one_and_update(
        {"id": branch_id},
        {"$inc": {counter_field: 1}},
        return_document=True,
        projection={"_id": 0, counter_field: 1}
    )
    
    counter = result.get(counter_field, 1) if result else 1
    
    # Format: PBPTKL0001
    return f"{state_code}{city_code}{id_type}{counter:04d}"

async def get_whatsapp_settings():
    """Get WhatsApp notification settings"""
    settings = await db.whatsapp_settings.find_one({}, {"_id": 0})
    if not settings:
        # Create default settings
        default_settings = WhatsAppSettings().model_dump()
        default_settings['updated_at'] = default_settings['updated_at'].isoformat()
        await db.whatsapp_settings.insert_one(default_settings)
        return WhatsAppSettings()
    
    # Ensure the events structure exists and has all required events
    default_events = WhatsAppSettings().events
    existing_events = settings.get('events', {})
    
    # Merge existing events with defaults (keep user-configured values)
    merged_events = {}
    for event_key, default_config in default_events.items():
        if event_key in existing_events:
            merged_events[event_key] = {**default_config, **existing_events[event_key]}
        else:
            merged_events[event_key] = default_config
    
    settings['events'] = merged_events
    
    if isinstance(settings.get('updated_at'), str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])
    
    return WhatsAppSettings(**settings)

async def send_whatsapp_notification(phone_number: str, event_type: str, template_data: dict):
    """Send WhatsApp notification via MSG91 template API with per-event configuration"""
    # TEST MODE: override recipient to a single test number when WHATSAPP_TEST_NUMBER is set
    test_number = os.environ.get('WHATSAPP_TEST_NUMBER', '').strip()
    if test_number:
        logging.info(f"[WHATSAPP TEST MODE] Redirecting message for '{event_type}' from {phone_number} -> {test_number}")
        phone_number = test_number
    
    settings = await get_whatsapp_settings()
    
    if not settings.enabled:
        logging.info("WhatsApp notifications are disabled globally")
        return {"success": False, "reason": "Notifications disabled"}
    
    # Get event configuration
    event_config = settings.events.get(event_type, {})
    if not event_config:
        logging.info(f"No configuration found for event '{event_type}'")
        return {"success": False, "reason": f"Event '{event_type}' not configured"}
    
    # Check if this specific event is enabled
    if not event_config.get('enabled', False):
        logging.info(f"WhatsApp notification for event '{event_type}' is disabled")
        return {"success": False, "reason": f"Event '{event_type}' is disabled"}
    
    # Check if template is configured
    template_name = event_config.get('template_name', '')
    namespace = event_config.get('namespace', '')
    if not template_name or not namespace:
        logging.info(f"Template not configured for event '{event_type}'")
        return {"success": False, "reason": f"Template not configured for '{event_type}'"}
    
    # Build component variables based on event type
    variables = event_config.get('variables', [])
    components = {}
    for i, var_name in enumerate(variables, 1):
        var_value = template_data.get(var_name, '')
        if var_value:
            components[f"body_{i}"] = {"type": "text", "value": str(var_value)}
    
    # Send via MSG91 template API
    return await send_whatsapp_template_with_config(
        phone_number=phone_number,
        template_name=template_name,
        namespace=namespace,
        integrated_number=settings.integrated_number,
        components=components
    )

async def send_whatsapp_template_with_config(phone_number: str, template_name: str, namespace: str, 
                                              integrated_number: str, components: dict):
    """Send WhatsApp message via MSG91 Template API with full configuration"""
    MSG91_KEY = os.environ.get('MSG91_AUTH_KEY', '')
    
    if not MSG91_KEY:
        logging.warning("MSG91_AUTH_KEY not configured. Skipping WhatsApp message.")
        return {"success": False, "error": "MSG91_AUTH_KEY not configured"}
    
    try:
        url = "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/"
        
        headers = {
            "authkey": MSG91_KEY,
            "Content-Type": "application/json"
        }
        
        # Format phone number (ensure it has country code)
        formatted_phone = phone_number.strip()
        if not formatted_phone.startswith('+'):
            if not formatted_phone.startswith('91'):
                formatted_phone = '91' + formatted_phone
        formatted_phone = formatted_phone.replace('+', '')
        
        payload = {
            "integrated_number": integrated_number,
            "content_type": "template",
            "payload": {
                "messaging_product": "whatsapp",
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {
                        "code": "en",
                        "policy": "deterministic"
                    },
                    "namespace": namespace,
                    "to_and_components": [
                        {
                            "to": [formatted_phone],
                            "components": components
                        }
                    ]
                }
            }
        }
        
        async with httpx.AsyncClient() as client:
            logging.info(
                f"MSG91 send → template={template_name} to={formatted_phone} components={components}"
            )
            response = await client.post(url, headers=headers, json=payload, timeout=30.0)
            logging.info(f"MSG91 WhatsApp API response: {response.status_code} - {response.text}")
            
            if response.status_code == 200:
                return {"success": True, "response": response.json()}
            else:
                return {"success": False, "error": response.text}
                
    except Exception as e:
        logging.error(f"Failed to send WhatsApp message: {str(e)}")
        return {"success": False, "error": str(e)}

async def send_whatsapp_template(phone_number: str, body_value: str, settings: WhatsAppSettings):
    """Legacy: Send WhatsApp message via MSG91 Template API (keeping for backwards compatibility)"""
    MSG91_KEY = os.environ.get('MSG91_AUTH_KEY', '')
    
    if not MSG91_KEY:
        logging.warning("MSG91_AUTH_KEY not configured. Skipping WhatsApp message.")
        return {"success": False, "error": "MSG91_AUTH_KEY not configured"}
    
    try:
        url = "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/"
        
        headers = {
            "authkey": MSG91_KEY,
            "Content-Type": "application/json"
        }
        
        # Format phone number (ensure it has country code)
        formatted_phone = phone_number.strip()
        if not formatted_phone.startswith('+'):
            if not formatted_phone.startswith('91'):
                formatted_phone = '91' + formatted_phone
        formatted_phone = formatted_phone.replace('+', '')
        
        # Use first event's template as default
        template_name = "crmwelcome"
        namespace = "73fda5e9_77e9_445f_82ac_9c2e532b32f4"
        
        payload = {
            "integrated_number": settings.integrated_number,
            "content_type": "template",
            "payload": {
                "messaging_product": "whatsapp",
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {
                        "code": "en",
                        "policy": "deterministic"
                    },
                    "namespace": namespace,
                    "to_and_components": [
                        {
                            "to": [formatted_phone],
                            "components": {
                                "body_1": {
                                    "type": "text",
                                    "value": body_value
                                }
                            }
                        }
                    ]
                }
            }
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(url, headers=headers, json=payload, timeout=30.0)
            logging.info(f"MSG91 WhatsApp API response: {response.status_code} - {response.text}")
            
            if response.status_code == 200:
                return {"success": True, "response": response.json()}
            else:
                return {"success": False, "error": response.text}
                
    except Exception as e:
        logging.error(f"Failed to send WhatsApp message: {str(e)}")
        return {"success": False, "error": str(e)}

async def send_whatsapp_message(phone_number: str, message_text: str, lead_name: str):
    """Send WhatsApp message via MSG91 API (legacy text-based)"""
    MSG91_KEY = os.environ.get('MSG91_AUTH_KEY', '')
    
    if not MSG91_KEY:
        logging.warning("MSG91_AUTH_KEY not configured. Skipping WhatsApp message.")
        return {"success": False, "error": "MSG91_AUTH_KEY not configured"}
    
    try:
        url = "https://control.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/"
        
        headers = {
            "authkey": MSG91_KEY,
            "Content-Type": "application/json"
        }
        
        # Format phone number (ensure it has country code)
        formatted_phone = phone_number.strip()
        if not formatted_phone.startswith('+'):
            if not formatted_phone.startswith('91'):
                formatted_phone = '91' + formatted_phone
        formatted_phone = formatted_phone.replace('+', '')
        
        payload = {
            "integrated_number": "919876543210",  # Replace with your MSG91 registered number
            "content_type": "text",
            "payload": {
                "messaging_product": "whatsapp",
                "type": "text",
                "text": {
                    "body": message_text
                }
            },
            "recipients": [
                {
                    "mobiles": formatted_phone,
                    "name": lead_name
                }
            ]
        }
        
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            response = await http_client.post(url, headers=headers, json=payload)
            
            if response.status_code == 200:
                logging.info(f"WhatsApp message sent successfully to {phone_number}")
                return {"success": True, "message": "Message sent successfully"}
            else:
                logging.error(f"MSG91 API error: {response.status_code} - {response.text}")
                return {"success": False, "error": f"API returned status {response.status_code}"}
                
    except Exception as e:
        logging.error(f"Error sending WhatsApp message: {str(e)}")
        return {"success": False, "error": str(e)}

# Audit Logging Helper Function
async def create_audit_log(
    user: User,
    action: str,
    entity_type: str,
    entity_id: str = None,
    entity_name: str = None,
    changes: dict = None,
    request: Request = None
):
    """Create an audit log entry for tracking user actions"""
    try:
        log = AuditLog(
            user_id=user.id,
            user_email=user.email,
            user_name=user.name,
            user_role=user.role.value if hasattr(user.role, 'value') else str(user.role),
            branch_id=user.branch_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            changes=changes,
            ip_address=request.client.host if request else None,
            user_agent=request.headers.get("user-agent", "")[:200] if request else None
        )
        log_dict = log.model_dump()
        log_dict['created_at'] = log_dict['created_at'].isoformat()
        await db.audit_logs.insert_one(log_dict)
    except Exception as e:
        logging.error(f"Failed to create audit log: {e}")

# Authentication Endpoints
@api_router.post("/auth/register", response_model=UserResponse)
async def register(user: UserCreate):
    existing_user = await db.users.find_one({"email": user.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Wizbang role: enforce single global account + require opening_balance
    if user.role == UserRole.WIZBANG.value or user.role == UserRole.WIZBANG:
        if user.opening_balance is None:
            raise HTTPException(status_code=400, detail="Opening bank balance is required for Wizbang role")
        if user.opening_balance < 0:
            raise HTTPException(status_code=400, detail="Opening bank balance cannot be negative")
        existing_wizbang_account = await db.wizbang_account.find_one({}, {"_id": 0})
        if existing_wizbang_account:
            raise HTTPException(status_code=400, detail="A Wizbang account already exists. Only one Wizbang user is allowed.")
    
    new_user = User(
        email=user.email,
        name=user.name,
        role=user.role,
        branch_id=user.branch_id,
        phone=user.phone,
        alternate_phone=user.alternate_phone,
        address=user.address,
        city=user.city,
        state=user.state,
        pincode=user.pincode,
        date_of_birth=user.date_of_birth,
        designation=user.designation,
        photo_url=user.photo_url,
        hashed_password=get_password_hash(user.password)
    )
    
    user_dict = new_user.model_dump()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # Create Wizbang account with locked opening balance
    if user.role == UserRole.WIZBANG.value or user.role == UserRole.WIZBANG:
        await db.wizbang_account.insert_one({
            "id": str(uuid.uuid4()),
            "owner_user_id": new_user.id,
            "owner_email": new_user.email,
            "opening_balance": float(user.opening_balance),
            "opening_balance_set_at": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"Wizbang account created for {new_user.email} with opening balance ₹{user.opening_balance} (LOCKED)")
    
    # Auto-create fixed batches for Trainer role
    if user.role == UserRole.TRAINER.value and user.branch_id:
        for slot in FIXED_BATCH_TIMINGS:
            batch = {
                "id": str(uuid.uuid4()),
                "name": f"{new_user.name} - {slot['name']} ({slot['timing']})",
                "program_id": None,  # Will be assigned later
                "program_name": None,
                "trainer_id": new_user.id,
                "trainer_name": new_user.name,
                "branch_id": user.branch_id,
                "timing": slot['timing'],
                "slot_number": slot['slot'],
                "max_students": 30,
                "status": "Active",
                "created_by": "system",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.batches.insert_one(batch)
        logger.info(f"Auto-created {len(FIXED_BATCH_TIMINGS)} fixed batches for trainer: {new_user.name}")
    
    return UserResponse(**{k: v for k, v in new_user.model_dump().items() if k != 'hashed_password'})

@api_router.post("/auth/login", response_model=Token)
async def login(form_data: OAuth2PasswordRequestForm = Depends(), session: Optional[str] = Form(None)):
    user_doc = await db.users.find_one({"email": form_data.username}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    
    user = User(**user_doc)
    if not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Your account has been deactivated. Please contact your administrator.")
    
    # Use provided session or default to current session
    selected_session = session if session else get_current_academic_session()
    
    # Include session in JWT token
    access_token = create_access_token(data={"sub": user.email, "session": selected_session})
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(**{k: v for k, v in user.model_dump().items() if k != 'hashed_password'}),
        session=selected_session
    )

# Get available academic sessions
@api_router.get("/auth/sessions")
async def get_sessions():
    """Get list of available academic sessions."""
    return {
        "sessions": get_available_sessions(),
        "current_session": get_current_academic_session()
    }

@api_router.get("/auth/session-stats/{session_year}")
async def get_session_stats(session_year: str):
    """Get stats for a specific session (public endpoint for login page)"""
    try:
        year = int(session_year)
        # Session runs from April of year to March of year+1
        start_date = datetime(year, 4, 1)
        end_date = datetime(year + 1, 3, 31, 23, 59, 59)
        
        date_query = {
            "created_at": {
                "$gte": start_date.isoformat(),
                "$lte": end_date.isoformat()
            }
        }
        
        # Count enquiries (leads) for this session
        total_enquiries = await db.leads.count_documents({
            **date_query,
            "is_deleted": {"$ne": True}
        })
        
        # Count converted leads
        converted = await db.leads.count_documents({
            **date_query,
            "status": "Converted",
            "is_deleted": {"$ne": True}
        })
        
        # Count enrollments
        total_enrollments = await db.enrollments.count_documents(date_query)
        
        # Calculate total collections
        payments = await db.payments.find(date_query, {"_id": 0, "amount": 1}).to_list(10000)
        total_collections = sum(p.get("amount", 0) for p in payments)
        
        return {
            "session": f"{year}-{year+1}",
            "total_enquiries": total_enquiries,
            "converted": converted,
            "total_enrollments": total_enrollments,
            "total_collections": total_collections
        }
    except Exception as e:
        return {
            "session": session_year,
            "total_enquiries": 0,
            "converted": 0,
            "total_enrollments": 0,
            "total_collections": 0
        }

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: User = Depends(get_current_user)):
    return UserResponse(**{k: v for k, v in current_user.model_dump().items() if k != 'hashed_password'})

# Admin - Branch Management
@api_router.post("/admin/branches", response_model=Branch)
async def create_branch(branch: BranchCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    # Auto-generate state and city codes
    state_code = generate_state_code(branch.state)
    city_code = generate_city_code(branch.city)
    
    # Generate unique webhook key for external lead capture
    import secrets
    webhook_key = secrets.token_urlsafe(32)
    
    new_branch = Branch(
        **branch.model_dump(),
        state_code=state_code,
        city_code=city_code,
        lead_counter=0,
        enrollment_counter=0,
        receipt_counter=0,
        webhook_key=webhook_key
    )
    branch_dict = new_branch.model_dump()
    branch_dict['created_at'] = branch_dict['created_at'].isoformat()
    
    await db.branches.insert_one(branch_dict)
    return new_branch

@api_router.get("/admin/branches", response_model=List[Branch])
async def get_branches(current_user: User = Depends(get_current_user)):
    branches = await db.branches.find({}, {"_id": 0}).to_list(1000)
    for branch in branches:
        if isinstance(branch.get('created_at'), str):
            branch['created_at'] = datetime.fromisoformat(branch['created_at'])
    return [Branch(**b) for b in branches]

# Admin - Program Management
@api_router.post("/admin/programs", response_model=Program)
async def create_program(program: ProgramCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    new_program = Program(**program.model_dump())
    program_dict = new_program.model_dump()
    program_dict['created_at'] = program_dict['created_at'].isoformat()
    
    await db.programs.insert_one(program_dict)
    return new_program

@api_router.get("/programs", response_model=List[Program])
async def get_programs(current_user: User = Depends(get_current_user)):
    programs = await db.programs.find({}, {"_id": 0}).to_list(1000)
    for program in programs:
        if isinstance(program.get('created_at'), str):
            program['created_at'] = datetime.fromisoformat(program['created_at'])
    return [Program(**p) for p in programs]

@api_router.get("/admin/programs", response_model=List[Program])
async def get_programs_admin(current_user: User = Depends(require_role([UserRole.ADMIN]))):
    return await get_programs(current_user)

@api_router.put("/admin/programs/{program_id}", response_model=Program)
async def update_program(program_id: str, program_update: ProgramCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    result = await db.programs.update_one(
        {"id": program_id},
        {"$set": program_update.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    
    updated_program = await db.programs.find_one({"id": program_id}, {"_id": 0})
    if isinstance(updated_program.get('created_at'), str):
        updated_program['created_at'] = datetime.fromisoformat(updated_program['created_at'])
    return Program(**updated_program)

@api_router.delete("/admin/programs/{program_id}")
async def delete_program(program_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    result = await db.programs.delete_one({"id": program_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    return {"message": "Program deleted successfully"}

# ============ ACADEMIC SESSION MANAGEMENT ============

@api_router.get("/admin/sessions")
async def get_all_sessions(current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Get all academic sessions including custom ones"""
    # Get default sessions
    default_sessions = get_available_sessions()
    
    # Get custom sessions from DB
    custom_sessions = await db.academic_sessions.find({}, {"_id": 0}).to_list(100)
    
    # Merge - custom sessions override default ones
    session_dict = {s["value"]: s for s in default_sessions}
    for cs in custom_sessions:
        session_dict[str(cs["year"])] = {
            "value": str(cs["year"]),
            "label": cs.get("label", f"{cs['year']}-{str(cs['year']+1)[2:]}"),
            "is_active": cs.get("is_active", True),
            "is_custom": True
        }
    
    # Sort by year descending (newest first)
    sessions = sorted(session_dict.values(), key=lambda x: int(x["value"]), reverse=True)
    return sessions

@api_router.post("/admin/sessions")
async def create_session(session_data: AcademicSessionCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Create a new academic session (Super Admin only)"""
    # Check if session already exists
    existing = await db.academic_sessions.find_one({"year": session_data.year})
    if existing:
        raise HTTPException(status_code=400, detail=f"Session {session_data.year}-{session_data.year+1} already exists")
    
    label = session_data.label or f"{session_data.year}-{session_data.year+1}"
    
    new_session = {
        "id": str(uuid.uuid4()),
        "year": session_data.year,
        "label": label,
        "is_active": session_data.is_active,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    
    await db.academic_sessions.insert_one(new_session)
    return {"message": f"Session {label} created successfully", "session": {"value": str(session_data.year), "label": label}}

@api_router.delete("/admin/sessions/{year}")
async def delete_session(year: int, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Delete a custom academic session (Super Admin only)"""
    result = await db.academic_sessions.delete_one({"year": year})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Custom session not found")
    return {"message": f"Session {year}-{year+1} deleted successfully"}

@api_router.delete("/admin/branches/{branch_id}")
async def delete_branch(branch_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    result = await db.branches.delete_one({"id": branch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    return {"message": "Branch deleted successfully"}

@api_router.put("/admin/branches/{branch_id}", response_model=Branch)
async def update_branch(branch_id: str, branch_update: BranchUpdate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    # Only update non-None fields
    update_data = {k: v for k, v in branch_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.branches.update_one(
        {"id": branch_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    updated_branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if isinstance(updated_branch.get('created_at'), str):
        updated_branch['created_at'] = datetime.fromisoformat(updated_branch['created_at'])
    return Branch(**updated_branch)

@api_router.get("/admin/branches/{branch_id}", response_model=Branch)
async def get_branch_details(branch_id: str, current_user: User = Depends(get_current_user)):
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    if isinstance(branch.get('created_at'), str):
        branch['created_at'] = datetime.fromisoformat(branch['created_at'])
    return Branch(**branch)

@api_router.post("/upload/photo", response_model=dict)
async def upload_photo(file: bytes = None, current_user: User = Depends(get_current_user)):
    """Simple base64 photo upload endpoint"""
    if not file:
        raise HTTPException(status_code=400, detail="No file provided")
    
    import base64
    # In production, upload to S3/cloud storage
    # For now, store as base64
    photo_data = base64.b64encode(file).decode('utf-8')
    photo_url = f"data:image/jpeg;base64,{photo_data[:100]}..."  # Truncated for demo
    
    return {"photo_url": photo_url, "message": "Photo uploaded successfully"}

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Delete a user - Super Admin only"""
    # Prevent deleting self
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

@api_router.put("/admin/users/{user_id}/password")
async def change_user_password(user_id: str, password_data: PasswordChange, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Change any user's password - Super Admin only"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    hashed_password = pwd_context.hash(password_data.new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"hashed_password": hashed_password}})
    return {"message": "Password changed successfully"}

@api_router.put("/admin/users/{user_id}/status")
async def update_user_status(user_id: str, status_data: UserStatusUpdate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Mark user as active/inactive - Super Admin only"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one({"id": user_id}, {"$set": {"is_active": status_data.is_active}})
    return {"message": f"User {'activated' if status_data.is_active else 'deactivated'} successfully"}


@api_router.put("/admin/users/{user_id}/target")
async def update_user_target(
    user_id: str,
    payload: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ADMIN])),
):
    """Set per-user monthly target for hero progress bar.
    Body accepts any of:
      - monthly_target            (₹ amount, for Counsellor)
      - monthly_target_leads      (count, for FDE)
      - monthly_target_placements (count, for Placement Manager)
    Super Admin only."""
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    update: dict = {}
    for key in ("monthly_target", "monthly_target_leads", "monthly_target_placements"):
        if key in payload and payload[key] is not None:
            try:
                update[key] = int(payload[key])
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"{key} must be a number")
    if not update:
        raise HTTPException(status_code=400, detail="No target field provided")
    await db.users.update_one({"id": user_id}, {"$set": update})
    return {"ok": True, "user_id": user_id, "updated": update}

@api_router.put("/auth/change-password")
async def change_own_password(password_data: PasswordChange, current_user: User = Depends(get_current_user)):
    """Change own password - requires current password"""
    if not password_data.current_password:
        raise HTTPException(status_code=400, detail="Current password is required")
    
    user = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not pwd_context.verify(password_data.current_password, user['hashed_password']):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    hashed_password = pwd_context.hash(password_data.new_password)
    await db.users.update_one({"id": current_user.id}, {"$set": {"hashed_password": hashed_password}})
    return {"message": "Password changed successfully"}

# Admin - User Management
@api_router.post("/admin/users", response_model=UserResponse)
async def create_user(user: UserCreate, current_user: User = Depends(get_current_user)):
    """Create a new user - Super Admin can create all roles, Branch Admin can only create Trainers"""
    if current_user.role == UserRole.ADMIN:
        # Super Admin can create any role
        pass
    elif current_user.role == UserRole.BRANCH_ADMIN:
        # Branch Admin can only create Trainer role
        if user.role != UserRole.TRAINER.value:
            raise HTTPException(status_code=403, detail="Branch Admin can only create Trainer users")
        # Set branch_id to Branch Admin's branch
        user.branch_id = current_user.branch_id
    else:
        raise HTTPException(status_code=403, detail="Only Admin or Branch Admin can create users")
    
    return await register(user)

@api_router.get("/admin/users", response_model=List[UserResponse])
async def get_users(current_user: User = Depends(get_current_user)):
    """Get users - Super Admin sees all, Branch Admin sees their branch users"""
    if current_user.role == UserRole.ADMIN:
        users = await db.users.find({}, {"_id": 0, "hashed_password": 0}).to_list(1000)
    elif current_user.role == UserRole.BRANCH_ADMIN:
        users = await db.users.find(
            {"branch_id": current_user.branch_id}, 
            {"_id": 0, "hashed_password": 0}
        ).to_list(1000)
    else:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return [UserResponse(**u) for u in users]

@api_router.get("/branch/users", response_model=List[UserResponse])
async def get_branch_users(current_user: User = Depends(get_current_user)):
    """Get users in the current user's branch - For Branch Admin and Counsellor task assignment"""
    if current_user.role == UserRole.ADMIN:
        # Super Admin sees all users
        users = await db.users.find({}, {"_id": 0, "hashed_password": 0}).to_list(1000)
    elif current_user.role == UserRole.BRANCH_ADMIN:
        # Branch Admin sees users in their branch only
        users = await db.users.find(
            {"branch_id": current_user.branch_id}, 
            {"_id": 0, "hashed_password": 0}
        ).to_list(1000)
    elif current_user.role == UserRole.COUNSELLOR:
        # Counsellor can only see Trainers and FDEs in their branch for task assignment
        users = await db.users.find(
            {
                "branch_id": current_user.branch_id,
                "role": {"$in": [UserRole.TRAINER.value, UserRole.FRONT_DESK.value]}
            }, 
            {"_id": 0, "hashed_password": 0}
        ).to_list(1000)
    else:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return [UserResponse(**u) for u in users]

@api_router.get("/branch/counsellors")
async def get_branch_counsellors(current_user: User = Depends(get_current_user)):
    """Return active Counsellors visible to the current user (for Lead Owner picker).

    - Admin: all active counsellors across all branches.
    - Branch Admin / Front Desk / Counsellor: counsellors in their own branch.
    """
    if current_user.role == UserRole.ADMIN:
        query: Dict[str, Any] = {"role": UserRole.COUNSELLOR.value, "is_active": True}
    elif current_user.role in [UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK, UserRole.COUNSELLOR]:
        if not current_user.branch_id:
            return []
        query = {"role": UserRole.COUNSELLOR.value, "is_active": True, "branch_id": current_user.branch_id}
    else:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    docs = await db.users.find(query, {"_id": 0, "id": 1, "name": 1, "email": 1, "branch_id": 1}).to_list(500)
    return [{"id": d["id"], "name": d.get("name") or d.get("email", "Counsellor"), "email": d.get("email"), "branch_id": d.get("branch_id")} for d in docs]


@api_router.get("/branch/lead-owners")
async def get_branch_lead_owners(current_user: User = Depends(get_current_user)):
    """Return active users eligible to own a lead — Counsellors AND Branch Admins.
    Used by the FDE / Branch Admin / Admin "Assign Lead to" picker.

    - Admin: all active counsellors + branch admins across all branches.
    - Branch Admin / Front Desk / Counsellor: counsellors + branch admins in their own branch.
    """
    eligible_roles = [UserRole.COUNSELLOR.value, UserRole.BRANCH_ADMIN.value]
    if current_user.role == UserRole.ADMIN:
        query: Dict[str, Any] = {"role": {"$in": eligible_roles}, "is_active": True}
    elif current_user.role in [UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK, UserRole.COUNSELLOR]:
        if not current_user.branch_id:
            return []
        query = {"role": {"$in": eligible_roles}, "is_active": True, "branch_id": current_user.branch_id}
    else:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    docs = await db.users.find(
        query,
        {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1, "branch_id": 1},
    ).to_list(500)
    # Sort Counsellors first, then Branch Admins, then by name
    role_order = {UserRole.COUNSELLOR.value: 0, UserRole.BRANCH_ADMIN.value: 1}
    docs.sort(key=lambda d: (role_order.get(d.get("role"), 9), (d.get("name") or d.get("email") or "").lower()))
    return [
        {
            "id": d["id"],
            "name": d.get("name") or d.get("email", "User"),
            "email": d.get("email"),
            "role": d.get("role"),
            "branch_id": d.get("branch_id"),
        }
        for d in docs
    ]

# Lead Management
@api_router.post("/leads", response_model=Lead)
async def create_lead(lead: LeadCreate, request: Request, current_user: User = Depends(get_current_user)):
    program = await db.programs.find_one({"id": lead.program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    # Use user's branch_id, or accept one supplied in the payload (e.g. admin
    # creating a lead on behalf of a specific branch). Never silently fall back
    # to a "default" sentinel — that creates orphan records.
    branch_id = current_user.branch_id or lead.branch_id
    if not branch_id:
        raise HTTPException(status_code=400, detail="branch_id is required (your user has no branch assigned)")
    branch_doc = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch_doc:
        raise HTTPException(status_code=400, detail=f"Branch {branch_id} not found")
    
    # Generate custom lead ID
    custom_lead_id = await generate_custom_id(branch_id, "L")
    
    payload_data = lead.model_dump()
    payload_data.pop("branch_id", None)  # don't pass twice
    requested_counsellor_id = payload_data.pop("counsellor_id", None)

    # Determine the Lead Owner (counsellor / branch admin):
    # - Counsellor / Branch Admin creating directly → auto-assign to themselves.
    # - FDE / Admin creating → if no explicit owner, leave unassigned (counsellor_id_final = None);
    #   if an explicit owner_id supplied, validate it's an active Counsellor or Branch Admin in this branch.
    counsellor_id_final: Optional[str] = None
    counsellor_name_final: Optional[str] = None
    if current_user.role in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN]:
        # Auto-assign to creator
        counsellor_id_final = current_user.id
        counsellor_name_final = current_user.name
    if requested_counsellor_id:
        # Only Admin / Branch Admin / Front Desk are allowed to set an arbitrary owner.
        if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
            raise HTTPException(status_code=403, detail="Only Admin / Branch Admin / Front Desk can assign Lead Owner")
        counsellor_doc = await db.users.find_one(
            {
                "id": requested_counsellor_id,
                "role": {"$in": [UserRole.COUNSELLOR.value, UserRole.BRANCH_ADMIN.value]},
                "branch_id": branch_id,
                "is_active": True,
            },
            {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1},
        )
        if not counsellor_doc:
            raise HTTPException(status_code=400, detail="Selected Lead Owner must be an active Counsellor or Branch Admin in this branch")
        counsellor_id_final = counsellor_doc["id"]
        counsellor_name_final = counsellor_doc.get("name") or counsellor_doc.get("email", "Counsellor")

    new_lead = Lead(
        **payload_data,
        lead_id=custom_lead_id,
        branch_id=branch_id,
        counsellor_id=counsellor_id_final,
        counsellor_name=counsellor_name_final,
        program_name=program['name']
    )
    # Auto-compute priority based on initial state
    new_lead.priority = compute_lead_priority(new_lead.model_dump())
    lead_dict = new_lead.model_dump()
    lead_dict['created_at'] = lead_dict['created_at'].isoformat()
    lead_dict['updated_at'] = lead_dict['updated_at'].isoformat()
    
    await db.leads.insert_one(lead_dict)
    
    # Audit Log
    await create_audit_log(
        user=current_user,
        action="create",
        entity_type="lead",
        entity_id=new_lead.id,
        entity_name=f"{new_lead.name} ({custom_lead_id})",
        changes={"name": new_lead.name, "phone": new_lead.number, "program": program['name']},
        request=request
    )
    
    # Send WhatsApp notification for new enquiry/lead
    await send_whatsapp_notification(
        new_lead.number, 
        "enquiry_saved",  # Updated event name 
        {"name": new_lead.name, "course": program['name']}
    )
    
    return new_lead

@api_router.get("/leads/recent")
async def get_recent_leads(
    since: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    """Polled by branch admin / counsellor dashboards to surface new leads as toasts.
    Returns leads created after `since` (ISO timestamp). Defaults to last 5 minutes.
    Scoped: Branch Admin -> their branch; Counsellor -> leads assigned to them."""
    if not since:
        since_dt = datetime.now(timezone.utc) - timedelta(minutes=5)
    else:
        try:
            since_dt = datetime.fromisoformat(since.replace("Z", "+00:00"))
        except Exception:
            since_dt = datetime.now(timezone.utc) - timedelta(minutes=5)

    query: dict = {
        "is_deleted": {"$ne": True},
        "created_at": {"$gt": since_dt.isoformat()},
    }

    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        query["branch_id"] = current_user.branch_id
    elif current_user.role == UserRole.COUNSELLOR:
        # Counsellor sees only leads assigned to them (or unassigned 'system')
        if current_user.branch_id:
            query["branch_id"] = current_user.branch_id
        query["counsellor_id"] = {"$in": [current_user.id, "system"]}
    elif current_user.role == UserRole.ADMIN:
        pass  # all branches
    else:
        return {"leads": [], "server_time": datetime.now(timezone.utc).isoformat()}

    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    return {"leads": leads, "server_time": datetime.now(timezone.utc).isoformat()}



@api_router.get("/leads", response_model=List[Lead])
async def get_leads(
    request: Request,
    status: Optional[str] = None,
    source: Optional[str] = None,
    program_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    session: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"is_deleted": {"$ne": True}}  # Exclude soft-deleted leads
    
    # Only filter by branch if user is not admin AND has a branch_id
    if current_user.role != UserRole.ADMIN and current_user.branch_id:
        query["branch_id"] = current_user.branch_id
    
    if status:
        query["status"] = status
    else:
        # Hide Lost leads from the main list unless explicitly filtered for them (case-insensitive)
        query["status"] = {"$not": {"$regex": "^lost$", "$options": "i"}}
    if source:
        query["lead_source"] = source
    if program_id:
        query["program_id"] = program_id
    if branch_id and current_user.role == UserRole.ADMIN:
        query["branch_id"] = branch_id
    
    # Session filtering (takes precedence over manual date filters)
    session_val = session or await get_session_from_request(request)
    if session_val and session_val != "all":
        session_filter = get_session_filter(session_val, "created_at")
        query.update(session_filter)
    elif start_date or end_date:
        # Manual date filter only if no session filter
        if start_date:
            query["created_at"] = {"$gte": start_date}
        if end_date:
            if "created_at" in query:
                query["created_at"]["$lte"] = end_date
            else:
                query["created_at"] = {"$lte": end_date}
    
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for lead in leads:
        if isinstance(lead.get('created_at'), str):
            lead['created_at'] = datetime.fromisoformat(lead['created_at'])
        if isinstance(lead.get('updated_at'), str):
            lead['updated_at'] = datetime.fromisoformat(lead['updated_at'])
    return [Lead(**lead) for lead in leads]

# Converted Leads for Enrollment - must be before /leads/{lead_id}
@api_router.get("/leads/converted")
async def get_converted_leads(current_user: User = Depends(get_current_user)):
    """Get converted leads for enrollment - FDA sees only their branch"""
    query = {"status": LeadStatus.CONVERTED.value, "is_deleted": {"$ne": True}}
    
    if current_user.role != UserRole.ADMIN:
        query["branch_id"] = current_user.branch_id
    
    # Check if already enrolled
    enrolled_lead_ids = await db.enrollments.find({}, {"lead_id": 1, "_id": 0}).to_list(1000)
    enrolled_ids = [e["lead_id"] for e in enrolled_lead_ids]
    
    if enrolled_ids:
        query["id"] = {"$nin": enrolled_ids}
    
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for lead in leads:
        if isinstance(lead.get('created_at'), str):
            lead['created_at'] = datetime.fromisoformat(lead['created_at'])
        if isinstance(lead.get('updated_at'), str):
            lead['updated_at'] = datetime.fromisoformat(lead['updated_at'])
    return leads

# Deleted Leads - must be before /leads/{lead_id}
@api_router.get("/leads/deleted", response_model=List[Lead])
async def get_deleted_leads(current_user: User = Depends(require_role([UserRole.BRANCH_ADMIN]))):
    """Get soft-deleted leads - Only Branch Admin can see deleted leads from their branch"""
    query = {"is_deleted": True, "branch_id": current_user.branch_id}
    
    leads = await db.leads.find(query, {"_id": 0}).sort("deleted_at", -1).to_list(1000)
    for lead in leads:
        if isinstance(lead.get('created_at'), str):
            lead['created_at'] = datetime.fromisoformat(lead['created_at'])
        if isinstance(lead.get('updated_at'), str):
            lead['updated_at'] = datetime.fromisoformat(lead['updated_at'])
        if isinstance(lead.get('deleted_at'), str):
            lead['deleted_at'] = datetime.fromisoformat(lead['deleted_at'])
    return [Lead(**lead) for lead in leads]

# Lost Leads - must be before /leads/{lead_id}
@api_router.get("/leads/lost")
async def get_lost_leads(current_user: User = Depends(get_current_user)):
    """Get leads marked as Lost. Visible to Admin, Branch Admin, and Counsellor.
    
    Returns raw lead dicts (not Lead model) so that legacy records missing some optional
    fields still show up. Case-insensitive status match catches 'lost', 'LOST', 'Lost'.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.COUNSELLOR]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {
        "status": {"$regex": "^lost$", "$options": "i"},
        "is_deleted": {"$ne": True},
    }
    
    if current_user.role != UserRole.ADMIN and current_user.branch_id:
        query["branch_id"] = current_user.branch_id
    
    # Counsellor sees own lost leads OR unassigned ones in their branch
    if current_user.role == UserRole.COUNSELLOR:
        query["$or"] = [
            {"created_by": current_user.id},
            {"counsellor_id": current_user.id},
            {"counsellor_id": {"$in": [None, ""]}},
        ]
    
    leads = await db.leads.find(query, {"_id": 0}).sort("updated_at", -1).to_list(2000)
    # Normalize: ensure 'status' is always 'Lost' in the response so the UI tag is consistent
    for lead in leads:
        lead["status"] = "Lost"
        # Ensure datetimes are serialised
        for f in ("created_at", "updated_at"):
            v = lead.get(f)
            if isinstance(v, datetime):
                lead[f] = v.isoformat()
    return leads

@api_router.put("/leads/{lead_id}/restore-from-lost")
async def restore_lead_from_lost(lead_id: str, new_status: Optional[str] = "New", current_user: User = Depends(get_current_user)):
    """Restore a Lost lead back to active. Visible to Admin, Branch Admin, Counsellor."""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.COUNSELLOR]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if lead.get('status') != LeadStatus.LOST.value:
        raise HTTPException(status_code=400, detail="Lead is not in Lost status")
    
    if current_user.role != UserRole.ADMIN and lead.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied for this branch")
    
    # Counsellor: only own leads
    if current_user.role == UserRole.COUNSELLOR:
        if lead.get('created_by') != current_user.id and lead.get('counsellor_id') != current_user.id:
            raise HTTPException(status_code=403, detail="You can only restore your own leads")
    
    valid_statuses = {s.value for s in LeadStatus} - {LeadStatus.LOST.value, LeadStatus.CONVERTED.value}
    target_status = new_status if new_status in valid_statuses else "New"
    
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {
            "status": target_status,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"message": f"Lead restored to '{target_status}'", "status": target_status}


# ============== STUDENT PORTAL ENDPOINTS ==============
@api_router.post("/student-auth/login")
async def student_login(payload: StudentLoginRequest):
    """Student login. Username = enrollment_number, password = enrollment_number.
    
    Returns a JWT token and a list of all the student's enrolled programs.
    """
    enrollment_number = (payload.enrollment_number or "").strip()
    password = (payload.password or "").strip()
    
    if not enrollment_number or not password:
        raise HTTPException(status_code=400, detail="Enrollment number and password are required")
    if enrollment_number != password:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    enrollment = await db.enrollments.find_one(
        {"$or": [{"enrollment_id": enrollment_number}, {"id": enrollment_number}]},
        {"_id": 0}
    )
    if not enrollment:
        raise HTTPException(status_code=401, detail="No enrollment found for this number")
    if enrollment.get("status") == EnrollmentStatus.CANCELLED.value:
        raise HTTPException(status_code=403, detail="This enrollment has been cancelled. Please contact your branch.")
    
    primary_id = enrollment.get("enrollment_id") or enrollment.get("id")
    token = create_access_token(data={"sub": primary_id, "role": "Student"})
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "student": {
            "enrollment_id": primary_id,
            "lead_id": enrollment.get("lead_id"),
            "name": enrollment.get("student_name"),
            "email": enrollment.get("email"),
            "phone": enrollment.get("phone"),
            "branch_id": enrollment.get("branch_id"),
        }
    }


@api_router.get("/student/me")
async def student_me(student: dict = Depends(get_current_student)):
    return student


@api_router.get("/student/my-curricula")
async def student_my_curricula(student: dict = Depends(get_current_student)):
    """Return curriculum + per-topic progress for every program the student is enrolled in."""
    # All enrollments for this student (matched by lead_id — same person can have multiple programs)
    enrollments = await db.enrollments.find(
        {"lead_id": student["lead_id"], "status": {"$ne": EnrollmentStatus.CANCELLED.value}},
        {"_id": 0}
    ).to_list(50)
    
    results = []
    for enr in enrollments:
        prog_id = enr.get("program_id")
        # Pick the most recent active curriculum for this program (prefer same branch first)
        curriculum = await db.curricula.find_one(
            {"program_id": prog_id, "branch_id": student["branch_id"]},
            {"_id": 0},
            sort=[("created_at", -1)],
        )
        if not curriculum:
            curriculum = await db.curricula.find_one(
                {"program_id": prog_id}, {"_id": 0}, sort=[("created_at", -1)]
            )
        
        # Fetch progress for this enrollment
        progress_rows = []
        if curriculum:
            progress_rows = await db.student_topic_progress.find(
                {"lead_id": student["lead_id"], "curriculum_id": curriculum.get("id")},
                {"_id": 0}
            ).to_list(500)
        progress_map = {p["topic_index"]: p for p in progress_rows}
        
        topics_with_progress = []
        if curriculum:
            for idx, t in enumerate(curriculum.get("topics", []) or []):
                p = progress_map.get(idx)
                topics_with_progress.append({
                    "topic_index": idx,
                    "topic_name": t,
                    "done": bool(p and p.get("done")),
                    "updated_at": p.get("updated_at").isoformat() if p and isinstance(p.get("updated_at"), datetime) else (p.get("updated_at") if p else None),
                })
        
        total_topics = len(topics_with_progress)
        done_count = sum(1 for t in topics_with_progress if t["done"])
        results.append({
            "enrollment": {
                "enrollment_id": enr.get("enrollment_id") or enr.get("id"),
                "program_id": prog_id,
                "program_name": enr.get("program_name"),
                "enrollment_date": str(enr.get("enrollment_date")) if enr.get("enrollment_date") else None,
                "status": enr.get("status"),
                "final_fee": enr.get("final_fee"),
                "total_paid": enr.get("total_paid"),
            },
            "curriculum": {
                "id": curriculum.get("id") if curriculum else None,
                "title": curriculum.get("title") if curriculum else None,
                "description": curriculum.get("description") if curriculum else None,
                "duration_weeks": curriculum.get("duration_weeks") if curriculum else None,
                "available": bool(curriculum),
            },
            "topics": topics_with_progress,
            "progress_percent": round((done_count / total_topics) * 100) if total_topics else 0,
            "topics_done": done_count,
            "topics_total": total_topics,
        })
    
    return {"programs": results, "student": student}


@api_router.post("/student/topic-progress")
async def student_mark_topic(update: StudentTopicProgressUpdate, student: dict = Depends(get_current_student)):
    """Toggle a curriculum topic as done/undone for the current student."""
    # Verify the enrollment belongs to this student
    enr = await db.enrollments.find_one(
        {"$or": [{"enrollment_id": update.enrollment_id}, {"id": update.enrollment_id}],
         "lead_id": student["lead_id"]},
        {"_id": 0}
    )
    if not enr:
        raise HTTPException(status_code=403, detail="Not your enrollment")
    
    primary_enrollment_id = enr.get("enrollment_id") or enr.get("id")
    now_iso = datetime.now(timezone.utc).isoformat()
    
    existing = await db.student_topic_progress.find_one({
        "lead_id": student["lead_id"],
        "curriculum_id": update.curriculum_id,
        "topic_index": update.topic_index,
    }, {"_id": 0})
    
    if existing:
        await db.student_topic_progress.update_one(
            {"id": existing["id"]},
            {"$set": {"done": update.done, "topic_name": update.topic_name, "updated_at": now_iso}}
        )
    else:
        rec = StudentTopicProgress(
            lead_id=student["lead_id"],
            enrollment_id=primary_enrollment_id,
            program_id=update.program_id,
            curriculum_id=update.curriculum_id,
            topic_index=update.topic_index,
            topic_name=update.topic_name,
            done=update.done,
        ).model_dump()
        rec["updated_at"] = now_iso
        await db.student_topic_progress.insert_one(rec)
    
    return {"success": True, "topic_index": update.topic_index, "done": update.done}


@api_router.post("/student/feedback")
async def student_submit_feedback(payload: StudentFeedbackCreate, student: dict = Depends(get_current_student)):
    """Submit feedback that goes to the Branch Admin of the student's branch."""
    if not payload.message or not payload.message.strip():
        raise HTTPException(status_code=400, detail="Feedback message is required")
    if len(payload.message) > 5000:
        raise HTTPException(status_code=400, detail="Feedback too long (max 5000 chars)")
    
    enr = await db.enrollments.find_one(
        {"$or": [{"enrollment_id": payload.enrollment_id}, {"id": payload.enrollment_id}],
         "lead_id": student["lead_id"]},
        {"_id": 0}
    )
    if not enr:
        raise HTTPException(status_code=403, detail="Not your enrollment")
    
    primary_enrollment_id = enr.get("enrollment_id") or enr.get("id")
    
    fb = StudentFeedback(
        lead_id=student["lead_id"],
        enrollment_id=primary_enrollment_id,
        student_name=student["name"] or enr.get("student_name") or "Unknown",
        student_email=student.get("email"),
        student_phone=student.get("phone"),
        branch_id=student["branch_id"] or enr.get("branch_id"),
        program_id=payload.program_id or enr.get("program_id"),
        program_name=enr.get("program_name"),
        message=payload.message.strip(),
        rating=payload.rating,
    ).model_dump()
    fb["created_at"] = fb["created_at"].isoformat()
    await db.student_feedback.insert_one(fb)
    return {"success": True, "id": fb["id"], "message": "Feedback sent to branch admin"}


@api_router.get("/branch-admin/student-feedback")
async def branch_admin_get_student_feedback(
    only_unread: bool = False,
    current_user: User = Depends(get_current_user)
):
    """List student feedback for the current Branch Admin (or all if Super Admin)."""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        query["branch_id"] = current_user.branch_id
    if only_unread:
        query["is_read"] = False
    
    feedback = await db.student_feedback.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    unread_count = sum(1 for f in feedback if not f.get("is_read"))
    return {"feedback": feedback, "count": len(feedback), "unread": unread_count}


@api_router.put("/branch-admin/student-feedback/{feedback_id}/read")
async def branch_admin_mark_feedback_read(
    feedback_id: str,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    fb = await db.student_feedback.find_one({"id": feedback_id}, {"_id": 0})
    if not fb:
        raise HTTPException(status_code=404, detail="Feedback not found")
    if current_user.role == UserRole.BRANCH_ADMIN and fb.get("branch_id") != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Not your branch")
    
    await db.student_feedback.update_one(
        {"id": feedback_id},
        {"$set": {
            "is_read": True,
            "read_by": current_user.id,
            "read_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"success": True}


@api_router.get("/leads/{lead_id}", response_model=Lead)
async def get_lead(lead_id: str, current_user: User = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if current_user.role != UserRole.ADMIN and lead.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if isinstance(lead.get('created_at'), str):
        lead['created_at'] = datetime.fromisoformat(lead['created_at'])
    if isinstance(lead.get('updated_at'), str):
        lead['updated_at'] = datetime.fromisoformat(lead['updated_at'])
    
    return Lead(**lead)

@api_router.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, lead_update: LeadUpdate, request: Request, current_user: User = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN] and lead.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Allow Branch Admin and Admin to edit converted leads
    # Counsellor cannot edit converted leads
    if lead.get('status') == 'Converted' and current_user.role == UserRole.COUNSELLOR:
        raise HTTPException(
            status_code=400, 
            detail="Counsellors cannot edit converted leads. Contact Branch Admin."
        )
    
    old_status = lead.get('status')
    
    update_data = {k: v for k, v in lead_update.model_dump(exclude_unset=True).items() if v is not None}
    
    # If program_id is being updated, also update program_name
    if 'program_id' in update_data:
        program = await db.programs.find_one({"id": update_data['program_id']}, {"_id": 0})
        if program:
            update_data['program_name'] = program.get('name', '')

    # If counsellor_id (Lead Owner) is being changed, validate and cache the name
    if 'counsellor_id' in update_data:
        if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
            raise HTTPException(status_code=403, detail="Only Admin / Branch Admin / Front Desk can reassign Lead Owner")
        counsellor_doc = await db.users.find_one(
            {
                "id": update_data['counsellor_id'],
                "role": UserRole.COUNSELLOR.value,
                "branch_id": lead.get('branch_id'),
                "is_active": True
            },
            {"_id": 0, "id": 1, "name": 1, "email": 1}
        )
        if not counsellor_doc:
            raise HTTPException(status_code=400, detail="Lead Owner must be an active Counsellor in this branch")
        update_data['counsellor_name'] = counsellor_doc.get("name") or counsellor_doc.get("email", "Counsellor")

    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()

    # Phase C — Recompute lead priority if status / fee / parent changed
    if any(k in update_data for k in ('status', 'fee_quoted', 'parent_status')):
        merged = {**lead, **update_data}
        update_data['priority'] = compute_lead_priority(merged)

    await db.leads.update_one({"id": lead_id}, {"$set": update_data})
    
    updated_lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if isinstance(updated_lead.get('created_at'), str):
        updated_lead['created_at'] = datetime.fromisoformat(updated_lead['created_at'])
    if isinstance(updated_lead.get('updated_at'), str):
        updated_lead['updated_at'] = datetime.fromisoformat(updated_lead['updated_at'])
    
    updated_lead_obj = Lead(**updated_lead)
    
    # Send WhatsApp notifications on status change
    if lead_update.status and lead_update.status != old_status:
        if lead_update.status == LeadStatus.DEMO_BOOKED:
            # Send demo booked notification with full details
            await send_whatsapp_notification(
                updated_lead_obj.number, 
                "demo_booked", 
                {
                    "name": updated_lead_obj.name,
                    "demo_date": updated_lead_obj.demo_date or "",
                    "demo_time": updated_lead_obj.demo_time or "",
                    "trainer": updated_lead_obj.trainer_name or ""
                }
            )
        
        # Notify lead's assigned counsellor when someone else changes the status
        try:
            assigned_counsellor_id = lead.get('counsellor_id')
            new_status_str = lead_update.status.value if hasattr(lead_update.status, 'value') else str(lead_update.status)
            if assigned_counsellor_id and assigned_counsellor_id != current_user.id:
                status_change_note = Notification(
                    user_id=assigned_counsellor_id,
                    branch_id=lead.get('branch_id'),
                    type="lead_status_changed_by_other",
                    title="Lead status changed",
                    message=f"{updated_lead_obj.name}: {old_status} → {new_status_str} (by {current_user.name})",
                    data={
                        "lead_id": lead_id,
                        "lead_name": updated_lead_obj.name,
                        "old_status": old_status,
                        "new_status": new_status_str,
                        "changed_by_id": current_user.id,
                        "changed_by_name": current_user.name,
                    },
                    play_audio=False,
                )
                await db.notifications.insert_one(status_change_note.model_dump())
        except Exception as e:
            logger.warning(f"Failed to record status-change notification: {e}")

        # Create notification for FDE when lead is converted
        if lead_update.status == LeadStatus.CONVERTED:
            branch_id = lead.get('branch_id')
            # Find all FDEs in this branch
            fdes = await db.users.find(
                {"branch_id": branch_id, "role": UserRole.FRONT_DESK.value, "is_active": True},
                {"_id": 0, "id": 1}
            ).to_list(100)
            
            for fde in fdes:
                notification = Notification(
                    user_id=fde['id'],
                    branch_id=branch_id,
                    type="lead_converted",
                    title="New Lead Converted!",
                    message=f"{updated_lead_obj.name} has been converted. Please proceed with enrollment.",
                    data={
                        "lead_id": lead_id,
                        "lead_name": updated_lead_obj.name,
                        "lead_number": updated_lead_obj.number,
                        "program": updated_lead_obj.program_name
                    },
                    play_audio=True
                )
                await db.notifications.insert_one(notification.model_dump())
                logger.info(f"Created lead converted notification for FDE {fde['id']}")
    
    # Audit Log
    await create_audit_log(
        user=current_user,
        action="update",
        entity_type="lead",
        entity_id=lead_id,
        entity_name=f"{updated_lead_obj.name} ({lead.get('lead_id', '')})",
        changes={"updated_fields": list(update_data.keys()), "old_status": old_status, "new_status": lead_update.status},
        request=request
    )
    
    return updated_lead_obj

class LeadDeleteRequest(BaseModel):
    reason: Optional[str] = None

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, reason: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Soft delete a lead - all users can delete leads from their branch"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if lead.get('is_deleted'):
        raise HTTPException(status_code=400, detail="Lead is already deleted")
    
    # Users can only delete leads from their branch (Admin can delete any)
    if current_user.role != UserRole.ADMIN and lead.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only delete leads from your branch")
    
    # Counsellor can only delete leads they created or are assigned to
    if current_user.role == UserRole.COUNSELLOR:
        if lead.get('created_by') != current_user.id and lead.get('counsellor_id') != current_user.id:
            raise HTTPException(status_code=403, detail="You can only delete leads assigned to you")
    
    # Soft delete - mark as deleted instead of removing
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id,
            "deleted_by_name": current_user.name,
            "deletion_reason": reason
        }}
    )
    return {"message": "Lead deleted successfully"}

@api_router.put("/leads/{lead_id}/restore")
async def restore_lead(lead_id: str, current_user: User = Depends(require_role([UserRole.BRANCH_ADMIN, UserRole.ADMIN]))):
    """Restore a soft-deleted lead"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if not lead.get('is_deleted'):
        raise HTTPException(status_code=400, detail="Lead is not deleted")
    
    # Branch Admin can only restore leads from their branch
    if current_user.role == UserRole.BRANCH_ADMIN and lead.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only restore leads from your branch")
    
    # Restore the lead
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {
            "is_deleted": False,
            "deleted_at": None,
            "deleted_by": None,
            "deleted_by_name": None,
            "deletion_reason": None,
            "restored_at": datetime.now(timezone.utc).isoformat(),
            "restored_by": current_user.id,
            "restored_by_name": current_user.name
        }}
    )
    return {"message": "Lead restored successfully"}


# Follow-up Management
@api_router.post("/followups", response_model=FollowUp)
async def create_followup(followup: FollowUpCreate, current_user: User = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": followup.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    new_followup = FollowUp(
        **followup.model_dump(),
        branch_id=lead.get('branch_id'),  # Store branch_id from lead
        created_by=current_user.id,
        created_by_name=current_user.name,
        lead_name=lead['name'],
        lead_number=lead['number']
    )
    
    followup_dict = new_followup.model_dump()
    followup_dict['created_at'] = followup_dict['created_at'].isoformat()
    followup_dict['followup_date'] = followup_dict['followup_date'].isoformat()
    
    await db.followups.insert_one(followup_dict)
    return new_followup

@api_router.get("/followups/pending")
async def get_pending_followups(current_user: User = Depends(get_current_user)):
    """Get all pending followups (today and future) - optimized"""
    query = {
        "status": FollowUpStatus.PENDING
    }
    
    # For counsellors, only show their own followups
    if current_user.role == UserRole.COUNSELLOR:
        query["created_by"] = current_user.id
    # For branch admin, show all followups in their branch
    elif current_user.role == UserRole.BRANCH_ADMIN:
        query["branch_id"] = current_user.branch_id
    # Admin sees all
    
    followups = await db.followups.find(query, {"_id": 0}).sort("followup_date", 1).to_list(1000)
    
    if not followups:
        return []
    
    # Bulk fetch lead info
    lead_ids = list(set(fu.get('lead_id') for fu in followups if fu.get('lead_id')))
    leads_list = await db.leads.find({"id": {"$in": lead_ids}}, {"_id": 0, "id": 1, "name": 1, "number": 1}).to_list(len(lead_ids))
    leads_map = {l['id']: l for l in leads_list}
    
    # Bulk fetch creator names
    creator_ids = list(set(fu.get('created_by') for fu in followups if fu.get('created_by')))
    creators_list = await db.users.find({"id": {"$in": creator_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(len(creator_ids))
    creators_map = {c['id']: c for c in creators_list}
    
    # Enrich with lead info
    for fu in followups:
        if isinstance(fu.get('created_at'), str):
            try:
                fu['created_at'] = datetime.fromisoformat(fu['created_at'].replace('Z', '+00:00'))
            except:
                pass
        if isinstance(fu.get('followup_date'), str):
            try:
                fu['followup_date'] = datetime.fromisoformat(fu['followup_date'].replace('Z', '+00:00'))
            except:
                pass
        
        # Get lead info from map
        if fu.get('lead_id') and fu['lead_id'] in leads_map:
            lead = leads_map[fu['lead_id']]
            fu['lead_name'] = lead.get('name', 'Unknown')
            fu['lead_number'] = lead.get('number', '')
        
        # Get creator name from map
        if fu.get('created_by') and fu['created_by'] in creators_map:
            creator = creators_map[fu['created_by']]
            fu['created_by_name'] = creator.get('name', creator.get('email', 'Unknown'))
    
    return followups

@api_router.get("/followups/pending/count")
async def get_pending_followups_count(current_user: User = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date()
    today_start = datetime.combine(today, datetime.min.time()).isoformat()
    today_end = datetime.combine(today, datetime.max.time()).isoformat()
    
    query = {
        "status": FollowUpStatus.PENDING,
        "followup_date": {"$gte": today_start, "$lte": today_end}
    }
    
    if current_user.role != UserRole.ADMIN:
        query["created_by"] = current_user.id
    
    count = await db.followups.count_documents(query)
    return {"count": count}

@api_router.get("/leads/{lead_id}/followups")
async def get_lead_followups(lead_id: str, current_user: User = Depends(get_current_user)):
    followups = await db.followups.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for fu in followups:
        if isinstance(fu.get('created_at'), str):
            fu['created_at'] = datetime.fromisoformat(fu['created_at'])
        if isinstance(fu.get('followup_date'), str):
            fu['followup_date'] = datetime.fromisoformat(fu['followup_date'])
    return followups

@api_router.put("/followups/{followup_id}/status")
async def update_followup_status(followup_id: str, status: FollowUpStatus, current_user: User = Depends(get_current_user)):
    result = await db.followups.update_one(
        {"id": followup_id},
        {"$set": {"status": status}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    return {"message": "Follow-up status updated"}

@api_router.post("/followups/{followup_id}/log")
async def log_followup_outcome(followup_id: str, log_data: FollowUpLogCreate, current_user: User = Depends(get_current_user)):
    """Log a follow-up call outcome with trail history"""
    # Get the follow-up
    followup = await db.followups.find_one({"id": followup_id}, {"_id": 0})
    if not followup:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    
    # Create the log entry
    log_entry = FollowUpLog(
        followup_id=followup_id,
        lead_id=followup.get('lead_id'),
        branch_id=followup.get('branch_id'),
        outcome=log_data.outcome,
        notes=log_data.notes,
        next_action=log_data.next_action,
        next_followup_date=log_data.next_followup_date,
        logged_by=current_user.id,
        logged_by_name=current_user.name
    )
    
    log_dict = log_entry.model_dump()
    log_dict['logged_at'] = log_dict['logged_at'].isoformat()
    if log_dict.get('next_followup_date'):
        log_dict['next_followup_date'] = log_dict['next_followup_date'].isoformat()
    
    await db.followup_logs.insert_one(log_dict)
    
    # Update the follow-up with latest outcome
    update_data = {
        "call_attempts": (followup.get('call_attempts', 0) or 0) + 1,
        "last_outcome": log_data.outcome.value,
        "last_call_at": datetime.now(timezone.utc).isoformat()
    }
    
    # If connected and completed, mark as completed
    if log_data.outcome == FollowUpOutcome.CONNECTED and not log_data.next_followup_date:
        update_data["status"] = FollowUpStatus.COMPLETED
    
    await db.followups.update_one({"id": followup_id}, {"$set": update_data})
    
    # If next follow-up date specified, create a new follow-up
    if log_data.next_followup_date:
        new_followup = FollowUp(
            lead_id=followup.get('lead_id'),
            branch_id=followup.get('branch_id'),
            note=f"Follow-up from previous call: {log_data.notes}. Next action: {log_data.next_action or 'Call again'}",
            followup_date=log_data.next_followup_date,
            created_by=current_user.id,
            created_by_name=current_user.name,
            lead_name=followup.get('lead_name'),
            lead_number=followup.get('lead_number')
        )
        new_fu_dict = new_followup.model_dump()
        new_fu_dict['created_at'] = new_fu_dict['created_at'].isoformat()
        new_fu_dict['followup_date'] = new_fu_dict['followup_date'].isoformat()
        await db.followups.insert_one(new_fu_dict)
        
        # Mark current follow-up as completed since new one is created
        await db.followups.update_one({"id": followup_id}, {"$set": {"status": FollowUpStatus.COMPLETED}})
    
    return {"message": "Follow-up logged successfully", "log_id": log_entry.id}

@api_router.get("/followups/{followup_id}/logs")
async def get_followup_logs(followup_id: str, current_user: User = Depends(get_current_user)):
    """Get all log entries (trail) for a follow-up"""
    logs = await db.followup_logs.find({"followup_id": followup_id}, {"_id": 0}).sort("logged_at", -1).to_list(100)
    return logs

@api_router.get("/leads/{lead_id}/followup-trail")
async def get_lead_followup_trail(lead_id: str, current_user: User = Depends(get_current_user)):
    """Get complete follow-up trail for a lead including all calls and outcomes"""
    # Get all follow-ups for this lead
    followups = await db.followups.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Get all logs for this lead
    logs = await db.followup_logs.find({"lead_id": lead_id}, {"_id": 0}).sort("logged_at", -1).to_list(500)
    
    # Build timeline
    timeline = []
    for fu in followups:
        timeline.append({
            "type": "followup_created",
            "date": fu.get('created_at'),
            "data": fu
        })
    
    for log in logs:
        timeline.append({
            "type": "call_logged",
            "date": log.get('logged_at'),
            "data": log
        })
    
    # Sort by date
    timeline.sort(key=lambda x: x['date'] if x['date'] else '', reverse=True)
    
    return {
        "lead_id": lead_id,
        "total_followups": len(followups),
        "total_calls": len(logs),
        "timeline": timeline
    }

# ========== NOTIFICATIONS ==========

@api_router.get("/notifications")
async def get_notifications(current_user: User = Depends(get_current_user)):
    """Get all notifications for the current user"""
    notifications = await db.notifications.find(
        {"user_id": current_user.id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return notifications

@api_router.get("/notifications/unread")
async def get_unread_notifications(current_user: User = Depends(get_current_user)):
    """Get unread notifications for the current user"""
    notifications = await db.notifications.find(
        {"user_id": current_user.id, "is_read": False},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return notifications

@api_router.put("/notifications/{notification_id}/read")
async def mark_user_notification_read(notification_id: str, current_user: User = Depends(get_current_user)):
    """Mark a notification as read"""
    # Try both user_id and recipient_ids schemas
    result = await db.notifications.update_one(
        {"id": notification_id, "$or": [{"user_id": current_user.id}, {"recipient_ids": current_user.id}]},
        {"$set": {"is_read": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/mark-all-read")
async def mark_all_user_notifications_read(current_user: User = Depends(get_current_user)):
    """Mark all notifications as read for the current user"""
    # Try both schemas
    await db.notifications.update_many(
        {"$or": [{"user_id": current_user.id}, {"recipient_ids": current_user.id}], "is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"message": "All notifications marked as read"}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, current_user: User = Depends(get_current_user)):
    """Delete/dismiss a notification"""
    result = await db.notifications.delete_one({
        "id": notification_id,
        "$or": [{"user_id": current_user.id}, {"recipient_ids": current_user.id}]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification dismissed"}

@api_router.get("/notifications/followup-reminders")
async def get_followup_reminders(current_user: User = Depends(get_current_user)):
    """Get follow-ups that are due in the next 10 minutes for audio reminders"""
    now = datetime.now(timezone.utc)
    in_10_min = now + timedelta(minutes=10)
    
    query = {
        "created_by": current_user.id,
        "status": "Pending"
    }
    
    followups = await db.followups.find(query, {"_id": 0}).to_list(1000)
    
    upcoming = []
    for fu in followups:
        fu_date = fu.get('followup_date')
        if isinstance(fu_date, str):
            fu_date = datetime.fromisoformat(fu_date.replace('Z', '+00:00'))
        
        # Check if followup is within the next 10 minutes
        if fu_date and now <= fu_date <= in_10_min:
            upcoming.append(fu)
    
    return upcoming


# AI-Powered Lead Analytics for Counsellors and Branch Admins
@api_router.get("/analytics/ai-leads-insights")
async def get_ai_leads_insights(current_user: User = Depends(get_current_user)):
    """Get AI-powered insights for leads management - For Counsellors and Branch Admins"""
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Counsellors and Branch Admins can access this")
    
    branch_filter = {"branch_id": current_user.branch_id, "is_deleted": {"$ne": True}}
    
    # Get all leads for analysis
    leads = await db.leads.find(branch_filter, {"_id": 0}).to_list(10000)
    
    # Basic counts
    total_leads = len(leads)
    status_counts = {}
    source_counts = {}
    program_counts = {}
    
    # Date-based analysis
    today = datetime.now(timezone.utc).date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    leads_this_week = 0
    leads_this_month = 0
    converted_this_month = 0
    lost_this_month = 0
    
    # Followup analysis
    pending_followups = 0
    overdue_followups = 0
    
    for lead in leads:
        status = lead.get('status', 'New')
        source = lead.get('lead_source', 'Unknown')
        program = lead.get('program_name', 'Unknown')
        
        status_counts[status] = status_counts.get(status, 0) + 1
        source_counts[source] = source_counts.get(source, 0) + 1
        program_counts[program] = program_counts.get(program, 0) + 1
        
        # Date parsing
        created_str = lead.get('created_at', '')
        if isinstance(created_str, str) and created_str:
            try:
                created_date = datetime.fromisoformat(created_str.replace('Z', '+00:00')).date()
                if created_date >= week_ago:
                    leads_this_week += 1
                if created_date >= month_ago:
                    leads_this_month += 1
                    if status == 'Converted':
                        converted_this_month += 1
                    elif status in ['Not Interested', 'Lost']:
                        lost_this_month += 1
            except ValueError:
                pass
    
    # Get followups
    followups = await db.followups.find(
        {"branch_id": current_user.branch_id, "status": "Pending"},
        {"_id": 0}
    ).to_list(1000)
    
    pending_followups = len(followups)
    today_str = today.isoformat()
    overdue_followups = len([f for f in followups if f.get('followup_date', '') < today_str])
    
    # Calculate conversion rate
    conversion_rate = round((status_counts.get('Converted', 0) / total_leads * 100) if total_leads > 0 else 0, 1)
    monthly_conversion_rate = round((converted_this_month / leads_this_month * 100) if leads_this_month > 0 else 0, 1)
    
    # Best performing source
    best_source = max(source_counts.items(), key=lambda x: x[1])[0] if source_counts else "N/A"
    
    # Most popular program
    popular_program = max(program_counts.items(), key=lambda x: x[1])[0] if program_counts else "N/A"
    
    # Generate AI insights
    insights = []
    recommendations = []
    
    # Conversion insights
    if conversion_rate < 20:
        insights.append({
            "type": "warning",
            "title": "Low Conversion Rate",
            "message": f"Your conversion rate is {conversion_rate}%. Industry average is 20-30%.",
            "priority": "high"
        })
        recommendations.append("Focus on follow-up quality and timing to improve conversions")
    elif conversion_rate > 30:
        insights.append({
            "type": "success",
            "title": "Excellent Conversion Rate",
            "message": f"Your conversion rate is {conversion_rate}%! Keep up the great work.",
            "priority": "low"
        })
    
    # Followup insights
    if overdue_followups > 0:
        insights.append({
            "type": "alert",
            "title": "Overdue Follow-ups",
            "message": f"You have {overdue_followups} overdue follow-ups. These leads may go cold!",
            "priority": "high"
        })
        recommendations.append(f"Clear {overdue_followups} overdue follow-ups today to prevent lead loss")
    
    if pending_followups > 10:
        insights.append({
            "type": "warning",
            "title": "High Pending Follow-ups",
            "message": f"{pending_followups} follow-ups pending. Consider prioritizing hot leads.",
            "priority": "medium"
        })
    
    # Status analysis
    new_leads = status_counts.get('New', 0)
    if new_leads > total_leads * 0.4:
        insights.append({
            "type": "warning",
            "title": "Too Many New Leads Untouched",
            "message": f"{new_leads} leads ({round(new_leads/total_leads*100)}%) are still in 'New' status.",
            "priority": "high"
        })
        recommendations.append("Contact new leads within 24 hours for best results")
    
    # Demo booked analysis
    demo_booked = status_counts.get('Demo Booked', 0)
    if demo_booked > 0:
        insights.append({
            "type": "info",
            "title": "Demo Sessions Scheduled",
            "message": f"{demo_booked} leads have demos scheduled. Prepare well!",
            "priority": "medium"
        })
    
    # Source insights
    for source, count in source_counts.items():
        if count >= total_leads * 0.3:
            insights.append({
                "type": "info",
                "title": f"Strong Lead Source: {source}",
                "message": f"{source} contributes {round(count/total_leads*100)}% of leads. Consider investing more here.",
                "priority": "low"
            })
    
    # Weekly trend
    if leads_this_week == 0:
        insights.append({
            "type": "warning",
            "title": "No New Leads This Week",
            "message": "No new leads in the past 7 days. Check marketing efforts.",
            "priority": "high"
        })
        recommendations.append("Review and boost marketing campaigns")
    
    # Monthly trend recommendations
    if monthly_conversion_rate < conversion_rate:
        recommendations.append("This month's conversion is below average. Analyze what changed.")
    
    if lost_this_month > converted_this_month:
        recommendations.append(f"Lost {lost_this_month} leads this month vs {converted_this_month} conversions. Review lost lead feedback.")
    
    # Calculate health score (rule-based)
    health_score = min(100, max(0, 
        50 + 
        (conversion_rate - 20) * 2 +
        (10 if overdue_followups == 0 else -10) +
        (10 if new_leads < total_leads * 0.3 else -5)
    ))
    
    # Try to get AI-powered insights using real LLM
    ai_insights = None
    ai_recommendations = None
    
    if LLM_AVAILABLE and os.environ.get('EMERGENT_LLM_KEY'):
        try:
            # Prepare data summary for LLM
            lead_data_summary = f"""
Lead Management Data Summary:
- Total Leads: {total_leads}
- Leads This Week: {leads_this_week}
- Leads This Month: {leads_this_month}
- Overall Conversion Rate: {conversion_rate}%
- Monthly Conversion Rate: {monthly_conversion_rate}%
- Pending Follow-ups: {pending_followups}
- Overdue Follow-ups: {overdue_followups}
- Best Performing Source: {best_source}
- Most Popular Program: {popular_program}

Status Breakdown: {json.dumps(status_counts)}
Lead Sources: {json.dumps(source_counts)}
Programs: {json.dumps(program_counts)}

New Leads: {new_leads} ({round(new_leads/total_leads*100) if total_leads > 0 else 0}% of total)
Converted This Month: {converted_this_month}
Lost This Month: {lost_this_month}
"""
            
            # Initialize LLM Chat
            chat = LlmChat(
                api_key=os.environ.get('EMERGENT_LLM_KEY'),
                session_id=f"ai-insights-{current_user.id}-{datetime.now().strftime('%Y%m%d')}",
                system_message="""You are an expert CRM analyst and sales coach. Analyze the lead management data and provide actionable insights.

Your response MUST be valid JSON with this exact structure:
{
  "insights": [
    {"type": "warning|success|alert|info", "title": "Short Title", "message": "Detailed insight message", "priority": "high|medium|low"}
  ],
  "recommendations": ["Specific actionable recommendation 1", "Recommendation 2", "Recommendation 3"],
  "summary": "One sentence summary of the overall lead health"
}

Provide 3-5 insights and 3-4 recommendations. Focus on:
1. Conversion rate analysis and benchmarks
2. Follow-up effectiveness
3. Lead source ROI
4. Actionable improvements
5. Trends and patterns"""
            ).with_model("openai", "gpt-4o")
            
            # Send message to LLM
            user_message = UserMessage(text=f"Analyze this lead data and provide insights:\n\n{lead_data_summary}")
            response = await chat.send_message(user_message)
            
            # Parse LLM response
            try:
                # Try to extract JSON from response
                response_text = response.strip()
                if response_text.startswith("```json"):
                    response_text = response_text[7:]
                if response_text.startswith("```"):
                    response_text = response_text[3:]
                if response_text.endswith("```"):
                    response_text = response_text[:-3]
                
                ai_response = json.loads(response_text.strip())
                ai_insights = ai_response.get('insights', [])
                ai_recommendations = ai_response.get('recommendations', [])
                logging.info(f"AI Insights generated successfully for user {current_user.email}")
            except json.JSONDecodeError as e:
                logging.warning(f"Failed to parse LLM response as JSON: {e}")
                # Fall back to rule-based
                ai_insights = None
                ai_recommendations = None
        except Exception as e:
            logging.error(f"LLM API error: {e}")
            # Fall back to rule-based insights
            ai_insights = None
            ai_recommendations = None
    
    # Use AI insights if available, otherwise use rule-based
    final_insights = ai_insights if ai_insights else insights
    final_recommendations = ai_recommendations if ai_recommendations else recommendations
    
    return {
        "summary": {
            "total_leads": total_leads,
            "leads_this_week": leads_this_week,
            "leads_this_month": leads_this_month,
            "conversion_rate": conversion_rate,
            "monthly_conversion_rate": monthly_conversion_rate,
            "pending_followups": pending_followups,
            "overdue_followups": overdue_followups,
            "best_source": best_source,
            "popular_program": popular_program
        },
        "status_breakdown": status_counts,
        "source_breakdown": source_counts,
        "program_breakdown": program_counts,
        "insights": final_insights,
        "recommendations": final_recommendations,
        "health_score": health_score,
        "ai_powered": ai_insights is not None
    }


@api_router.get("/analytics/ai-branch-insights")
async def get_ai_branch_insights(current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Get comprehensive AI-powered branch analytics - For Branch Admins only"""
    branch_id = current_user.branch_id
    branch_filter = {"branch_id": branch_id} if branch_id else {}
    
    today = datetime.now(timezone.utc)
    today_str = today.strftime('%Y-%m-%d')
    month_start = today.replace(day=1).strftime('%Y-%m-%d')
    last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d')
    last_month_end = (today.replace(day=1) - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # === TRAINER WORKLOAD ANALYSIS ===
    trainers = await db.users.find(
        {**branch_filter, "role": UserRole.TRAINER.value, "is_active": True},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(100)
    
    trainer_workload = []
    for trainer in trainers:
        # Get batches assigned to this trainer
        batches = await db.batches.find(
            {**branch_filter, "trainer_id": trainer['id']},
            {"_id": 0, "id": 1, "name": 1, "schedule": 1}
        ).to_list(100)
        
        # Get students in these batches
        batch_ids = [b['id'] for b in batches]
        students_count = await db.enrollments.count_documents({"batch_id": {"$in": batch_ids}, "status": "Active"})
        
        trainer_workload.append({
            "trainer_name": trainer.get('name', trainer['email']),
            "trainer_id": trainer['id'],
            "total_batches": len(batches),
            "total_students": students_count,
            "batches": [{"name": b.get('name', 'Unnamed'), "schedule": b.get('schedule', 'Not set')} for b in batches[:5]]
        })
    
    # === INCOME ANALYSIS ===
    # This month's income
    this_month_payments = await db.payments.find(
        {**branch_filter, "payment_date": {"$gte": month_start, "$lte": today_str}},
        {"_id": 0, "amount": 1, "payment_mode": 1}
    ).to_list(10000)
    
    this_month_income = sum(p.get('amount', 0) for p in this_month_payments)
    income_by_mode = {}
    for p in this_month_payments:
        mode = p.get('payment_mode', 'Other')
        income_by_mode[mode] = income_by_mode.get(mode, 0) + p.get('amount', 0)
    
    # Last month's income for comparison
    last_month_payments = await db.payments.find(
        {**branch_filter, "payment_date": {"$gte": last_month_start, "$lte": last_month_end}},
        {"_id": 0, "amount": 1}
    ).to_list(10000)
    last_month_income = sum(p.get('amount', 0) for p in last_month_payments)
    
    income_growth = ((this_month_income - last_month_income) / last_month_income * 100) if last_month_income > 0 else 0
    
    # === STUDENT PERFORMANCE ===
    active_students = await db.enrollments.count_documents({**branch_filter, "status": "Active"})
    completed_students = await db.enrollments.count_documents({**branch_filter, "status": "Completed"})
    
    # Fee collection efficiency
    total_fee = await db.enrollments.aggregate([
        {"$match": {**branch_filter, "status": "Active"}},
        {"$group": {"_id": None, "total": {"$sum": "$final_fee"}, "paid": {"$sum": "$total_paid"}}}
    ]).to_list(1)
    
    fee_efficiency = 0
    total_pending = 0
    if total_fee and total_fee[0].get('total', 0) > 0:
        fee_efficiency = round(total_fee[0].get('paid', 0) / total_fee[0].get('total', 1) * 100, 1)
        total_pending = total_fee[0].get('total', 0) - total_fee[0].get('paid', 0)
    
    # === LEAD FUNNEL ===
    total_leads = await db.leads.count_documents({**branch_filter, "is_deleted": {"$ne": True}})
    converted_leads = await db.leads.count_documents({**branch_filter, "status": "Converted", "is_deleted": {"$ne": True}})
    conversion_rate = round(converted_leads / total_leads * 100, 1) if total_leads > 0 else 0
    
    # === BUILD SUMMARY FOR AI ===
    analytics_summary = f"""
Branch Analytics Summary:

TRAINER WORKLOAD:
{json.dumps(trainer_workload, indent=2)}

INCOME ANALYSIS:
- This Month Income: ₹{this_month_income:,.0f}
- Last Month Income: ₹{last_month_income:,.0f}
- Growth: {income_growth:.1f}%
- Income by Payment Mode: {json.dumps(income_by_mode)}

STUDENT METRICS:
- Active Students: {active_students}
- Completed Students: {completed_students}
- Fee Collection Efficiency: {fee_efficiency}%
- Total Pending Fees: ₹{total_pending:,.0f}

LEAD CONVERSION:
- Total Leads: {total_leads}
- Converted: {converted_leads}
- Conversion Rate: {conversion_rate}%
"""

    # === AI ANALYSIS ===
    ai_analysis = None
    if LLM_AVAILABLE and os.environ.get('EMERGENT_LLM_KEY'):
        try:
            chat = LlmChat(
                api_key=os.environ.get('EMERGENT_LLM_KEY'),
                session_id=f"branch-analytics-{branch_id}-{today.strftime('%Y%m%d')}",
                system_message="""You are a business analyst for an educational institute. Analyze the branch data and provide strategic insights.

Your response MUST be valid JSON with this structure:
{
  "trainer_analysis": {
    "overloaded": ["trainer names with too many students/batches"],
    "underutilized": ["trainer names with capacity"],
    "recommendation": "specific action to balance workload"
  },
  "income_insights": {
    "trend": "growing|stable|declining",
    "top_payment_mode": "Cash/Online/etc",
    "forecast": "projected trend for next month",
    "recommendation": "how to improve collections"
  },
  "student_insights": {
    "retention_risk": "low|medium|high",
    "fee_collection_status": "healthy|needs attention|critical",
    "recommendation": "specific action for students"
  },
  "overall_health": {
    "score": 1-10,
    "status": "excellent|good|needs improvement|critical",
    "top_priority": "most important action to take",
    "summary": "2-3 sentence branch health summary"
  }
}

Be specific and actionable. Use actual numbers from the data."""
            ).with_model("openai", "gpt-4o")
            
            user_message = UserMessage(text=f"Analyze this branch data and provide insights:\n\n{analytics_summary}")
            ai_response = await chat.send_message(user_message)
            
            # Parse AI response
            try:
                ai_text = ai_response.strip()
                if ai_text.startswith("```"):
                    ai_text = ai_text.split("```")[1]
                    if ai_text.startswith("json"):
                        ai_text = ai_text[4:]
                ai_analysis = json.loads(ai_text)
            except json.JSONDecodeError:
                ai_analysis = {"raw_response": ai_response}
                
        except Exception as e:
            logger.error(f"AI analysis failed: {e}")
            ai_analysis = None
    
    return {
        "generated_at": today.isoformat(),
        "trainer_workload": trainer_workload,
        "income": {
            "this_month": this_month_income,
            "last_month": last_month_income,
            "growth_percent": round(income_growth, 1),
            "by_payment_mode": income_by_mode
        },
        "students": {
            "active": active_students,
            "completed": completed_students,
            "fee_efficiency_percent": fee_efficiency,
            "pending_fees": total_pending
        },
        "leads": {
            "total": total_leads,
            "converted": converted_leads,
            "conversion_rate": conversion_rate
        },
        "ai_analysis": ai_analysis,
        "ai_powered": ai_analysis is not None
    }

@api_router.get("/analytics/user-efficiency")
async def get_user_efficiency_analysis(current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Get AI-powered user efficiency analysis - For Branch Admins only"""
    branch_id = current_user.branch_id
    branch_filter = {"branch_id": branch_id} if branch_id else {}
    
    today = datetime.now(timezone.utc)
    today_str = today.strftime('%Y-%m-%d')
    week_ago = (today - timedelta(days=7)).strftime('%Y-%m-%d')
    month_start = today.replace(day=1).strftime('%Y-%m-%d')
    
    # === COUNSELLOR EFFICIENCY ===
    counsellors = await db.users.find(
        {**branch_filter, "role": UserRole.COUNSELLOR.value, "is_active": True},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(100)
    
    counsellor_efficiency = []
    for counsellor in counsellors:
        # Count leads assigned this month
        leads_assigned = await db.leads.count_documents({
            "assigned_to": counsellor['id'],
            "is_deleted": {"$ne": True}
        })
        
        # Count leads converted this month
        leads_converted = await db.leads.count_documents({
            "assigned_to": counsellor['id'],
            "status": "Converted",
            "is_deleted": {"$ne": True}
        })
        
        # Count follow-ups scheduled vs completed
        followups_scheduled = await db.followups.count_documents({
            "created_by": counsellor['id'],
            "followup_date": {"$gte": month_start, "$lte": today_str}
        })
        
        followups_completed = await db.followups.count_documents({
            "created_by": counsellor['id'],
            "status": "Completed"
        })
        
        # Average response time (time from lead creation to first followup)
        # This is a simplified calculation
        conversion_rate = round(leads_converted / leads_assigned * 100, 1) if leads_assigned > 0 else 0
        followup_rate = round(followups_completed / followups_scheduled * 100, 1) if followups_scheduled > 0 else 0
        
        # Calculate efficiency score (weighted average)
        efficiency_score = round((conversion_rate * 0.5 + followup_rate * 0.5), 1)
        
        counsellor_efficiency.append({
            "user_id": counsellor['id'],
            "name": counsellor.get('name', counsellor['email']),
            "email": counsellor['email'],
            "role": "Counsellor",
            "leads_assigned": leads_assigned,
            "leads_converted": leads_converted,
            "conversion_rate": conversion_rate,
            "followups_scheduled": followups_scheduled,
            "followups_completed": followups_completed,
            "followup_completion_rate": followup_rate,
            "efficiency_score": efficiency_score
        })
    
    # === FDE EFFICIENCY ===
    fdes = await db.users.find(
        {**branch_filter, "role": UserRole.FRONT_DESK.value, "is_active": True},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(100)
    
    fde_efficiency = []
    for fde in fdes:
        # Count enrollments processed this month
        enrollments_processed = await db.enrollments.count_documents({
            **branch_filter,
            "enrolled_by": fde['id'],
            "enrolled_on": {"$gte": month_start}
        })
        
        # Count payments collected this month
        payments_collected = await db.payments.find(
            {**branch_filter, "collected_by": fde['id'], "payment_date": {"$gte": month_start}},
            {"_id": 0, "amount": 1}
        ).to_list(10000)
        
        total_payments = len(payments_collected)
        total_amount = sum(p.get('amount', 0) for p in payments_collected)
        
        # Count tasks completed
        tasks_assigned = await db.tasks.count_documents({"assigned_to": fde['id']})
        tasks_completed = await db.tasks.count_documents({"assigned_to": fde['id'], "status": "Completed"})
        task_completion_rate = round(tasks_completed / tasks_assigned * 100, 1) if tasks_assigned > 0 else 100
        
        # Calculate efficiency score
        efficiency_score = round((task_completion_rate * 0.3 + min(enrollments_processed * 10, 100) * 0.4 + min(total_payments * 5, 100) * 0.3), 1)
        efficiency_score = min(efficiency_score, 100)
        
        fde_efficiency.append({
            "user_id": fde['id'],
            "name": fde.get('name', fde['email']),
            "email": fde['email'],
            "role": "Front Desk Executive",
            "enrollments_processed": enrollments_processed,
            "payments_collected": total_payments,
            "amount_collected": total_amount,
            "tasks_assigned": tasks_assigned,
            "tasks_completed": tasks_completed,
            "task_completion_rate": task_completion_rate,
            "efficiency_score": efficiency_score
        })
    
    # === TRAINER EFFICIENCY ===
    trainers = await db.users.find(
        {**branch_filter, "role": UserRole.TRAINER.value, "is_active": True},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(100)
    
    trainer_efficiency = []
    for trainer in trainers:
        # Get batches assigned
        batches = await db.batches.find(
            {"trainer_id": trainer['id']},
            {"_id": 0, "id": 1}
        ).to_list(100)
        batch_ids = [b['id'] for b in batches]
        
        # Count active students
        active_students = await db.enrollments.count_documents({
            "batch_id": {"$in": batch_ids},
            "status": "Active"
        })
        
        # Count completed students (marked by this trainer)
        completed_students = await db.enrollments.count_documents({
            "batch_id": {"$in": batch_ids},
            "status": "Completed"
        })
        
        # Attendance marking rate (last 7 days)
        attendance_days = await db.attendance.distinct("date", {
            "batch_id": {"$in": batch_ids},
            "date": {"$gte": week_ago}
        })
        attendance_rate = round(len(attendance_days) / 7 * 100, 1) if batch_ids else 0
        
        # Feedback scores (if available)
        feedbacks = await db.feedbacks.find(
            {"trainer_id": trainer['id']},
            {"_id": 0, "trainer_rating": 1}
        ).to_list(1000)
        avg_rating = round(sum(f.get('trainer_rating', 0) for f in feedbacks) / len(feedbacks), 1) if feedbacks else 0
        
        # Calculate efficiency score
        efficiency_score = round((attendance_rate * 0.3 + avg_rating * 10 * 0.4 + min(completed_students * 10, 100) * 0.3), 1)
        efficiency_score = min(efficiency_score, 100)
        
        trainer_efficiency.append({
            "user_id": trainer['id'],
            "name": trainer.get('name', trainer['email']),
            "email": trainer['email'],
            "role": "Trainer",
            "total_batches": len(batches),
            "active_students": active_students,
            "completed_students": completed_students,
            "attendance_rate": attendance_rate,
            "avg_feedback_rating": avg_rating,
            "efficiency_score": efficiency_score
        })
    
    # Combine all users and sort by efficiency
    all_users = counsellor_efficiency + fde_efficiency + trainer_efficiency
    all_users.sort(key=lambda x: x['efficiency_score'], reverse=True)
    
    # === AI ANALYSIS ===
    ai_analysis = None
    if LLM_AVAILABLE and os.environ.get('EMERGENT_LLM_KEY'):
        try:
            efficiency_summary = f"""
User Efficiency Data for Branch Analysis:

COUNSELLORS:
{json.dumps(counsellor_efficiency, indent=2)}

FRONT DESK EXECUTIVES:
{json.dumps(fde_efficiency, indent=2)}

TRAINERS:
{json.dumps(trainer_efficiency, indent=2)}

Analysis Period: {month_start} to {today_str}
"""
            
            chat = LlmChat(
                api_key=os.environ.get('EMERGENT_LLM_KEY'),
                session_id=f"user-efficiency-{branch_id}-{today.strftime('%Y%m%d')}",
                system_message="""You are an HR analyst for an educational institute. Analyze user efficiency data and provide actionable insights.

Your response MUST be valid JSON with this structure:
{
  "top_performers": [
    {"name": "user name", "role": "role", "highlight": "what makes them excel"}
  ],
  "needs_attention": [
    {"name": "user name", "role": "role", "issue": "what needs improvement", "suggestion": "how to improve"}
  ],
  "team_insights": {
    "counsellors": "brief analysis of counsellor team performance",
    "fdes": "brief analysis of FDE team performance", 
    "trainers": "brief analysis of trainer team performance"
  },
  "recommendations": [
    "specific actionable recommendation 1",
    "specific actionable recommendation 2",
    "specific actionable recommendation 3"
  ],
  "overall_efficiency": {
    "score": 1-100,
    "status": "excellent|good|average|needs improvement",
    "summary": "2-3 sentence summary of overall team efficiency"
  }
}

Be specific and use actual data. Focus on actionable insights."""
            ).with_model("openai", "gpt-4o")
            
            user_message = UserMessage(text=f"Analyze this user efficiency data:\n\n{efficiency_summary}")
            ai_response = await chat.send_message(user_message)
            
            try:
                ai_text = ai_response.strip()
                if ai_text.startswith("```"):
                    ai_text = ai_text.split("```")[1]
                    if ai_text.startswith("json"):
                        ai_text = ai_text[4:]
                ai_analysis = json.loads(ai_text)
            except json.JSONDecodeError:
                ai_analysis = {"raw_response": ai_response}
                
        except Exception as e:
            logger.error(f"User efficiency AI analysis failed: {e}")
    
    return {
        "generated_at": today.isoformat(),
        "analysis_period": {"start": month_start, "end": today_str},
        "counsellors": counsellor_efficiency,
        "fdes": fde_efficiency,
        "trainers": trainer_efficiency,
        "all_users_ranked": all_users,
        "ai_analysis": ai_analysis,
        "ai_powered": ai_analysis is not None
    }


# Analytics
@api_router.get("/analytics/overview")
async def get_analytics_overview(current_user: User = Depends(get_current_user)):
    # Base query - EXCLUDE deleted leads
    base_query = {"is_deleted": {"$ne": True}}
    if current_user.role != UserRole.ADMIN:
        base_query["branch_id"] = current_user.branch_id
    
    # Count active (non-deleted) leads
    total_leads = await db.leads.count_documents(base_query)
    
    # Count deleted leads separately
    deleted_query = {"is_deleted": True}
    if current_user.role != UserRole.ADMIN:
        deleted_query["branch_id"] = current_user.branch_id
    deleted_count = await db.leads.count_documents(deleted_query)
    
    # Status counts - only for non-deleted leads
    status_counts = {}
    for lead_status in LeadStatus:
        count = await db.leads.count_documents({**base_query, "status": lead_status.value})
        status_counts[lead_status.value] = count
    
    # Add deleted count to status breakdown
    status_counts["Deleted"] = deleted_count
    
    pipeline = [
        {"$match": base_query},
        {"$group": {"_id": "$lead_source", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    source_stats = await db.leads.aggregate(pipeline).to_list(100)
    
    pipeline = [
        {"$match": base_query},
        {"$group": {"_id": "$program_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    program_stats = await db.leads.aggregate(pipeline).to_list(100)
    
    return {
        "total_leads": total_leads,
        "deleted_leads": deleted_count,
        "status_breakdown": status_counts,
        "source_performance": [{"source": s["_id"], "count": s["count"]} for s in source_stats],
        "program_performance": [{"program": p["_id"], "count": p["count"]} for p in program_stats]
    }

@api_router.get("/analytics/branch-wise")
async def get_branch_wise_analytics(current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Get analytics grouped by branch - Admin only"""
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    
    branch_analytics = []
    for branch in branches:
        branch_id = branch["id"]
        
        # Base query excludes deleted leads
        base_query = {"branch_id": branch_id, "is_deleted": {"$ne": True}}
        
        # Total leads (excluding deleted)
        total_leads = await db.leads.count_documents(base_query)
        
        # Deleted leads count
        deleted_count = await db.leads.count_documents({"branch_id": branch_id, "is_deleted": True})
        
        # Status counts (excluding deleted)
        new_count = await db.leads.count_documents({**base_query, "status": "New"})
        contacted_count = await db.leads.count_documents({**base_query, "status": "Contacted"})
        demo_count = await db.leads.count_documents({**base_query, "status": "Demo Booked"})
        followup_count = await db.leads.count_documents({**base_query, "status": "Follow-up"})
        converted_count = await db.leads.count_documents({**base_query, "status": "Converted"})
        lost_count = await db.leads.count_documents({**base_query, "status": "Lost"})
        
        # Conversion rate
        conversion_rate = (converted_count / total_leads * 100) if total_leads > 0 else 0
        
        # Active counsellors
        counsellors_count = await db.users.count_documents({"branch_id": branch_id, "role": "Counsellor"})
        
        branch_analytics.append({
            "branch_id": branch_id,
            "branch_name": branch["name"],
            "branch_location": branch["location"],
            "total_leads": total_leads,
            "deleted_leads": deleted_count,
            "new_leads": new_count,
            "contacted": contacted_count,
            "demo_booked": demo_count,
            "followup": followup_count,
            "converted": converted_count,
            "lost": lost_count,
            "conversion_rate": round(conversion_rate, 2),
            "active_counsellors": counsellors_count
        })
    
    # Sort by total leads (descending)
    branch_analytics.sort(key=lambda x: x["total_leads"], reverse=True)
    
    return branch_analytics

# Reports
@api_router.get("/reports/leads")
async def generate_leads_report(
    status: Optional[str] = None,
    source: Optional[str] = None,
    program_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "csv",
    current_user: User = Depends(get_current_user)
):
    query = {}
    
    if current_user.role != UserRole.ADMIN:
        query["branch_id"] = current_user.branch_id
    
    if status:
        query["status"] = status
    if source:
        query["lead_source"] = source
    if program_id:
        query["program_id"] = program_id
    if branch_id and current_user.role == UserRole.ADMIN:
        query["branch_id"] = branch_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Name', 'Email', 'Phone', 'Program', 'Source', 'Status', 'City', 'Fee Quoted', 'Created At'])
        
        for lead in leads:
            writer.writerow([
                lead.get('name'),
                lead.get('email'),
                lead.get('number'),
                lead.get('program_name'),
                lead.get('lead_source'),
                lead.get('status'),
                lead.get('city', ''),
                lead.get('fee_quoted', ''),
                lead.get('created_at', '')
            ])
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=leads_report.csv"}
        )
    
    return leads

# Unified Reports Endpoint
@api_router.get("/reports/generate")
async def generate_report(
    request: Request,
    report_type: str = "leads",
    branch_id: Optional[str] = None,
    status: Optional[str] = None,
    source: Optional[str] = None,
    program_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "csv",
    session: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Generate various types of reports"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Determine branch filter
    branch_filter = {}
    if current_user.role not in [UserRole.ADMIN]:
        branch_filter["branch_id"] = current_user.branch_id
    elif branch_id and branch_id != "All":
        branch_filter["branch_id"] = branch_id
    
    # Session filter
    session_val = session or await get_session_from_request(request)
    
    # Date filter helper - uses session if provided, otherwise manual dates
    def get_date_query(date_field):
        if session_val and session_val != "all":
            return get_session_filter(session_val, date_field)
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        return {date_field: date_query} if date_query else {}
    
    if report_type == "leads":
        query = {**branch_filter, "is_deleted": {"$ne": True}}
        if status and status != "All":
            query["status"] = status
        if source and source != "All":
            query["lead_source"] = source
        if program_id and program_id != "All":
            query["program_id"] = program_id
        if start_date or end_date:
            query.update(get_date_query("created_at"))
        
        leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
        # Columns: Lead ID, Name, Phone Number, Course, Source
        writer.writerow(['Lead ID', 'Name', 'Phone Number', 'Course', 'Source'])
        for lead in leads:
            writer.writerow([
                lead.get('lead_id', lead.get('id', '')[:8] if lead.get('id') else ''),
                lead.get('name', ''),
                lead.get('number', ''),
                lead.get('program_name', ''),
                lead.get('lead_source', '')
            ])
        filename = "leads_report.csv"
    
    elif report_type == "enrollments":
        query = {**branch_filter}
        if start_date or end_date:
            query.update(get_date_query("enrollment_date"))
        
        enrollments = await db.enrollments.find(query, {"_id": 0}).sort("enrollment_date", -1).to_list(10000)
        writer.writerow(['Enrollment ID', 'Student Name', 'Email', 'Phone', 'Program', 'Fee Quoted', 'Discount', 'Final Fee', 'Total Paid', 'Payment Status', 'Enrollment Date', 'Status'])
        for e in enrollments:
            writer.writerow([
                e.get('enrollment_id', ''), e.get('student_name', ''), e.get('email', ''), e.get('phone', ''),
                e.get('program_name', ''), e.get('fee_quoted', ''), 
                e.get('discount_amount') or f"{e.get('discount_percent', 0)}%",
                e.get('final_fee', ''), e.get('total_paid', 0), e.get('payment_status', ''),
                e.get('enrollment_date', ''), e.get('status', 'Active')
            ])
        filename = "enrollments_report.csv"
    
    elif report_type == "income":
        query = {**branch_filter}
        if start_date or end_date:
            query.update(get_date_query("payment_date"))
        
        payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(10000)
        # Columns: Receipt No, Student Name, Number, Amount, Mode, Payment Date
        writer.writerow(['Receipt No', 'Student Name', 'Number', 'Amount', 'Mode', 'Payment Date'])
        for p in payments:
            writer.writerow([
                p.get('receipt_number', ''),
                p.get('student_name', ''),
                p.get('student_phone', p.get('phone', '')),
                p.get('amount', ''),
                p.get('payment_mode', ''),
                p.get('payment_date', '')
            ])
        filename = "monthly_collection_report.csv"
    
    elif report_type == "expenses":
        query = {**branch_filter}
        if start_date or end_date:
            query.update(get_date_query("expense_date"))
        
        expenses = await db.expenses.find(query, {"_id": 0}).sort("expense_date", -1).to_list(10000)
        # Columns: Name, Date, Amount, Description
        writer.writerow(['Name', 'Date', 'Amount', 'Description'])
        for e in expenses:
            writer.writerow([
                e.get('name', ''),
                e.get('expense_date', ''),
                e.get('amount', ''),
                e.get('remarks', e.get('description', ''))
            ])
        filename = "expenses_report.csv"
    
    elif report_type == "pending_payments":
        # Get enrollments with pending fees
        query = {**branch_filter}
        if start_date or end_date:
            query.update(get_date_query("enrollment_date"))
        
        enrollments = await db.enrollments.find(query, {"_id": 0}).to_list(10000)
        
        # Filter for pending payments (final_fee > total_paid)
        pending = []
        for e in enrollments:
            final_fee = e.get('final_fee') or e.get('fee_quoted', 0) or 0
            total_paid = e.get('total_paid', 0) or 0
            pending_fee = final_fee - total_paid
            if pending_fee > 0:
                pending.append({
                    'name': e.get('student_name', ''),
                    'course': e.get('program_name', ''),
                    'number': e.get('phone', e.get('student_phone', '')),
                    'final_fee': final_fee,
                    'paid_fee': total_paid,
                    'pending_fee': pending_fee
                })
        
        # Columns: Name, Course, Number, Final Fee, Paid Fee, Pending Fee
        writer.writerow(['Name', 'Course', 'Number', 'Final Fee', 'Paid Fee', 'Pending Fee'])
        for p in pending:
            writer.writerow([
                p['name'],
                p['course'],
                p['number'],
                p['final_fee'],
                p['paid_fee'],
                p['pending_fee']
            ])
        filename = "pending_payments_report.csv"
    
    elif report_type == "fee_collection":
        # Monthly fee collection report - shows installments due in a specific month
        from calendar import monthrange
        
        # Default to current month if no date specified
        if start_date:
            year = int(start_date[:4])
            month = int(start_date[5:7])
        else:
            now = datetime.now()
            year = now.year
            month = now.month
        
        # Get first and last day of the month
        _, last_day = monthrange(year, month)
        month_start = f"{year}-{str(month).zfill(2)}-01"
        month_end = f"{year}-{str(month).zfill(2)}-{str(last_day).zfill(2)}"
        
        # Get all installments due in the selected month
        installments = await db.installment_schedule.find({
            "due_date": {"$gte": month_start, "$lte": month_end}
        }, {"_id": 0}).to_list(10000)
        
        due_installments = []
        
        for installment in installments:
            # Get the payment plan
            plan = await db.payment_plans.find_one({"id": installment.get('payment_plan_id')}, {"_id": 0})
            if not plan:
                continue
            
            # Check branch filter
            if branch_filter.get('branch_id') and plan.get('branch_id') != branch_filter.get('branch_id'):
                continue
                
            enrollment = await db.enrollments.find_one({"id": plan.get('enrollment_id')}, {"_id": 0})
            if not enrollment:
                continue
                
            due_installments.append({
                'enrollment_id': enrollment.get('enrollment_id', ''),
                'student_name': enrollment.get('student_name', ''),
                'phone': enrollment.get('phone') or enrollment.get('student_phone', ''),
                'program_name': enrollment.get('program_name', ''),
                'total_fee': enrollment.get('final_fee') or enrollment.get('fee_quoted', 0),
                'paid_fee': enrollment.get('total_paid', 0),
                'due_amount': installment.get('amount', 0),
                'due_date': installment.get('due_date', ''),
                'installment_number': installment.get('installment_number', 1),
                'status': installment.get('status', 'Pending')
            })
        
        # Sort by due date
        due_installments.sort(key=lambda x: x['due_date'])
        
        writer.writerow(['Student ID', 'Student Name', 'Phone', 'Course', 'Total Fee', 'Paid Fee', 'Due Amount', 'Due Date', 'Installment #', 'Status'])
        for item in due_installments:
            writer.writerow([
                item['enrollment_id'], item['student_name'], item['phone'], item['program_name'],
                item['total_fee'], item['paid_fee'], item['due_amount'], item['due_date'],
                item['installment_number'], item['status']
            ])
        
        month_name = datetime(year, month, 1).strftime('%B_%Y')
        filename = f"fee_collection_{month_name}.csv"
    
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Expense Category Management (Admin)
@api_router.post("/admin/expense-categories", response_model=ExpenseCategory)
async def create_expense_category(category: ExpenseCategoryCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    new_category = ExpenseCategory(**category.model_dump())
    category_dict = new_category.model_dump()
    category_dict['created_at'] = category_dict['created_at'].isoformat()
    
    await db.expense_categories.insert_one(category_dict)
    return new_category

@api_router.delete("/admin/expense-categories/{category_id}")
async def delete_expense_category(category_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Delete an expense category - Admin only"""
    # Check if category is used in any expense
    expense_using = await db.expenses.find_one({"category_id": category_id}, {"_id": 0})
    if expense_using:
        raise HTTPException(status_code=400, detail="Cannot delete category - it is being used in expenses")
    
    result = await db.expense_categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted successfully"}

@api_router.get("/expense-categories", response_model=List[ExpenseCategory])
async def get_expense_categories(current_user: User = Depends(get_current_user)):
    categories = await db.expense_categories.find({}, {"_id": 0}).to_list(1000)
    for cat in categories:
        if isinstance(cat.get('created_at'), str):
            cat['created_at'] = datetime.fromisoformat(cat['created_at'])
    return [ExpenseCategory(**c) for c in categories]

# Lead Sources Management (Admin configurable)
@api_router.post("/admin/lead-sources", response_model=LeadSource)
async def create_lead_source(source: LeadSourceCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Create a new lead source - Admin only"""
    # Check if source already exists
    existing = await db.lead_sources.find_one({"name": source.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Lead source with this name already exists")
    
    new_source = LeadSource(**source.model_dump())
    source_dict = new_source.model_dump()
    source_dict['created_at'] = source_dict['created_at'].isoformat()
    
    await db.lead_sources.insert_one(source_dict)
    return new_source

@api_router.get("/lead-sources", response_model=List[LeadSource])
async def get_lead_sources(current_user: User = Depends(get_current_user)):
    """Get all active lead sources"""
    sources = await db.lead_sources.find({"is_active": True}, {"_id": 0}).to_list(100)
    for src in sources:
        if isinstance(src.get('created_at'), str):
            src['created_at'] = datetime.fromisoformat(src['created_at'])
    return [LeadSource(**s) for s in sources]

@api_router.delete("/admin/lead-sources/{source_id}")
async def delete_lead_source(source_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Delete a lead source - Admin only (soft delete by setting inactive)"""
    result = await db.lead_sources.update_one(
        {"id": source_id},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead source not found")
    return {"message": "Lead source deleted successfully"}

# Expense Management (FDA)
@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense: ExpenseCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        raise HTTPException(status_code=403, detail="Only Branch Admin or Front Desk Executive can add expenses")
    
    if not current_user.branch_id and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="User must be assigned to a branch")
    
    category = await db.expense_categories.find_one({"id": expense.category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Expense category not found")
    
    expense_data = expense.model_dump()
    expense_data.pop('expense_date', None)  # Remove to avoid duplicate argument
    
    new_expense = Expense(
        **expense_data,
        branch_id=current_user.branch_id or "admin",
        category_name=category['name'],
        created_by=current_user.id,
        expense_date=datetime.fromisoformat(expense.expense_date).date()
    )
    expense_dict = new_expense.model_dump()
    expense_dict['created_at'] = expense_dict['created_at'].isoformat()
    expense_dict['expense_date'] = expense_dict['expense_date'].isoformat()
    
    await db.expenses.insert_one(expense_dict)
    return new_expense

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(current_user: User = Depends(get_current_user)):
    # FDE should not see expenses
    if current_user.role == UserRole.FRONT_DESK:
        raise HTTPException(status_code=403, detail="Front Desk Executive cannot access expenses")
    
    query = {}
    if current_user.role not in [UserRole.ADMIN]:
        query["branch_id"] = current_user.branch_id
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("expense_date", -1).to_list(1000)
    for exp in expenses:
        if isinstance(exp.get('created_at'), str):
            exp['created_at'] = datetime.fromisoformat(exp['created_at'])
        if isinstance(exp.get('expense_date'), str):
            exp['expense_date'] = datetime.fromisoformat(exp['expense_date']).date()
    return [Expense(**e) for e in expenses]

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: User = Depends(get_current_user)):
    """Delete an expense - Branch Admin only for their branch"""
    if current_user.role != UserRole.BRANCH_ADMIN:
        raise HTTPException(status_code=403, detail="Only Branch Admin can delete expenses")
    
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    # Branch Admin can only delete expenses from their branch
    if expense.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only delete expenses from your branch")
    
    await db.expenses.delete_one({"id": expense_id})
    return {"message": "Expense deleted successfully"}

# File Upload for Student Photos
@api_router.post("/upload/image")
async def upload_image(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Upload student photo or document image and return base64 data URL"""
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Only image files are allowed")
    
    # Read file content
    content = await file.read()
    
    # Check file size (max 5MB)
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be less than 5MB")
    
    # Convert to base64 data URL
    base64_data = base64.b64encode(content).decode('utf-8')
    data_url = f"data:{file.content_type};base64,{base64_data}"
    
    return {"url": data_url, "filename": file.filename}

# Enrollment Management (FDA)
@api_router.post("/enrollments", response_model=Enrollment)
async def create_enrollment(enrollment: EnrollmentCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.FRONT_DESK, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Front Desk Executive or Branch Admin can create enrollments")
    
    # Verify lead exists and is converted
    lead = await db.leads.find_one({"id": enrollment.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if lead["status"] != LeadStatus.CONVERTED.value:
        raise HTTPException(status_code=400, detail="Only converted leads can be enrolled")
    
    # Check if already enrolled
    existing = await db.enrollments.find_one({"lead_id": enrollment.lead_id}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Lead already enrolled")
    
    # Get program details
    program = await db.programs.find_one({"id": enrollment.program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    # Calculate final fee - support both percentage and direct amount discount
    if enrollment.discount_amount and enrollment.discount_amount > 0:
        # Use direct discount amount
        discount = enrollment.discount_amount
    else:
        # Use percentage discount
        discount = (enrollment.fee_quoted * (enrollment.discount_percent or 0)) / 100
    final_fee = enrollment.fee_quoted - discount
    
    # Generate custom enrollment ID
    # Source of truth for branch is the lead (the enrollment inherits its branch).
    # Only fall back to the current user's branch if the lead has none.
    branch_id = lead.get("branch_id") or current_user.branch_id
    if not branch_id:
        raise HTTPException(status_code=400, detail="Cannot determine branch for enrollment")
    branch_doc = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch_doc:
        raise HTTPException(status_code=400, detail=f"Branch {branch_id} not found; cannot generate enrollment ID")
    custom_enrollment_id = await generate_custom_id(branch_id, "E")
    if not custom_enrollment_id:
        raise HTTPException(status_code=500, detail="Failed to generate enrollment ID")
    
    enrollment_data = enrollment.model_dump()
    enrollment_data.pop('enrollment_date', None)  # Remove to avoid duplicate argument
    
    new_enrollment = Enrollment(
        **enrollment_data,
        enrollment_id=custom_enrollment_id,
        branch_id=branch_id,
        program_name=program['name'],
        final_fee=final_fee,
        enrollment_date=datetime.fromisoformat(enrollment.enrollment_date).date(),
        created_by=current_user.id
    )
    
    enrollment_dict = new_enrollment.model_dump()
    enrollment_dict['created_at'] = enrollment_dict['created_at'].isoformat()
    enrollment_dict['enrollment_date'] = enrollment_dict['enrollment_date'].isoformat()
    
    await db.enrollments.insert_one(enrollment_dict)
    
    # Send WhatsApp notification for enrollment confirmation
    await send_whatsapp_notification(
        new_enrollment.phone,
        "enrollment_confirmed",
        {
            "name": new_enrollment.student_name, 
            "enrollment_number": custom_enrollment_id or new_enrollment.id,
            "course": new_enrollment.program_name
        }
    )
    
    return new_enrollment

@api_router.get("/enrollments", response_model=List[Enrollment])
async def get_enrollments(request: Request, session: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.ADMIN:
        query["branch_id"] = current_user.branch_id
    
    # Session filtering
    session_val = session or await get_session_from_request(request)
    if session_val and session_val != "all":
        session_filter = get_session_filter(session_val, "enrollment_date")
        query.update(session_filter)
    
    enrollments = await db.enrollments.find(query, {"_id": 0}).sort("enrollment_date", -1).to_list(1000)
    for enr in enrollments:
        if isinstance(enr.get('created_at'), str):
            enr['created_at'] = datetime.fromisoformat(enr['created_at'])
        if isinstance(enr.get('enrollment_date'), str):
            enr['enrollment_date'] = datetime.fromisoformat(enr['enrollment_date']).date()
    return [Enrollment(**e) for e in enrollments]

# Payment Plan Management
@api_router.post("/payment-plans", response_model=PaymentPlan)
async def create_payment_plan(plan: PaymentPlanCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    # Check if enrollment exists
    enrollment = await db.enrollments.find_one({"id": plan.enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    # Validate max 6 installments
    if plan.plan_type == PaymentPlanType.INSTALLMENTS:
        if plan.installments_count and plan.installments_count > 6:
            raise HTTPException(status_code=400, detail="Maximum 6 installments allowed")
        if plan.installments and len(plan.installments) > 6:
            raise HTTPException(status_code=400, detail="Maximum 6 installments allowed")
    
    # Check if payment plan already exists
    existing_plan = await db.payment_plans.find_one({"enrollment_id": plan.enrollment_id}, {"_id": 0})
    if existing_plan:
        raise HTTPException(status_code=400, detail="Payment plan already exists for this enrollment")
    
    new_plan = PaymentPlan(
        enrollment_id=plan.enrollment_id,
        plan_type=plan.plan_type,
        total_amount=plan.total_amount,
        installments_count=plan.installments_count if plan.plan_type == PaymentPlanType.INSTALLMENTS else None
    )
    
    plan_dict = new_plan.model_dump()
    plan_dict['created_at'] = plan_dict['created_at'].isoformat()
    
    await db.payment_plans.insert_one(plan_dict)
    
    # If installments, create installment schedule
    if plan.plan_type == PaymentPlanType.INSTALLMENTS and plan.installments:
        for idx, inst in enumerate(plan.installments):
            await db.installment_schedule.insert_one({
                "id": str(uuid.uuid4()),
                "payment_plan_id": new_plan.id,
                "enrollment_id": plan.enrollment_id,
                "installment_number": inst.get("installment_number", idx + 1),
                "amount": inst["amount"],
                "due_date": inst["due_date"],
                "status": "Pending"
            })
    
    return new_plan

# Payment Recording
@api_router.post("/payments", response_model=Payment)
async def create_payment(payment: PaymentCreate, request: Request, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.FRONT_DESK, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Front Desk Executive or Branch Admin can record payments")
    
    # Get enrollment to get branch_id and validate payment amount
    enrollment = await db.enrollments.find_one({"id": payment.enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    # Calculate total paid so far
    existing_payments = await db.payments.find({"enrollment_id": payment.enrollment_id}, {"_id": 0, "amount": 1}).to_list(1000)
    total_paid = sum(p.get('amount', 0) for p in existing_payments)
    
    # Check if payment exceeds remaining fee
    final_fee = enrollment.get('final_fee', 0)
    remaining_fee = final_fee - total_paid
    
    if payment.amount > remaining_fee:
        raise HTTPException(
            status_code=400, 
            detail=f"Payment amount (₹{payment.amount}) exceeds remaining fee (₹{remaining_fee}). Total fee: ₹{final_fee}, Already paid: ₹{total_paid}"
        )
    
    # Validate installment payment - must pay exact or more than installment amount
    if payment.installment_number and payment.payment_plan_id:
        installment = await db.installment_schedule.find_one(
            {
                "payment_plan_id": payment.payment_plan_id,
                "installment_number": payment.installment_number
            },
            {"_id": 0}
        )
        if installment:
            installment_amount = installment.get('amount', 0)
            # Branch Admin, Admin, AND Front Desk can record payments below the installment
            # amount (FDE handles real-world cases where student brings partial cash today
            # and rest tomorrow). Other roles (e.g. Counsellor) must pay full amount.
            if payment.amount < installment_amount and current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Payment amount (₹{payment.amount}) is less than installment amount (₹{installment_amount}). Please pay the full installment amount or contact Branch Admin for special cases."
                )
    
    # Generate custom receipt number
    branch_id = enrollment.get('branch_id') or current_user.branch_id
    if not branch_id:
        raise HTTPException(status_code=400, detail="Cannot determine branch for receipt generation")
    custom_receipt_id = await generate_custom_id(branch_id, "R")
    
    payment_data = payment.model_dump()
    payment_data.pop('payment_date', None)  # Remove to avoid duplicate argument
    
    new_payment = Payment(
        **payment_data,
        receipt_number=custom_receipt_id or f"RCP-{str(uuid.uuid4())[:8].upper()}",
        branch_id=branch_id,
        payment_date=datetime.fromisoformat(payment.payment_date).date(),
        student_name=enrollment.get('student_name', ''),
        program_name=enrollment.get('program_name', ''),
        created_by=current_user.id
    )
    
    payment_dict = new_payment.model_dump()
    payment_dict['created_at'] = payment_dict['created_at'].isoformat()
    payment_dict['payment_date'] = payment_dict['payment_date'].isoformat()
    
    await db.payments.insert_one(payment_dict)
    
    # Update installment status if applicable
    if payment.installment_number:
        # Get the installment details
        installment = await db.installment_schedule.find_one(
            {
                "payment_plan_id": payment.payment_plan_id,
                "installment_number": payment.installment_number
            },
            {"_id": 0}
        )
        
        if installment:
            installment_amount = installment.get('amount', 0)
            
            # Check if this is a partial payment for this installment
            if payment.amount < installment_amount:
                shortfall = installment_amount - payment.amount
                
                # Get remaining unpaid installments (after the current one)
                remaining_installments = await db.installment_schedule.find(
                    {
                        "payment_plan_id": payment.payment_plan_id,
                        "installment_number": {"$gt": payment.installment_number},
                        "status": {"$ne": "Paid"}
                    },
                    {"_id": 0}
                ).to_list(100)
                
                # Distribute shortfall equally among remaining installments
                if remaining_installments:
                    shortfall_per_installment = round(shortfall / len(remaining_installments), 2)
                    
                    for remaining_inst in remaining_installments:
                        new_amount = remaining_inst.get('amount', 0) + shortfall_per_installment
                        await db.installment_schedule.update_one(
                            {
                                "payment_plan_id": payment.payment_plan_id,
                                "installment_number": remaining_inst.get('installment_number')
                            },
                            {"$set": {"amount": new_amount}}
                        )
                    
                    logger.info(f"Redistributed shortfall of ₹{shortfall} (₹{shortfall_per_installment} each) to {len(remaining_installments)} remaining installments")
                else:
                    # This is the LAST installment and partial payment - create a new installment
                    # Get total number of installments
                    all_installments = await db.installment_schedule.find(
                        {"payment_plan_id": payment.payment_plan_id},
                        {"_id": 0}
                    ).to_list(100)
                    new_installment_number = len(all_installments) + 1
                    
                    # Calculate due date (30 days from now)
                    new_due_date = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
                    
                    # Create new installment for remaining amount
                    new_installment = {
                        "id": str(uuid.uuid4()),
                        "payment_plan_id": payment.payment_plan_id,
                        "enrollment_id": payment.enrollment_id,
                        "installment_number": new_installment_number,
                        "amount": shortfall,
                        "due_date": new_due_date,
                        "status": "Pending"
                    }
                    await db.installment_schedule.insert_one(new_installment)
                    
                    # Update payment plan total installments
                    await db.payment_plans.update_one(
                        {"id": payment.payment_plan_id},
                        {"$set": {"total_installments": new_installment_number}}
                    )
                    
                    logger.info(f"Created new installment #{new_installment_number} for ₹{shortfall} due on {new_due_date}")
            
            # Mark current installment as Paid
            await db.installment_schedule.update_one(
                {
                    "payment_plan_id": payment.payment_plan_id,
                    "installment_number": payment.installment_number
                },
                {"$set": {"status": "Paid", "paid_date": payment.payment_date, "paid_amount": payment.amount}}
            )
    
    # Update total_paid and payment_status in enrollment
    new_total_paid = total_paid + payment.amount
    new_payment_status = "Paid" if new_total_paid >= final_fee else ("Partial" if new_total_paid > 0 else "Pending")
    await db.enrollments.update_one(
        {"id": payment.enrollment_id},
        {"$set": {
            "total_paid": new_total_paid,
            "payment_status": new_payment_status
        }}
    )
    
    # Send WhatsApp notification for payment received with full details
    pending_fee = final_fee - new_total_paid
    await send_whatsapp_notification(
        enrollment.get('phone', ''),
        "payment_received",
        {
            "name": enrollment.get('student_name', ''),
            "amount": str(payment.amount),
            "total_fee": str(final_fee),
            "paid_fee": str(new_total_paid),
            "pending_fee": str(pending_fee),
            "receipt_number": new_payment.receipt_number
        }
    )
    
    # Audit Log
    await create_audit_log(
        user=current_user,
        action="create",
        entity_type="payment",
        entity_id=new_payment.id,
        entity_name=f"₹{payment.amount} for {enrollment.get('student_name', '')} ({new_payment.receipt_number})",
        changes={"amount": payment.amount, "student": enrollment.get('student_name'), "receipt": new_payment.receipt_number},
        request=request
    )
    
    return new_payment

@api_router.get("/enrollments/{enrollment_id}/payments")
async def get_enrollment_payments(enrollment_id: str, current_user: User = Depends(get_current_user)):
    payments = await db.payments.find({"enrollment_id": enrollment_id}, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    for pay in payments:
        if isinstance(pay.get('created_at'), str):
            pay['created_at'] = datetime.fromisoformat(pay['created_at'])
        if isinstance(pay.get('payment_date'), str):
            pay['payment_date'] = datetime.fromisoformat(pay['payment_date']).date()
    return payments

@api_router.get("/enrollments/{enrollment_id}/payment-plan")
async def get_enrollment_payment_plan(enrollment_id: str, current_user: User = Depends(get_current_user)):
    plan = await db.payment_plans.find_one({"enrollment_id": enrollment_id}, {"_id": 0})
    if not plan:
        return None
    
    if isinstance(plan.get('created_at'), str):
        plan['created_at'] = datetime.fromisoformat(plan['created_at'])
    
    # Get installment schedule if applicable
    if plan.get('plan_type') == PaymentPlanType.INSTALLMENTS:
        schedule = await db.installment_schedule.find({"payment_plan_id": plan['id']}, {"_id": 0}).to_list(1000)
        plan['installments'] = schedule
    
    return plan

@api_router.get("/payments/{payment_id}/receipt")
async def generate_receipt(payment_id: str, current_user: User = Depends(get_current_user)):
    """Generate payment receipt with full professional details for printing."""
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    enrollment = await db.enrollments.find_one({"id": payment['enrollment_id']}, {"_id": 0})
    branch = await db.branches.find_one({"id": enrollment.get('branch_id')}, {"_id": 0}) if enrollment else None
    
    # Calculate paid before this receipt + total paid till date
    all_payments = await db.payments.find(
        {"enrollment_id": payment['enrollment_id']},
        {"_id": 0, "amount": 1, "payment_date": 1, "id": 1}
    ).to_list(1000)
    total_paid = sum(p.get('amount', 0) for p in all_payments)
    # Previous payment = all paid before THIS payment (ordered by date+id)
    pay_amount = payment.get('amount', 0)
    previous_payment = max(0, total_paid - pay_amount)
    final_fee = enrollment.get('final_fee', 0) if enrollment else 0
    balance_fee = max(0, final_fee - total_paid)
    
    # Pull lead for parent/guardian + counsellor + lead_source
    lead = None
    if enrollment and enrollment.get('lead_id'):
        lead = await db.leads.find_one({"id": enrollment['lead_id']}, {"_id": 0})
    
    # Batch info
    batch = None
    if enrollment and enrollment.get('batch_id'):
        batch = await db.batches.find_one({"id": enrollment['batch_id']}, {"_id": 0})
    
    # Next installment due date (next unpaid installment)
    next_due_date = None
    fee_head = "Tuition"
    if enrollment:
        installments = await db.installments.find(
            {"enrollment_id": enrollment['id'], "status": {"$ne": "Paid"}},
            {"_id": 0}
        ).sort("due_date", 1).to_list(50) if hasattr(db, 'installments') else []
        # installments collection may or may not exist; safely handle
        try:
            for inst in installments:
                if inst.get('due_date'):
                    next_due_date = inst['due_date']
                    break
        except Exception:
            pass
        if payment.get('installment_number'):
            fee_head = f"Installment #{payment.get('installment_number')}"
    
    # Amount in words (Indian numbering)
    def _num_to_words_inr(n: int) -> str:
        if n == 0:
            return "Zero Rupees Only"
        ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
                "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
                "Seventeen", "Eighteen", "Nineteen"]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
        def two(num):
            if num < 20:
                return ones[num]
            t, o = divmod(num, 10)
            return tens[t] + (" " + ones[o] if o else "")
        def three(num):
            h, rest = divmod(num, 100)
            return (ones[h] + " Hundred" + (" " + two(rest) if rest else "")) if h else two(rest)
        crore, rem = divmod(n, 10000000)
        lakh, rem = divmod(rem, 100000)
        thou, rem = divmod(rem, 1000)
        hund = rem
        parts = []
        if crore: parts.append(three(crore) + " Crore")
        if lakh: parts.append(three(lakh) + " Lakh")
        if thou: parts.append(three(thou) + " Thousand")
        if hund: parts.append(three(hund))
        return " ".join(parts).strip() + " Rupees Only"
    
    amount_received = int(round(pay_amount))
    amount_in_words = _num_to_words_inr(amount_received)
    
    receipt_data = {
        "receipt_number": payment.get('receipt_number', payment_id[:8].upper()),
        "payment_id": payment_id,
        "payment_date": payment['payment_date'],
        "student_name": enrollment['student_name'] if enrollment else 'N/A',
        "student_id": enrollment.get('enrollment_id') or (enrollment.get('id', '')[:8].upper() if enrollment else 'N/A'),
        "student_email": enrollment.get('email', ''),
        "phone": enrollment.get('phone', ''),
        "father_name": (lead.get('parent_name') if lead else None) or enrollment.get('parent_name', '') if enrollment else '',
        "program_name": enrollment['program_name'] if enrollment else 'N/A',
        "enrollment_id": enrollment.get('enrollment_id') or (enrollment.get('id', '')[:8].upper() if enrollment else 'N/A'),
        "batch_name": (batch.get('name') if batch else None) or enrollment.get('batch_name', '') if enrollment else '',
        "session": enrollment.get('academic_session') or enrollment.get('session') or get_current_academic_session() if enrollment else get_current_academic_session(),
        "amount": pay_amount,
        "amount_in_words": amount_in_words,
        "payment_mode": payment['payment_mode'],
        "transaction_ref": payment.get('transaction_id') or payment.get('reference_number') or '',
        "installment_number": payment.get('installment_number'),
        "fee_head": fee_head,
        "remarks": payment.get('remarks', ''),
        "received_by": payment.get('received_by_name') or payment.get('received_by') or current_user.name,
        "counsellor_name": (lead.get('counsellor_name') if lead else None) or '',
        "total_fee": final_fee,
        "previous_payment": previous_payment,
        "total_paid": total_paid,
        "balance_fee": balance_fee,
        "next_installment_due_date": next_due_date,
        "branch_name": branch.get('name', 'ETI Educom') if branch else 'ETI Educom',
        "branch_address": branch.get('address', '') if branch else '',
        "branch_city": branch.get('city', '') if branch else '',
        "branch_phone": branch.get('branch_phone', '') or (branch.get('phone', '') if branch else '') if branch else '',
        "branch_email": branch.get('email', '') if branch else '',
        "branch_gstin": branch.get('gstin', '') if branch else '',
        "branch_website": branch.get('website', '') if branch else '',
        "institute_name": "ETI Educom™",
        "legal_entity": "ETI Learning Systems Private Limited",
        "legal_tagline": "Operating Under the Brand Name: ETI Educom™",
        "institute_tagline": "Professional Training & Skill Development",
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    
    return receipt_data

# All Payments Page with filters
@api_router.get("/payments/all")
async def get_all_payments(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    student_name: Optional[str] = None,
    contact_number: Optional[str] = None,
    payment_mode: Optional[str] = None,
    branch_id: Optional[str] = None,
    session: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all payments with filters"""
    query = {}
    
    # Branch filter based on role
    if current_user.role != UserRole.ADMIN:
        query["branch_id"] = current_user.branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    # Session filtering (takes precedence over manual date filters)
    session_val = session or await get_session_from_request(request)
    if session_val and session_val != "all":
        session_filter = get_session_filter(session_val, "payment_date")
        query.update(session_filter)
    elif start_date or end_date:
        if start_date:
            query["payment_date"] = {"$gte": start_date}
        if end_date:
            if "payment_date" in query:
                query["payment_date"]["$lte"] = end_date
            else:
                query["payment_date"] = {"$lte": end_date}
    
    if payment_mode:
        query["payment_mode"] = payment_mode
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(10000)
    
    # Enrich with enrollment data and filter by student name/contact if provided
    result = []
    for pay in payments:
        enrollment = await db.enrollments.find_one({"id": pay.get('enrollment_id')}, {"_id": 0})
        if enrollment:
            pay['student_name'] = enrollment.get('student_name', '')
            pay['student_email'] = enrollment.get('email', '')
            pay['student_phone'] = enrollment.get('phone', '')
            pay['program_name'] = enrollment.get('program_name', '')
            pay['final_fee'] = enrollment.get('final_fee', 0)
            
            # Filter by student name if provided
            if student_name and student_name.lower() not in pay['student_name'].lower():
                continue
            # Filter by contact number if provided
            if contact_number and contact_number not in pay.get('student_phone', ''):
                continue
        
        if isinstance(pay.get('created_at'), str):
            pay['created_at'] = datetime.fromisoformat(pay['created_at'])
        if isinstance(pay.get('payment_date'), str):
            pay['payment_date'] = datetime.fromisoformat(pay['payment_date']).date()
        
        result.append(pay)
    
    return result

# Pending Payments (Both One-time and Installments)
@api_router.get("/payments/pending")
async def get_pending_payments(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    student_name: Optional[str] = None,
    contact_number: Optional[str] = None,
    branch_id: Optional[str] = None,
    session: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all pending payments - both one-time and installments (optimized)"""
    pending_payments = []
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Determine branch filter
    branch_filter = {}
    if current_user.role != UserRole.ADMIN:
        branch_filter["branch_id"] = current_user.branch_id
    elif branch_id:
        branch_filter["branch_id"] = branch_id
    
    # Session filtering for enrollments
    session_val = session or await get_session_from_request(request)
    if session_val and session_val != "all":
        session_filter = get_session_filter(session_val, "enrollment_date")
        branch_filter.update(session_filter)
    
    # Exclude Dropped / Cancelled / Inactive enrollments from pending payments —
    # their pending fee should NOT show up as an outstanding due.
    branch_filter["status"] = {"$nin": ["Dropped", "Cancelled", "Inactive"]}

    # Get all enrollments for the branch
    enrollments = await db.enrollments.find(branch_filter, {"_id": 0}).to_list(10000)
    
    if not enrollments:
        return []
    
    # Filter by student name and contact in Python (avoid extra DB calls)
    if student_name:
        enrollments = [e for e in enrollments if student_name.lower() in e.get('student_name', '').lower()]
    if contact_number:
        enrollments = [e for e in enrollments if contact_number in e.get('phone', '')]
    
    if not enrollments:
        return []
    
    enrollment_ids = [e['id'] for e in enrollments]
    
    # Bulk fetch payment totals
    payment_agg = await db.payments.aggregate([
        {"$match": {"enrollment_id": {"$in": enrollment_ids}}},
        {"$group": {"_id": "$enrollment_id", "total_paid": {"$sum": "$amount"}}}
    ]).to_list(10000)
    payments_map = {p['_id']: p['total_paid'] for p in payment_agg}
    
    # Bulk fetch payment plans
    all_plans = await db.payment_plans.find(
        {"enrollment_id": {"$in": enrollment_ids}}, {"_id": 0}
    ).to_list(10000)
    plans_map = {p['enrollment_id']: p for p in all_plans}
    
    # Bulk fetch installment schedules for plans with installments
    installment_plan_ids = [p['id'] for p in all_plans if p.get('plan_type') == PaymentPlanType.INSTALLMENTS.value]
    all_schedules = []
    if installment_plan_ids:
        all_schedules = await db.installment_schedule.find(
            {"payment_plan_id": {"$in": installment_plan_ids}, "status": {"$ne": "Paid"}},
            {"_id": 0}
        ).to_list(10000)
    schedules_map = {}
    for s in all_schedules:
        pid = s.get('payment_plan_id')
        if pid not in schedules_map:
            schedules_map[pid] = []
        schedules_map[pid].append(s)
    
    for enrollment in enrollments:
        final_fee = enrollment.get('final_fee', 0)
        enrollment_id = enrollment.get('id')
        total_paid = payments_map.get(enrollment_id, 0)
        pending_amount = final_fee - total_paid
        
        if pending_amount <= 0:
            continue
        
        payment_plan = plans_map.get(enrollment_id)
        
        if payment_plan and payment_plan.get('plan_type') == PaymentPlanType.INSTALLMENTS.value:
            schedule = schedules_map.get(payment_plan['id'], [])
            
            for inst in schedule:
                due_date = inst.get('due_date', '')
                if start_date and due_date < start_date:
                    continue
                if end_date and due_date > end_date:
                    continue
                
                is_overdue = due_date < today
                
                pending_payments.append({
                    "type": "installment",
                    "enrollment_id": enrollment_id,
                    "student_name": enrollment.get('student_name', ''),
                    "student_phone": enrollment.get('phone', ''),
                    "student_email": enrollment.get('email', ''),
                    "program_name": enrollment.get('program_name', ''),
                    "installment_number": inst.get('installment_number'),
                    "amount": inst.get('amount'),
                    "due_date": due_date,
                    "is_overdue": is_overdue,
                    "payment_plan_id": payment_plan['id'],
                    "total_installments": payment_plan.get('installments_count', 0),
                    "total_fee": final_fee,
                    "total_paid": total_paid,
                    "pending_amount": pending_amount
                })
        else:
            enrollment_date = enrollment.get('enrollment_date', today)
            if start_date and enrollment_date < start_date:
                continue
            if end_date and enrollment_date > end_date:
                continue
            
            enrollment_date_obj = datetime.fromisoformat(enrollment_date).date() if isinstance(enrollment_date, str) else enrollment_date
            today_obj = datetime.fromisoformat(today).date() if isinstance(today, str) else today
            days_since_enrollment = (today_obj - enrollment_date_obj).days if hasattr(enrollment_date_obj, 'days') or isinstance(enrollment_date_obj, date) else 0
            is_overdue = days_since_enrollment > 30 and pending_amount > 0
            
            pending_payments.append({
                "type": "one_time",
                "enrollment_id": enrollment_id,
                "student_name": enrollment.get('student_name', ''),
                "student_phone": enrollment.get('phone', ''),
                "student_email": enrollment.get('email', ''),
                "program_name": enrollment.get('program_name', ''),
                "installment_number": None,
                "amount": pending_amount,
                "due_date": enrollment_date,
                "is_overdue": is_overdue,
                "payment_plan_id": None,
                "total_installments": 0,
                "total_fee": final_fee,
                "total_paid": total_paid,
                "pending_amount": pending_amount
            })
    
    # Sort by overdue first, then by due date
    pending_payments.sort(key=lambda x: (not x['is_overdue'], x['due_date']))
    
    return pending_payments

# Financial Analytics
@api_router.get("/analytics/financial/monthly")
async def get_monthly_financial_analytics(year: int = None, current_user: User = Depends(get_current_user)):
    """Get monthly income and expense data for charts"""
    if year is None:
        year = datetime.now().year
    
    # Build query based on user role
    branch_query = {}
    if current_user.role != UserRole.ADMIN:
        branch_query["branch_id"] = current_user.branch_id
    
    # Get all payments for the year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Get payments (income)
    payments = await db.payments.find({
        **branch_query,
        "payment_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(10000)
    
    # Get expenses
    expenses = await db.expenses.find({
        **branch_query,
        "expense_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(10000)
    
    # Aggregate by month
    monthly_data = {}
    for month in range(1, 13):
        month_str = f"{year}-{str(month).zfill(2)}"
        monthly_data[month_str] = {"month": month, "income": 0, "expenses": 0}
    
    for payment in payments:
        pay_date = payment.get('payment_date', '')
        if isinstance(pay_date, str):
            month_key = pay_date[:7]
        else:
            month_key = pay_date.strftime('%Y-%m')
        if month_key in monthly_data:
            monthly_data[month_key]["income"] += payment.get('amount', 0)
    
    for expense in expenses:
        exp_date = expense.get('expense_date', '')
        if isinstance(exp_date, str):
            month_key = exp_date[:7]
        else:
            month_key = exp_date.strftime('%Y-%m')
        if month_key in monthly_data:
            monthly_data[month_key]["expenses"] += expense.get('amount', 0)
    
    # Convert to list sorted by month
    result = sorted(monthly_data.values(), key=lambda x: x["month"])
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for item in result:
        item["month_name"] = month_names[item["month"] - 1]
    
    return {
        "year": year,
        "monthly_data": result,
        "total_income": sum(p.get('amount', 0) for p in payments),
        "total_expenses": sum(e.get('amount', 0) for e in expenses)
    }

@api_router.get("/analytics/admissions/monthly")
async def get_monthly_admission_stats(year: int = None, current_user: User = Depends(get_current_user)):
    """Get monthly admission (enrollment) statistics for charts - Branch Admin and Admin only"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Admin or Branch Admin can access this")
    
    if year is None:
        year = datetime.now().year
    
    # Build query based on user role
    branch_query = {}
    if current_user.role != UserRole.ADMIN:
        branch_query["branch_id"] = current_user.branch_id
    
    # Date range for the year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Get all enrollments for the year
    enrollments = await db.enrollments.find({
        **branch_query,
        "enrollment_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0, "enrollment_date": 1, "status": 1, "program_name": 1, "fee_quoted": 1}).to_list(10000)
    
    # Aggregate by month
    monthly_data = {}
    for month in range(1, 13):
        month_str = f"{year}-{str(month).zfill(2)}"
        monthly_data[month_str] = {
            "month": month, 
            "admissions": 0, 
            "active": 0,
            "completed": 0,
            "total_fee": 0
        }
    
    for enrollment in enrollments:
        enroll_date = enrollment.get('enrollment_date', '')
        if isinstance(enroll_date, str):
            month_key = enroll_date[:7]
        else:
            month_key = enroll_date.strftime('%Y-%m')
        
        if month_key in monthly_data:
            monthly_data[month_key]["admissions"] += 1
            monthly_data[month_key]["total_fee"] += enrollment.get('fee_quoted', 0)
            status = enrollment.get('status', 'Active')
            if status == 'Active':
                monthly_data[month_key]["active"] += 1
            elif status == 'Completed':
                monthly_data[month_key]["completed"] += 1
    
    # Convert to list sorted by month
    result = sorted(monthly_data.values(), key=lambda x: x["month"])
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for item in result:
        item["month_name"] = month_names[item["month"] - 1]
    
    # Get program-wise breakdown
    program_breakdown = {}
    for enrollment in enrollments:
        program = enrollment.get('program_name', 'Unknown')
        if program not in program_breakdown:
            program_breakdown[program] = 0
        program_breakdown[program] += 1
    
    return {
        "year": year,
        "monthly_data": result,
        "total_admissions": len(enrollments),
        "program_breakdown": program_breakdown
    }

@api_router.get("/analytics/financial/branch-wise")
async def get_branch_wise_financial(current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Get income and expenses for all branches - Admin only"""
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    
    branch_financials = []
    for branch in branches:
        branch_id = branch["id"]
        
        # Get total income from payments
        payments = await db.payments.find({"branch_id": branch_id}, {"_id": 0, "amount": 1}).to_list(10000)
        total_income = sum(p.get('amount', 0) for p in payments)
        
        # Get total expenses
        expenses = await db.expenses.find({"branch_id": branch_id}, {"_id": 0, "amount": 1}).to_list(10000)
        total_expenses = sum(e.get('amount', 0) for e in expenses)
        
        # Get enrollments count
        enrollments_count = await db.enrollments.count_documents({"branch_id": branch_id})
        
        branch_financials.append({
            "branch_id": branch_id,
            "branch_name": branch["name"],
            "branch_location": branch.get("location", ""),
            "total_income": total_income,
            "total_expenses": total_expenses,
            "net_profit": total_income - total_expenses,
            "enrollments_count": enrollments_count
        })
    
    # Sort by income descending
    branch_financials.sort(key=lambda x: x["total_income"], reverse=True)
    
    return branch_financials

# Marketing Resources Management
@api_router.post("/admin/resources", response_model=MarketingResource)
async def create_resource(resource: MarketingResourceCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Create a marketing resource - Admin only"""
    new_resource = MarketingResource(
        **resource.model_dump(),
        created_by=current_user.id
    )
    resource_dict = new_resource.model_dump()
    resource_dict['created_at'] = resource_dict['created_at'].isoformat()
    
    await db.marketing_resources.insert_one(resource_dict)
    return new_resource

@api_router.get("/resources", response_model=List[MarketingResource])
async def get_resources(current_user: User = Depends(get_current_user)):
    """Get all marketing resources - accessible to all users"""
    resources = await db.marketing_resources.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for res in resources:
        if isinstance(res.get('created_at'), str):
            res['created_at'] = datetime.fromisoformat(res['created_at'])
    return [MarketingResource(**r) for r in resources]

@api_router.delete("/admin/resources/{resource_id}")
async def delete_resource(resource_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Delete a marketing resource - Admin only"""
    result = await db.marketing_resources.delete_one({"id": resource_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Resource not found")
    return {"message": "Resource deleted successfully"}

# WhatsApp Settings Management
@api_router.get("/admin/whatsapp-settings")
async def get_whatsapp_settings_api(current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Get WhatsApp notification settings - Super Admin only"""
    settings = await get_whatsapp_settings()
    return settings

@api_router.post("/admin/whatsapp-test")
async def whatsapp_test_send(
    event_type: str = "enquiry_saved",
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Manually fire a WhatsApp test message for a given event.
    Honours WHATSAPP_TEST_NUMBER override. Returns the raw MSG91 response so we can debug.
    """
    sample_data = {
        "name": "Test User",
        "course": "Basics of Computer",
        "demo_date": datetime.now().strftime("%d %b %Y"),
        "demo_time": "11:00 AM",
        "trainer": "Test Trainer",
        "enrollment_number": "ETI-TEST-001",
        "amount": "5000",
        "total_fee": "15000",
        "paid_fee": "5000",
        "pending_fee": "10000",
        "receipt_number": "RCPT-TEST-001",
        "amount_due": "10000",
        "due_date": (datetime.now() + timedelta(days=7)).strftime("%d %b %Y"),
        "certificate_id": "ETI-CERT-TEST-001",
    }
    
    settings = await get_whatsapp_settings()
    event_config = settings.events.get(event_type, {})
    
    debug_info = {
        "event_type": event_type,
        "global_enabled": settings.enabled,
        "event_enabled": event_config.get("enabled", False),
        "template_name": event_config.get("template_name", ""),
        "namespace": event_config.get("namespace", ""),
        "integrated_number": settings.integrated_number,
        "msg91_key_configured": bool(os.environ.get("MSG91_AUTH_KEY")),
        "test_number_override": os.environ.get("WHATSAPP_TEST_NUMBER", "(not set)"),
        "recipient_used": os.environ.get("WHATSAPP_TEST_NUMBER", "917973706853"),
    }
    
    # Force-send even if global enabled is false, by calling send_whatsapp_template_with_config directly
    if not settings.enabled:
        # Temporarily flip enabled in memory just to allow this manual test
        settings.enabled = True
    
    result = await send_whatsapp_notification(
        phone_number=os.environ.get("WHATSAPP_TEST_NUMBER", "917973706853"),
        event_type=event_type,
        template_data=sample_data,
    )
    
    return {"debug": debug_info, "msg91_result": result}

@api_router.put("/admin/whatsapp-settings")
async def update_whatsapp_settings(
    settings_update: WhatsAppSettingsUpdate, 
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Update WhatsApp notification settings - Super Admin only"""
    existing = await db.whatsapp_settings.find_one({}, {"_id": 0})
    
    update_data = {k: v for k, v in settings_update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_data['updated_by'] = current_user.id
    
    if existing:
        await db.whatsapp_settings.update_one({}, {"$set": update_data})
    else:
        new_settings = WhatsAppSettings(**update_data)
        settings_dict = new_settings.model_dump()
        settings_dict['updated_at'] = settings_dict['updated_at'].isoformat()
        await db.whatsapp_settings.insert_one(settings_dict)
    
    return {"message": "WhatsApp settings updated successfully"}

# Super Admin Dashboard Analytics
@api_router.get("/analytics/super-admin-dashboard")
async def get_super_admin_dashboard(request: Request, session: Optional[str] = None, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Get comprehensive dashboard data for Super Admin"""
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    
    # Get session filter
    session_val = session or await get_session_from_request(request)
    leads_session_filter = get_session_filter(session_val, "created_at") if session_val and session_val != "all" else {}
    enrollment_session_filter = get_session_filter(session_val, "enrollment_date") if session_val and session_val != "all" else {}
    payment_session_filter = get_session_filter(session_val, "payment_date") if session_val and session_val != "all" else {}
    
    branch_data = []
    for branch in branches:
        branch_id = branch["id"]
        
        # Leads count with session filter
        leads_query = {"branch_id": branch_id, "is_deleted": {"$ne": True}, **leads_session_filter}
        leads_count = await db.leads.count_documents(leads_query)
        
        # Enrollments count with session filter
        enrollments_query = {"branch_id": branch_id, **enrollment_session_filter}
        enrollments_count = await db.enrollments.count_documents(enrollments_query)
        
        # Unique students count (distinct lead_ids) with session filter
        enrollments_for_students = await db.enrollments.find(
            enrollments_query, 
            {"_id": 0, "lead_id": 1}
        ).to_list(10000)
        unique_students = len(set(e.get("lead_id") for e in enrollments_for_students if e.get("lead_id")))
        
        # Total income with session filter
        payments_query = {"branch_id": branch_id, **payment_session_filter}
        payments = await db.payments.find(payments_query, {"_id": 0, "amount": 1}).to_list(10000)
        total_income = sum(p.get('amount', 0) for p in payments)
        
        # Converted leads with session filter
        converted_query = {"branch_id": branch_id, "status": "Converted", "is_deleted": {"$ne": True}, **leads_session_filter}
        converted_count = await db.leads.count_documents(converted_query)
        
        # Conversion rate
        conversion_rate = (converted_count / leads_count * 100) if leads_count > 0 else 0
        
        branch_data.append({
            "branch_id": branch_id,
            "branch_name": branch["name"],
            "branch_location": branch.get("location", ""),
            "leads_count": leads_count,
            "enrollments_count": enrollments_count,
            "unique_students": unique_students,
            "total_income": total_income,
            "converted_count": converted_count,
            "conversion_rate": round(conversion_rate, 1)
        })
    
    # Sort by income to determine performance
    branch_data.sort(key=lambda x: x["total_income"], reverse=True)
    
    # Calculate totals
    total_leads = sum(b["leads_count"] for b in branch_data)
    total_enrollments = sum(b["enrollments_count"] for b in branch_data)
    total_unique_students = sum(b["unique_students"] for b in branch_data)
    total_income = sum(b["total_income"] for b in branch_data)
    avg_income = total_income / len(branch_data) if branch_data else 0
    
    # Mark performance
    for branch in branch_data:
        if branch["total_income"] >= avg_income * 1.2:
            branch["performance"] = "outperforming"
        elif branch["total_income"] <= avg_income * 0.8:
            branch["performance"] = "underperforming"
        else:
            branch["performance"] = "average"
    
    return {
        "branches": branch_data,
        "totals": {
            "total_leads": total_leads,
            "total_enrollments": total_enrollments,
            "total_students": total_unique_students,
            "total_income": total_income,
            "average_income_per_branch": avg_income
        }
    }

@api_router.get("/analytics/session-comparison")
async def get_session_comparison(request: Request, session: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get session comparison data - current vs previous session metrics"""
    
    # Get current session from request or use default
    session_val = session or await get_session_from_request(request)
    if not session_val or session_val == "all":
        session_val = get_current_academic_session()
    
    current_session = int(session_val)
    prev_session = current_session - 1
    
    # Branch filter
    branch_filter = {}
    if current_user.role != UserRole.ADMIN:
        branch_filter["branch_id"] = current_user.branch_id
    
    # Get date ranges
    current_start, current_end = get_session_date_range(str(current_session))
    prev_start, prev_end = get_session_date_range(str(prev_session))
    
    # Current session metrics
    current_leads_query = {**branch_filter, "is_deleted": {"$ne": True}, "created_at": {"$gte": current_start.isoformat(), "$lte": current_end.isoformat()}}
    current_leads = await db.leads.count_documents(current_leads_query)
    
    current_converted_query = {**current_leads_query, "status": "Converted"}
    current_converted = await db.leads.count_documents(current_converted_query)
    
    current_enrollments_query = {**branch_filter, "enrollment_date": {"$gte": current_start.isoformat(), "$lte": current_end.isoformat()}}
    current_enrollments = await db.enrollments.count_documents(current_enrollments_query)
    
    current_payments_query = {**branch_filter, "payment_date": {"$gte": current_start.isoformat(), "$lte": current_end.isoformat()}}
    current_payments = await db.payments.find(current_payments_query, {"_id": 0, "amount": 1}).to_list(10000)
    current_income = sum(p.get('amount', 0) for p in current_payments)
    
    # Previous session metrics
    prev_leads_query = {**branch_filter, "is_deleted": {"$ne": True}, "created_at": {"$gte": prev_start.isoformat(), "$lte": prev_end.isoformat()}}
    prev_leads = await db.leads.count_documents(prev_leads_query)
    
    prev_converted_query = {**prev_leads_query, "status": "Converted"}
    prev_converted = await db.leads.count_documents(prev_converted_query)
    
    prev_enrollments_query = {**branch_filter, "enrollment_date": {"$gte": prev_start.isoformat(), "$lte": prev_end.isoformat()}}
    prev_enrollments = await db.enrollments.count_documents(prev_enrollments_query)
    
    prev_payments_query = {**branch_filter, "payment_date": {"$gte": prev_start.isoformat(), "$lte": prev_end.isoformat()}}
    prev_payments = await db.payments.find(prev_payments_query, {"_id": 0, "amount": 1}).to_list(10000)
    prev_income = sum(p.get('amount', 0) for p in prev_payments)
    
    # Calculate percentage changes
    def calc_change(current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return round(((current - previous) / previous) * 100, 1)
    
    # Conversion rates
    current_conversion_rate = round((current_converted / current_leads * 100), 1) if current_leads > 0 else 0
    prev_conversion_rate = round((prev_converted / prev_leads * 100), 1) if prev_leads > 0 else 0
    
    return {
        "current_session": {
            "year": current_session,
            "label": f"{current_session}-{str(current_session + 1)[2:]}",
            "period": f"Apr {current_session} - Mar {current_session + 1}",
            "leads": current_leads,
            "converted": current_converted,
            "conversion_rate": current_conversion_rate,
            "enrollments": current_enrollments,
            "income": current_income
        },
        "previous_session": {
            "year": prev_session,
            "label": f"{prev_session}-{str(prev_session + 1)[2:]}",
            "period": f"Apr {prev_session} - Mar {prev_session + 1}",
            "leads": prev_leads,
            "converted": prev_converted,
            "conversion_rate": prev_conversion_rate,
            "enrollments": prev_enrollments,
            "income": prev_income
        },
        "changes": {
            "leads": calc_change(current_leads, prev_leads),
            "converted": calc_change(current_converted, prev_converted),
            "conversion_rate": round(current_conversion_rate - prev_conversion_rate, 1),
            "enrollments": calc_change(current_enrollments, prev_enrollments),
            "income": calc_change(current_income, prev_income)
        }
    }

@api_router.get("/analytics/fde-dashboard")
async def get_fde_dashboard(request: Request, session: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get FDE (Front Desk Executive) dashboard data"""
    if current_user.role not in [UserRole.FRONT_DESK, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    branch_id = current_user.branch_id
    today = datetime.now().strftime('%Y-%m-%d')
    current_month = datetime.now().strftime('%Y-%m')
    
    # Get session filter
    session_val = session or await get_session_from_request(request)
    enrollment_session_filter = get_session_filter(session_val, "enrollment_date") if session_val and session_val != "all" else {}
    
    # 1. Fee due for the month (installments due this month)
    from calendar import monthrange
    year, month = datetime.now().year, datetime.now().month
    _, last_day = monthrange(year, month)
    month_start = f"{year}-{str(month).zfill(2)}-01"
    month_end = f"{year}-{str(month).zfill(2)}-{str(last_day).zfill(2)}"
    
    installments_due = await db.installment_schedule.find({
        "due_date": {"$gte": month_start, "$lte": month_end},
        "status": {"$ne": "Paid"}
    }, {"_id": 0}).to_list(1000)
    
    # Filter by branch if not super admin
    fee_due_amount = 0
    fee_due_count = 0
    for inst in installments_due:
        plan = await db.payment_plans.find_one({"id": inst.get('payment_plan_id')}, {"_id": 0, "branch_id": 1})
        if plan and (current_user.role == UserRole.ADMIN or plan.get('branch_id') == branch_id):
            fee_due_amount += inst.get('amount', 0)
            fee_due_count += 1
    
    # 2. Students not assigned to batches (with session filter)
    branch_filter = {} if current_user.role == UserRole.ADMIN else {"branch_id": branch_id}
    students_without_batch = await db.enrollments.count_documents({
        **branch_filter,
        **enrollment_session_filter,
        "status": "Active",
        "$or": [{"batch_id": {"$exists": False}}, {"batch_id": None}, {"batch_id": ""}]
    })
    
    # 3. Cash handling status for today
    cash_record = await db.cash_handling.find_one({
        "date": today,
        "branch_id": branch_id if current_user.role != UserRole.ADMIN else {"$exists": True}
    }, {"_id": 0})
    
    cash_updated_today = cash_record is not None and cash_record.get('status') == 'Deposited'
    
    # Get today's cash total from payments
    today_payments = await db.payments.find({
        **branch_filter,
        "payment_date": today,
        "payment_mode": "Cash"
    }, {"_id": 0, "amount": 1}).to_list(1000)
    today_cash_total = sum(p.get('amount', 0) for p in today_payments)
    
    # 4. Tasks not responded/pending
    pending_tasks = await db.tasks.count_documents({
        **branch_filter,
        "status": {"$in": ["Pending", "In Progress"]}
    })
    
    overdue_tasks = await db.tasks.count_documents({
        **branch_filter,
        "status": {"$in": ["Pending", "In Progress"]},
        "due_date": {"$lt": today}
    })
    
    return {
        "fee_due": {
            "amount": fee_due_amount,
            "count": fee_due_count,
            "month": datetime.now().strftime('%B %Y')
        },
        "students_without_batch": students_without_batch,
        "cash_handling": {
            "updated_today": cash_updated_today,
            "today_cash_total": today_cash_total
        },
        "tasks": {
            "pending": pending_tasks,
            "overdue": overdue_tasks
        },
        "overdue_payments": [],  # Will be populated below
        "ready_to_enroll": [],   # Leads that are converted but not enrolled
        "pending_exams": []      # Course completed but exam pending
    }

@api_router.get("/analytics/fde-dashboard-enhanced")
async def get_fde_dashboard_enhanced(request: Request, session: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Enhanced FDE dashboard with overdue payments, ready-to-enroll leads, and pending exams"""
    if current_user.role not in [UserRole.FRONT_DESK, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    branch_id = current_user.branch_id
    today = datetime.now().strftime('%Y-%m-%d')
    today_date = datetime.now().date()
    
    branch_filter = {} if current_user.role == UserRole.ADMIN else {"branch_id": branch_id}
    
    # 1. Students with overdue payments (with days overdue)
    overdue_payments = []
    installments = await db.installment_schedule.find({
        "due_date": {"$lt": today},
        "status": {"$ne": "Paid"}
    }, {"_id": 0}).to_list(500)
    
    for inst in installments:
        plan = await db.payment_plans.find_one({"id": inst.get('payment_plan_id')}, {"_id": 0})
        if plan and (current_user.role == UserRole.ADMIN or plan.get('branch_id') == branch_id):
            enrollment = await db.enrollments.find_one({"id": plan.get('enrollment_id')}, {"_id": 0})
            if enrollment:
                due_date = datetime.strptime(inst.get('due_date'), '%Y-%m-%d').date()
                days_overdue = (today_date - due_date).days
                overdue_payments.append({
                    "student_name": enrollment.get('student_name', 'Unknown'),
                    "contact_no": enrollment.get('contact_no', ''),
                    "amount": inst.get('amount', 0),
                    "due_date": inst.get('due_date'),
                    "days_overdue": days_overdue,
                    "enrollment_id": enrollment.get('id'),
                    "program_name": enrollment.get('program_name', '')
                })
    
    # Sort by days overdue (highest first)
    overdue_payments.sort(key=lambda x: x['days_overdue'], reverse=True)
    
    # 2. Leads ready to enroll (Converted status but no enrollment)
    ready_to_enroll = []
    converted_leads = await db.leads.find({
        **branch_filter,
        "status": "Converted",
        "is_deleted": {"$ne": True}
    }, {"_id": 0}).to_list(500)
    
    for lead in converted_leads:
        # Check if this lead has an enrollment
        enrollment = await db.enrollments.find_one({
            "lead_id": lead.get('id')
        }, {"_id": 0})
        if not enrollment:
            ready_to_enroll.append({
                "lead_id": lead.get('id'),
                "student_name": lead.get('name', 'Unknown'),
                "contact_no": lead.get('contact_no', ''),
                "email": lead.get('email', ''),
                "program_name": lead.get('program_name', ''),
                "converted_date": lead.get('updated_at', lead.get('created_at', ''))
            })
    
    # 3. Students with course completed but exam pending
    pending_exams = []
    completed_courses = await db.enrollments.find({
        **branch_filter,
        "course_status": "Completed"
    }, {"_id": 0}).to_list(500)
    
    for enrollment in completed_courses:
        # Check if exam is scheduled/completed
        exam_booking = await db.exam_bookings.find_one({
            "enrollment_id": enrollment.get('id'),
            "status": {"$in": ["Completed", "Passed"]}
        }, {"_id": 0})
        if not exam_booking:
            pending_exams.append({
                "student_name": enrollment.get('student_name', 'Unknown'),
                "contact_no": enrollment.get('contact_no', ''),
                "program_name": enrollment.get('program_name', ''),
                "enrollment_id": enrollment.get('id'),
                "completion_date": enrollment.get('updated_at', '')
            })
    
    return {
        "overdue_payments": overdue_payments[:20],  # Top 20
        "ready_to_enroll": ready_to_enroll[:15],    # Top 15
        "pending_exams": pending_exams[:15]          # Top 15
    }

@api_router.get("/analytics/counsellor-dashboard")
async def get_counsellor_dashboard(current_user: User = Depends(get_current_user)):
    """Get Counsellor dashboard data with follow-up info"""
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    branch_id = current_user.branch_id
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Build query based on role
    if current_user.role == UserRole.COUNSELLOR:
        followup_query = {"created_by": current_user.id, "status": "Pending"}
        today_query = {"created_by": current_user.id, "status": "Pending"}
    elif current_user.role == UserRole.BRANCH_ADMIN:
        followup_query = {"branch_id": branch_id, "status": "Pending"}
        today_query = {"branch_id": branch_id, "status": "Pending"}
    else:  # Admin
        followup_query = {"status": "Pending"}
        today_query = {"status": "Pending"}
    
    # Total pending follow-ups
    total_pending = await db.followups.count_documents(followup_query)
    
    # Today's follow-ups
    today_followups = await db.followups.find({
        **today_query,
        "followup_date": {"$regex": f"^{today}"}
    }, {"_id": 0}).to_list(100)
    
    # Overdue follow-ups (before today)
    overdue_followups = await db.followups.count_documents({
        **followup_query,
        "followup_date": {"$lt": today}
    })
    
    # This week's follow-ups
    from datetime import timedelta
    week_end = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    this_week_count = await db.followups.count_documents({
        **followup_query,
        "followup_date": {"$gte": today, "$lte": week_end}
    })
    
    # Get detailed today's follow-ups with lead info
    today_followups_detailed = []
    for fu in today_followups:
        lead = await db.leads.find_one({"id": fu.get('lead_id')}, {"_id": 0, "name": 1, "number": 1, "status": 1})
        if lead:
            today_followups_detailed.append({
                **fu,
                "lead_name": lead.get('name'),
                "lead_number": lead.get('number'),
                "lead_status": lead.get('status')
            })
    
    return {
        "total_pending": total_pending,
        "today_count": len(today_followups),
        "today_followups": today_followups_detailed,
        "overdue_count": overdue_followups,
        "this_week_count": this_week_count
    }

@api_router.get("/analytics/counsellor-dashboard-enhanced")
async def get_counsellor_dashboard_enhanced(request: Request, current_user: User = Depends(get_current_user)):
    """Enhanced Counsellor dashboard with lead stats, pending feedbacks, missed tasks.
    Respects X-Academic-Session header / session query param. "all" = no session filter."""
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")

    branch_id = current_user.branch_id
    today = datetime.now().strftime('%Y-%m-%d')

    # Session filters
    session_val = await get_session_from_request(request)
    leads_session_filter = get_session_filter(session_val, "created_at")
    enroll_session_filter = get_session_filter(session_val, "enrollment_date")

    # Build query based on role
    if current_user.role == UserRole.COUNSELLOR:
        # Counsellor sees leads assigned to them OR created by them
        lead_query = {
            "$or": [
                {"counsellor_id": current_user.id},
                {"created_by": current_user.id}
            ],
            "is_deleted": {"$ne": True},
            **leads_session_filter,
        }
        followup_query = {"created_by": current_user.id, "status": "Pending"}
        task_query = {"assigned_to": current_user.id}
    elif current_user.role == UserRole.BRANCH_ADMIN:
        lead_query = {"branch_id": branch_id, "is_deleted": {"$ne": True}, **leads_session_filter}
        followup_query = {"branch_id": branch_id, "status": "Pending"}
        task_query = {"branch_id": branch_id}
    else:  # Admin
        lead_query = {"is_deleted": {"$ne": True}, **leads_session_filter}
        followup_query = {"status": "Pending"}
        task_query = {}
    
    # 1. Lead Stats
    total_leads = await db.leads.count_documents(lead_query)
    total_converted = await db.leads.count_documents({**lead_query, "status": "Converted"})
    total_lost = await db.leads.count_documents({**lead_query, "status": "Lost"})
    conversion_rate = round((total_converted / total_leads * 100), 1) if total_leads > 0 else 0
    
    # 2. Today's Follow-ups
    today_followups = await db.followups.find({
        **followup_query,
        "followup_date": {"$regex": f"^{today}"}
    }, {"_id": 0}).to_list(100)
    
    today_followups_detailed = []
    for fu in today_followups:
        lead = await db.leads.find_one({"id": fu.get('lead_id')}, {"_id": 0, "name": 1, "number": 1, "status": 1, "program_name": 1})
        if lead:
            today_followups_detailed.append({
                "id": fu.get('id'),
                "lead_id": fu.get('lead_id'),
                "lead_name": lead.get('name'),
                "lead_number": lead.get('number'),
                "lead_status": lead.get('status'),
                "program": lead.get('program_name', ''),
                "note": fu.get('note', ''),
                "followup_date": fu.get('followup_date')
            })
    
    # 3. Overdue/Missed Follow-ups
    overdue_followups = await db.followups.find({
        **followup_query,
        "followup_date": {"$lt": today}
    }, {"_id": 0}).to_list(50)
    
    missed_followups = []
    for fu in overdue_followups:
        lead = await db.leads.find_one({"id": fu.get('lead_id')}, {"_id": 0, "name": 1, "number": 1, "status": 1})
        if lead:
            due_date = datetime.strptime(fu.get('followup_date', today)[:10], '%Y-%m-%d').date()
            days_missed = (datetime.now().date() - due_date).days
            missed_followups.append({
                "id": fu.get('id'),
                "lead_name": lead.get('name'),
                "lead_number": lead.get('number'),
                "lead_status": lead.get('status'),
                "followup_date": fu.get('followup_date'),
                "days_missed": days_missed,
                "note": fu.get('note', '')
            })
    
    missed_followups.sort(key=lambda x: x['days_missed'], reverse=True)
    
    # 4. Pending Feedback (students enrolled but feedback not submitted today)
    pending_feedbacks = []
    # Get enrollments by this counsellor where feedback might be due
    if current_user.role == UserRole.COUNSELLOR:
        recent_enrollments = await db.enrollments.find({
            "counsellor_id": current_user.id,
            "status": "Active",
            **enroll_session_filter,
        }, {"_id": 0}).sort("enrollment_date", -1).to_list(50)
    else:
        recent_enrollments = await db.enrollments.find({
            "branch_id": branch_id,
            "status": "Active",
            **enroll_session_filter,
        }, {"_id": 0}).sort("enrollment_date", -1).to_list(50)
    
    for enrollment in recent_enrollments:
        # Check if feedback exists for this student this month
        current_month_start = datetime.now().replace(day=1).strftime('%Y-%m-%d')
        feedback = await db.feedback.find_one({
            "enrollment_id": enrollment.get('id'),
            "created_at": {"$gte": current_month_start}
        }, {"_id": 0})
        if not feedback:
            # Check if student has been enrolled for at least 7 days
            enrollment_date = enrollment.get('enrollment_date', '')
            if enrollment_date:
                try:
                    enroll_dt = datetime.fromisoformat(enrollment_date[:10]) if isinstance(enrollment_date, str) else enrollment_date
                    days_enrolled = (datetime.now() - enroll_dt).days if hasattr(enroll_dt, 'days') else (datetime.now().date() - enroll_dt).days if hasattr(enroll_dt, 'month') else 0
                    if days_enrolled >= 7:
                        pending_feedbacks.append({
                            "student_name": enrollment.get('student_name', 'Unknown'),
                            "contact_no": enrollment.get('contact_no', ''),
                            "program_name": enrollment.get('program_name', ''),
                            "enrollment_id": enrollment.get('id'),
                            "days_enrolled": days_enrolled
                        })
                except:
                    pass
    
    # 5. Missed/Incomplete Tasks (from responsibilities)
    missed_tasks = []
    tasks = await db.responsibilities.find({
        **task_query,
        "status": {"$nin": ["Completed", "Cancelled"]},
        "$or": [
            {"due_date": {"$lt": today}},
            {"due_date": {"$exists": False}}
        ]
    }, {"_id": 0}).to_list(50)
    
    for task in tasks:
        due_date = task.get('due_date', '')
        days_overdue = 0
        if due_date:
            try:
                due_dt = datetime.strptime(due_date[:10], '%Y-%m-%d').date()
                days_overdue = (datetime.now().date() - due_dt).days
            except:
                pass
        missed_tasks.append({
            "id": task.get('id'),
            "title": task.get('title', ''),
            "description": task.get('description', ''),
            "priority": task.get('priority', 'Medium'),
            "due_date": due_date,
            "days_overdue": days_overdue
        })
    
    missed_tasks.sort(key=lambda x: x.get('days_overdue', 0), reverse=True)
    
    # 6. Counsellor Incentives (if applicable)
    incentive_data = None
    if current_user.role == UserRole.COUNSELLOR:
        # Get incentive stats for this counsellor (bookings are tracked via created_by)
        bookings = await db.exam_bookings.find({
            "created_by": current_user.id
        }, {"_id": 0}).to_list(1000)
        
        total_bookings = len(bookings)
        completed_exams = len([b for b in bookings if b.get('status') == 'Completed'])
        
        def _calc_incentive(b):
            amt = b.get('counsellor_incentive') or 0
            if not amt:
                amt = round((b.get('exam_price') or 0) * 0.10, 2)
            return amt
        
        # Earned incentive = all completed exams (both Earned & Paid statuses belong to the counsellor)
        earned_incentive = sum(
            _calc_incentive(b) for b in bookings
            if b.get('incentive_status') in ('Earned', 'Paid')
        )
        # Released incentive = amount actually paid out to the counsellor
        released_incentive = sum(
            _calc_incentive(b) for b in bookings
            if b.get('incentive_status') == 'Paid'
        )
        # Pending incentive = from bookings that are not yet completed (Pending/Confirmed)
        pending_incentive = sum(
            round((b.get('exam_price') or 0) * 0.10, 2)
            for b in bookings if b.get('status') in ('Pending', 'Confirmed')
        )
        
        incentive_data = {
            "total_bookings": total_bookings,
            "completed_exams": completed_exams,
            "earned_incentive": earned_incentive,
            "released_incentive": released_incentive,
            "pending_incentive": pending_incentive
        }
    
    pipeline_stages = await _build_counsellor_pipeline(lead_query)
    today_activity = await _build_counsellor_today_activity(current_user, lead_query, followup_query, today)

    # 7. Scheduled Demos (next 7 days — leads with demo_date set)
    week_end = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    scheduled_demos_raw = await db.leads.find({
        **lead_query,
        "demo_date": {"$gte": today, "$lte": week_end},
    }, {"_id": 0, "id": 1, "name": 1, "number": 1, "program_name": 1, "demo_date": 1,
        "demo_time": 1, "status": 1, "trainer_name": 1, "assigned_trainer": 1}).sort("demo_date", 1).to_list(100)

    scheduled_demos = []
    for l in scheduled_demos_raw:
        d_date = (l.get('demo_date') or '')[:10]
        try:
            days_away = (datetime.strptime(d_date, '%Y-%m-%d').date() - datetime.now().date()).days
        except Exception:
            days_away = None
        scheduled_demos.append({
            "lead_id": l.get('id'),
            "name": l.get('name'),
            "number": l.get('number'),
            "program": l.get('program_name', ''),
            "demo_date": d_date,
            "demo_time": l.get('demo_time', ''),
            "status": l.get('status', ''),
            "trainer": l.get('trainer_name') or l.get('assigned_trainer') or '',
            "days_away": days_away,
        })

    # 8. Upcoming Follow-ups (next 7 days, excluding today which is in today_followups)
    upcoming_followups_raw = await db.followups.find({
        **followup_query,
        "followup_date": {"$gt": today, "$lte": week_end},
    }, {"_id": 0}).sort("followup_date", 1).to_list(100)

    upcoming_followups = []
    for fu in upcoming_followups_raw:
        lead = await db.leads.find_one(
            {"id": fu.get('lead_id')},
            {"_id": 0, "name": 1, "number": 1, "status": 1, "program_name": 1}
        )
        if not lead:
            continue
        fu_date = (fu.get('followup_date') or '')[:10]
        try:
            days_away = (datetime.strptime(fu_date, '%Y-%m-%d').date() - datetime.now().date()).days
        except Exception:
            days_away = None
        upcoming_followups.append({
            "id": fu.get('id'),
            "lead_id": fu.get('lead_id'),
            "lead_name": lead.get('name'),
            "lead_number": lead.get('number'),
            "lead_status": lead.get('status'),
            "program": lead.get('program_name', ''),
            "note": fu.get('note', ''),
            "followup_date": fu.get('followup_date'),
            "days_away": days_away,
        })

    return {
        "lead_stats": {
            "total_leads": total_leads,
            "total_converted": total_converted,
            "total_lost": total_lost,
            "conversion_rate": conversion_rate
        },
        "today_followups": today_followups_detailed[:10],
        "upcoming_followups": upcoming_followups[:20],
        "scheduled_demos": scheduled_demos[:20],
        "missed_followups": missed_followups[:10],
        "pending_feedbacks": pending_feedbacks[:10],
        "missed_tasks": missed_tasks[:10],
        "incentive": incentive_data,
        "pipeline": pipeline_stages,
        "funnel": _build_funnel_from_pipeline(pipeline_stages),
        "hot_leads": await _build_counsellor_hot_leads(lead_query),
        "today_activity": today_activity,
        "weekly_trend": await _build_counsellor_weekly_trend(lead_query),
        "streak": _augment_streak_with_target(
            await _build_counsellor_streak(current_user),
            await _resolve_user_streak_target(current_user.id, "Counsellor"),
        ),
        "monthly_revenue": await _build_counsellor_monthly_revenue(current_user),
        "today_priority": await _build_counsellor_today_priority(current_user, lead_query, followup_query, today),
        "lead_aging": await _build_counsellor_lead_aging(lead_query),
        "scorecard": await _build_counsellor_scorecard(current_user, lead_query, followup_query, today, today_activity),
    }


@api_router.get("/analytics/counsellor-leaderboard")
async def get_counsellor_leaderboard(
    period: str = "month",
    metric: str = "admissions",
    current_user: User = Depends(get_current_user),
):
    """Branch-wide counsellor leaderboard. Period: month|week. Metric: admissions|revenue|conversion.
    Returns sorted list with `is_current_user` flag for the calling counsellor.
    """
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    if metric not in ("admissions", "revenue", "conversion"):
        metric = "admissions"
    if period not in ("month", "week"):
        period = "month"

    now = datetime.now(timezone.utc)
    if period == "week":
        start_str = (now.date() - timedelta(days=7)).isoformat()
    else:
        start_str = now.replace(day=1).date().isoformat()

    branch_filter = {}
    if current_user.role != UserRole.ADMIN:
        branch_filter["branch_id"] = current_user.branch_id

    counsellors = await db.users.find(
        {**branch_filter, "role": UserRole.COUNSELLOR.value, "is_active": {"$ne": False}},
        {"_id": 0, "id": 1, "name": 1, "email": 1},
    ).to_list(200)

    rows = []
    for c in counsellors:
        cid = c["id"]
        enrolls = await db.enrollments.find(
            {"counsellor_id": cid, "enrollment_date": {"$gte": start_str}},
            {"_id": 0, "final_fee": 1},
        ).to_list(2000)
        admissions = len(enrolls)
        revenue = sum((e.get("final_fee") or 0) for e in enrolls)
        leads_total = await db.leads.count_documents({
            "$or": [{"counsellor_id": cid}, {"created_by": cid}],
            "is_deleted": {"$ne": True},
            "created_at": {"$gte": start_str},
        })
        converted = await db.leads.count_documents({
            "$or": [{"counsellor_id": cid}, {"created_by": cid}],
            "status": {"$in": ["Converted", "Admitted"]},
            "updated_at": {"$gte": start_str},
        })
        conversion = round((converted / leads_total * 100), 1) if leads_total > 0 else 0.0
        rows.append({
            "counsellor_id": cid,
            "name": c.get("name"),
            "admissions": admissions,
            "revenue": revenue,
            "conversion_rate": conversion,
            "is_current_user": cid == current_user.id,
        })

    sort_key = {"admissions": "admissions", "revenue": "revenue", "conversion": "conversion_rate"}[metric]
    rows.sort(key=lambda r: r[sort_key], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return {"period": period, "metric": metric, "leaderboard": rows}


@api_router.get("/leads-search/universal")
async def universal_search_leads(
    q: str = "",
    limit: int = 20,
    current_user: User = Depends(get_current_user),
):
    """Counsellor-scoped universal search: name, phone, email, program, lead_id."""
    q = (q or "").strip()
    if len(q) < 2:
        return {"results": [], "query": q}

    base: dict = {"is_deleted": {"$ne": True}}
    if current_user.role == UserRole.COUNSELLOR:
        base["$or"] = [{"counsellor_id": current_user.id}, {"created_by": current_user.id}]
    elif current_user.role == UserRole.BRANCH_ADMIN:
        base["branch_id"] = current_user.branch_id
    # Admin: all leads

    import re
    safe = re.escape(q)
    text_or = [
        {"name": {"$regex": safe, "$options": "i"}},
        {"number": {"$regex": safe, "$options": "i"}},
        {"email": {"$regex": safe, "$options": "i"}},
        {"program_name": {"$regex": safe, "$options": "i"}},
        {"lead_id": {"$regex": safe, "$options": "i"}},
        {"id": q},
    ]
    if "$or" in base:
        # combine assignment filter with text search via $and
        query = {"$and": [{"$or": base.pop("$or")}, {"$or": text_or}], **base}
    else:
        query = {**base, "$or": text_or}

    leads = await db.leads.find(query, {"_id": 0, "id": 1, "lead_id": 1, "name": 1, "number": 1, "email": 1, "program_name": 1, "status": 1, "priority": 1, "lead_source": 1, "created_at": 1}).sort("updated_at", -1).limit(min(limit, 50)).to_list(min(limit, 50))
    return {"results": leads, "query": q, "count": len(leads)}


# --- Lead Conversion Management System: Stage Transitions + Kanban --------------

# Canonical pipeline stages (subset of LeadStatus enum) in counsellor-facing order
PIPELINE_STAGES = [
    "New",
    "Contact Attempted",
    "Connected",
    "Interested",
    "Counselling Scheduled",
    "Demo Scheduled",
    "Demo Attended",
    "Admission Likely",
    "Converted",
    "Lost",
]

# Required-field rules per target stage. Each entry: list of field names that
# must be present (and non-empty) on the lead (post-merge with payload).
STAGE_REQUIRED_FIELDS = {
    "Demo Scheduled":   ["demo_date", "program_id"],
    "Demo Booked":      ["demo_date", "program_id"],
    "Counselling Scheduled": ["counselling_date"],
    "Admission Likely": ["fee_quoted"],
    "Converted":        ["final_fee", "enrollment_date"],
    "Admitted":         ["final_fee", "enrollment_date"],
    "Lost":             ["lost_reason"],
    "Not Interested":   ["lost_reason"],
}


class LeadReengagePayload(BaseModel):
    reason: Optional[str] = None
    assign_to: Optional[str] = None  # Optionally reassign to a different counsellor
    target_stage: Optional[str] = "New"  # Default re-entry point


@api_router.post("/leads/{lead_id}/reengage")
async def reengage_lost_lead(
    lead_id: str,
    payload: LeadReengagePayload,
    current_user: User = Depends(get_current_user),
):
    """Branch Admin / Admin only: re-open a Lost or Not Interested lead so the team
    can pursue it again. Writes a `reengaged` audit entry.
    """
    if current_user.role not in [UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admins can re-engage lost leads")

    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id and lead.get("branch_id") != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Lead belongs to a different branch")

    if lead.get("status") not in ("Lost", "Not Interested"):
        raise HTTPException(status_code=400, detail=f"Only Lost / Not Interested leads can be re-engaged (current: {lead.get('status')})")

    target = (payload.target_stage or "New").strip()
    if target not in {s.value for s in LeadStatus}:
        raise HTTPException(status_code=400, detail=f"Invalid target stage: {target}")

    now_iso = datetime.now(timezone.utc).isoformat()
    update: dict = {"status": target, "updated_at": now_iso, "reengaged_at": now_iso, "lost_reason": None}
    if payload.assign_to:
        owner = await db.users.find_one({"id": payload.assign_to, "is_active": {"$ne": False}}, {"_id": 0, "id": 1, "name": 1, "role": 1, "branch_id": 1})
        if not owner:
            raise HTTPException(status_code=400, detail="assign_to user not found / inactive")
        if owner["role"] not in (UserRole.COUNSELLOR.value, UserRole.BRANCH_ADMIN.value):
            raise HTTPException(status_code=400, detail="assign_to must be a Counsellor or Branch Admin")
        update["counsellor_id"] = owner["id"]
        update["counsellor_name"] = owner["name"]

    # Recompute priority post-reopen
    update["priority"] = compute_lead_priority({**lead, **update})

    await db.leads.update_one({"id": lead_id}, {"$set": update})

    await db.leads_audit.insert_one(
        {
            "id": str(uuid.uuid4()),
            "lead_id": lead_id,
            "old_status": lead.get("status"),
            "new_status": target,
            "event_type": "reengaged",
            "changed_by": current_user.id,
            "changed_by_name": current_user.name or current_user.email,
            "reason": payload.reason or "",
            "assigned_to": payload.assign_to,
            "changed_at": now_iso,
            "created_at": now_iso,
        }
    )

    updated = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return {"ok": True, "lead": updated, "reengaged_at": now_iso, "new_status": target}


@api_router.get("/leads-search/lost")
async def list_lost_leads(
    days_min: int = 0,
    days_max: int = 365,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
):
    """List Lost / Not Interested leads scoped by role, optionally filtered by lost-age.
    Branch Admins use this to find candidates for re-engagement.
    """
    base: dict = {"is_deleted": {"$ne": True}, "status": {"$in": ["Lost", "Not Interested"]}}
    if current_user.role == UserRole.COUNSELLOR:
        base["$or"] = [{"counsellor_id": current_user.id}, {"created_by": current_user.id}]
    elif current_user.role == UserRole.BRANCH_ADMIN:
        base["branch_id"] = current_user.branch_id
    elif current_user.role not in [UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")

    leads = await db.leads.find(
        base,
        {"_id": 0, "id": 1, "lead_id": 1, "name": 1, "number": 1, "email": 1, "program_name": 1, "status": 1, "lost_reason": 1, "counsellor_name": 1, "counsellor_id": 1, "lead_source": 1, "updated_at": 1, "created_at": 1},
    ).sort("updated_at", -1).limit(min(limit, 500)).to_list(min(limit, 500))

    # Compute lost_days from updated_at
    now = datetime.now(timezone.utc)
    out = []
    for l in leads:
        try:
            dt = datetime.fromisoformat(str(l.get("updated_at") or "")[:19].replace("Z", ""))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            days = max(0, int((now - dt).total_seconds() // 86400))
        except Exception:
            days = 0
        if days_min <= days <= days_max:
            l["lost_days"] = days
            out.append(l)
    return {"results": out, "count": len(out)}


class LeadStageTransitionPayload(BaseModel):
    target_stage: str
    fields: Dict[str, Any] = {}
    note: Optional[str] = None


@api_router.post("/leads/{lead_id}/transition")
async def transition_lead_stage(
    lead_id: str,
    payload: LeadStageTransitionPayload,
    current_user: User = Depends(get_current_user),
):
    """Atomic stage transition with required-field validation + audit log."""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    # Authorisation — same rules as update_lead
    if current_user.role == UserRole.COUNSELLOR and lead.get("counsellor_id") != current_user.id and lead.get("created_by") != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id and lead.get("branch_id") != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")

    target = (payload.target_stage or "").strip()
    valid = {s.value for s in LeadStatus}
    if target not in valid:
        raise HTTPException(status_code=400, detail=f"Invalid target stage: {target}")

    # Counsellors cannot transition AWAY from terminal stages (Converted/Admitted/Lost equivalents)
    # Check this BEFORE required-field validation for clearer error messages.
    locked_for_counsellor = {"Converted", "Admitted"}
    if lead.get("status") in locked_for_counsellor and current_user.role == UserRole.COUNSELLOR:
        raise HTTPException(status_code=400, detail="Counsellors cannot edit converted leads. Contact Branch Admin.")

    # Validate required fields
    incoming = {k: v for k, v in (payload.fields or {}).items() if v not in (None, "")}
    merged = {**lead, **incoming}
    required = STAGE_REQUIRED_FIELDS.get(target, [])
    missing = [f for f in required if not merged.get(f)]
    if missing:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "missing_required_fields",
                "target_stage": target,
                "missing_fields": missing,
                "message": f"Stage '{target}' requires: {', '.join(missing)}",
            },
        )

    old_status = lead.get("status") or "New"
    now_iso = datetime.now(timezone.utc).isoformat()

    update_payload = {**incoming, "status": target, "updated_at": now_iso}
    # If program_id is being set, also cache program_name
    if "program_id" in incoming and incoming["program_id"]:
        prog = await db.programs.find_one({"id": incoming["program_id"]}, {"_id": 0})
        if prog:
            update_payload["program_name"] = prog.get("name", "")

    # Recompute priority if relevant fields changed
    if any(k in update_payload for k in ("status", "fee_quoted", "parent_status")):
        update_payload["priority"] = compute_lead_priority({**lead, **update_payload})

    await db.leads.update_one({"id": lead_id}, {"$set": update_payload})

    # Write audit entry
    await db.leads_audit.insert_one(
        {
            "id": str(uuid.uuid4()),
            "lead_id": lead_id,
            "old_status": old_status,
            "new_status": target,
            "changed_by": current_user.id,
            "changed_by_name": current_user.name or current_user.email,
            "reason": payload.note or "",
            "fields_set": list(incoming.keys()),
            "changed_at": now_iso,
            "created_at": now_iso,
        }
    )

    updated = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return {
        "ok": True,
        "lead_id": lead_id,
        "old_status": old_status,
        "new_status": target,
        "lead": updated,
        "transitioned_at": now_iso,
    }


@api_router.get("/leads-kanban/board")
async def get_leads_kanban(
    counsellor_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    """Kanban board: leads grouped by pipeline stage. Scoped by role."""
    base: Dict[str, Any] = {"is_deleted": {"$ne": True}}

    if current_user.role == UserRole.COUNSELLOR:
        base["$or"] = [{"counsellor_id": current_user.id}, {"created_by": current_user.id}]
    elif current_user.role == UserRole.BRANCH_ADMIN:
        base["branch_id"] = current_user.branch_id
        if counsellor_id:
            base["counsellor_id"] = counsellor_id
    elif current_user.role == UserRole.ADMIN:
        if counsellor_id:
            base["counsellor_id"] = counsellor_id

    project = {
        "_id": 0,
        "id": 1,
        "lead_id": 1,
        "name": 1,
        "number": 1,
        "email": 1,
        "program_name": 1,
        "status": 1,
        "priority": 1,
        "lead_source": 1,
        "counsellor_name": 1,
        "demo_date": 1,
        "fee_quoted": 1,
        "final_fee": 1,
        "created_at": 1,
        "updated_at": 1,
        "next_followup_date": 1,
    }
    leads = await db.leads.find(base, project).sort("updated_at", -1).to_list(2000)

    grouped: Dict[str, List[Dict[str, Any]]] = {s: [] for s in PIPELINE_STAGES}
    other: List[Dict[str, Any]] = []
    for l in leads:
        s = l.get("status") or "New"
        # Map legacy aliases
        if s == "Admitted":
            s = "Converted"
        if s == "Demo Booked":
            s = "Demo Scheduled"
        if s == "Not Interested":
            s = "Lost"
        if s in grouped:
            grouped[s].append(l)
        else:
            other.append(l)

    stages = [
        {
            "status": stage,
            "count": len(grouped.get(stage, [])),
            "leads": grouped.get(stage, [])[:50],  # cap per stage for payload size
            "is_won": stage == "Converted",
            "is_lost": stage == "Lost",
        }
        for stage in PIPELINE_STAGES
    ]
    return {
        "stages": stages,
        "total": len(leads),
        "other_count": len(other),
        "required_fields": STAGE_REQUIRED_FIELDS,
    }


async def _build_counsellor_streak(current_user: User) -> dict:
    """Consecutive-day streak: longest run of days ending today (or yesterday) on which
    this counsellor logged at least one activity (lead created, lead updated, followup
    created, enrollment created). Activity dates pulled from the last 60 days."""
    if current_user.role != UserRole.COUNSELLOR:
        return {"current": 0, "longest_30d": 0, "active_today": False}
    today = datetime.now(timezone.utc).date()
    since = (today - timedelta(days=60)).isoformat()

    def _dates_from(items: list, fields: list) -> set:
        out = set()
        for it in items:
            for f in fields:
                v = it.get(f)
                if not v:
                    continue
                s = str(v)[:10]
                if s >= since[:10]:
                    out.add(s)
        return out

    activity_dates: set = set()
    # Leads created or last updated by counsellor
    leads = await db.leads.find(
        {"$or": [{"counsellor_id": current_user.id}, {"created_by": current_user.id}], "is_deleted": {"$ne": True}},
        {"_id": 0, "created_at": 1, "updated_at": 1},
    ).to_list(2000)
    activity_dates |= _dates_from(leads, ["created_at", "updated_at"])
    # Followups created by counsellor
    followups = await db.followups.find({"created_by": current_user.id}, {"_id": 0, "created_at": 1, "completed_at": 1}).to_list(2000)
    activity_dates |= _dates_from(followups, ["created_at", "completed_at"])
    # Enrollments created by counsellor
    enrolls = await db.enrollments.find({"counsellor_id": current_user.id}, {"_id": 0, "enrollment_date": 1, "created_at": 1}).to_list(2000)
    activity_dates |= _dates_from(enrolls, ["enrollment_date", "created_at"])

    today_str = today.isoformat()
    active_today = today_str in activity_dates

    # Current streak — count backwards from today (or yesterday if not active today)
    cur = 0
    cursor_day = today if active_today else (today - timedelta(days=1))
    while cursor_day.isoformat() in activity_dates:
        cur += 1
        cursor_day = cursor_day - timedelta(days=1)

    # Longest streak in last 30 days
    longest = 0
    run = 0
    for i in range(29, -1, -1):
        d = (today - timedelta(days=i)).isoformat()
        if d in activity_dates:
            run += 1
            longest = max(longest, run)
        else:
            run = 0

    return {"current": cur, "longest_30d": longest, "active_today": active_today}


async def _build_counsellor_monthly_revenue(current_user: User) -> dict:
    """This-month enrollment revenue handled by this counsellor + monthly target.
    Target is read from user.monthly_target (defaults to 500000)."""
    if current_user.role != UserRole.COUNSELLOR:
        return {"this_month": 0, "target": 0, "percent": 0}
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).date().isoformat()
    enrolls = await db.enrollments.find({
        "counsellor_id": current_user.id,
        "enrollment_date": {"$gte": month_start},
    }, {"_id": 0, "final_fee": 1}).to_list(2000)
    revenue = sum((e.get("final_fee") or 0) for e in enrolls)
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0, "monthly_target": 1})
    target = int((user_doc or {}).get("monthly_target") or 500000)
    pct = round((revenue / target * 100), 1) if target > 0 else 0
    return {"this_month": revenue, "target": target, "percent": pct, "enrollments": len(enrolls)}


async def _build_user_streak(user_id: str) -> dict:
    """Generic activity streak: any of lead create/update, followup, enrollment.
    Same logic as counsellor streak but role-agnostic — works for FDE & Placement too."""
    today = datetime.now(timezone.utc).date()
    since = (today - timedelta(days=60)).isoformat()[:10]

    def _dates_from(items: list, fields: list) -> set:
        out = set()
        for it in items:
            for f in fields:
                v = it.get(f)
                if not v:
                    continue
                s = str(v)[:10]
                if s >= since:
                    out.add(s)
        return out

    activity_dates: set = set()
    leads = await db.leads.find({"$or": [{"counsellor_id": user_id}, {"created_by": user_id}], "is_deleted": {"$ne": True}}, {"_id": 0, "created_at": 1, "updated_at": 1}).to_list(2000)
    activity_dates |= _dates_from(leads, ["created_at", "updated_at"])
    followups = await db.followups.find({"created_by": user_id}, {"_id": 0, "created_at": 1, "completed_at": 1}).to_list(2000)
    activity_dates |= _dates_from(followups, ["created_at", "completed_at"])
    enrolls = await db.enrollments.find({"counsellor_id": user_id}, {"_id": 0, "enrollment_date": 1, "created_at": 1}).to_list(2000)
    activity_dates |= _dates_from(enrolls, ["enrollment_date", "created_at"])

    today_str = today.isoformat()
    active_today = today_str in activity_dates
    cur = 0
    cursor_day = today if active_today else (today - timedelta(days=1))
    while cursor_day.isoformat() in activity_dates:
        cur += 1
        cursor_day = cursor_day - timedelta(days=1)
    longest = 0
    run = 0
    for i in range(29, -1, -1):
        d = (today - timedelta(days=i)).isoformat()
        if d in activity_dates:
            run += 1
            longest = max(longest, run)
        else:
            run = 0
    return {"current": cur, "longest_30d": longest, "active_today": active_today}


# ========== STREAK TARGETS (Phase F) ==========
_STREAK_TARGET_DEFAULTS = {
    "Counsellor": {"target_days": 7, "longest_target": 30},
    "Branch Admin": {"target_days": 14, "longest_target": 30},
    "Front Desk Executive": {"target_days": 5, "longest_target": 30},
    "Placement Manager": {"target_days": 5, "longest_target": 30},
}


async def _get_role_streak_target(role: str) -> dict:
    """Role-default streak target read from db (falls back to defaults)."""
    doc = await db.streak_targets.find_one({"role": role}, {"_id": 0})
    if doc:
        return {
            "target_days": int(doc.get("target_days") or 7),
            "longest_target": int(doc.get("longest_target") or 30),
        }
    return _STREAK_TARGET_DEFAULTS.get(role, {"target_days": 7, "longest_target": 30})


async def _resolve_user_streak_target(user_id: str, role: str) -> dict:
    """Per-user override (user.streak_target_days) takes priority over role default."""
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0, "streak_target_days": 1}) or {}
    role_default = await _get_role_streak_target(role)
    if user_doc.get("streak_target_days"):
        return {
            "target_days": int(user_doc["streak_target_days"]),
            "longest_target": role_default["longest_target"],
        }
    return role_default


def _augment_streak_with_target(streak: dict, target: dict) -> dict:
    """Attach target_days, target_met, target_progress_pct fields to a streak dict."""
    out = dict(streak or {})
    target_days = int(target.get("target_days") or 0)
    cur = int(out.get("current") or 0)
    out["target_days"] = target_days
    out["target_met"] = cur >= target_days if target_days > 0 else False
    out["target_progress_pct"] = (
        min(100, round((cur / target_days) * 100)) if target_days > 0 else 0
    )
    return out


async def _build_branch_streak(branch_id: str) -> dict:
    """Branch-wide activity streak — counts days on which any user did something at this branch.
    Activities considered: leads created/updated, enrollments, payments, follow-ups."""
    if not branch_id:
        return {"current": 0, "longest_30d": 0, "active_today": False}
    today = datetime.now(timezone.utc).date()
    since = (today - timedelta(days=60)).isoformat()[:10]
    activity_dates: set = set()

    def _add(items, fields):
        for it in items:
            for f in fields:
                v = it.get(f)
                if not v:
                    continue
                s = str(v)[:10]
                if s >= since:
                    activity_dates.add(s)

    leads = await db.leads.find(
        {"branch_id": branch_id, "is_deleted": {"$ne": True}},
        {"_id": 0, "created_at": 1, "updated_at": 1},
    ).to_list(5000)
    _add(leads, ["created_at", "updated_at"])
    enrolls = await db.enrollments.find(
        {"branch_id": branch_id},
        {"_id": 0, "enrollment_date": 1, "created_at": 1},
    ).to_list(5000)
    _add(enrolls, ["enrollment_date", "created_at"])
    payments = await db.payments.find(
        {"branch_id": branch_id},
        {"_id": 0, "payment_date": 1, "created_at": 1},
    ).to_list(5000)
    _add(payments, ["payment_date", "created_at"])
    followups = await db.followups.find(
        {"branch_id": branch_id},
        {"_id": 0, "created_at": 1, "completed_at": 1},
    ).to_list(5000)
    _add(followups, ["created_at", "completed_at"])

    today_str = today.isoformat()
    active_today = today_str in activity_dates
    cur = 0
    cursor_day = today if active_today else (today - timedelta(days=1))
    while cursor_day.isoformat() in activity_dates:
        cur += 1
        cursor_day = cursor_day - timedelta(days=1)
    longest = 0
    run = 0
    for i in range(29, -1, -1):
        d = (today - timedelta(days=i)).isoformat()
        if d in activity_dates:
            run += 1
            longest = max(longest, run)
        else:
            run = 0
    return {"current": cur, "longest_30d": longest, "active_today": active_today}


@api_router.get("/admin/streak-targets")
async def get_all_streak_targets(current_user: User = Depends(get_current_user)):
    """List all role-default streak targets. Super Admin only."""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Super Admin only")
    docs = await db.streak_targets.find({}, {"_id": 0}).to_list(50)
    docs_by_role = {d["role"]: d for d in docs if d.get("role")}
    targetable_roles = ["Counsellor", "Branch Admin", "Front Desk Executive", "Placement Manager"]
    result = []
    for role in targetable_roles:
        if role in docs_by_role:
            d = docs_by_role[role]
            result.append({
                "role": role,
                "target_days": int(d.get("target_days") or _STREAK_TARGET_DEFAULTS[role]["target_days"]),
                "longest_target": int(d.get("longest_target") or 30),
                "updated_by_name": d.get("updated_by_name"),
                "updated_at": d.get("updated_at"),
            })
        else:
            result.append({
                "role": role,
                "target_days": _STREAK_TARGET_DEFAULTS[role]["target_days"],
                "longest_target": _STREAK_TARGET_DEFAULTS[role]["longest_target"],
                "updated_by_name": None,
                "updated_at": None,
            })
    return {"targets": result}


class StreakTargetUpdate(BaseModel):
    target_days: int = Field(ge=1, le=365)
    longest_target: Optional[int] = Field(default=30, ge=1, le=365)


@api_router.put("/admin/streak-targets/{role}")
async def update_role_streak_target(role: str, payload: StreakTargetUpdate, current_user: User = Depends(get_current_user)):
    """Update streak target for a role. Super Admin only."""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Super Admin only")
    if role not in _STREAK_TARGET_DEFAULTS:
        raise HTTPException(status_code=400, detail=f"Streak targets not supported for role: {role}")
    await db.streak_targets.update_one(
        {"role": role},
        {"$set": {
            "role": role,
            "target_days": payload.target_days,
            "longest_target": payload.longest_target or 30,
            "updated_by": current_user.id,
            "updated_by_name": current_user.name,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    return {"message": "Streak target updated", "role": role, "target_days": payload.target_days}


class UserStreakOverride(BaseModel):
    streak_target_days: Optional[int] = Field(default=None, ge=0, le=365)


@api_router.put("/admin/users/{user_id}/streak-target")
async def set_user_streak_target_override(user_id: str, payload: UserStreakOverride, current_user: User = Depends(get_current_user)):
    """Per-user streak target override. Pass null/0 to remove. Super Admin only."""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Super Admin only")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "id": 1, "role": 1})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.streak_target_days and payload.streak_target_days > 0:
        await db.users.update_one({"id": user_id}, {"$set": {"streak_target_days": payload.streak_target_days}})
    else:
        await db.users.update_one({"id": user_id}, {"$unset": {"streak_target_days": ""}})
    return {"message": "User streak override saved", "user_id": user_id, "streak_target_days": payload.streak_target_days}


@api_router.get("/admin/users/streak-overrides")
async def list_user_streak_overrides(current_user: User = Depends(get_current_user)):
    """List all users who have an individual streak override. Super Admin only."""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Super Admin only")
    users = await db.users.find(
        {"streak_target_days": {"$exists": True, "$ne": None}},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1, "streak_target_days": 1, "branch_id": 1},
    ).to_list(500)
    return {"overrides": users}


@api_router.get("/branch-admin/branch-streak")
async def get_branch_streak(current_user: User = Depends(get_current_user)):
    """Branch-wide activity streak for the caller's branch. Branch Admin / Admin only."""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Branch Admin only")
    branch_id = current_user.branch_id
    if not branch_id and current_user.role == UserRole.ADMIN:
        first_branch = await db.branches.find_one({}, {"_id": 0, "id": 1})
        branch_id = first_branch["id"] if first_branch else None
    streak = await _build_branch_streak(branch_id)
    tgt = await _resolve_user_streak_target(current_user.id, "Branch Admin")
    return _augment_streak_with_target(streak, tgt)


@api_router.get("/analytics/role-hero")
async def get_role_hero(current_user: User = Depends(get_current_user)):
    """Lightweight hero data tailored to the caller's role — used by FDE & Placement Manager
    dashboards (counsellor + branch admin have richer dedicated endpoints).
    Returns: { kpis: [...], target: {label, current, target, percent}, streak: {...} }
    """
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).date().isoformat()
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0}) or {}

    streak = await _build_user_streak(current_user.id)
    streak = _augment_streak_with_target(
        streak, await _resolve_user_streak_target(current_user.id, current_user.role)
    )

    if current_user.role == UserRole.FRONT_DESK:
        # FDE — focus on lead capture
        scope = {"branch_id": current_user.branch_id, "is_deleted": {"$ne": True}} if current_user.branch_id else {"is_deleted": {"$ne": True}}
        new_today = await db.leads.count_documents({**scope, "created_by": current_user.id, "created_at": {"$regex": f"^{today}"}})
        new_month = await db.leads.count_documents({**scope, "created_by": current_user.id, "created_at": {"$gte": month_start}})
        demos_scheduled_today = await db.leads.count_documents({**scope, "demo_date": today})
        converted_month = await db.leads.count_documents({**scope, "created_by": current_user.id, "status": {"$in": ["Converted", "Admitted"]}, "updated_at": {"$gte": month_start}})
        target = int(user_doc.get("monthly_target_leads") or 50)
        pct = round((new_month / target * 100), 1) if target > 0 else 0
        return {
            "role": "Front Desk Executive",
            "kpis": [
                {"label": "New leads today", "value": new_today, "icon": "user-plus", "color": "sky"},
                {"label": "New leads this month", "value": new_month, "icon": "users", "color": "blue"},
                {"label": "Demos today (branch)", "value": demos_scheduled_today, "icon": "calendar", "color": "amber"},
                {"label": "Converted this month", "value": converted_month, "icon": "trophy", "color": "emerald"},
            ],
            "target": {"label": "Leads target this month", "current": new_month, "target": target, "percent": pct, "unit": "count"},
            "streak": streak,
        }

    if current_user.role == UserRole.PLACEMENT_MANAGER:
        scope = {"branch_id": current_user.branch_id} if current_user.branch_id else {}
        placed_today = await db.placements.count_documents({**scope, "placement_date": {"$regex": f"^{today}"}}) if "placements" in await db.list_collection_names() else 0
        placed_month = 0
        avg_ctc = 0
        try:
            placements_this_month = await db.placements.find({**scope, "placement_date": {"$gte": month_start}}, {"_id": 0, "salary_lpa": 1}).to_list(2000)
            placed_month = len(placements_this_month)
            if placed_month:
                avg_ctc = round(sum((p.get("salary_lpa") or 0) for p in placements_this_month) / placed_month, 2)
        except Exception:
            pass
        pending_eligible = await db.enrollments.count_documents({**scope, "status": "Completed", "placement_status": {"$ne": "Placed"}})
        target = int(user_doc.get("monthly_target_placements") or 10)
        pct = round((placed_month / target * 100), 1) if target > 0 else 0
        return {
            "role": "Placement Manager",
            "kpis": [
                {"label": "Placed today", "value": placed_today, "icon": "trophy", "color": "amber"},
                {"label": "Placed this month", "value": placed_month, "icon": "graduation-cap", "color": "emerald"},
                {"label": "Avg CTC (LPA)", "value": f"₹{avg_ctc}L" if avg_ctc else "—", "icon": "dollar-sign", "color": "sky"},
                {"label": "Pending eligible", "value": pending_eligible, "icon": "users", "color": "violet"},
            ],
            "target": {"label": "Placements target this month", "current": placed_month, "target": target, "percent": pct, "unit": "count"},
            "streak": streak,
        }

    # Default fallback for any other role calling this
    return {
        "role": current_user.role,
        "kpis": [],
        "target": None,
        "streak": streak,
    }




async def _build_counsellor_pipeline(lead_query: dict) -> list:
    """Returns a list of {status, count} for the 7 main pipeline stages — used for the funnel viz."""
    stages = ["New", "Contacted", "Interested", "Demo Booked", "Demo Attended", "Admission Likely", "Converted"]
    out = []
    for stage in stages:
        count = await db.leads.count_documents({**lead_query, "status": {"$regex": f"^{stage}$", "$options": "i"}})
        out.append({"status": stage, "count": count})
    return out


async def _build_counsellor_hot_leads(lead_query: dict) -> list:
    """Top 5 high-priority leads needing immediate attention (priority=High, not yet converted/lost)."""
    leads = await db.leads.find({
        **lead_query,
        "priority": "High",
        "status": {"$nin": ["Converted", "Admitted", "Lost", "Not Interested"]},
    }, {"_id": 0}).sort("updated_at", -1).limit(5).to_list(5)
    return [
        {
            "id": l.get("id"),
            "lead_id": l.get("lead_id"),
            "name": l.get("name"),
            "number": l.get("number"),
            "program_name": l.get("program_name"),
            "status": l.get("status"),
            "lead_source": l.get("lead_source"),
            "updated_at": l.get("updated_at"),
        }
        for l in leads
    ]


async def _build_counsellor_today_activity(current_user: User, lead_query: dict, followup_query: dict, today: str) -> dict:
    """Quick 'what I did today' counters for the activity strip."""
    new_leads_today = await db.leads.count_documents({
        **lead_query,
        "created_at": {"$regex": f"^{today}"},
    })
    followups_completed_today = await db.followups.count_documents({
        **{k: v for k, v in followup_query.items() if k != "status"},  # ignore status filter
        "status": {"$in": ["Completed", "Done"]},
        "completed_at": {"$regex": f"^{today}"},
    })
    demos_today = await db.leads.count_documents({
        **lead_query,
        "demo_date": today,
    })
    conversions_today = await db.leads.count_documents({
        **lead_query,
        "status": {"$in": ["Converted", "Admitted"]},
        "updated_at": {"$regex": f"^{today}"},
    })
    return {
        "new_leads_today": new_leads_today,
        "followups_completed_today": followups_completed_today,
        "demos_today": demos_today,
        "conversions_today": conversions_today,
    }


async def _build_counsellor_weekly_trend(lead_query: dict) -> dict:
    """Last-7-days vs prior-7-days conversion comparison for a momentum indicator."""
    today = datetime.now(timezone.utc).date()
    week_start = today - timedelta(days=7)
    prev_start = today - timedelta(days=14)

    def _range_q(start, end):
        return {**lead_query, "updated_at": {"$gte": start.isoformat(), "$lt": end.isoformat()}}

    this_week_conv = await db.leads.count_documents({
        **_range_q(week_start, today + timedelta(days=1)),
        "status": {"$in": ["Converted", "Admitted"]},
    })
    last_week_conv = await db.leads.count_documents({
        **_range_q(prev_start, week_start),
        "status": {"$in": ["Converted", "Admitted"]},
    })
    delta = this_week_conv - last_week_conv
    pct = round(((this_week_conv - last_week_conv) / last_week_conv * 100), 1) if last_week_conv > 0 else None
    return {
        "this_week_conversions": this_week_conv,
        "last_week_conversions": last_week_conv,
        "delta": delta,
        "percent_change": pct,
    }


# --- Counsellor Dashboard Enhancements (Today's Priority, Lead Aging, Funnel, Scorecard) ---

DEFAULT_DAILY_CALL_TARGET = 30
DEFAULT_MONTHLY_ADMISSION_TARGET = 10
ACTIVE_LEAD_STATUSES_EXCLUDE = ["Converted", "Admitted", "Lost", "Not Interested"]


async def _build_counsellor_lead_aging(lead_query: dict) -> dict:
    """Bucket open leads (not converted/lost) by created_at age into 0-24h, 1-3d, 3-7d, 7+d."""
    now = datetime.now(timezone.utc)
    open_q = {**lead_query, "status": {"$nin": ACTIVE_LEAD_STATUSES_EXCLUDE}}
    leads = await db.leads.find(open_q, {"_id": 0, "id": 1, "lead_id": 1, "name": 1, "number": 1, "program_name": 1, "status": 1, "created_at": 1, "priority": 1}).sort("created_at", -1).to_list(2000)
    buckets = {"bucket_0_24h": [], "bucket_1_3d": [], "bucket_3_7d": [], "bucket_7_plus": []}
    for l in leads:
        created = l.get("created_at")
        if not created:
            continue
        try:
            created_dt = datetime.fromisoformat(str(created)[:19].replace("Z", ""))
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=timezone.utc)
        except Exception:
            continue
        age_hours = (now - created_dt).total_seconds() / 3600
        if age_hours < 24:
            buckets["bucket_0_24h"].append(l)
        elif age_hours < 72:
            buckets["bucket_1_3d"].append(l)
        elif age_hours < 168:
            buckets["bucket_3_7d"].append(l)
        else:
            buckets["bucket_7_plus"].append(l)
    return {
        "bucket_0_24h": len(buckets["bucket_0_24h"]),
        "bucket_1_3d": len(buckets["bucket_1_3d"]),
        "bucket_3_7d": len(buckets["bucket_3_7d"]),
        "bucket_7_plus": len(buckets["bucket_7_plus"]),
        "sample_7_plus": buckets["bucket_7_plus"][:5],
        "sample_3_7d": buckets["bucket_3_7d"][:5],
    }


def _build_funnel_from_pipeline(pipeline: list) -> dict:
    """Adds inter-stage drop-off % to pipeline counts."""
    stages = []
    for i, s in enumerate(pipeline):
        entry = {"status": s["status"], "count": s["count"], "drop_off_pct": None, "drop_off_count": 0}
        if i > 0:
            prev = pipeline[i - 1]["count"]
            if prev > 0:
                drop = max(0, prev - s["count"])
                entry["drop_off_count"] = drop
                entry["drop_off_pct"] = round((drop / prev) * 100, 1)
        stages.append(entry)
    top = pipeline[0]["count"] if pipeline else 0
    bottom = pipeline[-1]["count"] if pipeline else 0
    overall_pct = round((bottom / top) * 100, 1) if top > 0 else 0
    return {"stages": stages, "overall_conversion_pct": overall_pct}


async def _build_counsellor_today_priority(current_user: User, lead_query: dict, followup_query: dict, today: str) -> dict:
    """Top-of-dashboard actionable counters for the Counsellor."""
    # 1. Hot leads pending (High priority, not converted/lost, not touched today)
    hot_pending = await db.leads.count_documents({
        **lead_query,
        "priority": "High",
        "status": {"$nin": ACTIVE_LEAD_STATUSES_EXCLUDE},
        "updated_at": {"$not": {"$regex": f"^{today}"}},
    })
    # 2. Follow-ups due today
    followups_due = await db.followups.count_documents({**followup_query, "followup_date": {"$regex": f"^{today}"}})
    # Also include overdue
    overdue_count = await db.followups.count_documents({**followup_query, "followup_date": {"$lt": today}})
    # 3. Calls pending = open leads (New/Contacted/Interested) not touched today
    calls_pending = await db.leads.count_documents({
        **lead_query,
        "status": {"$in": ["New", "Contacted", "Interested", "Follow-up", "Follow Up"]},
        "updated_at": {"$not": {"$regex": f"^{today}"}},
    })
    # 4. Admissions needed = monthly_admission_target - admissions_this_month
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0, "monthly_admission_target": 1}) or {}
    target_admissions = int(user_doc.get("monthly_admission_target") or DEFAULT_MONTHLY_ADMISSION_TARGET)
    month_start = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    if current_user.role == UserRole.COUNSELLOR:
        admissions_this_month = await db.enrollments.count_documents({
            "counsellor_id": current_user.id,
            "enrollment_date": {"$gte": month_start},
        })
    else:
        admissions_this_month = 0
    admissions_needed = max(0, target_admissions - admissions_this_month)
    return {
        "hot_leads_pending": hot_pending,
        "calls_pending": calls_pending,
        "followups_due_today": followups_due,
        "followups_overdue": overdue_count,
        "admissions_needed": admissions_needed,
        "admissions_this_month": admissions_this_month,
        "monthly_admission_target": target_admissions,
    }


async def _build_counsellor_scorecard(current_user: User, lead_query: dict, followup_query: dict, today: str, today_activity: dict) -> dict:
    """Today's KPIs vs daily targets for the counsellor performance scorecard."""
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0, "daily_call_target": 1}) or {}
    call_target = int(user_doc.get("daily_call_target") or DEFAULT_DAILY_CALL_TARGET)
    # Calls made today = total followups touched today (Pending touched + Completed)
    calls_made_today = await db.followups.count_documents({
        **{k: v for k, v in followup_query.items() if k != "status"},
        "$or": [
            {"updated_at": {"$regex": f"^{today}"}},
            {"created_at": {"$regex": f"^{today}"}},
            {"completed_at": {"$regex": f"^{today}"}},
        ],
    })
    return {
        "calls": {"made": calls_made_today, "target": call_target, "pct": min(100, round((calls_made_today / call_target * 100), 1)) if call_target > 0 else 0},
        "followups": {"completed": today_activity.get("followups_completed_today", 0), "target": max(5, call_target // 3), "pct": 0},
        "demos": {"scheduled": today_activity.get("demos_today", 0), "target": 2, "pct": 0},
        "admissions": {"today": today_activity.get("conversions_today", 0), "target": 1, "pct": 0},
        "new_leads_today": today_activity.get("new_leads_today", 0),
    }



# ----------------------------------------------------------------------------
# Phase B — Counsellor Action-Required Dashboard
# Returns Lead Summary, Follow-Up Summary, Admission Summary, Conversion Summary,
# Action-Required lists, and a Daily Work Queue. Built so Counsellors land on a
# screen that answers "what should I do next to convert leads?".
# ----------------------------------------------------------------------------
ADMISSION_TARGET_PCT = 0.70  # global default — 70% of leads assigned this month

def _classify_lead_score(lead: dict) -> str:
    """Hot / Warm / Cold / Lost — simple heuristic based on existing fields.

    Hot:  Demo Booked OR has fee_quoted (fee discussion happened) OR Converted
    Warm: Contacted, Follow-up
    Cold: New
    Lost: Lost
    """
    status = lead.get('status') or ''
    if status == 'Lost':
        return 'Lost'
    if status in ('Demo Booked', 'Demo Scheduled', 'Demo Attended', 'Admission Likely', 'Counselling Completed', 'Converted', 'Admitted') or (lead.get('fee_quoted') or 0) > 0:
        return 'Hot'
    if status in ('Contacted', 'Connected', 'Interested', 'Follow-up', 'Counselling Scheduled'):
        return 'Warm'
    return 'Cold'


# Phase C — Lead Priority (used for 🔴/🟡/🟢 badges)
HIGH_PRIORITY_STATUSES = {
    'Demo Attended', 'Admission Likely', 'Counselling Completed',
    'Parent Discussion Pending', 'Demo Booked', 'Demo Scheduled',
}
MEDIUM_PRIORITY_STATUSES = {
    'Interested', 'Connected', 'Counselling Scheduled', 'Contacted', 'Follow-up',
}
LOW_PRIORITY_STATUSES = {
    'New', 'Contact Attempted', 'Future Prospect',
}
TERMINAL_STATUSES = {'Lost', 'Not Interested', 'Converted', 'Admitted'}


def compute_lead_priority(lead: dict) -> str:
    """Auto-derive lead priority from status + fee_quoted + parent involvement.

    Per spec:
      High: Asking Fee | Wants Immediate Joining | Parent Involved | Demo Attended
      Medium: Interested | Exploring Options
      Low: Future Prospect | No Immediate Requirement
    """
    status = lead.get('status') or 'New'
    if status in TERMINAL_STATUSES:
        return 'Low'  # terminal leads aren't actionable
    if status in HIGH_PRIORITY_STATUSES:
        return 'High'
    # Fee discussion happening?
    if (lead.get('fee_quoted') or 0) > 0:
        return 'High'
    # Parent involved (any of these fields means parent is in the loop)
    parent_status = lead.get('parent_status')
    if parent_status and parent_status not in ('Not Contacted', None):
        return 'High'
    if status in MEDIUM_PRIORITY_STATUSES:
        return 'Medium'
    return 'Low'


@api_router.get("/counsellor/action-required")
async def get_counsellor_action_required(current_user: User = Depends(get_current_user)):
    """Action-Required dashboard for Counsellors (also visible to Branch Admin/Admin).

    The payload is intentionally one round-trip so the Counsellor dashboard can
    render summary cards + action lists without N+1 calls.
    """
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")

    now = datetime.now(timezone.utc)
    now_naive = datetime.now()
    today_str = now_naive.strftime('%Y-%m-%d')
    week_start = (now_naive - timedelta(days=now_naive.weekday())).strftime('%Y-%m-%d')
    month_start = now_naive.replace(day=1).strftime('%Y-%m-%d')

    # ---- Scope: leads visible to this user ----
    base_lead_query: Dict[str, Any] = {"is_deleted": {"$ne": True}}
    if current_user.role == UserRole.COUNSELLOR:
        base_lead_query["$or"] = [
            {"counsellor_id": current_user.id},
            {"created_by": current_user.id},
        ]
    elif current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        base_lead_query["branch_id"] = current_user.branch_id

    leads = await db.leads.find(base_lead_query, {"_id": 0}).to_list(2000)

    # ---- Lead Summary ----
    new_today = 0
    new_this_week = 0
    active_count = 0
    hot = warm = cold = lost = 0
    leads_this_month = 0
    for ld in leads:
        created = (ld.get('lead_date') or (ld.get('created_at') or '')[:10] or '')
        if created and created >= today_str:
            new_today += 1
        if created and created >= week_start:
            new_this_week += 1
        if created and created >= month_start:
            leads_this_month += 1
        status = ld.get('status') or 'New'
        if status not in ('Converted', 'Lost'):
            active_count += 1
        score = _classify_lead_score(ld)
        if score == 'Hot':
            hot += 1
        elif score == 'Warm':
            warm += 1
        elif score == 'Cold':
            cold += 1
        elif score == 'Lost':
            lost += 1

    # ---- Follow-Up Summary (uses followups collection) ----
    if current_user.role == UserRole.COUNSELLOR:
        followup_scope = {"created_by": current_user.id}
    elif current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        followup_scope = {"branch_id": current_user.branch_id}
    else:
        followup_scope = {}

    followups = await db.followups.find(followup_scope, {"_id": 0}).to_list(2000)
    fu_due_today = 0
    fu_overdue = 0
    fu_upcoming = 0
    fu_completed_today = 0
    fu_due_today_list: List[Dict[str, Any]] = []
    fu_overdue_list: List[Dict[str, Any]] = []
    for fu in followups:
        date10 = (fu.get('followup_date') or '')[:10]
        status = fu.get('status') or 'Pending'
        if status == 'Completed':
            completed_at = (fu.get('completed_at') or fu.get('updated_at') or '')[:10]
            if completed_at == today_str:
                fu_completed_today += 1
            continue
        if not date10:
            continue
        if date10 == today_str:
            fu_due_today += 1
            fu_due_today_list.append(fu)
        elif date10 < today_str:
            fu_overdue += 1
            fu_overdue_list.append(fu)
        else:
            fu_upcoming += 1

    # ---- Admission Summary (Converted leads) ----
    admissions_this_month = 0
    admissions_this_week = 0
    for ld in leads:
        if (ld.get('status') or '') != 'Converted':
            continue
        when = (ld.get('updated_at') or ld.get('lead_date') or (ld.get('created_at') or ''))[:10]
        if when and when >= month_start:
            admissions_this_month += 1
        if when and when >= week_start:
            admissions_this_week += 1
    target = max(1, int(round(ADMISSION_TARGET_PCT * leads_this_month))) if leads_this_month else 0
    achievement_pct = round((admissions_this_month / target * 100), 1) if target > 0 else 0.0

    # ---- Conversion Summary ----
    total_leads = len(leads)
    counseled = sum(1 for ld in leads if (ld.get('status') or '') in ('Contacted', 'Follow-up', 'Demo Booked', 'Converted'))
    converted = sum(1 for ld in leads if (ld.get('status') or '') == 'Converted')
    lead_to_counseling_pct = round((counseled / total_leads * 100), 1) if total_leads else 0.0
    counseling_to_admission_pct = round((converted / counseled * 100), 1) if counseled else 0.0
    overall_conversion_pct = round((converted / total_leads * 100), 1) if total_leads else 0.0

    # ---- Build a per-lead "last contact" map from followups so we can flag
    # "not contacted" and "interested without counselling" ----
    last_contact_map: Dict[str, str] = {}
    for fu in followups:
        lid = fu.get('lead_id')
        if not lid:
            continue
        when = (fu.get('completed_at') or fu.get('updated_at') or fu.get('created_at') or fu.get('followup_date') or '')[:19]
        if not when:
            continue
        prev = last_contact_map.get(lid)
        if not prev or when > prev:
            last_contact_map[lid] = when

    # ---- Action Required Lists ----
    hot_not_contacted: List[Dict[str, Any]] = []
    interested_no_counseling: List[Dict[str, Any]] = []
    admission_likely: List[Dict[str, Any]] = []
    parent_pending: List[Dict[str, Any]] = []  # Phase C — placeholder (parent_status not yet on Lead model)
    for ld in leads:
        lid = ld.get('id')
        status = ld.get('status') or 'New'
        score = _classify_lead_score(ld)
        created_at_raw = ld.get('created_at') or ''
        try:
            created_dt = datetime.fromisoformat(created_at_raw.replace('Z', '+00:00')) if created_at_raw else None
        except Exception:
            created_dt = None
        minutes_since_creation = None
        if created_dt:
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=timezone.utc)
            minutes_since_creation = int((now - created_dt).total_seconds() // 60)

        last_contact = last_contact_map.get(lid)
        # Hot Lead Not Contacted: new/fresh, no followup logged, > 10 minutes old
        if status == 'New' and not last_contact and minutes_since_creation is not None and minutes_since_creation >= 10:
            hot_not_contacted.append({
                "id": lid,
                "lead_id": ld.get('lead_id'),
                "name": ld.get('name'),
                "number": ld.get('number'),
                "program_name": ld.get('program_name'),
                "minutes_since_creation": minutes_since_creation,
                "lead_source": ld.get('lead_source'),
            })

        # Interested without counselling: Contacted / Follow-up for 3+ days, never Demo Booked / Converted
        if status in ('Contacted', 'Follow-up'):
            since_ref = last_contact or created_at_raw
            try:
                ref_dt = datetime.fromisoformat(since_ref.replace('Z', '+00:00')) if since_ref else None
            except Exception:
                ref_dt = None
            if ref_dt:
                if ref_dt.tzinfo is None:
                    ref_dt = ref_dt.replace(tzinfo=timezone.utc)
                days_in_stage = (now - ref_dt).days
                if days_in_stage >= 3:
                    interested_no_counseling.append({
                        "id": lid,
                        "lead_id": ld.get('lead_id'),
                        "name": ld.get('name'),
                        "number": ld.get('number'),
                        "program_name": ld.get('program_name'),
                        "status": status,
                        "days_in_stage": days_in_stage,
                    })

        # Admission Likely: Demo Booked, OR fee discussed with active status
        if status == 'Demo Booked' or (score == 'Hot' and status not in ('Converted', 'Lost')):
            admission_likely.append({
                "id": lid,
                "lead_id": ld.get('lead_id'),
                "name": ld.get('name'),
                "number": ld.get('number'),
                "program_name": ld.get('program_name'),
                "status": status,
                "fee_quoted": ld.get('fee_quoted'),
            })

    hot_not_contacted.sort(key=lambda x: x.get('minutes_since_creation') or 0, reverse=True)
    interested_no_counseling.sort(key=lambda x: x.get('days_in_stage') or 0, reverse=True)

    # ---- Overdue follow-ups (with lead info) ----
    overdue_list: List[Dict[str, Any]] = []
    leads_by_id = {ld.get('id'): ld for ld in leads}
    for fu in fu_overdue_list:
        ld = leads_by_id.get(fu.get('lead_id'))
        if not ld:
            continue
        try:
            due = datetime.strptime((fu.get('followup_date') or '')[:10], '%Y-%m-%d').date()
            days_overdue = (now_naive.date() - due).days
        except Exception:
            days_overdue = 0
        overdue_list.append({
            "id": fu.get('id'),
            "lead_id": ld.get('id'),
            "lead_custom_id": ld.get('lead_id'),
            "name": ld.get('name'),
            "number": ld.get('number'),
            "program_name": ld.get('program_name'),
            "followup_date": fu.get('followup_date'),
            "days_overdue": days_overdue,
            "note": fu.get('note', ''),
        })
    overdue_list.sort(key=lambda x: x.get('days_overdue', 0), reverse=True)

    # ---- Daily Work Queue ----
    # to_call: active leads with no contact in the last 24h (or never), capped at 15
    to_call: List[Dict[str, Any]] = []
    for ld in leads:
        status = ld.get('status') or 'New'
        if status in ('Converted', 'Lost'):
            continue
        last_contact = last_contact_map.get(ld.get('id'))
        stale = True
        if last_contact:
            try:
                lc_dt = datetime.fromisoformat(last_contact.replace('Z', '+00:00'))
                if lc_dt.tzinfo is None:
                    lc_dt = lc_dt.replace(tzinfo=timezone.utc)
                stale = (now - lc_dt).total_seconds() > 24 * 3600
            except Exception:
                stale = True
        if stale:
            to_call.append({
                "id": ld.get('id'),
                "lead_id": ld.get('lead_id'),
                "name": ld.get('name'),
                "number": ld.get('number'),
                "program_name": ld.get('program_name'),
                "status": status,
                "last_contact": last_contact,
            })
    to_call.sort(key=lambda x: (x.get('last_contact') or ''))

    # Today's follow-ups (decorated with lead)
    followups_due_today: List[Dict[str, Any]] = []
    for fu in fu_due_today_list:
        ld = leads_by_id.get(fu.get('lead_id'))
        if not ld:
            continue
        followups_due_today.append({
            "id": fu.get('id'),
            "lead_id": ld.get('id'),
            "name": ld.get('name'),
            "number": ld.get('number'),
            "program_name": ld.get('program_name'),
            "status": ld.get('status'),
            "followup_date": fu.get('followup_date'),
            "note": fu.get('note', ''),
        })

    return {
        "lead_summary": {
            "new_today": new_today,
            "new_this_week": new_this_week,
            "active": active_count,
            "hot": hot,
            "warm": warm,
            "cold": cold,
            "lost": lost,
            "total": total_leads,
        },
        "followup_summary": {
            "due_today": fu_due_today,
            "overdue": fu_overdue,
            "upcoming": fu_upcoming,
            "completed_today": fu_completed_today,
        },
        "admission_summary": {
            "admissions_this_month": admissions_this_month,
            "admissions_this_week": admissions_this_week,
            "target": target,
            "target_pct": int(ADMISSION_TARGET_PCT * 100),
            "achievement_pct": achievement_pct,
            "leads_this_month": leads_this_month,
        },
        "conversion_summary": {
            "lead_to_counseling_pct": lead_to_counseling_pct,
            "counseling_to_admission_pct": counseling_to_admission_pct,
            "overall_conversion_pct": overall_conversion_pct,
        },
        "action_required": {
            "hot_leads_not_contacted": hot_not_contacted[:10],
            "overdue_followups": overdue_list[:10],
            "interested_no_counseling": interested_no_counseling[:10],
            "admission_likely": admission_likely[:10],
            "parent_pending": parent_pending,  # Phase C
        },
        "daily_queue": {
            "to_call": to_call[:15],
            "followups_due_today": followups_due_today[:15],
        },
    }


# ----------------------------------------------------------------------------
# Phase C — Lead Detail Page extensions: Communication Timeline & Aging
# ----------------------------------------------------------------------------
@api_router.get("/leads/{lead_id}/timeline")
async def get_lead_timeline(lead_id: str, current_user: User = Depends(get_current_user)):
    """Unified communication timeline for a single lead.

    Aggregates: follow-ups (calls, WA, notes), status changes (audit), and demo events.
    Sorted by timestamp DESC for "most recent first" UX.
    """
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    # Authorisation — Counsellors can only see their own leads; Branch Admin sees branch; Admin sees all
    if current_user.role == UserRole.COUNSELLOR and lead.get('counsellor_id') != current_user.id and lead.get('created_by') != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id and lead.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")

    events: List[Dict[str, Any]] = []

    # Follow-ups (calls, WA, notes, meetings)
    followups = await db.followups.find({"lead_id": lead_id}, {"_id": 0}).to_list(200)
    for fu in followups:
        timestamp = fu.get('completed_at') or fu.get('updated_at') or fu.get('created_at') or fu.get('followup_date')
        events.append({
            "type": fu.get('followup_type') or 'followup',
            "title": f"Follow-up ({fu.get('status') or 'Pending'})",
            "note": fu.get('note') or '',
            "timestamp": timestamp,
            "actor_id": fu.get('created_by'),
            "actor_name": fu.get('created_by_name') or fu.get('counsellor_name') or '',
            "extra": {
                "outcome": fu.get('outcome'),
                "next_followup_date": fu.get('next_followup_date'),
                "followup_date": fu.get('followup_date'),
            },
        })

    # Status changes (if a leads_audit collection exists, otherwise skip silently)
    try:
        audits = await db.leads_audit.find({"lead_id": lead_id}, {"_id": 0}).to_list(200)
        for a in audits:
            events.append({
                "type": "status_change",
                "title": f"Status: {a.get('old_status', '—')} → {a.get('new_status', '—')}",
                "note": a.get('reason') or '',
                "timestamp": a.get('changed_at') or a.get('created_at'),
                "actor_id": a.get('changed_by'),
                "actor_name": a.get('changed_by_name') or '',
            })
    except Exception:
        pass

    # Lead creation event
    events.append({
        "type": "created",
        "title": f"Lead created via {lead.get('lead_source') or 'unknown'}",
        "note": f"Counsellor: {lead.get('counsellor_name') or lead.get('counsellor_id') or '—'}",
        "timestamp": lead.get('created_at'),
        "actor_id": lead.get('created_by'),
        "actor_name": lead.get('counsellor_name') or '',
    })

    # Demo events
    if lead.get('demo_date'):
        events.append({
            "type": "demo_scheduled",
            "title": f"Demo scheduled with {lead.get('trainer_name') or 'trainer'}",
            "note": f"{lead.get('demo_date')} {lead.get('demo_time') or ''}",
            "timestamp": lead.get('demo_date'),
            "actor_id": None,
            "actor_name": '',
        })

    # Sort newest first; treat None as oldest
    def _ts_key(e):
        t = e.get('timestamp') or ''
        return str(t)
    events.sort(key=_ts_key, reverse=True)

    # Lead aging bucket
    created_raw = lead.get('created_at') or ''
    try:
        created_dt = datetime.fromisoformat(str(created_raw).replace('Z', '+00:00')) if created_raw else None
        if created_dt and created_dt.tzinfo is None:
            created_dt = created_dt.replace(tzinfo=timezone.utc)
        age_days = (datetime.now(timezone.utc) - created_dt).days if created_dt else 0
    except Exception:
        age_days = 0
    if age_days <= 3:
        aging_bucket = "Fresh Leads"
    elif age_days <= 7:
        aging_bucket = "Needs Follow-Up"
    elif age_days <= 15:
        aging_bucket = "At Risk"
    else:
        aging_bucket = "Recovery Required"

    return {
        "lead_id": lead_id,
        "age_days": age_days,
        "aging_bucket": aging_bucket,
        "priority": lead.get('priority') or compute_lead_priority(lead),
        "events": events,
    }


# ----------------------------------------------------------------------------
# Phase D — Performance Dashboard, Lead Funnel, Lost Lead Recovery
# ----------------------------------------------------------------------------
ADMITTED_STATUSES = {"Converted", "Admitted"}
LOST_STATUSES = {"Lost", "Not Interested"}
COUNSELLED_STATUSES = {
    "Contacted", "Connected", "Interested", "Follow-up",
    "Counselling Scheduled", "Counselling Completed",
    "Demo Booked", "Demo Scheduled", "Demo Attended",
    "Admission Likely", "Converted", "Admitted",
}
DEMO_STATUSES = {"Demo Booked", "Demo Scheduled", "Demo Attended", "Admission Likely", "Converted", "Admitted"}


@api_router.get("/counsellor/performance")
async def get_counsellor_performance(current_user: User = Depends(get_current_user)):
    """Personal performance + funnel data for the logged-in Counsellor.

    Admin / Branch Admin can also call this to view aggregated branch numbers
    (Branch Admin scoped to their branch).
    """
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")

    base_query: Dict[str, Any] = {"is_deleted": {"$ne": True}}
    if current_user.role == UserRole.COUNSELLOR:
        base_query["$or"] = [{"counsellor_id": current_user.id}, {"created_by": current_user.id}]
    elif current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        base_query["branch_id"] = current_user.branch_id
    leads = await db.leads.find(base_query, {"_id": 0}).to_list(5000)
    total = len(leads)

    # Followups grouped by lead id
    followup_scope: Dict[str, Any] = {}
    if current_user.role == UserRole.COUNSELLOR:
        followup_scope = {"created_by": current_user.id}
    elif current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        followup_scope = {"branch_id": current_user.branch_id}
    followups = await db.followups.find(followup_scope, {"_id": 0}).to_list(5000)
    contacted_lead_ids = {fu.get('lead_id') for fu in followups if fu.get('lead_id')}
    fu_completed = sum(1 for fu in followups if (fu.get('status') or '') == 'Completed')
    today = datetime.now().strftime('%Y-%m-%d')
    fu_missed = sum(
        1 for fu in followups
        if (fu.get('status') or 'Pending') != 'Completed'
        and (fu.get('followup_date') or '')[:10] < today
        and (fu.get('followup_date') or '')[:10] != ''
    )

    # Funnel counts (each stage = leads that ever reached that stage or beyond)
    def _has_reached(status: str, statuses_to_match: set) -> bool:
        return status in statuses_to_match

    leads_count = total
    connected_count = sum(1 for ld in leads if ld.get('id') in contacted_lead_ids or _has_reached(ld.get('status') or '', {'Connected', 'Contacted', 'Interested', 'Follow-up', 'Counselling Scheduled', 'Counselling Completed', 'Demo Booked', 'Demo Scheduled', 'Demo Attended', 'Admission Likely', 'Converted', 'Admitted'}))
    interested_count = sum(1 for ld in leads if (ld.get('status') or '') in {'Interested', 'Counselling Scheduled', 'Counselling Completed', 'Demo Booked', 'Demo Scheduled', 'Demo Attended', 'Admission Likely', 'Converted', 'Admitted'})
    counselling_count = sum(1 for ld in leads if (ld.get('status') or '') in {'Counselling Completed', 'Demo Booked', 'Demo Scheduled', 'Demo Attended', 'Admission Likely', 'Converted', 'Admitted'})
    parent_discussion_count = sum(1 for ld in leads if (ld.get('parent_status') or '') in {'Contacted', 'Interested', 'Approval Pending', 'Approved'} or (ld.get('status') or '') == 'Parent Discussion Pending')
    demo_attended_count = sum(1 for ld in leads if (ld.get('status') or '') in {'Demo Attended', 'Admission Likely', 'Converted', 'Admitted'})
    admission_likely_count = sum(1 for ld in leads if (ld.get('status') or '') in {'Admission Likely', 'Converted', 'Admitted'})
    admitted_count = sum(1 for ld in leads if (ld.get('status') or '') in ADMITTED_STATUSES)
    lost_count = sum(1 for ld in leads if (ld.get('status') or '') in LOST_STATUSES)

    def _pct(num, den):
        return round((num / den * 100), 1) if den else 0.0

    funnel = [
        {"stage": "Leads Assigned",     "count": leads_count,         "pct_of_top": 100.0},
        {"stage": "Connected",           "count": connected_count,    "pct_of_top": _pct(connected_count, leads_count)},
        {"stage": "Interested",          "count": interested_count,   "pct_of_top": _pct(interested_count, leads_count)},
        {"stage": "Counselling Completed","count": counselling_count, "pct_of_top": _pct(counselling_count, leads_count)},
        {"stage": "Parent Discussion",   "count": parent_discussion_count, "pct_of_top": _pct(parent_discussion_count, leads_count)},
        {"stage": "Demo Attended",       "count": demo_attended_count, "pct_of_top": _pct(demo_attended_count, leads_count)},
        {"stage": "Admission Likely",    "count": admission_likely_count, "pct_of_top": _pct(admission_likely_count, leads_count)},
        {"stage": "Admitted",            "count": admitted_count,      "pct_of_top": _pct(admitted_count, leads_count)},
    ]

    # Aging buckets
    now_utc = datetime.now(timezone.utc)
    aging = {"fresh": 0, "needs_followup": 0, "at_risk": 0, "recovery": 0}
    for ld in leads:
        if (ld.get('status') or '') in (ADMITTED_STATUSES | LOST_STATUSES):
            continue
        try:
            created_dt = datetime.fromisoformat(str(ld.get('created_at') or '').replace('Z', '+00:00'))
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=timezone.utc)
            days = (now_utc - created_dt).days
        except Exception:
            days = 0
        if days <= 3:
            aging["fresh"] += 1
        elif days <= 7:
            aging["needs_followup"] += 1
        elif days <= 15:
            aging["at_risk"] += 1
        else:
            aging["recovery"] += 1

    return {
        "scope": "branch" if current_user.role == UserRole.BRANCH_ADMIN else ("self" if current_user.role == UserRole.COUNSELLOR else "all"),
        "metrics": {
            "leads_assigned": leads_count,
            "leads_contacted": len(contacted_lead_ids),
            "followups_completed": fu_completed,
            "followups_missed": fu_missed,
            "admissions_generated": admitted_count,
            "conversion_rate": _pct(admitted_count, leads_count),
            "lost_count": lost_count,
        },
        "funnel": funnel,
        "aging_buckets": aging,
    }


@api_router.get("/counsellor/lost-leads")
async def get_lost_leads(current_user: User = Depends(get_current_user)):
    """List Lost / Not Interested leads with their captured reason for recovery."""
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    q: Dict[str, Any] = {"status": {"$in": list(LOST_STATUSES)}, "is_deleted": {"$ne": True}}
    if current_user.role == UserRole.COUNSELLOR:
        q["$or"] = [{"counsellor_id": current_user.id}, {"created_by": current_user.id}]
    elif current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        q["branch_id"] = current_user.branch_id
    leads = await db.leads.find(q, {"_id": 0}).to_list(500)
    return [
        {
            "id": ld.get("id"),
            "lead_id": ld.get("lead_id"),
            "name": ld.get("name"),
            "number": ld.get("number"),
            "email": ld.get("email"),
            "program_name": ld.get("program_name"),
            "status": ld.get("status"),
            "lost_reason": ld.get("lost_reason") or ld.get("delete_reason"),
            "lost_at": ld.get("updated_at") or ld.get("lost_at"),
            "counsellor_name": ld.get("counsellor_name"),
        }
        for ld in leads
    ]


class LeadReopenRequest(BaseModel):
    notes: Optional[str] = None


@api_router.post("/counsellor/lost-leads/{lead_id}/reopen")
async def reopen_lost_lead(lead_id: str, payload: LeadReopenRequest, current_user: User = Depends(get_current_user)):
    """Move a Lost/Not-Interested lead back to 'New' so the counsellor can re-engage."""
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if current_user.role == UserRole.COUNSELLOR and lead.get('counsellor_id') != current_user.id:
        raise HTTPException(status_code=403, detail="Not your lead")
    if (lead.get('status') or '') not in LOST_STATUSES:
        raise HTTPException(status_code=400, detail="Only Lost / Not Interested leads can be reopened")
    update = {
        "status": "New",
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "reopened_at": datetime.now(timezone.utc).isoformat(),
        "reopened_by": current_user.id,
        "reopen_notes": payload.notes or "",
    }
    update["priority"] = compute_lead_priority({**lead, **update})
    await db.leads.update_one({"id": lead_id}, {"$set": update})
    return {"success": True, "lead_id": lead_id, "status": "New"}


# ----------------------------------------------------------------------------
# Phase E — Smart Follow-Up Suggestions + Notifications
# ----------------------------------------------------------------------------
@api_router.get("/leads/{lead_id}/suggestions")
async def get_lead_suggestions(lead_id: str, current_user: User = Depends(get_current_user)):
    """Rule-based suggestions for the next-best-action on a single lead.

    Examples (per spec):
      • Fee inquiry received      -> Send fee structure + schedule parent discussion
      • Demo attended             -> Admission follow-up within 24 hours
      • Inactive for 3+ days      -> Re-engagement call
      • Parent not contacted      -> Initiate parent discussion
    """
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    # Auth — same as timeline
    if current_user.role == UserRole.COUNSELLOR and lead.get('counsellor_id') != current_user.id and lead.get('created_by') != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id and lead.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")

    suggestions: List[Dict[str, Any]] = []
    status = lead.get('status') or 'New'
    fee_quoted = lead.get('fee_quoted') or 0
    parent_status = lead.get('parent_status') or 'Not Contacted'

    # Most-recent contact (for inactivity check)
    last_fu = await db.followups.find_one(
        {"lead_id": lead_id},
        {"_id": 0},
        sort=[("created_at", -1)]
    )
    last_contact_ts = (last_fu or {}).get('completed_at') or (last_fu or {}).get('created_at')
    days_inactive = None
    if last_contact_ts:
        try:
            lc = datetime.fromisoformat(str(last_contact_ts).replace('Z', '+00:00'))
            if lc.tzinfo is None:
                lc = lc.replace(tzinfo=timezone.utc)
            days_inactive = (datetime.now(timezone.utc) - lc).days
        except Exception:
            days_inactive = None

    if fee_quoted > 0 and parent_status in ('Not Contacted', None):
        suggestions.append({
            "id": "fee-and-parent",
            "priority": "high",
            "title": "Send fee structure & schedule parent discussion",
            "rationale": "Fee has been discussed but parent is not yet looped in.",
            "action": "send_fee_then_parent",
        })

    if status in ('Demo Attended',):
        suggestions.append({
            "id": "post-demo",
            "priority": "high",
            "title": "Admission follow-up within 24 hours",
            "rationale": "Lead just attended the demo — strike while the iron is hot.",
            "action": "schedule_admission_call",
        })

    if status in ('Counselling Completed', 'Demo Attended') and parent_status == 'Not Contacted':
        suggestions.append({
            "id": "parent-loop-in",
            "priority": "medium",
            "title": "Initiate parent discussion",
            "rationale": "Counselling done but parent has not been contacted yet.",
            "action": "call_parent",
        })

    if days_inactive is not None and days_inactive >= 3 and status not in ('Converted', 'Admitted', 'Lost', 'Not Interested'):
        suggestions.append({
            "id": "reengage",
            "priority": "medium",
            "title": "Re-engagement call",
            "rationale": f"Lead has been inactive for {days_inactive} days.",
            "action": "reengage",
        })

    if status == 'Admission Likely':
        suggestions.append({
            "id": "close",
            "priority": "high",
            "title": "Close the admission",
            "rationale": "Lead is admission-likely — push for the final commitment.",
            "action": "close",
        })

    if status == 'New' and not last_fu:
        suggestions.append({
            "id": "first-contact",
            "priority": "high",
            "title": "Make the first contact",
            "rationale": "Brand new lead with no contact attempt logged yet.",
            "action": "first_call",
        })

    return {"lead_id": lead_id, "suggestions": suggestions}


@api_router.get("/counsellor/notifications")
async def get_counsellor_notifications(current_user: User = Depends(get_current_user)):
    """In-app notifications for the Counsellor.

    Derived signals (no separate notifications collection):
      • New Lead Assigned       (lead.created_at within last 24h)
      • Hot Lead Received       (created within last 24h AND fee_quoted OR Demo Booked)
      • Follow-Up Due Today
      • Follow-Up Overdue
      • Parent Discussion Pending (status='Parent Discussion Pending' OR parent_status='Approval Pending')
      • Admission Likely
    """
    if current_user.role != UserRole.COUNSELLOR:
        # Only Counsellors get notifications for now; Admin/BranchAdmin can use dashboards instead
        return {"notifications": []}

    now = datetime.now(timezone.utc)
    today = datetime.now().strftime('%Y-%m-%d')
    yesterday = now - timedelta(hours=24)

    leads = await db.leads.find(
        {
            "$or": [{"counsellor_id": current_user.id}, {"created_by": current_user.id}],
            "is_deleted": {"$ne": True},
        },
        {"_id": 0}
    ).to_list(2000)
    followups = await db.followups.find({"created_by": current_user.id}, {"_id": 0}).to_list(2000)

    notes: List[Dict[str, Any]] = []

    for ld in leads:
        try:
            created_dt = datetime.fromisoformat(str(ld.get('created_at') or '').replace('Z', '+00:00'))
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=timezone.utc)
        except Exception:
            created_dt = None

        status = ld.get('status') or 'New'
        # New lead assigned (last 24h, no follow-up yet)
        if created_dt and created_dt >= yesterday and status == 'New':
            notes.append({
                "id": f"new-{ld.get('id')}",
                "kind": "new_lead",
                "icon": "user-plus",
                "tone": "blue",
                "title": "New lead assigned",
                "body": f"{ld.get('name')} • {ld.get('program_name')}",
                "lead_id": ld.get('id'),
                "timestamp": ld.get('created_at'),
            })
        # Hot lead
        if (status in ('Demo Booked', 'Demo Scheduled', 'Demo Attended', 'Admission Likely')
                or (ld.get('fee_quoted') or 0) > 0):
            if created_dt and created_dt >= yesterday:
                notes.append({
                    "id": f"hot-{ld.get('id')}",
                    "kind": "hot_lead",
                    "icon": "flame",
                    "tone": "rose",
                    "title": "Hot lead received",
                    "body": f"{ld.get('name')} • {status}",
                    "lead_id": ld.get('id'),
                    "timestamp": ld.get('created_at'),
                })
        # Parent discussion pending
        if status == 'Parent Discussion Pending' or (ld.get('parent_status') == 'Approval Pending'):
            notes.append({
                "id": f"parent-{ld.get('id')}",
                "kind": "parent_pending",
                "icon": "users",
                "tone": "pink",
                "title": "Parent discussion pending",
                "body": f"{ld.get('name')} — loop in parent",
                "lead_id": ld.get('id'),
                "timestamp": ld.get('updated_at') or ld.get('created_at'),
            })
        # Admission likely
        if status == 'Admission Likely':
            notes.append({
                "id": f"likely-{ld.get('id')}",
                "kind": "admission_likely",
                "icon": "graduation-cap",
                "tone": "emerald",
                "title": "Admission likely",
                "body": f"{ld.get('name')} — close the loop",
                "lead_id": ld.get('id'),
                "timestamp": ld.get('updated_at') or ld.get('created_at'),
            })

    leads_by_id = {ld.get('id'): ld for ld in leads}
    for fu in followups:
        if (fu.get('status') or 'Pending') == 'Completed':
            continue
        date10 = (fu.get('followup_date') or '')[:10]
        if not date10:
            continue
        ld = leads_by_id.get(fu.get('lead_id'))
        if not ld:
            continue
        if date10 == today:
            notes.append({
                "id": f"fud-{fu.get('id')}",
                "kind": "followup_due",
                "icon": "calendar-clock",
                "tone": "violet",
                "title": "Follow-up due today",
                "body": f"{ld.get('name')} — {fu.get('note') or 'no note'}",
                "lead_id": ld.get('id'),
                "timestamp": fu.get('followup_date'),
            })
        elif date10 < today:
            notes.append({
                "id": f"fuo-{fu.get('id')}",
                "kind": "followup_overdue",
                "icon": "alert-triangle",
                "tone": "amber",
                "title": "Follow-up overdue",
                "body": f"{ld.get('name')} • due {date10}",
                "lead_id": ld.get('id'),
                "timestamp": fu.get('followup_date'),
            })

    # De-dup (by id) and sort newest first
    seen = set()
    deduped: List[Dict[str, Any]] = []
    for n in notes:
        if n['id'] in seen:
            continue
        seen.add(n['id'])
        deduped.append(n)

    # Pull persisted "lead_status_changed_by_other" notifications from db.notifications
    persisted = await db.notifications.find(
        {
            "user_id": current_user.id,
            "type": "lead_status_changed_by_other",
        },
        {"_id": 0},
    ).sort("created_at", -1).to_list(50)
    for p in persisted:
        deduped.append({
            "id": f"sc-{p.get('id')}",
            "kind": "status_changed",
            "icon": "shuffle",
            "tone": "sky",
            "title": p.get("title") or "Lead status changed",
            "body": p.get("message") or "",
            "lead_id": (p.get("data") or {}).get("lead_id"),
            "timestamp": p.get("created_at"),
            "is_read": bool(p.get("is_read")),
            "notification_id": p.get("id"),
        })

    deduped.sort(key=lambda x: str(x.get('timestamp') or ''), reverse=True)
    return {"notifications": deduped[:50]}


@api_router.post("/counsellor/notifications/{notification_id}/read")
async def mark_counsellor_notification_read(notification_id: str, current_user: User = Depends(get_current_user)):
    """Mark a persisted counsellor notification as read."""
    if current_user.role != UserRole.COUNSELLOR:
        raise HTTPException(status_code=403, detail="Counsellor only")
    await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user.id},
        {"$set": {"is_read": True}},
    )
    return {"message": "Marked read"}


@api_router.post("/counsellor/notifications/mark-all-read")
async def mark_all_counsellor_notifications_read(current_user: User = Depends(get_current_user)):
    """Mark all of this counsellor's persisted notifications as read."""
    if current_user.role != UserRole.COUNSELLOR:
        raise HTTPException(status_code=403, detail="Counsellor only")
    result = await db.notifications.update_many(
        {"user_id": current_user.id, "is_read": {"$ne": True}},
        {"$set": {"is_read": True}},
    )
    return {"message": "All marked read", "updated": result.modified_count}


# Branch Admin specific endpoints
@api_router.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    """Delete a payment - Branch Admin only for their branch"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can delete payments")
    
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Branch Admin can only delete payments from their branch
    if current_user.role == UserRole.BRANCH_ADMIN and payment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only delete payments from your branch")
    
    enrollment_id = payment.get('enrollment_id')
    installment_number = payment.get('installment_number')
    payment_plan_id = payment.get('payment_plan_id')
    
    # Delete the payment
    await db.payments.delete_one({"id": payment_id})
    
    # Recalculate enrollment totals
    if enrollment_id:
        enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
        if enrollment:
            # Get remaining payments for this enrollment
            remaining_payments = await db.payments.find(
                {"enrollment_id": enrollment_id}, 
                {"_id": 0, "amount": 1}
            ).to_list(1000)
            new_total_paid = sum(p.get('amount', 0) for p in remaining_payments)
            final_fee = enrollment.get('final_fee', 0)
            
            # Determine new status
            new_status = "Paid" if new_total_paid >= final_fee else (
                "Partial" if new_total_paid > 0 else "Pending"
            )
            
            # Update enrollment
            await db.enrollments.update_one(
                {"id": enrollment_id},
                {"$set": {
                    "total_paid": new_total_paid,
                    "payment_status": new_status
                }}
            )
            
            logger.info(f"Deleted payment {payment_id}. Updated enrollment {enrollment_id}: total_paid={new_total_paid}, status={new_status}")
    
    # Reset installment status if this was an installment payment
    if installment_number and payment_plan_id:
        # Check if there are any other payments for this installment
        other_inst_payments = await db.payments.find(
            {
                "payment_plan_id": payment_plan_id,
                "installment_number": installment_number
            },
            {"_id": 0, "amount": 1}
        ).to_list(100)
        
        total_inst_paid = sum(p.get('amount', 0) for p in other_inst_payments)
        
        if total_inst_paid == 0:
            # No payments left for this installment - reset to Pending
            await db.installment_schedule.update_one(
                {
                    "payment_plan_id": payment_plan_id,
                    "installment_number": installment_number
                },
                {"$set": {"status": "Pending", "paid_amount": 0, "paid_date": None}}
            )
        else:
            # Update the paid_amount
            await db.installment_schedule.update_one(
                {
                    "payment_plan_id": payment_plan_id,
                    "installment_number": installment_number
                },
                {"$set": {"paid_amount": total_inst_paid}}
            )
    
    return {"message": "Payment deleted successfully"}

@api_router.put("/payments/{payment_id}")
async def update_payment(payment_id: str, payment_update: dict, current_user: User = Depends(get_current_user)):
    """Update a payment - Branch Admin only for their branch"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can update payments")
    
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Branch Admin can only update payments from their branch
    if current_user.role == UserRole.BRANCH_ADMIN and payment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only update payments from your branch")
    
    # Only allow certain fields to be updated
    allowed_fields = ['amount', 'payment_mode', 'payment_date', 'remarks']
    update_data = {k: v for k, v in payment_update.items() if k in allowed_fields}
    
    old_amount = payment.get('amount', 0)
    new_amount = update_data.get('amount', old_amount)
    
    if update_data:
        await db.payments.update_one({"id": payment_id}, {"$set": update_data})
    
    # Recalculate enrollment totals if amount changed
    if 'amount' in update_data and old_amount != new_amount:
        enrollment_id = payment.get('enrollment_id')
        if enrollment_id:
            enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
            if enrollment:
                # Recalculate from all payments
                all_payments = await db.payments.find(
                    {"enrollment_id": enrollment_id}, 
                    {"_id": 0, "amount": 1}
                ).to_list(1000)
                new_total_paid = sum(p.get('amount', 0) for p in all_payments)
                final_fee = enrollment.get('final_fee', 0)
                
                # Determine new status
                new_status = "Paid" if new_total_paid >= final_fee else (
                    "Partial" if new_total_paid > 0 else "Pending"
                )
                
                # Update enrollment
                await db.enrollments.update_one(
                    {"id": enrollment_id},
                    {"$set": {
                        "total_paid": new_total_paid,
                        "payment_status": new_status
                    }}
                )
                
                logger.info(f"Updated payment {payment_id} amount from {old_amount} to {new_amount}. Enrollment {enrollment_id}: total_paid={new_total_paid}, status={new_status}")
        
        # Update installment paid_amount if applicable
        installment_number = payment.get('installment_number')
        payment_plan_id = payment.get('payment_plan_id')
        
        if installment_number and payment_plan_id:
            # Recalculate total paid for this installment
            inst_payments = await db.payments.find(
                {
                    "payment_plan_id": payment_plan_id,
                    "installment_number": installment_number
                },
                {"_id": 0, "amount": 1}
            ).to_list(100)
            
            total_inst_paid = sum(p.get('amount', 0) for p in inst_payments)
            
            await db.installment_schedule.update_one(
                {
                    "payment_plan_id": payment_plan_id,
                    "installment_number": installment_number
                },
                {"$set": {"paid_amount": total_inst_paid}}
            )
    
    return {"message": "Payment updated successfully"}


@api_router.post("/payments/recalculate-enrollment/{enrollment_id}")
async def recalculate_enrollment_payments(enrollment_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Recalculate total_paid and payment_status for a specific enrollment - Admin/Branch Admin only"""
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    # Branch Admin can only recalculate for their branch
    if current_user.role == UserRole.BRANCH_ADMIN and enrollment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only recalculate payments for your branch")
    
    # Get all payments for this enrollment
    payments = await db.payments.find({"enrollment_id": enrollment_id}, {"_id": 0}).to_list(1000)
    actual_total_paid = sum(p.get('amount', 0) for p in payments)
    final_fee = enrollment.get('final_fee', 0)
    
    # Determine correct status
    new_status = "Paid" if actual_total_paid >= final_fee else (
        "Partial" if actual_total_paid > 0 else "Pending"
    )
    
    old_total = enrollment.get('total_paid', 0)
    old_status = enrollment.get('payment_status', 'Unknown')
    
    # Update enrollment
    await db.enrollments.update_one(
        {"id": enrollment_id},
        {"$set": {
            "total_paid": actual_total_paid,
            "payment_status": new_status
        }}
    )
    
    logger.info(f"Recalculated enrollment {enrollment_id}: total_paid {old_total} -> {actual_total_paid}, status {old_status} -> {new_status}")
    
    return {
        "enrollment_id": enrollment_id,
        "student_name": enrollment.get('student_name'),
        "final_fee": final_fee,
        "old_total_paid": old_total,
        "new_total_paid": actual_total_paid,
        "old_status": old_status,
        "new_status": new_status,
        "payments_count": len(payments)
    }

@api_router.post("/payments/recalculate-all")
async def recalculate_all_enrollment_payments(
    branch_id: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Recalculate total_paid and payment_status for all enrollments - Super Admin only"""
    
    query = {}
    if branch_id:
        query["branch_id"] = branch_id
    
    enrollments = await db.enrollments.find(query, {"_id": 0, "id": 1, "final_fee": 1, "student_name": 1, "total_paid": 1, "payment_status": 1}).to_list(None)
    
    updated_count = 0
    issues_fixed = []
    
    for enrollment in enrollments:
        enrollment_id = enrollment.get('id')
        final_fee = enrollment.get('final_fee', 0) or 0
        old_total = enrollment.get('total_paid', 0) or 0
        old_status = enrollment.get('payment_status', 'Pending')
        
        # Calculate actual total from payments
        payments = await db.payments.find({"enrollment_id": enrollment_id}, {"_id": 0, "amount": 1}).to_list(None)
        actual_total = sum(p.get('amount', 0) or 0 for p in payments)
        
        # Determine correct status
        new_status = "Paid" if actual_total >= final_fee else (
            "Partial" if actual_total > 0 else "Pending"
        )
        
        # Check if update needed
        if abs(old_total - actual_total) > 1 or old_status != new_status:
            await db.enrollments.update_one(
                {"id": enrollment_id},
                {"$set": {
                    "total_paid": actual_total,
                    "payment_status": new_status
                }}
            )
            updated_count += 1
            issues_fixed.append({
                "enrollment_id": enrollment_id,
                "student_name": enrollment.get('student_name'),
                "old_total": old_total,
                "new_total": actual_total,
                "old_status": old_status,
                "new_status": new_status
            })
    
    return {
        "total_enrollments": len(enrollments),
        "updated_count": updated_count,
        "issues_fixed": issues_fixed[:50]  # Return first 50 for readability
    }


# Push Notifications Endpoints
@api_router.post("/notifications")
async def create_notification(notification: PushNotificationCreate, current_user: User = Depends(get_current_user)):
    """Create and send a push notification
    - Super Admin can send to all Branch Admins
    - Branch Admin can send to FDEs and Counsellors in their branch
    """
    recipient_ids = []
    
    if current_user.role == UserRole.ADMIN:
        # Super Admin can send to Branch Admins
        if notification.recipient_role:
            users = await db.users.find({"role": notification.recipient_role}, {"_id": 0, "id": 1}).to_list(1000)
            recipient_ids = [u['id'] for u in users]
        elif notification.recipient_ids:
            recipient_ids = notification.recipient_ids
    elif current_user.role == UserRole.BRANCH_ADMIN:
        # Branch Admin can only send to their branch's FDEs and Counsellors
        query = {
            "branch_id": current_user.branch_id,
            "role": {"$in": [UserRole.FRONT_DESK.value, UserRole.COUNSELLOR.value]}
        }
        if notification.recipient_ids:
            query["id"] = {"$in": notification.recipient_ids}
        users = await db.users.find(query, {"_id": 0, "id": 1}).to_list(1000)
        recipient_ids = [u['id'] for u in users]
    else:
        raise HTTPException(status_code=403, detail="Only Admin or Branch Admin can send notifications")
    
    if not recipient_ids:
        raise HTTPException(status_code=400, detail="No recipients found")
    
    new_notification = PushNotification(
        sender_id=current_user.id,
        sender_name=current_user.name,
        sender_role=current_user.role.value,
        recipient_ids=recipient_ids,
        recipient_role=notification.recipient_role,
        branch_id=current_user.branch_id,
        title=notification.title,
        message=notification.message,
        notification_type=notification.notification_type
    )
    
    notif_dict = new_notification.model_dump()
    notif_dict['created_at'] = notif_dict['created_at'].isoformat()
    
    await db.notifications.insert_one(notif_dict)
    return {"message": f"Notification sent to {len(recipient_ids)} recipients", "notification_id": new_notification.id}

@api_router.get("/notifications")
async def get_my_notifications(current_user: User = Depends(get_current_user)):
    """Get notifications for the current user"""
    notifications = await db.notifications.find(
        {"recipient_ids": current_user.id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    for n in notifications:
        if isinstance(n.get('created_at'), str):
            n['created_at'] = datetime.fromisoformat(n['created_at'])
    
    return notifications

@api_router.get("/notifications/unread-count")
async def get_unread_count(current_user: User = Depends(get_current_user)):
    """Get count of unread notifications"""
    count = await db.notifications.count_documents({
        "$or": [{"user_id": current_user.id}, {"recipient_ids": current_user.id}],
        "is_read": False
    })
    return {"count": count}

# Students (Enrolled) Endpoints
@api_router.get("/students")
async def get_students(request: Request, session: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get enrolled students with full details - optimized with bulk queries"""
    query = {}
    if current_user.role not in [UserRole.ADMIN]:
        query["branch_id"] = current_user.branch_id
    
    # Session filtering
    session_val = session or await get_session_from_request(request)
    if session_val and session_val != "all":
        session_filter = get_session_filter(session_val, "enrollment_date")
        query.update(session_filter)
    
    enrollments = await db.enrollments.find(query, {"_id": 0}).sort("enrollment_date", -1).to_list(1000)
    
    if not enrollments:
        return []
    
    # Collect all enrollment IDs for bulk queries
    enrollment_ids = [e['id'] for e in enrollments]
    
    # Bulk fetch all payment plans for these enrollments
    all_plans = await db.payment_plans.find(
        {"enrollment_id": {"$in": enrollment_ids}}, {"_id": 0}
    ).to_list(1000)
    plans_map = {p['enrollment_id']: p for p in all_plans}
    
    # Bulk fetch all payments using aggregation for totals
    payment_agg = await db.payments.aggregate([
        {"$match": {"enrollment_id": {"$in": enrollment_ids}}},
        {"$group": {
            "_id": "$enrollment_id",
            "total_paid": {"$sum": "$amount"},
            "payments_count": {"$sum": 1}
        }}
    ]).to_list(1000)
    payments_map = {p['_id']: p for p in payment_agg}
    
    students = []
    for e in enrollments:
        eid = e['id']
        plan = plans_map.get(eid)
        payment_info = payments_map.get(eid, {"total_paid": 0, "payments_count": 0})
        
        total_paid = payment_info['total_paid']
        final_fee = e.get('final_fee', 0)
        pending_amount = max(0, final_fee - total_paid)
        
        # Parse dates
        if isinstance(e.get('created_at'), str):
            e['created_at'] = datetime.fromisoformat(e['created_at'])
        if isinstance(e.get('enrollment_date'), str):
            e['enrollment_date'] = datetime.fromisoformat(e['enrollment_date']).date()
        
        students.append({
            **e,
            "total_paid": total_paid,
            "pending_amount": pending_amount,
            "has_payment_plan": plan is not None,
            "payment_plan_type": plan.get('plan_type') if plan else None,
            "payments_count": payment_info['payments_count']
        })
    
    return students

@api_router.get("/students/{enrollment_id}")
async def get_student_details(enrollment_id: str, current_user: User = Depends(get_current_user)):
    """Get full details of an enrolled student including add-on courses and other enrollments"""
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Check branch access
    if current_user.role not in [UserRole.ADMIN] and enrollment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get payment plan, payments, and add-on courses
    payment_plan = await db.payment_plans.find_one({"enrollment_id": enrollment_id}, {"_id": 0})
    payments = await db.payments.find({"enrollment_id": enrollment_id}, {"_id": 0}).sort("payment_date", -1).to_list(100)
    total_paid = sum(p.get('amount', 0) for p in payments)
    
    # Get add-on courses ONLY for THIS specific enrollment
    addon_courses = await db.addon_courses.find({"enrollment_id": enrollment_id}, {"_id": 0}).to_list(100)
    addon_total_fee = sum(a.get('final_fee', 0) for a in addon_courses)
    
    # Get branch info
    branch = await db.branches.find_one({"id": enrollment.get('branch_id')}, {"_id": 0, "name": 1, "location": 1})
    
    # Get other enrollments for the SAME student (strict: exact name + phone match)
    student_name_raw = enrollment.get('student_name') or ''
    student_name = student_name_raw.strip().lower() if isinstance(student_name_raw, str) else ''
    student_phone = enrollment.get('phone') or enrollment.get('student_phone')
    
    other_enrollments = []
    if student_phone and student_name:
        query = {"id": {"$ne": enrollment_id}}
        if current_user.role not in [UserRole.ADMIN]:
            query["branch_id"] = current_user.branch_id
        
        # Match by phone AND exact student name
        phone_conditions = []
        if student_phone:
            phone_conditions.append({"phone": student_phone})
            phone_conditions.append({"student_phone": student_phone})
        
        if phone_conditions:
            query["$or"] = phone_conditions
            candidates = await db.enrollments.find(query, {"_id": 0}).to_list(200)
            
            # EXACT name match only (case-insensitive) to avoid false positives
            for candidate in candidates:
                cname_raw = candidate.get('student_name') or ''
                candidate_name = cname_raw.strip().lower() if isinstance(cname_raw, str) else ''
                if candidate_name and candidate_name == student_name:
                    other_enrollments.append(candidate)
    
    # Calculate totals across ONLY verified same-student enrollments
    all_enrollment_ids = [enrollment_id] + [e['id'] for e in other_enrollments]
    
    # Bulk fetch payments for all related enrollments
    all_payments_agg = await db.payments.aggregate([
        {"$match": {"enrollment_id": {"$in": all_enrollment_ids}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    grand_total_paid = all_payments_agg[0]['total'] if all_payments_agg else total_paid
    
    # Bulk fetch addon courses for other enrollments
    other_addon_fees = 0
    if other_enrollments:
        other_eid_list = [e['id'] for e in other_enrollments]
        other_addons_agg = await db.addon_courses.aggregate([
            {"$match": {"enrollment_id": {"$in": other_eid_list}}},
            {"$group": {"_id": None, "total": {"$sum": "$final_fee"}}}
        ]).to_list(1)
        other_addon_fees = other_addons_agg[0]['total'] if other_addons_agg else 0
    
    # The enrollment's final_fee already includes addon fees (updated when addons are added)
    # So current_enrollment_total_fee = just the enrollment's final_fee (already has addons baked in)
    current_enrollment_total_fee = enrollment.get('final_fee', 0)
    
    grand_total_fee = current_enrollment_total_fee
    for other in other_enrollments:
        grand_total_fee += other.get('final_fee', 0)
    grand_total_fee += other_addon_fees
    
    return {
        "enrollment": enrollment,
        "payment_plan": payment_plan,
        "payments": payments,
        "total_paid": total_paid,
        "pending_amount": max(0, current_enrollment_total_fee - total_paid),
        "current_enrollment_total_fee": current_enrollment_total_fee,
        "branch": branch,
        "addon_courses": addon_courses,
        "addon_total_fee": addon_total_fee,
        "other_enrollments": other_enrollments,
        "grand_total_fee": grand_total_fee,
        "grand_total_paid": grand_total_paid,
        "grand_pending": max(0, grand_total_fee - grand_total_paid)
    }

@api_router.put("/students/{enrollment_id}/cancel")
async def cancel_enrollment(enrollment_id: str, reason: str = "", current_user: User = Depends(get_current_user)):
    """Cancel/Drop an enrollment - Branch Admin only"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can cancel enrollments")
    
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    # Branch Admin can only cancel their branch's enrollments
    if current_user.role == UserRole.BRANCH_ADMIN and enrollment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only cancel enrollments from your branch")
    
    await db.enrollments.update_one(
        {"id": enrollment_id},
        {"$set": {
            "status": "Cancelled",
            "cancelled_at": datetime.now(timezone.utc).isoformat(),
            "cancelled_by": current_user.id,
            "cancellation_reason": reason
        }}
    )
    
    return {"message": "Enrollment cancelled successfully"}

class StudentUpdateModel(BaseModel):
    """Model for updating student details after enrollment"""
    student_name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    highest_qualification: Optional[str] = None
    institution_name: Optional[str] = None
    passing_year: Optional[str] = None
    percentage: Optional[float] = None
    student_photo_url: Optional[str] = None
    aadhar_photo_url: Optional[str] = None
    aadhar_documents: Optional[List[str]] = None
    # Personal details
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    alternate_phone: Optional[str] = None
    # Fee fields - editable by Branch Admin and FDE
    fee_quoted: Optional[float] = None
    discount_percent: Optional[float] = None
    discount_amount: Optional[float] = None
    final_fee: Optional[float] = None
    # Program change
    program_id: Optional[str] = None
    program_name: Optional[str] = None
    # Admission date - editable by Branch Admin only
    enrollment_date: Optional[str] = None

@api_router.put("/students/{enrollment_id}/update")
async def update_student_details(enrollment_id: str, data: StudentUpdateModel, current_user: User = Depends(get_current_user)):
    """Update student details after enrollment - FDE, Branch Admin, or Admin"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Branch access check
    if current_user.role != UserRole.ADMIN and enrollment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only update students from your branch")
    
    # Build update data (only include non-None values)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    # FDE restrictions
    if current_user.role == UserRole.FRONT_DESK:
        # FDE cannot edit student name
        if 'student_name' in update_data:
            current_name = enrollment.get('student_name', '')
            new_name = update_data.get('student_name', '')
            if new_name != current_name:
                raise HTTPException(status_code=403, detail="Front Desk cannot edit student name")
            del update_data['student_name']
        
        # FDE cannot edit enrollment_date (admission date)
        if 'enrollment_date' in update_data:
            raise HTTPException(status_code=403, detail="Front Desk cannot edit admission date. Please contact Branch Admin.")
    
    # If program_id is being updated, also update program_name
    if 'program_id' in update_data:
        program = await db.programs.find_one({"id": update_data['program_id']}, {"_id": 0})
        if program:
            update_data['program_name'] = program.get('name', '')
            # Update fee_quoted if not explicitly provided
            if 'fee_quoted' not in update_data:
                update_data['fee_quoted'] = program.get('fee', 0)
    
    # Recalculate final_fee if discount or fee is changed AND final_fee is NOT explicitly provided by the user.
    # This preserves manual final_fee edits by Branch Admin.
    final_fee_explicit = 'final_fee' in update_data
    if not final_fee_explicit and any(k in update_data for k in ['discount_percent', 'discount_amount', 'fee_quoted']):
        fee_quoted = update_data.get('fee_quoted', enrollment.get('fee_quoted', 0))
        discount_amount = update_data.get('discount_amount', enrollment.get('discount_amount', 0)) or 0
        discount_percent = update_data.get('discount_percent', enrollment.get('discount_percent', 0)) or 0
        
        # discount_amount takes priority over discount_percent
        if discount_amount > 0:
            update_data['final_fee'] = fee_quoted - discount_amount
            update_data['discount_percent'] = 0  # Clear percent if amount is used
        else:
            update_data['final_fee'] = fee_quoted * (1 - discount_percent / 100)
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    await db.enrollments.update_one(
        {"id": enrollment_id},
        {"$set": update_data}
    )
    
    return {"message": "Student details updated successfully"}


@api_router.put("/students/{enrollment_id}/status")
async def update_enrollment_status(enrollment_id: str, status: str, reason: str = "", current_user: User = Depends(get_current_user)):
    """Update enrollment status (Active, Dropped, Inactive, Cancelled) - Branch Admin only"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can update enrollment status")
    
    allowed_statuses = ["Active", "Dropped", "Inactive", "Cancelled"]
    if status not in allowed_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Allowed: {', '.join(allowed_statuses)}")
    
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    # Prevent changing status if already Completed
    if enrollment.get('status') == 'Completed':
        raise HTTPException(status_code=400, detail="Cannot change status of a completed course. Course completion is final.")
    
    # Branch Admin can only update their branch's enrollments
    if current_user.role == UserRole.BRANCH_ADMIN and enrollment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only update enrollments from your branch")
    
    update_data = {"status": status}
    if status in ["Dropped", "Cancelled", "Inactive"]:
        update_data["cancelled_at"] = datetime.now(timezone.utc).isoformat()
        update_data["cancelled_by"] = current_user.id
        update_data["cancellation_reason"] = reason
    
    await db.enrollments.update_one({"id": enrollment_id}, {"$set": update_data})
    
    return {"message": f"Enrollment status updated to {status}"}


@api_router.delete("/students/{enrollment_id}/permanent")
async def permanently_delete_student(
    enrollment_id: str,
    reason: str = "",
    current_user: User = Depends(get_current_user)
):
    """Permanently delete a student — Branch Admin/Admin only.
    Only allowed if NO fee has been received for this enrollment (total_paid == 0).
    Works regardless of enrollment status.
    Also removes related payment plans, installment schedules, batch assignments, attendance, and add-ons.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin or Admin can permanently delete a student")

    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        # Idempotent: enrollment already deleted (stale UI row, duplicate click, etc.).
        # Return 200 with a friendly message so the frontend can just refresh its list
        # instead of showing a scary "not found" error to the user.
        return {
            "message": "Student was already removed. List refreshed.",
            "already_deleted": True,
        }

    # Branch access check
    if current_user.role == UserRole.BRANCH_ADMIN and enrollment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="You can only delete students from your branch")

    # Reject if any fee has been received
    payments_agg = await db.payments.aggregate([
        {"$match": {"enrollment_id": enrollment_id}},
        {"$group": {"_id": None, "total_paid": {"$sum": "$amount"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    total_paid = (payments_agg[0]["total_paid"] if payments_agg else 0) or 0
    payment_count = (payments_agg[0]["count"] if payments_agg else 0) or 0

    if total_paid > 0 or payment_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete: ₹{int(total_paid)} has already been received against this student across {payment_count} payment(s). Refund/adjust first."
        )

    student_label = enrollment.get('student_name') or enrollment.get('enrollment_id') or enrollment_id

    # Cascade delete related records
    plans = await db.payment_plans.find({"enrollment_id": enrollment_id}, {"_id": 0, "id": 1}).to_list(100)
    plan_ids = [p['id'] for p in plans]
    if plan_ids:
        await db.installment_schedule.delete_many({"payment_plan_id": {"$in": plan_ids}})
    await db.payment_plans.delete_many({"enrollment_id": enrollment_id})
    await db.student_batch_assignments.delete_many({"enrollment_id": enrollment_id})
    await db.attendance.delete_many({"enrollment_id": enrollment_id})
    await db.addon_courses.delete_many({"enrollment_id": enrollment_id})
    await db.course_completions.delete_many({"enrollment_id": enrollment_id})

    # Finally, delete the enrollment
    await db.enrollments.delete_one({"id": enrollment_id})

    # Audit log
    try:
        await db.audit_logs.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": current_user.id,
            "user_email": current_user.email,
            "user_name": current_user.name,
            "user_role": current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role),
            "branch_id": enrollment.get('branch_id'),
            "action": "permanent_delete",
            "entity_type": "student",
            "entity_id": enrollment_id,
            "entity_name": student_label,
            "changes": {"reason": reason, "status_at_delete": enrollment.get('status')},
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass

    return {"message": f"Student '{student_label}' permanently deleted"}


# Add-on Course Endpoints
@api_router.post("/enrollments/{enrollment_id}/add-on-course")
async def add_addon_course(enrollment_id: str, addon: AddOnCourseCreate, current_user: User = Depends(get_current_user)):
    """Add an additional course to an existing enrollment"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Get enrollment
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    # Get program details
    program = await db.programs.find_one({"id": addon.program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    # Calculate final fee with discount (amount takes priority over percent)
    discount_amount = addon.discount_amount or 0
    discount_percent = addon.discount_percent or 0
    
    if discount_amount > 0:
        final_fee = addon.fee_quoted - discount_amount
    else:
        final_fee = addon.fee_quoted * (1 - discount_percent / 100)
    
    # Create add-on course record
    new_addon = AddOnCourse(
        enrollment_id=enrollment_id,
        program_id=addon.program_id,
        program_name=program['name'],
        fee_quoted=addon.fee_quoted,
        discount_percent=discount_percent if discount_amount == 0 else 0,
        discount_amount=discount_amount,
        final_fee=final_fee,
        added_by=current_user.id
    )
    
    addon_dict = new_addon.model_dump()
    addon_dict['added_at'] = addon_dict['added_at'].isoformat()
    
    await db.addon_courses.insert_one(addon_dict)
    
    # Update enrollment's final_fee
    current_final_fee = enrollment.get('final_fee', 0)
    new_total_fee = current_final_fee + final_fee
    
    await db.enrollments.update_one(
        {"id": enrollment_id},
        {"$set": {"final_fee": new_total_fee}}
    )
    
    return {
        "message": f"Add-on course '{program['name']}' added successfully",
        "addon_fee": final_fee,
        "new_total_fee": new_total_fee
    }

@api_router.get("/enrollments/{enrollment_id}/add-on-courses")
async def get_addon_courses(enrollment_id: str, current_user: User = Depends(get_current_user)):
    """Get all add-on courses for an enrollment"""
    addons = await db.addon_courses.find(
        {"enrollment_id": enrollment_id},
        {"_id": 0}
    ).to_list(100)
    return addons

# Schools/Colleges Outreach Management
@api_router.post("/organizations")
async def create_organization(org: OrganizationCreate, current_user: User = Depends(get_current_user)):
    """Create a new school/college organization"""
    new_org = Organization(
        organization_type=OrganizationType(org.organization_type),
        name=org.name,
        city=org.city,
        address=org.address,
        contact_person_name=org.contact_person_name,
        contact_number=org.contact_number,
        email=org.email,
        alternate_number=org.alternate_number,
        alternate_email=org.alternate_email,
        notes=org.notes,
        created_by=current_user.id,
        branch_id=current_user.branch_id
    )
    
    org_dict = new_org.model_dump()
    org_dict['created_at'] = org_dict['created_at'].isoformat()
    org_dict['organization_type'] = org_dict['organization_type'].value
    
    await db.organizations.insert_one(org_dict)
    
    return {"message": "Organization created successfully", "id": new_org.id}

@api_router.get("/organizations")
async def get_organizations(current_user: User = Depends(get_current_user)):
    """Get all organizations"""
    organizations = await db.organizations.find({}, {"_id": 0}).to_list(1000)
    
    # Get follow-up counts for each organization
    for org in organizations:
        followups = await db.organization_followups.find(
            {"organization_id": org['id']},
            {"_id": 0}
        ).to_list(100)
        org['followup_count'] = len(followups)
        org['last_followup'] = followups[-1] if followups else None
    
    return organizations

@api_router.get("/organizations/{org_id}")
async def get_organization(org_id: str, current_user: User = Depends(get_current_user)):
    """Get a single organization with its follow-ups"""
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    followups = await db.organization_followups.find(
        {"organization_id": org_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    org['followups'] = followups
    return org

@api_router.put("/organizations/{org_id}")
async def update_organization(org_id: str, org: OrganizationUpdate, current_user: User = Depends(get_current_user)):
    """Update an organization"""
    existing = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    update_data = {k: v for k, v in org.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.organizations.update_one({"id": org_id}, {"$set": update_data})
    
    return {"message": "Organization updated successfully"}

@api_router.delete("/organizations/{org_id}")
async def delete_organization(org_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Delete an organization"""
    result = await db.organizations.delete_one({"id": org_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    # Also delete related follow-ups
    await db.organization_followups.delete_many({"organization_id": org_id})
    
    return {"message": "Organization deleted successfully"}

@api_router.post("/organizations/{org_id}/followups")
async def create_organization_followup(org_id: str, followup: OrganizationFollowUpCreate, current_user: User = Depends(get_current_user)):
    """Add a follow-up to an organization"""
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    new_followup = OrganizationFollowUp(
        organization_id=org_id,
        follow_up_date=followup.follow_up_date,
        follow_up_time=followup.follow_up_time,
        notes=followup.notes,
        outcome=followup.outcome,
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    followup_dict = new_followup.model_dump()
    followup_dict['created_at'] = followup_dict['created_at'].isoformat()
    
    await db.organization_followups.insert_one(followup_dict)
    
    return {"message": "Follow-up added successfully", "id": new_followup.id}

@api_router.get("/organizations/{org_id}/followups")
async def get_organization_followups(org_id: str, current_user: User = Depends(get_current_user)):
    """Get all follow-ups for an organization"""
    followups = await db.organization_followups.find(
        {"organization_id": org_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return followups

# ========== BATCH MANAGEMENT ==========
@api_router.get("/trainers")
async def get_trainers(current_user: User = Depends(get_current_user)):
    """Get all trainers for batch assignment"""
    query = {"role": UserRole.TRAINER.value, "is_active": True}
    if current_user.role not in [UserRole.ADMIN]:
        query["branch_id"] = current_user.branch_id
    
    trainers = await db.users.find(query, {"_id": 0, "hashed_password": 0}).to_list(100)
    
    # Get student count for each trainer
    for trainer in trainers:
        assignments = await db.student_batch_assignments.find(
            {"trainer_id": trainer['id']},
            {"_id": 0, "enrollment_id": 1}
        ).to_list(1000)
        unique_students = len(set(a['enrollment_id'] for a in assignments))
        trainer['student_count'] = unique_students
    
    return trainers

@api_router.post("/batches")
async def create_batch(batch: BatchCreate, current_user: User = Depends(get_current_user)):
    """Create a new batch"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Admin or Branch Admin can create batches")
    
    # Get program and trainer names
    program = await db.programs.find_one({"id": batch.program_id}, {"_id": 0, "name": 1})
    trainer = await db.users.find_one({"id": batch.trainer_id}, {"_id": 0, "name": 1})
    
    new_batch = Batch(
        name=batch.name,
        program_id=batch.program_id,
        program_name=program['name'] if program else None,
        trainer_id=batch.trainer_id,
        trainer_name=trainer['name'] if trainer else None,
        branch_id=current_user.branch_id or "main",
        start_date=batch.start_date,
        end_date=batch.end_date,
        timing=batch.timing,
        max_students=batch.max_students,
        created_by=current_user.id
    )
    
    batch_dict = new_batch.model_dump()
    batch_dict['created_at'] = batch_dict['created_at'].isoformat()
    
    await db.batches.insert_one(batch_dict)
    return {"message": "Batch created successfully", "id": new_batch.id}

@api_router.get("/batches")
async def get_batches(current_user: User = Depends(get_current_user)):
    """Get all batches"""
    query = {}
    if current_user.role not in [UserRole.ADMIN]:
        query["branch_id"] = current_user.branch_id
    
    batches = await db.batches.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Get student count for each batch
    for batch in batches:
        assignments = await db.student_batch_assignments.find(
            {"batch_id": batch['id']},
            {"_id": 0}
        ).to_list(100)
        batch['student_count'] = len(assignments)
    
    return batches

@api_router.put("/batches/{batch_id}")
async def update_batch(batch_id: str, batch: BatchUpdate, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Update a batch"""
    existing = await db.batches.find_one({"id": batch_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    update_data = {k: v for k, v in batch.model_dump().items() if v is not None}
    
    # Update trainer name if trainer_id changed
    if 'trainer_id' in update_data:
        trainer = await db.users.find_one({"id": update_data['trainer_id']}, {"_id": 0, "name": 1})
        update_data['trainer_name'] = trainer['name'] if trainer else None
    
    await db.batches.update_one({"id": batch_id}, {"$set": update_data})
    return {"message": "Batch updated successfully"}

@api_router.delete("/batches/{batch_id}")
async def delete_batch(batch_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Delete a batch"""
    result = await db.batches.delete_one({"id": batch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Remove all assignments for this batch
    await db.student_batch_assignments.delete_many({"batch_id": batch_id})
    return {"message": "Batch deleted successfully"}

@api_router.post("/batches/{batch_id}/assign-student")
async def assign_student_to_batch(batch_id: str, data: StudentBatchAssignmentCreate, current_user: User = Depends(get_current_user)):
    """Assign a student to a batch - Front Desk or above"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Get batch details
    batch = await db.batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get enrollment details
    enrollment = await db.enrollments.find_one({"id": data.enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Student enrollment not found")
    
    # Check if already assigned to this batch
    existing = await db.student_batch_assignments.find_one({
        "enrollment_id": data.enrollment_id,
        "batch_id": batch_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Student already assigned to this batch")
    
    assignment = StudentBatchAssignment(
        enrollment_id=data.enrollment_id,
        student_name=enrollment.get('student_name'),
        batch_id=batch_id,
        batch_name=batch.get('name'),
        trainer_id=batch.get('trainer_id'),
        trainer_name=batch.get('trainer_name'),
        assigned_by=current_user.id
    )
    
    assignment_dict = assignment.model_dump()
    assignment_dict['assigned_at'] = assignment_dict['assigned_at'].isoformat()
    
    await db.student_batch_assignments.insert_one(assignment_dict)
    return {"message": f"Student assigned to batch '{batch.get('name')}' successfully"}

@api_router.delete("/batches/{batch_id}/remove-student/{enrollment_id}")
async def remove_student_from_batch(batch_id: str, enrollment_id: str, current_user: User = Depends(get_current_user)):
    """Remove a student from a batch.
    
    Allowed for Admin, Branch Admin, and Front Desk Executive (FDE can remove directly,
    no approval gate needed — same powers as Branch Admin for batch composition).
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.student_batch_assignments.delete_one({
        "batch_id": batch_id,
        "enrollment_id": enrollment_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    return {"message": "Student removed from batch successfully"}

# Approval Requests endpoints
@api_router.get("/approval-requests")
async def get_approval_requests(current_user: User = Depends(get_current_user)):
    """Get pending approval requests for Branch Admin"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can view approval requests")
    
    query = {"status": "pending"}
    if current_user.role == UserRole.BRANCH_ADMIN:
        query["branch_id"] = current_user.branch_id
    
    requests = await db.approval_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return requests

@api_router.post("/approval-requests/{request_id}/approve")
async def approve_request(request_id: str, current_user: User = Depends(get_current_user)):
    """Approve a pending request"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can approve requests")
    
    request = await db.approval_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    if request["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Process based on request type
    if request["type"] == "batch_removal":
        # Execute the batch removal
        await db.student_batch_assignments.delete_one({
            "batch_id": request["batch_id"],
            "enrollment_id": request["enrollment_id"]
        })
    
    # Update request status
    await db.approval_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user.id,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notify the requester
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": request["requested_by"],
        "type": "approval_result",
        "title": "Request Approved",
        "message": f"Your request to remove {request.get('student_name', 'student')} from batch {request.get('batch_name', '')} has been approved.",
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    
    return {"message": "Request approved successfully"}

@api_router.post("/approval-requests/{request_id}/reject")
async def reject_request(request_id: str, current_user: User = Depends(get_current_user)):
    """Reject a pending request"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can reject requests")
    
    request = await db.approval_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    if request["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Update request status
    await db.approval_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "rejected_by": current_user.id,
            "rejected_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notify the requester
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": request["requested_by"],
        "type": "approval_result",
        "title": "Request Rejected",
        "message": f"Your request to remove {request.get('student_name', 'student')} from batch {request.get('batch_name', '')} has been rejected.",
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    
    return {"message": "Request rejected"}

@api_router.get("/batches/{batch_id}/students")
async def get_batch_students(batch_id: str, current_user: User = Depends(get_current_user)):
    """Get all students in a batch"""
    assignments = await db.student_batch_assignments.find(
        {"batch_id": batch_id},
        {"_id": 0}
    ).to_list(100)
    
    students = []
    for a in assignments:
        enrollment = await db.enrollments.find_one({"id": a['enrollment_id']}, {"_id": 0})
        if enrollment:
            students.append({
                **a,
                "phone": enrollment.get('phone'),
                "email": enrollment.get('email'),
                "program_name": enrollment.get('program_name')
            })
    
    return students

@api_router.get("/students/{enrollment_id}/batches")
async def get_student_batches(enrollment_id: str, current_user: User = Depends(get_current_user)):
    """Get all batches a student is assigned to"""
    assignments = await db.student_batch_assignments.find(
        {"enrollment_id": enrollment_id},
        {"_id": 0}
    ).to_list(100)
    return assignments

@api_router.get("/trainer-stats")
async def get_trainer_stats(current_user: User = Depends(get_current_user)):
    """Get trainer-wise student statistics for Branch Admin dashboard"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.COUNSELLOR]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"role": UserRole.TRAINER.value}
    if current_user.role != UserRole.ADMIN:
        query["branch_id"] = current_user.branch_id
    
    trainers = await db.users.find(query, {"_id": 0, "hashed_password": 0}).to_list(100)
    
    stats = []
    for trainer in trainers:
        # Get all batches for this trainer
        batches = await db.batches.find(
            {"trainer_id": trainer['id']},
            {"_id": 0, "id": 1, "name": 1, "program_name": 1, "status": 1}
        ).to_list(100)
        
        batch_ids = [b['id'] for b in batches]
        
        # Get all students assigned to this trainer's batches
        assignments = await db.student_batch_assignments.find(
            {"batch_id": {"$in": batch_ids}},
            {"_id": 0, "enrollment_id": 1, "batch_id": 1}
        ).to_list(1000)
        
        unique_student_ids = set(a['enrollment_id'] for a in assignments)
        
        # Get active students count
        active_count = 0
        if unique_student_ids:
            active_count = await db.enrollments.count_documents({
                "id": {"$in": list(unique_student_ids)},
                "status": "Active"
            })
        
        # Get completed students count
        completed_count = 0
        if unique_student_ids:
            completed_count = await db.enrollments.count_documents({
                "id": {"$in": list(unique_student_ids)},
                "status": "Completed"
            })
        
        stats.append({
            "trainer_id": trainer['id'],
            "trainer_name": trainer.get('name', trainer.get('email', 'Unknown')),
            "email": trainer.get('email'),
            "total_students": len(unique_student_ids),
            "active_students": active_count,
            "completed_students": completed_count,
            "total_batches": len(batches),
            "active_batches": len([b for b in batches if b.get('status') == 'Active']),
            "batches": batches
        })
    
    return stats

# ========== TRAINER DASHBOARD ==========
@api_router.get("/trainer/dashboard")
async def get_trainer_dashboard(current_user: User = Depends(get_current_user)):
    """Get trainer dashboard data - batches, students, and recent attendance"""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(status_code=403, detail="Only trainers can access this endpoint")
    
    # Get trainer's batches
    batches = await db.batches.find(
        {"trainer_id": current_user.id},
        {"_id": 0}
    ).to_list(100)
    
    # Get students assigned to trainer's batches
    assignments = await db.student_batch_assignments.find(
        {"trainer_id": current_user.id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get unique students
    student_ids = list(set(a['enrollment_id'] for a in assignments))
    all_students = await db.enrollments.find(
        {"id": {"$in": student_ids}},
        {"_id": 0, "id": 1, "student_name": 1, "email": 1, "phone": 1, "program_name": 1, "dob": 1, "status": 1}
    ).to_list(1000)
    
    # Separate active and completed students
    active_students = [s for s in all_students if s.get('status') != 'Completed']
    passed_students = [s for s in all_students if s.get('status') == 'Completed']
    
    # Get completion details for passed students
    for student in passed_students:
        completion = await db.course_completions.find_one(
            {"enrollment_id": student['id']},
            {"_id": 0, "completion_date": 1, "exam_status": 1, "exam_score": 1}
        )
        if completion:
            student['completion_date'] = completion.get('completion_date')
            student['exam_status'] = completion.get('exam_status')
            student['exam_score'] = completion.get('exam_score')
    
    # Calculate upcoming birthdays (next 30 days) - only for active students
    today = datetime.now(timezone.utc)
    upcoming_birthdays = []
    for student in active_students:
        if student.get('dob'):
            try:
                dob = student['dob']
                if isinstance(dob, str):
                    dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
                else:
                    dob_date = dob.date() if hasattr(dob, 'date') else dob
                
                # Create this year's birthday
                this_year_bday = dob_date.replace(year=today.year)
                if this_year_bday < today.date():
                    this_year_bday = dob_date.replace(year=today.year + 1)
                
                days_until = (this_year_bday - today.date()).days
                if 0 <= days_until <= 30:
                    upcoming_birthdays.append({
                        "student_name": student.get('student_name'),
                        "enrollment_id": student.get('id'),
                        "birthday": this_year_bday.strftime('%Y-%m-%d'),
                        "days_until": days_until
                    })
            except:
                pass
    
    # Sort by days until birthday
    upcoming_birthdays.sort(key=lambda x: x['days_until'])
    
    # Get today's attendance
    today_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    today_attendance = await db.attendance.find(
        {"batch_id": {"$in": [b['id'] for b in batches]}, "date": today_str},
        {"_id": 0}
    ).to_list(1000)
    
    # Get curriculum for programs trainer is teaching (or all if no programs assigned)
    program_ids = list(set(b['program_id'] for b in batches if b.get('program_id')))
    if program_ids:
        curricula = await db.curricula.find(
            {"program_id": {"$in": program_ids}},
            {"_id": 0}
        ).to_list(100)
    else:
        # Show all curricula if trainer has no programs assigned yet
        curricula = await db.curricula.find({}, {"_id": 0}).to_list(100)
    
    return {
        "trainer": {
            "id": current_user.id,
            "name": current_user.name,
            "email": current_user.email
        },
        "batches": batches,
        "total_students": len(active_students),
        "total_passed": len(passed_students),
        "students": active_students,
        "passed_students": passed_students,
        "upcoming_birthdays": upcoming_birthdays,
        "today_attendance": today_attendance,
        "curricula": curricula
    }

@api_router.get("/trainer/batches")
async def get_trainer_batches(current_user: User = Depends(get_current_user)):
    """Get batches for the logged-in trainer"""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(status_code=403, detail="Only trainers can access this endpoint")
    
    batches = await db.batches.find(
        {"trainer_id": current_user.id},
        {"_id": 0}
    ).sort("slot_number", 1).to_list(100)
    
    # Get student count for each batch
    for batch in batches:
        count = await db.student_batch_assignments.count_documents({"batch_id": batch['id']})
        batch['student_count'] = count
    
    return batches


# ========== STUDENT FEEDBACK SYSTEM ==========

@api_router.get("/feedback/list")
async def get_feedback_list(month: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get list of students for feedback collection - Counsellor only"""
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Default to current month
    if not month:
        month = datetime.now(timezone.utc).strftime('%Y-%m')
    
    branch_filter = {}
    if current_user.role in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN]:
        branch_filter["branch_id"] = current_user.branch_id
    
    # Get or create feedback list for the month
    feedback_list = await db.feedback_lists.find_one(
        {"month": month, **branch_filter},
        {"_id": 0}
    )
    
    if not feedback_list:
        # Generate list from active students
        students = await db.enrollments.find(
            {"status": {"$nin": ["Cancelled", "Completed"]}, **branch_filter},
            {"_id": 0, "id": 1, "student_name": 1, "program_name": 1, "phone": 1, "student_phone": 1}
        ).to_list(10000)
        
        student_list = []
        for s in students:
            # Check if feedback already given
            existing_feedback = await db.student_feedbacks.find_one({
                "enrollment_id": s['id'],
                "month": month
            }, {"_id": 0})
            
            student_list.append({
                "enrollment_id": s['id'],
                "student_name": s.get('student_name', ''),
                "student_phone": s.get('phone') or s.get('student_phone', ''),
                "program_name": s.get('program_name', ''),
                "feedback_status": "Completed" if existing_feedback else "Pending"
            })
        
        feedback_list = {
            "id": str(uuid.uuid4()),
            "month": month,
            "branch_id": current_user.branch_id,
            "students": student_list,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.feedback_lists.insert_one(feedback_list)
    else:
        # Update feedback status for each student
        for student in feedback_list.get('students', []):
            existing_feedback = await db.student_feedbacks.find_one({
                "enrollment_id": student['enrollment_id'],
                "month": month
            }, {"_id": 0})
            student['feedback_status'] = "Completed" if existing_feedback else "Pending"
    
    return feedback_list

@api_router.post("/feedback")
async def submit_feedback(feedback: StudentFeedbackCreate, current_user: User = Depends(require_role([UserRole.COUNSELLOR]))):
    """Submit student feedback - Counsellor only"""
    # Get enrollment details
    enrollment = await db.enrollments.find_one({"id": feedback.enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Check branch access
    if enrollment.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    month = datetime.now(timezone.utc).strftime('%Y-%m')
    
    # Check if already submitted for this month
    existing = await db.student_feedbacks.find_one({
        "enrollment_id": feedback.enrollment_id,
        "month": month
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(status_code=400, detail="Feedback already submitted for this student this month")
    
    new_feedback = StudentFeedback(
        **feedback.model_dump(),
        student_name=enrollment.get('student_name', ''),
        student_phone=enrollment.get('phone') or enrollment.get('student_phone', ''),
        program_name=enrollment.get('program_name', ''),
        branch_id=current_user.branch_id,
        month=month,
        collected_by=current_user.id
    )
    
    await db.student_feedbacks.insert_one(new_feedback.model_dump())
    return {"message": "Feedback submitted successfully", "id": new_feedback.id}


@api_router.get("/feedback/all")
async def get_all_feedback(
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get all student feedbacks for the branch"""
    query = {}
    if current_user.role not in [UserRole.ADMIN]:
        query["branch_id"] = current_user.branch_id
    
    feedbacks = await db.student_feedbacks.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    
    # Enrich with student and course details
    for fb in feedbacks:
        enrollment = await db.enrollments.find_one({"id": fb.get("enrollment_id")}, {"_id": 0, "student_name": 1, "program_name": 1})
        if enrollment:
            fb["student_name"] = enrollment.get("student_name", "Unknown")
            fb["course_name"] = enrollment.get("program_name", "Unknown")
        
        # Calculate average rating
        ratings = [fb.get("doubt_clearance", 0), fb.get("trainer_rating", 0), fb.get("facility_rating", 0)]
        ratings = [r for r in ratings if r > 0]
        fb["rating"] = round(sum(ratings) / len(ratings), 1) if ratings else 0
    
    return feedbacks


@api_router.get("/feedback/summary")
async def get_feedback_summary(month: Optional[str] = None, current_user: User = Depends(require_role([UserRole.BRANCH_ADMIN, UserRole.ADMIN]))):
    """Get AI-analyzed feedback summary - Branch Admin only"""
    if not month:
        month = datetime.now(timezone.utc).strftime('%Y-%m')
    
    branch_filter = {}
    if current_user.role == UserRole.BRANCH_ADMIN:
        branch_filter["branch_id"] = current_user.branch_id
    
    # Get all feedbacks for the month
    feedbacks = await db.student_feedbacks.find(
        {"month": month, **branch_filter},
        {"_id": 0}
    ).to_list(10000)
    
    if not feedbacks:
        return {
            "month": month,
            "total_feedbacks": 0,
            "summary": "No feedbacks collected for this month",
            "ai_analysis": None,
            "average_ratings": {},
            "feedbacks": []
        }
    
    # Calculate averages
    total = len(feedbacks)
    avg_doubt = sum(f.get('doubt_clearance', 0) for f in feedbacks) / total
    avg_teacher = sum(f.get('teacher_behavior', 0) for f in feedbacks) / total
    avg_facilities = sum(f.get('facilities', 0) for f in feedbacks) / total
    avg_overall = sum(f.get('overall_rating', 0) for f in feedbacks) / total
    
    # Collect remarks for AI analysis
    remarks_text = "\n".join([f.get('remarks', '') for f in feedbacks if f.get('remarks')])
    
    # Try AI analysis
    ai_analysis = None
    if LLM_AVAILABLE and os.environ.get('EMERGENT_LLM_KEY') and remarks_text:
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            
            summary_data = f"""
Student Feedback Summary for {month}:
- Total Feedbacks: {total}
- Average Doubt Clearance: {avg_doubt:.1f}/5
- Average Teacher Behavior: {avg_teacher:.1f}/5
- Average Facilities: {avg_facilities:.1f}/5
- Average Overall: {avg_overall:.1f}/5

Student Remarks:
{remarks_text[:2000]}
"""
            
            chat = LlmChat(
                api_key=os.environ.get('EMERGENT_LLM_KEY'),
                session_id=f"feedback-analysis-{month}",
                system_message="""You are an education management analyst. Analyze the student feedback and provide:
1. Key themes and patterns in the remarks
2. Areas of improvement needed
3. Positive highlights
4. Specific actionable recommendations

Keep the response concise and actionable (max 300 words)."""
            ).with_model("openai", "gpt-4o")
            
            response = await chat.send_message(UserMessage(text=summary_data))
            ai_analysis = response.strip()
            
        except Exception as e:
            logging.error(f"AI feedback analysis error: {e}")
    
    return {
        "month": month,
        "total_feedbacks": total,
        "average_ratings": {
            "doubt_clearance": round(avg_doubt, 1),
            "teacher_behavior": round(avg_teacher, 1),
            "facilities": round(avg_facilities, 1),
            "overall": round(avg_overall, 1)
        },
        "ai_analysis": ai_analysis,
        "feedbacks": feedbacks
    }

@api_router.get("/feedback/months")
async def get_feedback_months(current_user: User = Depends(require_role([UserRole.BRANCH_ADMIN, UserRole.ADMIN]))):
    """Get list of months with feedback data"""
    branch_filter = {}
    if current_user.role == UserRole.BRANCH_ADMIN:
        branch_filter["branch_id"] = current_user.branch_id
    
    feedbacks = await db.student_feedbacks.find(branch_filter, {"_id": 0, "month": 1}).to_list(10000)
    months = sorted(list(set(f.get('month') for f in feedbacks)), reverse=True)
    return months



# ========== ATTENDANCE MANAGEMENT ==========
@api_router.post("/attendance")
async def mark_attendance(attendance: AttendanceCreate, current_user: User = Depends(get_current_user)):
    """Mark attendance for a student - Trainer only"""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(status_code=403, detail="Only trainers can mark attendance")
    
    # Verify batch belongs to trainer
    batch = await db.batches.find_one({"id": attendance.batch_id, "trainer_id": current_user.id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=403, detail="You can only mark attendance for your batches")
    
    # Check if already marked
    existing = await db.attendance.find_one({
        "batch_id": attendance.batch_id,
        "enrollment_id": attendance.enrollment_id,
        "date": attendance.date
    }, {"_id": 0})
    
    if existing:
        # Update existing
        await db.attendance.update_one(
            {"id": existing['id']},
            {"$set": {"status": attendance.status, "remarks": attendance.remarks}}
        )
        return {"message": "Attendance updated successfully"}
    
    # Get student name
    enrollment = await db.enrollments.find_one({"id": attendance.enrollment_id}, {"_id": 0, "student_name": 1})
    
    new_attendance = Attendance(
        batch_id=attendance.batch_id,
        enrollment_id=attendance.enrollment_id,
        student_name=enrollment.get('student_name') if enrollment else None,
        date=attendance.date,
        status=attendance.status,
        marked_by=current_user.id,
        remarks=attendance.remarks
    )
    
    att_dict = new_attendance.model_dump()
    att_dict['marked_at'] = att_dict['marked_at'].isoformat()
    await db.attendance.insert_one(att_dict)
    
    return {"message": "Attendance marked successfully"}

@api_router.post("/attendance/bulk")
async def mark_bulk_attendance(data: AttendanceBulkCreate, current_user: User = Depends(get_current_user)):
    """Mark attendance for multiple students at once - Trainer only"""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(status_code=403, detail="Only trainers can mark attendance")
    
    # Verify batch belongs to trainer
    batch = await db.batches.find_one({"id": data.batch_id, "trainer_id": current_user.id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=403, detail="You can only mark attendance for your batches")
    
    # Check if date is valid (only today's date allowed, no past dates)
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    if data.date != today:
        # Check if attendance was already marked for the past date
        existing_for_date = await db.attendance.count_documents({
            "batch_id": data.batch_id,
            "date": data.date
        })
        if existing_for_date == 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot mark attendance for {data.date}. Attendance can only be marked for today ({today})."
            )
        # If some attendance exists, allow editing
    
    marked_count = 0
    for record in data.attendance_records:
        # Check if already marked
        existing = await db.attendance.find_one({
            "batch_id": data.batch_id,
            "enrollment_id": record['enrollment_id'],
            "date": data.date
        }, {"_id": 0})
        
        # Get student name
        enrollment = await db.enrollments.find_one({"id": record['enrollment_id']}, {"_id": 0, "student_name": 1})
        
        if existing:
            await db.attendance.update_one(
                {"id": existing['id']},
                {"$set": {"status": record['status'], "remarks": record.get('remarks')}}
            )
        else:
            new_attendance = {
                "id": str(uuid.uuid4()),
                "batch_id": data.batch_id,
                "enrollment_id": record['enrollment_id'],
                "student_name": enrollment.get('student_name') if enrollment else None,
                "date": data.date,
                "status": record['status'],
                "marked_by": current_user.id,
                "marked_at": datetime.now(timezone.utc).isoformat(),
                "remarks": record.get('remarks')
            }
            await db.attendance.insert_one(new_attendance)
        marked_count += 1
    
    return {"message": f"Attendance marked for {marked_count} students"}

@api_router.get("/attendance/{batch_id}")
async def get_batch_attendance(batch_id: str, date: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get attendance records for a batch"""
    query = {"batch_id": batch_id}
    if date:
        query["date"] = date
    
    attendance = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return attendance

@api_router.get("/attendance/student/{enrollment_id}")
async def get_student_attendance(enrollment_id: str, current_user: User = Depends(get_current_user)):
    """Get attendance history for a student"""
    attendance = await db.attendance.find(
        {"enrollment_id": enrollment_id},
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Calculate attendance stats
    total = len(attendance)
    present = len([a for a in attendance if a['status'] == 'Present'])
    absent = len([a for a in attendance if a['status'] == 'Absent'])
    late = len([a for a in attendance if a['status'] == 'Late'])
    
    return {
        "records": attendance,
        "stats": {
            "total_days": total,
            "present": present,
            "absent": absent,
            "late": late,
            "attendance_percentage": round((present / total * 100) if total > 0 else 0, 1)
        }
    }

@api_router.get("/attendance/insights/missed")
async def get_missed_attendance_insights(current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Get missed attendance insights for Branch Admin - which trainers missed marking attendance"""
    branch_id = current_user.branch_id
    branch_filter = {"branch_id": branch_id} if branch_id else {}
    
    today = datetime.now(timezone.utc)
    week_ago = (today - timedelta(days=7)).strftime('%Y-%m-%d')
    today_str = today.strftime('%Y-%m-%d')
    
    # Get all active batches in branch
    batches = await db.batches.find(
        {**branch_filter, "status": "Active"},
        {"_id": 0, "id": 1, "name": 1, "trainer_id": 1, "schedule": 1}
    ).to_list(100)
    
    # Get trainers info
    trainer_ids = list(set(b.get('trainer_id') for b in batches if b.get('trainer_id')))
    trainers = await db.users.find(
        {"id": {"$in": trainer_ids}},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(100)
    trainer_map = {t['id']: t for t in trainers}
    
    # For each batch, check if attendance was marked for the last 7 days
    missed_days = []
    trainer_insights = {}
    
    for batch in batches:
        trainer_id = batch.get('trainer_id')
        if not trainer_id:
            continue
            
        trainer = trainer_map.get(trainer_id, {})
        
        if trainer_id not in trainer_insights:
            trainer_insights[trainer_id] = {
                "trainer_id": trainer_id,
                "trainer_name": trainer.get('name', 'Unknown'),
                "trainer_email": trainer.get('email', ''),
                "total_batches": 0,
                "total_expected_days": 0,
                "marked_days": 0,
                "missed_days": 0,
                "missed_dates": []
            }
        
        trainer_insights[trainer_id]["total_batches"] += 1
        
        # Get dates when attendance was marked for this batch in last 7 days
        marked_dates = await db.attendance.distinct("date", {
            "batch_id": batch['id'],
            "date": {"$gte": week_ago, "$lte": today_str}
        })
        
        # Generate expected dates (weekdays only) - simplified: assume 6 days a week
        expected_days = 7  # Last 7 days
        trainer_insights[trainer_id]["total_expected_days"] += expected_days
        trainer_insights[trainer_id]["marked_days"] += len(marked_dates)
        
        # Calculate missed days
        missed_count = max(0, expected_days - len(marked_dates))
        trainer_insights[trainer_id]["missed_days"] += missed_count
        
        if missed_count > 0:
            # Find specific missed dates
            for i in range(7):
                check_date = (today - timedelta(days=i)).strftime('%Y-%m-%d')
                if check_date not in marked_dates:
                    missed_days.append({
                        "date": check_date,
                        "batch_name": batch['name'],
                        "batch_id": batch['id'],
                        "trainer_name": trainer.get('name', 'Unknown'),
                        "trainer_id": trainer_id
                    })
                    if check_date not in trainer_insights[trainer_id]["missed_dates"]:
                        trainer_insights[trainer_id]["missed_dates"].append(check_date)
    
    # Sort missed days by date (most recent first)
    missed_days.sort(key=lambda x: x['date'], reverse=True)
    
    # Calculate compliance percentages
    trainer_list = []
    for t in trainer_insights.values():
        t["compliance_rate"] = round((t["marked_days"] / t["total_expected_days"] * 100) if t["total_expected_days"] > 0 else 100, 1)
        trainer_list.append(t)
    
    # Sort by compliance rate (worst first)
    trainer_list.sort(key=lambda x: x["compliance_rate"])
    
    return {
        "analysis_period": {"start": week_ago, "end": today_str},
        "total_batches": len(batches),
        "total_trainers": len(trainer_ids),
        "missed_days_list": missed_days[:50],  # Top 50 most recent missed
        "trainer_insights": trainer_list,
        "summary": {
            "total_missed_days": sum(t["missed_days"] for t in trainer_list),
            "avg_compliance_rate": round(sum(t["compliance_rate"] for t in trainer_list) / len(trainer_list), 1) if trainer_list else 100
        }
    }

# ========== CURRICULUM MANAGEMENT ==========
@api_router.post("/curricula")
async def create_curriculum(curriculum: CurriculumCreate, current_user: User = Depends(require_role([UserRole.ACADEMIC_CONTROLLER]))):
    """Create curriculum for a program - Academic Controller only"""
    program = await db.programs.find_one({"id": curriculum.program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    new_curriculum = Curriculum(
        program_id=curriculum.program_id,
        program_name=program.get('name'),
        title=curriculum.title,
        description=curriculum.description,
        topics=curriculum.topics,
        duration_weeks=curriculum.duration_weeks,
        created_by=current_user.id,
        branch_id=current_user.branch_id
    )
    
    curr_dict = new_curriculum.model_dump()
    curr_dict['created_at'] = curr_dict['created_at'].isoformat()
    await db.curricula.insert_one(curr_dict)
    
    return {"message": "Curriculum created successfully", "id": new_curriculum.id}

@api_router.get("/curricula")
async def get_curricula(program_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get curricula - optionally filter by program"""
    query = {}
    if program_id:
        query["program_id"] = program_id
    
    curricula = await db.curricula.find(query, {"_id": 0}).to_list(100)
    return curricula

@api_router.get("/curricula/{curriculum_id}")
async def get_curriculum(curriculum_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific curriculum"""
    curriculum = await db.curricula.find_one({"id": curriculum_id}, {"_id": 0})
    if not curriculum:
        raise HTTPException(status_code=404, detail="Curriculum not found")
    return curriculum

@api_router.put("/curricula/{curriculum_id}")
async def update_curriculum(curriculum_id: str, data: CurriculumCreate, current_user: User = Depends(require_role([UserRole.ACADEMIC_CONTROLLER]))):
    """Update curriculum - Academic Controller only"""
    existing = await db.curricula.find_one({"id": curriculum_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Curriculum not found")
    
    program = await db.programs.find_one({"id": data.program_id}, {"_id": 0})
    
    await db.curricula.update_one(
        {"id": curriculum_id},
        {"$set": {
            "program_id": data.program_id,
            "program_name": program.get('name') if program else None,
            "title": data.title,
            "description": data.description,
            "topics": data.topics,
            "duration_weeks": data.duration_weeks
        }}
    )
    return {"message": "Curriculum updated successfully"}

@api_router.delete("/curricula/{curriculum_id}")
async def delete_curriculum(curriculum_id: str, current_user: User = Depends(require_role([UserRole.ACADEMIC_CONTROLLER]))):
    """Delete curriculum - Academic Controller only"""
    result = await db.curricula.delete_one({"id": curriculum_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Curriculum not found")
    return {"message": "Curriculum deleted successfully"}

# ========== COURSE COMPLETION ==========
@api_router.post("/course-completion")
async def mark_course_completion(
    enrollment_id: str,
    exam_status: str = "Passed",
    exam_score: Optional[float] = None,
    remarks: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Mark course as complete for a student - Trainer only"""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(status_code=403, detail="Only trainers can mark course completion")
    
    # Check if student is assigned to this trainer
    assignment = await db.student_batch_assignments.find_one(
        {"enrollment_id": enrollment_id, "trainer_id": current_user.id},
        {"_id": 0}
    )
    if not assignment:
        raise HTTPException(status_code=403, detail="Student is not assigned to you")
    
    # Get enrollment details
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    # Check if fee is cleared before allowing course completion
    final_fee = enrollment.get('final_fee', 0) or 0
    total_paid = enrollment.get('total_paid', 0) or 0
    pending_fee = final_fee - total_paid
    if pending_fee > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot mark course complete. Student has pending fee of ₹{pending_fee:,.0f}. Fee must be cleared first."
        )
    
    # Check if already completed
    existing = await db.course_completions.find_one({"enrollment_id": enrollment_id}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Course already marked as complete for this student")
    
    completion = CourseCompletion(
        enrollment_id=enrollment_id,
        student_name=enrollment.get('student_name'),
        program_id=enrollment.get('program_id'),
        program_name=enrollment.get('program_name'),
        batch_id=assignment.get('batch_id'),
        completion_date=datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        exam_status=exam_status,
        exam_score=exam_score,
        marked_by=current_user.id,
        remarks=remarks
    )
    
    comp_dict = completion.model_dump()
    comp_dict['marked_at'] = comp_dict['marked_at'].isoformat()
    await db.course_completions.insert_one(comp_dict)
    
    # Update enrollment status and remove from batch
    await db.enrollments.update_one(
        {"id": enrollment_id},
        {"$set": {
            "status": "Completed", 
            "course_completion_date": completion.completion_date,
            "course_status": "Completed",
            "batch_id": None  # Remove from batch so they don't appear in attendance
        }}
    )
    
    # Remove student from batch assignment (they've completed the course)
    await db.student_batch_assignments.delete_many({
        "enrollment_id": enrollment_id
    })
    
    # Notify FDEs about exam pending for this student
    lead = await db.leads.find_one({"id": enrollment.get("lead_id")}, {"_id": 0})
    fdes = await db.users.find({
        "branch_id": enrollment.get("branch_id"),
        "role": UserRole.FRONT_DESK.value,
        "is_active": True
    }).to_list(10)
    
    for fde in fdes:
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": fde["id"],
            "type": "exam_pending",
            "title": "Exam Pending",
            "message": f"Course completed for {lead.get('name') if lead else enrollment.get('student_name', 'Student')}. Please schedule exam.",
            "data": {"enrollment_id": enrollment_id, "student_name": lead.get('name') if lead else enrollment.get('student_name')},
            "read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
    
    return {"message": "Course marked as complete", "completion_id": completion.id}

@api_router.get("/course-completions")
async def get_course_completions(
    batch_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get course completions"""
    query = {}
    if current_user.role == UserRole.TRAINER:
        query["marked_by"] = current_user.id
    if batch_id:
        query["batch_id"] = batch_id
    
    completions = await db.course_completions.find(query, {"_id": 0}).sort("marked_at", -1).to_list(1000)
    return completions

# ========== BRANCH ADMIN FINANCIAL STATS ==========
@api_router.get("/branch-admin/demos-today")
async def get_demos_today(current_user: User = Depends(get_current_user)):
    """Get all leads with a demo scheduled for today. Used by Branch Admin / Counsellor dashboard."""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.COUNSELLOR]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    today_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    query = {
        "status": LeadStatus.DEMO_BOOKED.value,
        "demo_date": today_str,
        "is_deleted": {"$ne": True},
    }
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id:
        query["branch_id"] = current_user.branch_id
    elif current_user.role == UserRole.COUNSELLOR:
        # Counsellor sees demos for their own leads only, within their branch
        query["counsellor_id"] = current_user.id
        if current_user.branch_id:
            query["branch_id"] = current_user.branch_id
    
    leads = await db.leads.find(
        query,
        {"_id": 0, "id": 1, "name": 1, "number": 1, "program_name": 1, "demo_date": 1,
         "demo_time": 1, "trainer_name": 1, "counsellor_id": 1, "counsellor_name": 1}
    ).sort("demo_time", 1).to_list(200)
    
    return {"date": today_str, "count": len(leads), "demos": leads}


def _parse_batch_timing(timing: str):
    """Parse free-form batch timing string into (start_hour, end_hour) in 24h.
    Supports: "10:00 AM - 12:00 PM", "10-12", "9:30am - 11:30am", "14:00-16:00".
    Returns (None, None) if unparseable.
    """
    if not timing or not isinstance(timing, str):
        return None, None
    import re
    s = timing.strip().lower().replace('.', '')
    parts = re.split(r'\s*(?:-|to|–|—)\s*', s)
    if len(parts) < 2:
        return None, None
    
    def to_hour(token: str):
        token = token.strip()
        m = re.search(r'(\d{1,2})(?::(\d{1,2}))?\s*(am|pm)?', token)
        if not m:
            return None
        hour = int(m.group(1))
        minute = int(m.group(2)) if m.group(2) else 0
        meridian = m.group(3)
        if meridian == 'pm' and hour < 12:
            hour += 12
        elif meridian == 'am' and hour == 12:
            hour = 0
        # Heuristic: if no meridian and hour < 8, assume PM (e.g., "3 - 5" -> 15-17)
        if meridian is None and hour < 8:
            hour += 12
        return hour + (minute / 60.0)
    
    start = to_hour(parts[0])
    end = to_hour(parts[1])
    if start is None or end is None:
        return None, None
    if end <= start:
        end += 12
    return start, end


@api_router.get("/branch-admin/trainer-heatmap")
async def get_trainer_heatmap(current_user: User = Depends(get_current_user)):
    """Trainer load / availability heatmap for the branch.
    
    Returns hourly grid (8AM-10PM) showing how many active batches each trainer is running
    in that hour, plus an AI summary of who is overloaded vs. who has open slots.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    branch_id = current_user.branch_id if current_user.role == UserRole.BRANCH_ADMIN else None
    
    trainer_query = {"role": UserRole.TRAINER.value, "is_active": True}
    if branch_id:
        trainer_query["branch_id"] = branch_id
    trainers = await db.users.find(
        trainer_query, {"_id": 0, "id": 1, "name": 1}
    ).to_list(200)
    
    batch_query = {"status": "Active"}
    if branch_id:
        batch_query["branch_id"] = branch_id
    batches = await db.batches.find(
        batch_query,
        {"_id": 0, "id": 1, "name": 1, "trainer_id": 1, "trainer_name": 1,
         "timing": 1, "program_name": 1, "max_students": 1}
    ).to_list(500)
    
    assignments = await db.student_batch_assignments.find(
        {}, {"_id": 0, "batch_id": 1}
    ).to_list(5000)
    students_per_batch = {}
    for a in assignments:
        bid = a.get('batch_id')
        students_per_batch[bid] = students_per_batch.get(bid, 0) + 1
    
    HOUR_START = 8
    HOUR_END = 22
    hours = list(range(HOUR_START, HOUR_END))
    
    parsed_batches = []
    for b in batches:
        s, e = _parse_batch_timing(b.get('timing'))
        parsed_batches.append({
            **b,
            "start_hour": s,
            "end_hour": e,
            "student_count": students_per_batch.get(b.get('id'), 0),
        })
    
    rows = []
    for t in trainers:
        tbatches = [pb for pb in parsed_batches if pb.get('trainer_id') == t['id']]
        cells = []
        for h in hours:
            active = [pb for pb in tbatches
                      if pb['start_hour'] is not None and pb['end_hour'] is not None
                      and pb['start_hour'] < (h + 1) and pb['end_hour'] > h]
            cells.append({
                "hour": h,
                "batch_count": len(active),
                "student_count": sum(pb['student_count'] for pb in active),
                "batches": [{"id": pb['id'], "name": pb.get('name'),
                             "program": pb.get('program_name'),
                             "students": pb['student_count'],
                             "timing": pb.get('timing')} for pb in active],
            })
        total_batches = len(tbatches)
        total_students = sum(pb['student_count'] for pb in tbatches)
        busy_hours = sum(1 for c in cells if c['batch_count'] > 0)
        rows.append({
            "trainer_id": t['id'],
            "trainer_name": t.get('name', 'Unknown'),
            "cells": cells,
            "total_batches": total_batches,
            "total_students": total_students,
            "busy_hours": busy_hours,
            "free_hours": len(hours) - busy_hours,
            "utilization_pct": round((busy_hours / len(hours)) * 100) if hours else 0,
        })
    
    rows.sort(key=lambda r: (-r['total_batches'], -r['busy_hours']))
    
    free_slots = []
    for r in rows:
        for c in r['cells']:
            if c['batch_count'] == 0:
                free_slots.append({
                    "trainer_id": r['trainer_id'],
                    "trainer_name": r['trainer_name'],
                    "hour": c['hour'],
                })
    
    def _fmt_hour(h):
        ampm = 'AM' if h < 12 else 'PM'
        hh = h if h <= 12 else h - 12
        if hh == 0:
            hh = 12
        return f"{hh}{ampm}"
    
    ai_summary = None
    used_ai = False
    if LLM_AVAILABLE and os.environ.get('EMERGENT_LLM_KEY') and rows:
        try:
            compact = []
            for r in rows:
                busy_ranges = []
                start = None
                for c in r['cells']:
                    if c['batch_count'] > 0:
                        if start is None:
                            start = c['hour']
                    else:
                        if start is not None:
                            busy_ranges.append(f"{_fmt_hour(start)}-{_fmt_hour(c['hour'])}")
                            start = None
                if start is not None:
                    busy_ranges.append(f"{_fmt_hour(start)}-{_fmt_hour(HOUR_END)}")
                compact.append({
                    "trainer": r['trainer_name'],
                    "batches": r['total_batches'],
                    "students": r['total_students'],
                    "busy": busy_ranges,
                    "utilization_pct": r['utilization_pct'],
                })
            
            chat = LlmChat(
                api_key=os.environ.get('EMERGENT_LLM_KEY'),
                session_id=f"trainer-heatmap-{current_user.id}-{datetime.now().strftime('%Y%m%d%H')}",
                system_message=(
                    "You are an operations analyst for an education institute. "
                    "Given trainer schedules, produce a short, friendly summary (3-5 bullets, max 90 words total) "
                    "highlighting: (1) which trainers are most loaded, (2) which trainers have open capacity, "
                    "(3) which time slots are the institute's bottleneck, and (4) one actionable suggestion. "
                    "Do NOT use markdown headings. Plain text bullets only, each starting with '• '."
                )
            ).with_model("openai", "gpt-4o")
            
            response = await chat.send_message(UserMessage(
                text=f"Active trainer schedule (hours 8AM-10PM):\n{json.dumps(compact, indent=2)}"
            ))
            if response and response.strip():
                ai_summary = response.strip()
                used_ai = True
        except Exception as e:
            logging.warning(f"Trainer heatmap AI summary failed: {e}")
    
    if not ai_summary and rows:
        most_loaded = rows[0]
        least_loaded = rows[-1] if len(rows) > 1 else None
        lines = []
        if most_loaded['total_batches'] > 0:
            lines.append(
                f"• {most_loaded['trainer_name']} is the busiest — "
                f"{most_loaded['total_batches']} active batches, {most_loaded['total_students']} students, "
                f"{most_loaded['utilization_pct']}% utilization."
            )
        if least_loaded and least_loaded['trainer_id'] != most_loaded['trainer_id']:
            lines.append(
                f"• {least_loaded['trainer_name']} has the most open capacity — "
                f"{least_loaded['free_hours']} free hour-slots available."
            )
        slot_load = []
        for i, h in enumerate(hours):
            busy_trainers = sum(1 for r in rows if r['cells'][i]['batch_count'] > 0)
            slot_load.append((h, busy_trainers))
        slot_load.sort(key=lambda x: -x[1])
        if slot_load and slot_load[0][1] > 0:
            lines.append(
                f"• Peak hour is around {_fmt_hour(slot_load[0][0])} — "
                f"{slot_load[0][1]}/{len(rows)} trainers occupied."
            )
        lines.append("• Consider scheduling new demos / batches in the green slots to balance load.")
        ai_summary = "\n".join(lines)
    
    return {
        "hours": hours,
        "trainers": rows,
        "free_slots": free_slots,
        "ai_summary": ai_summary or "No trainers or active batches found yet.",
        "ai_powered": used_ai,
        "totals": {
            "trainer_count": len(rows),
            "active_batches": sum(r['total_batches'] for r in rows),
            "total_students": sum(r['total_students'] for r in rows),
        }
    }


@api_router.get("/branch-admin/financial-stats")
async def get_branch_financial_stats(request: Request, current_user: User = Depends(get_current_user)):
    """Get financial statistics for Branch Admin filtered by academic session - OPTIMIZED"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Admin or Branch Admin can view financial stats")
    
    branch_filter = {}
    if current_user.role == UserRole.BRANCH_ADMIN:
        branch_filter["branch_id"] = current_user.branch_id
    
    # Get academic session from token/header
    session_val = await get_session_from_request(request)
    if not session_val or session_val == "all":
        session_val = get_current_academic_session()
    
    # Get session date range (e.g., "2025" = April 1, 2025 to March 31, 2026)
    session_start, session_end = get_session_date_range(session_val)
    session_label = f"{session_val}-{int(str(session_val).split('-')[0])+1}" if session_val else ""
    
    # Current month boundaries
    now = datetime.now(timezone.utc)
    current_month_prefix = now.strftime('%Y-%m')
    
    # Session date strings for MongoDB queries
    session_start_str = session_start.strftime('%Y-%m-%d') if session_start else None
    session_end_str = session_end.strftime('%Y-%m-%d') if session_end else None
    
    # Helper to build date range query
    def date_range_query(field, start_str, end_str):
        if start_str and end_str:
            return {field: {"$gte": start_str, "$lte": end_str}}
        return {}
    
    # Use parallel aggregations for better performance
    # SESSION-FILTERED: Payments
    payment_pipeline = [
        {"$match": {**branch_filter, **(date_range_query("payment_date", session_start_str, session_end_str) if session_start_str else {})}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
    ]
    session_payment_result = await db.payments.aggregate(payment_pipeline).to_list(1)
    session_revenue = session_payment_result[0]['total'] if session_payment_result else 0
    
    # Monthly payments
    monthly_payment_pipeline = [
        {"$match": {**branch_filter, "payment_date": {"$regex": f"^{current_month_prefix}"}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    monthly_payment_result = await db.payments.aggregate(monthly_payment_pipeline).to_list(1)
    monthly_revenue = monthly_payment_result[0]['total'] if monthly_payment_result else 0
    
    # Total collections (all time)
    total_payment_pipeline = [
        {"$match": branch_filter},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    total_payment_result = await db.payments.aggregate(total_payment_pipeline).to_list(1)
    total_collections = total_payment_result[0]['total'] if total_payment_result else 0
    
    # SESSION-FILTERED: Enrollments
    enrollment_query = {**branch_filter}
    if session_start_str and session_end_str:
        enrollment_query["enrollment_date"] = {"$gte": session_start_str, "$lte": session_end_str}
    session_admissions = await db.enrollments.count_documents(enrollment_query)
    
    # Monthly admissions
    monthly_enrollment_query = {**branch_filter, "enrollment_date": {"$regex": f"^{current_month_prefix}"}}
    monthly_admissions = await db.enrollments.count_documents(monthly_enrollment_query)
    
    # Pending amounts
    pending_pipeline = [
        {"$match": branch_filter},
        {"$project": {"pending": {"$subtract": [{"$ifNull": ["$final_fee", 0]}, {"$ifNull": ["$total_paid", 0]}]}}},
        {"$group": {"_id": None, "total": {"$sum": "$pending"}}}
    ]
    pending_result = await db.enrollments.aggregate(pending_pipeline).to_list(1)
    pending_amounts = pending_result[0]['total'] if pending_result else 0
    
    # SESSION-FILTERED: Exam Bookings
    exam_query = {**branch_filter, "status": {"$ne": "Cancelled"}}
    if session_start_str and session_end_str:
        exam_query["booking_date"] = {"$gte": session_start_str, "$lte": session_end_str}
    exam_pipeline = [
        {"$match": exam_query},
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$exam_price", 0]}}}}
    ]
    session_exam_result = await db.exam_bookings.aggregate(exam_pipeline).to_list(1)
    session_exam_revenue = session_exam_result[0]['total'] if session_exam_result else 0
    
    # Total exam revenue
    total_exam_pipeline = [
        {"$match": {**branch_filter, "status": {"$ne": "Cancelled"}}},
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$exam_price", 0]}}}}
    ]
    total_exam_result = await db.exam_bookings.aggregate(total_exam_pipeline).to_list(1)
    total_exam_revenue = total_exam_result[0]['total'] if total_exam_result else 0
    
    # SESSION-FILTERED: Expenses
    expense_query = {**branch_filter}
    if session_start_str and session_end_str:
        expense_query["expense_date"] = {"$gte": session_start_str, "$lte": session_end_str}
    expense_pipeline = [
        {"$match": expense_query},
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$amount", 0]}}}}
    ]
    session_expense_result = await db.expenses.aggregate(expense_pipeline).to_list(1)
    session_expenses_total = session_expense_result[0]['total'] if session_expense_result else 0
    
    # Total expenses
    total_expense_pipeline = [
        {"$match": branch_filter},
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$amount", 0]}}}}
    ]
    total_expense_result = await db.expenses.aggregate(total_expense_pipeline).to_list(1)
    total_expenses = total_expense_result[0]['total'] if total_expense_result else 0
    
    # SESSION-FILTERED: Leads
    lead_query = {**branch_filter, "is_deleted": {"$ne": True}}
    if session_start_str and session_end_str:
        lead_query["created_at"] = {"$gte": session_start_str, "$lte": session_end_str}
    session_leads_count = await db.leads.count_documents(lead_query)
    
    # Monthly leads
    monthly_lead_query = {**branch_filter, "is_deleted": {"$ne": True}, "created_at": {"$regex": f"^{current_month_prefix}"}}
    monthly_leads_count = await db.leads.count_documents(monthly_lead_query)
    
    # Total leads
    total_leads = await db.leads.count_documents({**branch_filter, "is_deleted": {"$ne": True}})
    
    # Calculate session net revenue
    session_net_revenue = session_revenue + session_exam_revenue - session_expenses_total
    
    # ACTIVE UNIQUE STUDENTS - count distinct lead_ids
    active_pipeline = [
        {"$match": {**branch_filter, "status": {"$nin": ["Completed", "Cancelled", "Dropped"]}}},
        {"$group": {"_id": "$lead_id"}},
        {"$count": "count"}
    ]
    active_result = await db.enrollments.aggregate(active_pipeline).to_list(1)
    active_unique_students = active_result[0]['count'] if active_result else 0
    
    # Trainer-wise student count (simplified - only get counts)
    trainers_query = {"role": UserRole.TRAINER.value}
    if current_user.role == UserRole.BRANCH_ADMIN:
        trainers_query["branch_id"] = current_user.branch_id
    
    trainers = await db.users.find(trainers_query, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    # Batch get all assignments at once
    trainer_ids = [t['id'] for t in trainers]
    all_assignments = await db.student_batch_assignments.find(
        {"trainer_id": {"$in": trainer_ids}},
        {"_id": 0, "trainer_id": 1, "enrollment_id": 1}
    ).to_list(10000)
    
    # Group by trainer
    trainer_student_map = {}
    for a in all_assignments:
        tid = a.get('trainer_id')
        if tid not in trainer_student_map:
            trainer_student_map[tid] = set()
        trainer_student_map[tid].add(a.get('enrollment_id'))
    
    trainer_stats = []
    for trainer in trainers:
        trainer_stats.append({
            "trainer_id": trainer['id'],
            "trainer_name": trainer.get('name', ''),
            "unique_student_count": len(trainer_student_map.get(trainer['id'], set()))
        })
    
    return {
        # Monthly stats (for "This Month" display)
        "monthly_leads": monthly_leads_count,
        "monthly_admissions": monthly_admissions,
        "monthly_revenue": monthly_revenue,
        
        # Session-filtered stats (for "This Session" display)
        "session_label": session_label,
        "session_leads": session_leads_count,
        "session_admissions": session_admissions,
        "session_revenue": session_revenue,
        "session_expenses": session_expenses_total,
        "session_exam_revenue": session_exam_revenue,
        "session_net_revenue": session_net_revenue,
        
        # Active students count
        "active_unique_students": active_unique_students,
        
        # All-time stats (for reference)
        "total_collections": total_collections,
        "total_expenses": total_expenses,
        "total_leads": total_leads,
        "exam_revenue": total_exam_revenue,
        "pending_amounts": pending_amounts,
        "net_revenue": total_collections + total_exam_revenue - total_expenses,
        
        "trainer_stats": trainer_stats,
        "total_students": await db.enrollments.count_documents(branch_filter),
        "total_trainers": len(trainers),
        "branch_streak": _augment_streak_with_target(
            await _build_branch_streak(branch_filter.get("branch_id") or current_user.branch_id),
            await _resolve_user_streak_target(current_user.id, "Branch Admin"),
        ),
    }

# ========== CAMPAIGN MANAGEMENT ==========

@api_router.post("/campaigns")
async def create_campaign(campaign: CampaignCreate, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Create a new marketing campaign - Branch Admin only"""
    branch_id = current_user.branch_id
    if current_user.role == UserRole.ADMIN:
        # For super admin, use first branch or require selection
        first_branch = await db.branches.find_one({}, {"_id": 0, "id": 1})
        branch_id = first_branch['id'] if first_branch else None
    
    campaign_data = Campaign(
        **campaign.model_dump(),
        branch_id=branch_id,
        created_by=current_user.id
    )
    
    await db.campaigns.insert_one(campaign_data.model_dump())
    return {"message": "Campaign created successfully", "id": campaign_data.id}

@api_router.get("/campaigns")
async def get_campaigns(current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Get all campaigns for the branch"""
    query = {}
    if current_user.role == UserRole.BRANCH_ADMIN:
        query["branch_id"] = current_user.branch_id
    
    campaigns = await db.campaigns.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return campaigns

@api_router.put("/campaigns/{campaign_id}")
async def update_campaign(campaign_id: str, data: CampaignCreate, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Update a campaign"""
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    if current_user.role == UserRole.BRANCH_ADMIN and campaign.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.campaigns.update_one({"id": campaign_id}, {"$set": data.model_dump()})
    return {"message": "Campaign updated successfully"}

@api_router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Delete a campaign"""
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    if current_user.role == UserRole.BRANCH_ADMIN and campaign.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.campaigns.delete_one({"id": campaign_id})
    return {"message": "Campaign deleted successfully"}

@api_router.get("/campaigns/{campaign_id}/analytics")
async def get_campaign_analytics(campaign_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Get analytics for a specific campaign"""
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    if current_user.role == UserRole.BRANCH_ADMIN and campaign.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    start_date = campaign.get('start_date')
    end_date = campaign.get('end_date')
    branch_id = campaign.get('branch_id')
    
    # Get leads during campaign period
    leads_query = {
        "branch_id": branch_id,
        "created_at": {
            "$gte": datetime.fromisoformat(start_date) if start_date else datetime.min,
            "$lte": datetime.fromisoformat(end_date) if end_date else datetime.max
        }
    }
    leads = await db.leads.find(leads_query, {"_id": 0, "status": 1}).to_list(10000)
    
    total_leads_acquired = len(leads)
    converted_leads = len([lead for lead in leads if lead.get('status') == 'Converted'])
    conversion_rate = (converted_leads / total_leads_acquired * 100) if total_leads_acquired > 0 else 0
    
    total_spend = campaign.get('total_spend', 0)
    cost_per_lead = (total_spend / total_leads_acquired) if total_leads_acquired > 0 else 0
    cost_per_conversion = (total_spend / converted_leads) if converted_leads > 0 else 0
    
    # Lead source breakdown (if source matches platform)
    platform = campaign.get('platform', '').lower()
    leads_from_platform = len([lead for lead in leads if platform in lead.get('source', '').lower()])
    
    return {
        "campaign": campaign,
        "analytics": {
            "total_leads_acquired": total_leads_acquired,
            "leads_from_platform": leads_from_platform,
            "converted_leads": converted_leads,
            "conversion_rate": round(conversion_rate, 2),
            "total_spend": total_spend,
            "cost_per_lead": round(cost_per_lead, 2),
            "cost_per_conversion": round(cost_per_conversion, 2),
            "roi_indicator": "Positive" if conversion_rate > 10 else "Needs Improvement"
        },
        "lead_status_breakdown": {
            "new": len([lead for lead in leads if lead.get('status') == 'New']),
            "contacted": len([lead for lead in leads if lead.get('status') == 'Contacted']),
            "demo_booked": len([lead for lead in leads if lead.get('status') == 'Demo Booked']),
            "follow_up": len([lead for lead in leads if lead.get('status') == 'Follow-up']),
            "converted": converted_leads,
            "lost": len([lead for lead in leads if lead.get('status') == 'Lost'])
        }
    }

# ========== CASH HANDLING ==========

@api_router.get("/cash-handling/today")
async def get_today_cash(current_user: User = Depends(get_current_user)):
    """Get today's cash total for FDE"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    branch_filter = {}
    if current_user.role in [UserRole.BRANCH_ADMIN, UserRole.FRONT_DESK]:
        branch_filter["branch_id"] = current_user.branch_id
    
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # Get all cash payments for today
    cash_payments = await db.payments.find(
        {**branch_filter, "payment_mode": "Cash", "payment_date": today},
        {"_id": 0, "amount": 1, "student_name": 1, "payment_date": 1}
    ).to_list(1000)
    
    total_cash = sum(p.get('amount', 0) for p in cash_payments)
    
    # Check if there's an existing cash handling record for today
    existing_record = await db.cash_handling.find_one(
        {**branch_filter, "date": today},
        {"_id": 0}
    )
    
    return {
        "date": today,
        "total_cash": total_cash,
        "payments": cash_payments,
        "record": existing_record
    }

@api_router.post("/cash-handling/submit")
async def submit_cash_handling(
    deposit_receipt_url: Optional[str] = None,
    remarks: Optional[str] = None,
    manual_total: Optional[float] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.FRONT_DESK]))
):
    """Submit cash handling record with deposit receipt or remarks"""
    branch_id = current_user.branch_id
    if not branch_id and current_user.role == UserRole.ADMIN:
        first_branch = await db.branches.find_one({}, {"_id": 0, "id": 1})
        branch_id = first_branch['id'] if first_branch else None
    
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # Get total cash for today from payments
    cash_payments = await db.payments.find(
        {"branch_id": branch_id, "payment_mode": "Cash", "payment_date": today},
        {"_id": 0, "amount": 1}
    ).to_list(1000)
    system_total = sum(p.get('amount', 0) for p in cash_payments)
    
    # Use manual total if provided, otherwise use system calculated total
    total_cash = manual_total if manual_total is not None else system_total
    
    # Check existing record
    existing = await db.cash_handling.find_one({"branch_id": branch_id, "date": today})
    
    if existing:
        # Update existing record
        await db.cash_handling.update_one(
            {"branch_id": branch_id, "date": today},
            {"$set": {
                "deposit_receipt_url": deposit_receipt_url,
                "remarks": remarks,
                "status": "Deposited" if deposit_receipt_url else "Pending",
                "submitted_by": current_user.id,
                "submitted_at": datetime.now(timezone.utc),
                "total_cash": total_cash
            }}
        )
    else:
        # Create new record
        cash_record = CashHandling(
            date=today,
            branch_id=branch_id,
            total_cash=total_cash,
            deposit_receipt_url=deposit_receipt_url,
            remarks=remarks,
            status="Deposited" if deposit_receipt_url else "Pending",
            submitted_by=current_user.id,
            submitted_at=datetime.now(timezone.utc)
        )
        await db.cash_handling.insert_one(cash_record.model_dump())
    
    return {"message": "Cash handling record submitted successfully"}

@api_router.get("/cash-handling/history")
async def get_cash_handling_history(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))
):
    """Get cash handling history for Branch Admin"""
    branch_filter = {}
    if current_user.role == UserRole.BRANCH_ADMIN:
        branch_filter["branch_id"] = current_user.branch_id
    
    date_filter = {}
    if start_date:
        date_filter["$gte"] = start_date
    if end_date:
        date_filter["$lte"] = end_date
    
    query = {**branch_filter}
    if date_filter:
        query["date"] = date_filter
    
    records = await db.cash_handling.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Get all cash payments grouped by date
    all_cash_payments = await db.payments.aggregate([
        {"$match": {**branch_filter, "payment_mode": "Cash"}},
        {"$group": {
            "_id": "$payment_date",
            "total": {"$sum": "$amount"}
        }}
    ]).to_list(1000)
    
    # Create a map of date -> current total from payments
    payment_totals = {p['_id']: p['total'] for p in all_cash_payments if p['_id']}
    
    # Update existing records with current payment totals
    record_dates = set()
    for record in records:
        record_date = record.get('date')
        record_dates.add(record_date)
        # Always show current payment total for the date
        if record_date in payment_totals:
            record['current_total'] = payment_totals[record_date]
        else:
            record['current_total'] = record.get('total_cash', 0)
    
    # Add days with cash payments but no submission record
    for payment_day in all_cash_payments:
        if payment_day['_id'] and payment_day['_id'] not in record_dates:
            records.append({
                "date": payment_day['_id'],
                "total_cash": payment_day['total'],
                "current_total": payment_day['total'],
                "status": "Pending",
                "deposit_receipt_url": None,
                "remarks": None
            })
    
    # Sort by date descending
    records.sort(key=lambda x: x.get('date', ''), reverse=True)
    
    return records

# ========== TASK MANAGEMENT ==========

@api_router.post("/tasks")
async def create_task(task: TaskCreate, current_user: User = Depends(get_current_user)):
    """Create a task - Counsellor can assign to Trainer/FDE, Branch Admin can assign to anyone"""
    
    # Check permissions
    if current_user.role == UserRole.COUNSELLOR:
        # Counsellor can only assign to Trainers and FDEs
        assignee = await db.users.find_one({"id": task.assigned_to}, {"_id": 0, "role": 1, "name": 1, "email": 1})
        if not assignee:
            raise HTTPException(status_code=404, detail="Assigned user not found")
        if assignee.get('role') not in [UserRole.TRAINER.value, UserRole.FRONT_DESK.value]:
            raise HTTPException(status_code=403, detail="Counsellors can only assign tasks to Trainers and Front Desk Executives")
    elif current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="You don't have permission to create tasks")
    
    # Get assignee info
    assignee = await db.users.find_one({"id": task.assigned_to}, {"_id": 0, "name": 1, "email": 1})
    if not assignee:
        raise HTTPException(status_code=404, detail="Assigned user not found")
    
    new_task = Task(
        title=task.title,
        description=task.description,
        assigned_to=task.assigned_to,
        assigned_to_name=assignee.get('name', assignee.get('email', 'Unknown')),
        assigned_by=current_user.id,
        assigned_by_name=current_user.name or current_user.email,
        branch_id=current_user.branch_id,
        priority=task.priority,
        due_date=task.due_date
    )
    
    await db.tasks.insert_one(new_task.model_dump())
    
    # Create notification for the assignee
    notification = Notification(
        user_id=task.assigned_to,
        branch_id=current_user.branch_id,
        type="new_task",
        title="New Task Assigned",
        message=f"You have been assigned a new task: {task.title}",
        data={"task_id": new_task.id},
        play_audio=True
    )
    await db.notifications.insert_one(notification.model_dump())
    
    return {"message": "Task created successfully", "id": new_task.id}

@api_router.get("/tasks")
async def get_tasks(status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get tasks - assigned to me or created by me"""
    query = {
        "$or": [
            {"assigned_to": current_user.id},
            {"assigned_by": current_user.id}
        ]
    }
    
    if status:
        query["status"] = status
    
    tasks = await db.tasks.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return tasks

@api_router.get("/tasks/team")
async def get_team_tasks(current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Get all tasks in the branch - Branch Admin only"""
    query = {}
    if current_user.role == UserRole.BRANCH_ADMIN:
        query["branch_id"] = current_user.branch_id
    
    tasks = await db.tasks.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return tasks

@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, update: TaskUpdate, current_user: User = Depends(get_current_user)):
    """Update task - assignee can update status, creator can update other fields"""
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Check permissions
    is_assignee = task.get('assigned_to') == current_user.id
    is_creator = task.get('assigned_by') == current_user.id
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]
    
    if not (is_assignee or is_creator or is_admin):
        raise HTTPException(status_code=403, detail="You don't have permission to update this task")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    if update_data.get('status') == 'Completed':
        update_data['completed_at'] = datetime.now(timezone.utc)
    
    await db.tasks.update_one({"id": task_id}, {"$set": update_data})
    return {"message": "Task updated successfully"}

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, current_user: User = Depends(get_current_user)):
    """Delete task - only creator or admin can delete"""
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    is_creator = task.get('assigned_by') == current_user.id
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]
    
    if not (is_creator or is_admin):
        raise HTTPException(status_code=403, detail="You don't have permission to delete this task")
    
    await db.tasks.delete_one({"id": task_id})
    return {"message": "Task deleted successfully"}

# ========== PAYMENT PLAN EDIT (Branch Admin) ==========
@api_router.put("/payment-plans/{plan_id}/edit")
async def edit_payment_plan(plan_id: str, edit_data: PaymentPlanEdit, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Edit or recreate a payment plan - Branch Admin only for critical cases"""
    existing_plan = await db.payment_plans.find_one({"id": plan_id}, {"_id": 0})
    if not existing_plan:
        raise HTTPException(status_code=404, detail="Payment plan not found")
    
    enrollment_id = existing_plan.get('enrollment_id')
    
    # If installments are provided, recreate the schedule
    if edit_data.installments:
        # Delete old installments
        await db.installment_schedule.delete_many({"payment_plan_id": plan_id})
        
        # Create new installments
        for idx, inst in enumerate(edit_data.installments, 1):
            new_inst = {
                "id": str(uuid.uuid4()),
                "payment_plan_id": plan_id,
                "enrollment_id": enrollment_id,
                "installment_number": idx,
                "amount": inst.get('amount'),
                "due_date": inst.get('due_date'),
                "status": "Pending"
            }
            await db.installment_schedule.insert_one(new_inst)
        
        # Update plan
        update_data = {
            "total_installments": len(edit_data.installments),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.id
        }
        if edit_data.plan_type:
            update_data['plan_type'] = edit_data.plan_type
        
        await db.payment_plans.update_one({"id": plan_id}, {"$set": update_data})
        
        logger.info(f"Payment plan {plan_id} recreated by {current_user.email} with {len(edit_data.installments)} installments")
    
    return {"message": "Payment plan updated successfully"}

@api_router.delete("/payment-plans/{plan_id}")
async def delete_payment_plan(plan_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Delete a payment plan and its installments - for recreation"""
    existing_plan = await db.payment_plans.find_one({"id": plan_id}, {"_id": 0})
    if not existing_plan:
        raise HTTPException(status_code=404, detail="Payment plan not found")
    
    # Delete installments
    await db.installment_schedule.delete_many({"payment_plan_id": plan_id})
    
    # Delete plan
    await db.payment_plans.delete_one({"id": plan_id})
    
    logger.info(f"Payment plan {plan_id} deleted by {current_user.email}")
    return {"message": "Payment plan deleted successfully. You can now create a new plan."}

# International Exams Management
@api_router.post("/admin/exams")
async def create_exam(exam_data: dict, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Create international exam type - Super Admin only"""
    new_exam = InternationalExam(
        name=exam_data['name'],
        description=exam_data.get('description'),
        price=exam_data['price']
    )
    
    exam_dict = new_exam.model_dump()
    exam_dict['created_at'] = exam_dict['created_at'].isoformat()
    
    await db.international_exams.insert_one(exam_dict)
    return new_exam

@api_router.get("/admin/exams")
async def get_exams(current_user: User = Depends(get_current_user)):
    """Get all international exam types"""
    exams = await db.international_exams.find({"is_active": True}, {"_id": 0}).to_list(1000)
    return exams

@api_router.delete("/admin/exams/{exam_id}")
async def delete_exam(exam_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Delete/deactivate an exam type - Super Admin only"""
    await db.international_exams.update_one({"id": exam_id}, {"$set": {"is_active": False}})
    return {"message": "Exam type deleted successfully"}

@api_router.post("/exam-bookings")
async def create_exam_booking(booking_data: dict, current_user: User = Depends(get_current_user)):
    """Book an international exam for a student"""
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Please use a branch account to create bookings")
    
    # Get exam details
    exam = await db.international_exams.find_one({"id": booking_data['exam_id']}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    branch_id = current_user.branch_id
    
    # Check if student has pending fees (if enrollment_id provided)
    enrollment_id = booking_data.get('enrollment_id')
    if enrollment_id:
        enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
        if enrollment:
            final_fee = enrollment.get('final_fee', 0) or 0
            total_paid = enrollment.get('total_paid', 0) or 0
            pending_fee = final_fee - total_paid
            if pending_fee > 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot book exam. Student has pending fee of ₹{pending_fee:,.0f}. Fee must be cleared first."
                )
    
    # Generate custom booking ID
    custom_booking_id = await generate_custom_id(branch_id, "X")
    
    new_booking = ExamBooking(
        booking_id=custom_booking_id,
        student_name=booking_data['student_name'],
        student_phone=booking_data['student_phone'],
        student_email=booking_data.get('student_email'),
        exam_id=booking_data['exam_id'],
        exam_name=exam['name'],
        exam_price=exam['price'],
        branch_id=branch_id,
        exam_date=booking_data.get('exam_date'),
        notes=booking_data.get('notes'),
        created_by=current_user.id
    )
    
    booking_dict = new_booking.model_dump()
    booking_dict['created_at'] = booking_dict['created_at'].isoformat()
    
    await db.exam_bookings.insert_one(booking_dict)
    return new_booking

@api_router.get("/exam-bookings")
async def get_exam_bookings(current_user: User = Depends(get_current_user)):
    """Get exam bookings"""
    query = {}
    if current_user.role not in [UserRole.ADMIN]:
        query["branch_id"] = current_user.branch_id
    
    bookings = await db.exam_bookings.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for b in bookings:
        if isinstance(b.get('created_at'), str):
            b['created_at'] = datetime.fromisoformat(b['created_at'])
    return bookings

@api_router.put("/exam-bookings/{booking_id}/status")
async def update_booking_status(booking_id: str, status: str, current_user: User = Depends(get_current_user)):
    """Update exam booking status - handles incentive and refund tracking"""
    booking = await db.exam_bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        booking = await db.exam_bookings.find_one({"booking_id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Check branch access
    if current_user.role not in [UserRole.ADMIN] and booking.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Prevent changing status once exam is completed
    if booking.get('status') == 'Completed':
        raise HTTPException(status_code=400, detail="Cannot change status of a completed exam")
    
    update_data = {"status": status}
    
    # Handle Completed status - Calculate counsellor incentive (10%)
    if status == "Completed":
        exam_price = booking.get('exam_price', 0)
        incentive_amount = round(exam_price * 0.10, 2)  # 10% incentive
        update_data["counsellor_incentive"] = incentive_amount
        update_data["incentive_status"] = "Earned"
        update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
    
    # Handle Cancelled status - Track refund
    elif status == "Cancelled":
        exam_price = booking.get('exam_price', 0)
        update_data["incentive_status"] = "Cancelled"
        update_data["counsellor_incentive"] = 0.0
        update_data["refund_status"] = "Pending"
        update_data["refund_amount"] = exam_price
        update_data["cancelled_at"] = datetime.now(timezone.utc).isoformat()
    
    if booking.get('id'):
        update_filter = {"id": booking['id']}
    else:
        update_filter = {"booking_id": booking.get('booking_id')}
    await db.exam_bookings.update_one(update_filter, {"$set": update_data})
    return {"message": "Booking status updated successfully", "incentive_status": update_data.get("incentive_status")}

@api_router.put("/exam-bookings/{booking_id}/incentive-paid")
async def mark_incentive_paid(booking_id: str, current_user: User = Depends(get_current_user)):
    """Mark counsellor incentive as Paid (Released) - Branch Admin / Super Admin only"""
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can mark incentives as paid")
    
    # Search by both 'id' and 'booking_id' to handle old/legacy data
    booking = None
    if booking_id:
        booking = await db.exam_bookings.find_one(
            {"$or": [{"id": booking_id}, {"booking_id": booking_id}]},
            {"_id": 0}
        )
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if current_user.role != UserRole.ADMIN and booking.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Allow marking as paid only if booking is Completed
    if booking.get('status') != 'Completed':
        raise HTTPException(status_code=400, detail="Incentive can only be marked paid for completed exams")
    
    if booking.get('incentive_status') == 'Paid':
        raise HTTPException(status_code=400, detail="Incentive is already marked as paid")
    
    # Calculate incentive if not already set
    incentive_amount = booking.get('counsellor_incentive') or 0
    if not incentive_amount:
        incentive_amount = round((booking.get('exam_price') or 0) * 0.10, 2)
    
    # Use whichever ID field exists for the update
    if booking.get('id'):
        update_filter = {"id": booking['id']}
    else:
        update_filter = {"booking_id": booking.get('booking_id')}
    
    await db.exam_bookings.update_one(
        update_filter,
        {"$set": {
            "incentive_status": "Paid",
            "counsellor_incentive": incentive_amount,
            "incentive_paid_at": datetime.now(timezone.utc).isoformat(),
            "incentive_released_at": datetime.now(timezone.utc).isoformat(),
            "incentive_paid_by": current_user.id
        }}
    )
    
    return {
        "message": "Incentive marked as paid",
        "incentive_status": "Paid",
        "incentive_amount": incentive_amount,
        "booking_id": booking.get('id') or booking.get('booking_id')
    }

# Update Branch counter for exam bookings
async def update_branch_exam_counter(branch_id: str):
    """Increment the exam booking counter for a branch"""
    result = await db.branches.find_one_and_update(
        {"id": branch_id},
        {"$inc": {"exam_counter": 1}},
        return_document=True,
        projection={"_id": 0, "exam_counter": 1}
    )
    return result.get('exam_counter', 1) if result else 1


# ============ Counsellor Incentive Endpoints ============

@api_router.get("/counsellor/incentives")
async def get_counsellor_incentives(current_user: User = Depends(get_current_user)):
    """Get incentive data for the logged-in counsellor"""
    if current_user.role not in [UserRole.COUNSELLOR, UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Build query based on role
    query = {"created_by": current_user.id}
    if current_user.role == UserRole.ADMIN:
        query = {}  # Super Admin sees all
    elif current_user.role == UserRole.BRANCH_ADMIN:
        query = {"branch_id": current_user.branch_id}  # Branch Admin sees all in their branch
    
    # Get all exam bookings by this counsellor
    bookings = await db.exam_bookings.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Calculate totals
    def _inc_amt(b):
        amt = b.get('counsellor_incentive') or 0
        if not amt:
            amt = round((b.get('exam_price') or 0) * 0.10, 2)
        return amt
    
    # Earned (counsellor has earned it: includes both Earned & Paid statuses)
    total_earned = sum(_inc_amt(b) for b in bookings if b.get('incentive_status') in ('Earned', 'Paid'))
    # Released (actually paid out)
    total_released = sum(_inc_amt(b) for b in bookings if b.get('incentive_status') == 'Paid')
    # Earned but not released yet
    total_earned_unreleased = sum(_inc_amt(b) for b in bookings if b.get('incentive_status') == 'Earned')
    total_pending = sum(round((b.get('exam_price') or 0) * 0.10, 2) for b in bookings if b.get('status') in ['Pending', 'Confirmed'])
    total_cancelled = sum(b.get('refund_amount', 0) for b in bookings if b.get('status') == 'Cancelled')
    
    # Group by status for detailed view
    earned_bookings = [b for b in bookings if b.get('incentive_status') in ('Earned', 'Paid')]
    released_bookings = [b for b in bookings if b.get('incentive_status') == 'Paid']
    pending_bookings = [b for b in bookings if b.get('status') in ['Pending', 'Confirmed']]
    cancelled_bookings = [b for b in bookings if b.get('status') == 'Cancelled']
    
    return {
        "summary": {
            "total_earned": total_earned,
            "total_released": total_released,
            "total_earned_unreleased": total_earned_unreleased,
            "total_pending": total_pending,
            "total_cancelled_refunds": total_cancelled,
            "total_bookings": len(bookings),
            "completed_count": len(earned_bookings),
            "released_count": len(released_bookings),
            "pending_count": len(pending_bookings),
            "cancelled_count": len(cancelled_bookings)
        },
        "earned_bookings": earned_bookings,
        "released_bookings": released_bookings,
        "pending_bookings": pending_bookings,
        "cancelled_bookings": cancelled_bookings
    }

@api_router.get("/branch-admin/incentive-stats")
async def get_branch_incentive_stats(current_user: User = Depends(get_current_user)):
    """Get incentive statistics for Branch Admin - shows all counsellors in their branch"""
    if current_user.role not in [UserRole.BRANCH_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Only Branch Admin can view incentive stats")
    
    branch_filter = {}
    if current_user.role == UserRole.BRANCH_ADMIN:
        branch_filter["branch_id"] = current_user.branch_id
    
    # Get all exam bookings in the branch
    bookings = await db.exam_bookings.find(branch_filter, {"_id": 0}).to_list(10000)
    
    # Get counsellors in the branch
    counsellor_query = {"role": UserRole.COUNSELLOR.value}
    if current_user.role == UserRole.BRANCH_ADMIN:
        counsellor_query["branch_id"] = current_user.branch_id
    counsellors = await db.users.find(counsellor_query, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(100)
    
    # Calculate per-counsellor stats
    def _inc_amt2(b):
        amt = b.get('counsellor_incentive') or 0
        if not amt:
            amt = round((b.get('exam_price') or 0) * 0.10, 2)
        return amt
    
    counsellor_stats = []
    for counsellor in counsellors:
        c_bookings = [b for b in bookings if b.get('created_by') == counsellor['id']]
        earned = sum(_inc_amt2(b) for b in c_bookings if b.get('incentive_status') in ('Earned', 'Paid'))
        released = sum(_inc_amt2(b) for b in c_bookings if b.get('incentive_status') == 'Paid')
        earned_unreleased = sum(_inc_amt2(b) for b in c_bookings if b.get('incentive_status') == 'Earned')
        pending = sum(round((b.get('exam_price') or 0) * 0.10, 2) for b in c_bookings if b.get('status') in ['Pending', 'Confirmed'])
        
        counsellor_stats.append({
            "counsellor_id": counsellor['id'],
            "counsellor_name": counsellor['name'],
            "counsellor_email": counsellor['email'],
            "total_bookings": len(c_bookings),
            "earned_incentive": earned,
            "released_incentive": released,
            "earned_unreleased_incentive": earned_unreleased,
            "pending_incentive": pending,
            "completed_exams": len([b for b in c_bookings if b.get('status') == 'Completed']),
            "cancelled_exams": len([b for b in c_bookings if b.get('status') == 'Cancelled'])
        })
    
    # Branch totals
    total_earned = sum(_inc_amt2(b) for b in bookings if b.get('incentive_status') in ('Earned', 'Paid'))
    total_released = sum(_inc_amt2(b) for b in bookings if b.get('incentive_status') == 'Paid')
    total_earned_unreleased = sum(_inc_amt2(b) for b in bookings if b.get('incentive_status') == 'Earned')
    total_pending = sum(round((b.get('exam_price') or 0) * 0.10, 2) for b in bookings if b.get('status') in ['Pending', 'Confirmed'])
    total_refunds = sum(b.get('refund_amount', 0) for b in bookings if b.get('status') == 'Cancelled')
    
    return {
        "branch_summary": {
            "total_earned_incentives": total_earned,
            "total_released_incentives": total_released,
            "total_earned_unreleased_incentives": total_earned_unreleased,
            "total_pending_incentives": total_pending,
            "total_refunds_pending": total_refunds,
            "total_exam_bookings": len(bookings),
            "completed_exams": len([b for b in bookings if b.get('status') == 'Completed']),
            "cancelled_exams": len([b for b in bookings if b.get('status') == 'Cancelled'])
        },
        "counsellor_stats": counsellor_stats
    }

@api_router.put("/exam-bookings/{booking_id}/refund")
async def mark_refund_processed(booking_id: str, current_user: User = Depends(require_role([UserRole.BRANCH_ADMIN, UserRole.ADMIN]))):
    """Mark a cancelled exam's refund as processed - Branch Admin only"""
    booking = await db.exam_bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get('status') != 'Cancelled':
        raise HTTPException(status_code=400, detail="Only cancelled bookings can have refunds processed")
    
    if current_user.role == UserRole.BRANCH_ADMIN and booking.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.exam_bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "refund_status": "Processed",
            "refund_processed_at": datetime.now(timezone.utc).isoformat(),
            "refund_processed_by": current_user.id
        }}
    )
    
    return {"message": "Refund marked as processed"}


# ============ Quiz-Based Exams Endpoints ============

@api_router.post("/quiz-exams")
async def create_quiz_exam(exam: QuizExamCreate, current_user: User = Depends(require_role([UserRole.ACADEMIC_CONTROLLER]))):
    """Create a new quiz exam - Academic Controller only. No cap on bank size."""
    if len(exam.questions) == 0:
        raise HTTPException(status_code=400, detail="At least one question is required")
    
    # Convert question dicts to QuizQuestion objects
    questions = []
    for i, q in enumerate(exam.questions, 1):
        questions.append(QuizQuestion(
            question_number=i,
            question_text=q.get('question_text', ''),
            option_a=q.get('option_a', ''),
            option_b=q.get('option_b', ''),
            option_c=q.get('option_c', ''),
            option_d=q.get('option_d', ''),
            correct_answer=q.get('correct_answer', 'A').upper()
        ))
    
    questions_per_attempt = max(1, min(exam.questions_per_attempt or 100, len(questions)))
    
    new_exam = QuizExam(
        name=exam.name,
        description=exam.description,
        duration_minutes=exam.duration_minutes,
        pass_percentage=exam.pass_percentage,
        questions_per_attempt=questions_per_attempt,
        questions=questions,
        created_by=current_user.id
    )
    
    exam_dict = new_exam.model_dump()
    exam_dict['created_at'] = exam_dict['created_at'].isoformat()
    exam_dict['questions'] = [q.model_dump() for q in questions]
    
    await db.quiz_exams.insert_one(exam_dict)
    return {"message": "Quiz exam created successfully", "exam_id": new_exam.id, "bank_size": len(questions), "questions_per_attempt": questions_per_attempt}

@api_router.get("/quiz-exams")
async def get_quiz_exams(current_user: User = Depends(get_current_user)):
    """Get all quiz exams - for Admin/FDE"""
    exams = await db.quiz_exams.find({"is_active": True}, {"_id": 0}).to_list(100)
    for exam in exams:
        if isinstance(exam.get('created_at'), str):
            exam['created_at'] = datetime.fromisoformat(exam['created_at'])
        # Add attempt count
        attempts = await db.quiz_attempts.count_documents({"exam_id": exam['id']})
        exam['total_attempts'] = attempts
        # Count questions and remove the actual questions list
        exam['total_questions'] = len(exam.get('questions', []))
        exam.pop('questions', None)
    return exams

@api_router.get("/quiz-exams/{exam_id}")
async def get_quiz_exam_details(exam_id: str, current_user: User = Depends(get_current_user)):
    """Get full quiz exam with questions - Admin and Academic Controller only"""
    if current_user.role not in [UserRole.ADMIN, UserRole.ACADEMIC_CONTROLLER]:
        raise HTTPException(status_code=403, detail="Only Super Admin or Academic Controller can view exam details")
    
    exam = await db.quiz_exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Quiz exam not found")
    return exam

@api_router.put("/quiz-exams/{exam_id}")
async def update_quiz_exam(exam_id: str, exam: QuizExamCreate, current_user: User = Depends(require_role([UserRole.ACADEMIC_CONTROLLER]))):
    """Update a quiz exam - Academic Controller only"""
    existing = await db.quiz_exams.find_one({"id": exam_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Quiz exam not found")
    
    # Convert question dicts to QuizQuestion objects
    questions = []
    for i, q in enumerate(exam.questions, 1):
        questions.append({
            "question_number": i,
            "question_text": q.get('question_text', ''),
            "option_a": q.get('option_a', ''),
            "option_b": q.get('option_b', ''),
            "option_c": q.get('option_c', ''),
            "option_d": q.get('option_d', ''),
            "correct_answer": q.get('correct_answer', 'A').upper()
        })
    
    await db.quiz_exams.update_one(
        {"id": exam_id},
        {"$set": {
            "name": exam.name,
            "description": exam.description,
            "duration_minutes": exam.duration_minutes,
            "pass_percentage": exam.pass_percentage,
            "questions": questions
        }}
    )
    return {"message": "Quiz exam updated successfully"}

@api_router.delete("/quiz-exams/{exam_id}")
async def delete_quiz_exam(exam_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.ACADEMIC_CONTROLLER]))):
    """Delete (deactivate) a quiz exam - Admin and Academic Controller"""
    result = await db.quiz_exams.update_one({"id": exam_id}, {"$set": {"is_active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Quiz exam not found")
    return {"message": "Quiz exam deleted successfully"}

@api_router.get("/quiz-exams/import/sample.xlsx")
async def download_sample_quiz_xlsx(current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.ACADEMIC_CONTROLLER]))):
    """Download a ready-to-fill sample Excel file for quiz question bulk upload."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(["question_text", "option_a", "option_b", "option_c", "option_d", "correct_answer"])
    ws.append(["What is the capital of France?", "London", "Berlin", "Paris", "Madrid", "C"])
    ws.append(["Which planet is closest to the Sun?", "Venus", "Mercury", "Mars", "Earth", "B"])
    ws.append(["What is 2 + 2?", "3", "4", "5", "6", "B"])
    ws.append(["Who wrote Romeo and Juliet?", "Charles Dickens", "William Shakespeare", "Jane Austen", "Mark Twain", "B"])
    # Widen columns
    for col, width in enumerate([50, 22, 22, 22, 22, 16], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="quiz_questions_sample.xlsx"'}
    )


@api_router.post("/quiz-exams/import")
async def import_quiz_questions(
    file: UploadFile = File(...),
    exam_name: str = Form(...),
    description: str = Form(""),
    duration_minutes: int = Form(30),
    pass_percentage: int = Form(60),
    questions_per_attempt: int = Form(100),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.ACADEMIC_CONTROLLER]))
):
    """Import quiz questions from CSV or Excel file.
    
    Expected columns (header row, case-insensitive):
        question_text, option_a, option_b, option_c, option_d, correct_answer
    correct_answer must be A/B/C/D.
    """
    import csv
    import io
    
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported")
    
    questions = []
    
    try:
        contents = await file.read()
        
        def _parse_row(row_lower):
            """Parse a row dict (keys already lower-cased) into QuizQuestion fields."""
            question_text = (row_lower.get('question_text') or row_lower.get('question') or '').strip()
            option_a = (row_lower.get('option_a') or row_lower.get('a') or '').strip()
            option_b = (row_lower.get('option_b') or row_lower.get('b') or '').strip()
            option_c = (row_lower.get('option_c') or row_lower.get('c') or '').strip()
            option_d = (row_lower.get('option_d') or row_lower.get('d') or '').strip()
            correct = (row_lower.get('correct_answer') or row_lower.get('correct') or row_lower.get('answer') or '').strip().upper()
            if not question_text or not option_a or not option_b or not option_c or not option_d:
                return None
            if correct not in ('A', 'B', 'C', 'D'):
                return None
            return {
                "question_text": question_text,
                "option_a": option_a,
                "option_b": option_b,
                "option_c": option_c,
                "option_d": option_d,
                "correct_answer": correct,
            }
        
        if file.filename.endswith('.csv'):
            # Try common encodings
            try:
                decoded = contents.decode('utf-8-sig')
            except UnicodeDecodeError:
                decoded = contents.decode('latin-1')
            reader = csv.DictReader(io.StringIO(decoded))
            for idx, row in enumerate(reader, start=2):
                row_lower = {(k or '').lower().strip(): (str(v) if v is not None else '') for k, v in row.items()}
                parsed = _parse_row(row_lower)
                if parsed:
                    parsed["question_number"] = len(questions) + 1
                    questions.append(parsed)
        else:
            # Excel using openpyxl
            try:
                import openpyxl
                from io import BytesIO
                
                wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
                ws = wb.active
                
                headers = [(cell.value or '').lower().strip() if cell.value else '' for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    row_dict = {headers[i]: ('' if row[i] is None else str(row[i])) for i in range(min(len(headers), len(row)))}
                    parsed = _parse_row(row_dict)
                    if parsed:
                        parsed["question_number"] = len(questions) + 1
                        questions.append(parsed)
            except ImportError:
                raise HTTPException(status_code=400, detail="Excel support requires openpyxl. Please use CSV format.")
        
        if not questions:
            raise HTTPException(status_code=400, detail="No valid questions found. Check that your file has columns: question_text, option_a, option_b, option_c, option_d, correct_answer (A/B/C/D).")
        
        effective_qpa = max(1, min(questions_per_attempt or 100, len(questions)))
        
        # Create the quiz exam using the QuizExam model (consistent schema)
        quiz_exam = QuizExam(
            name=exam_name,
            description=description or None,
            duration_minutes=duration_minutes,
            pass_percentage=pass_percentage,
            questions_per_attempt=effective_qpa,
            questions=[QuizQuestion(**q) for q in questions],
            created_by=current_user.id,
        )
        exam_data = quiz_exam.model_dump()
        exam_data['created_at'] = exam_data['created_at'].isoformat()
        
        await db.quiz_exams.insert_one(exam_data)
        
        return {
            "message": "Quiz created successfully",
            "exam_id": quiz_exam.id,
            "bank_size": len(questions),
            "questions_per_attempt": effective_qpa
        }
        
    except HTTPException:
        raise
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="File encoding error. Please save the CSV as UTF-8.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

def _get_public_base_url(request: Request) -> str:
    """Resolve the public-facing base URL for links shared with users (QR codes, webhooks).
    Priority:
      1. FRONTEND_URL env var (explicit deployment override — set this on production VPS)
      2. X-Forwarded-Proto + X-Forwarded-Host (when behind nginx/cloudflare proxy)
      3. Request host header
    Never falls back to a hardcoded preview/dev URL.
    """
    explicit = os.environ.get('FRONTEND_URL', '').strip()
    if explicit:
        return explicit.rstrip('/')
    proto = request.headers.get('x-forwarded-proto') or request.url.scheme or 'https'
    host = request.headers.get('x-forwarded-host') or request.headers.get('host') or request.url.netloc
    return f"{proto}://{host}".rstrip('/')



@api_router.get("/quiz-exams/{exam_id}/qr-code")
async def get_quiz_qr_code(request: Request, exam_id: str, current_user: User = Depends(get_current_user)):
    """Generate QR code for quiz link"""
    exam = await db.quiz_exams.find_one({"id": exam_id}, {"_id": 0, "name": 1, "title": 1})
    if not exam:
        raise HTTPException(status_code=404, detail="Quiz exam not found")
    
    # Get exam name (support both 'name' and 'title' fields)
    exam_name = exam.get('name') or exam.get('title') or 'Quiz'
    
    # Generate the public quiz URL - use /exam route which matches frontend App.js.
    # Priority: FRONTEND_URL env (if explicitly set in deployment) > auto-derive from request.
    frontend_url = _get_public_base_url(request)
    quiz_url = f"{frontend_url}/exam/{exam_id}"
    
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(quiz_url)
    qr.make(fit=True)
    
    # Create image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to bytes
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    
    # Return as base64 for easy embedding in frontend
    img_base64 = base64.b64encode(img_bytes.read()).decode('utf-8')
    
    return {
        "exam_name": exam_name,
        "quiz_url": quiz_url,
        "qr_code_base64": f"data:image/png;base64,{img_base64}"
    }

# Public endpoints for students to take exams
@api_router.get("/public/quiz/{exam_id}")
async def get_public_quiz(exam_id: str):
    """Get quiz exam metadata for students. Questions are NOT returned here;
    they are returned by /start with a random subset per attempt."""
    exam = await db.quiz_exams.find_one({"id": exam_id, "is_active": True}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Quiz exam not found or inactive")
    
    bank_size = len(exam.get('questions', []))
    questions_per_attempt = min(exam.get('questions_per_attempt') or 100, bank_size)
    
    return {
        "id": exam['id'],
        "name": exam.get('name') or exam.get('title') or 'Quiz',
        "description": exam.get('description'),
        "duration_minutes": exam.get('duration_minutes') or exam.get('duration', 30),
        "pass_percentage": exam.get('pass_percentage', 60),
        "total_questions": questions_per_attempt,
        "bank_size": bank_size,
    }

@api_router.post("/public/quiz/{exam_id}/start")
async def start_quiz_attempt(exam_id: str, data: QuizAttemptCreate):
    """Start a quiz attempt. Returns a RANDOM subset of questions (shuffled)
    based on exam.questions_per_attempt (default 100)."""
    import random as _random
    exam = await db.quiz_exams.find_one({"id": exam_id, "is_active": True}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Quiz exam not found or inactive")
    
    # Verify enrollment number
    enrollment = await db.enrollments.find_one(
        {"enrollment_id": data.enrollment_number}, {"_id": 0, "student_name": 1}
    )
    if not enrollment:
        enrollment = await db.enrollments.find_one(
            {"id": data.enrollment_number}, {"_id": 0, "student_name": 1}
        )
    student_name = enrollment.get('student_name') if enrollment else None
    
    # Randomly pick questions_per_attempt questions from the bank
    bank = exam.get('questions', [])
    qpa = min(exam.get('questions_per_attempt') or 100, len(bank))
    sampled = _random.sample(bank, qpa) if qpa > 0 else []
    # Shuffle order as well
    _random.shuffle(sampled)
    
    # Re-number for the student (1..qpa); keep original number for scoring
    questions_for_student = []
    student_to_original = {}  # student_qnum -> original qnum
    for idx, q in enumerate(sampled, start=1):
        orig_num = q.get('question_number')
        student_to_original[str(idx)] = orig_num
        questions_for_student.append({
            "question_number": idx,
            "question_text": q.get('question_text'),
            "option_a": q.get('option_a'),
            "option_b": q.get('option_b'),
            "option_c": q.get('option_c'),
            "option_d": q.get('option_d'),
        })
    
    attempt = QuizAttempt(
        exam_id=exam_id,
        exam_name=exam.get('name') or exam.get('title') or 'Quiz',
        enrollment_number=data.enrollment_number,
        student_name=student_name,
        question_numbers=[q.get('question_number') for q in sampled],
        total_questions=len(sampled),
    )
    attempt_dict = attempt.model_dump()
    attempt_dict['started_at'] = attempt_dict['started_at'].isoformat()
    attempt_dict['student_to_original'] = student_to_original
    
    await db.quiz_attempts.insert_one(attempt_dict)
    
    return {
        "attempt_id": attempt.id,
        "student_name": student_name,
        "total_questions": len(sampled),
        "duration_minutes": exam.get('duration_minutes') or exam.get('duration', 30),
        "questions": questions_for_student,
    }

@api_router.post("/public/quiz/attempt/{attempt_id}/submit")
async def submit_quiz_attempt(attempt_id: str, submission: QuizAttemptSubmit):
    """Submit quiz answers. Scoring uses ONLY the questions that were
    randomly picked for this attempt."""
    attempt = await db.quiz_attempts.find_one({"id": attempt_id}, {"_id": 0})
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    
    if attempt.get('completed_at'):
        raise HTTPException(status_code=400, detail="This attempt has already been submitted")
    
    exam = await db.quiz_exams.find_one({"id": attempt['exam_id']}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Quiz exam not found")
    
    # Map student_qnum -> original qnum; fallback for legacy attempts (use all exam questions)
    student_to_original = attempt.get('student_to_original') or {}
    bank_by_num = {q.get('question_number'): q for q in exam.get('questions', [])}
    
    # Determine the list of questions to score
    if student_to_original:
        scored_questions = []
        for student_qnum_str, orig_qnum in student_to_original.items():
            q = bank_by_num.get(orig_qnum)
            if q:
                scored_questions.append((student_qnum_str, q))
    else:
        # Legacy attempt: score against every bank question using original numbering
        scored_questions = [(str(q.get('question_number')), q) for q in exam.get('questions', [])]
    
    correct_count = 0
    for student_qnum_str, q in scored_questions:
        student_answer = (submission.answers.get(student_qnum_str) or '').upper()
        if student_answer == (q.get('correct_answer') or '').upper():
            correct_count += 1
    
    total_questions = len(scored_questions)
    percentage = (correct_count / total_questions * 100) if total_questions > 0 else 0
    passed = percentage >= exam.get('pass_percentage', 60)
    
    started_at = attempt.get('started_at')
    if isinstance(started_at, str):
        started_at = datetime.fromisoformat(started_at)
    time_taken = int((datetime.now(timezone.utc) - started_at).total_seconds())
    
    await db.quiz_attempts.update_one(
        {"id": attempt_id},
        {"$set": {
            "answers": submission.answers,
            "score": correct_count,
            "total_questions": total_questions,
            "percentage": round(percentage, 1),
            "passed": passed,
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "time_taken_seconds": time_taken
        }}
    )
    
    return {
        "score": correct_count,
        "total_questions": total_questions,
        "percentage": round(percentage, 1),
        "passed": passed,
        "result": "PASS" if passed else "FAIL",
        "time_taken_seconds": time_taken
    }

@api_router.get("/quiz-attempts")
async def get_quiz_attempts(exam_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get all quiz attempts - Admin/FDE can see all"""
    query = {}
    if exam_id:
        query["exam_id"] = exam_id
    
    attempts = await db.quiz_attempts.find(query, {"_id": 0}).sort("started_at", -1).to_list(1000)
    for a in attempts:
        if isinstance(a.get('started_at'), str):
            a['started_at'] = datetime.fromisoformat(a['started_at'])
        if isinstance(a.get('completed_at'), str):
            a['completed_at'] = datetime.fromisoformat(a['completed_at'])
    return attempts

# ============================================
# WEBHOOK ENDPOINTS - External Lead Capture
# ============================================

@api_router.post("/webhooks/leads/{webhook_key}")
async def webhook_lead_capture(webhook_key: str, lead_data: WebhookLeadCreate):
    """
    Public webhook endpoint for capturing leads from Google Ads, Meta (Facebook), etc.
    Each branch has a unique webhook_key for security.
    
    Example URLs:
    - Google Ads: https://yourapp.com/api/webhooks/leads/{branch_webhook_key}
    - Meta/Facebook: https://yourapp.com/api/webhooks/leads/{branch_webhook_key}
    """
    # Find branch by webhook key
    branch = await db.branches.find_one({"webhook_key": webhook_key}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Invalid webhook key")
    
    # Find or create a default program for webhook leads
    default_program = await db.programs.find_one({}, {"_id": 0})
    if not default_program:
        raise HTTPException(status_code=400, detail="No programs configured in the system")
    
    # Try to match program by name if provided
    program = default_program
    if lead_data.program_name:
        matched_program = await db.programs.find_one(
            {"name": {"$regex": lead_data.program_name, "$options": "i"}}, 
            {"_id": 0}
        )
        if matched_program:
            program = matched_program
    
    # Determine lead source
    source = lead_data.source or "Webhook"
    if lead_data.campaign:
        source = f"{source} - {lead_data.campaign}"
    
    # Generate custom lead ID
    custom_lead_id = await generate_custom_id(branch['id'], "L")
    
    # Create the lead
    new_lead = Lead(
        lead_id=custom_lead_id,
        name=lead_data.name,
        number=lead_data.phone,
        email=lead_data.email or f"webhook_{datetime.now().timestamp()}@placeholder.com",
        program_id=program['id'],
        program_name=program['name'],
        lead_source=source,
        branch_id=branch['id'],
        counsellor_id="system",  # System-generated lead
        city=lead_data.city,
        state=lead_data.state,
    )
    
    lead_dict = new_lead.model_dump()
    lead_dict['created_at'] = lead_dict['created_at'].isoformat()
    lead_dict['updated_at'] = lead_dict['updated_at'].isoformat()
    lead_dict['webhook_metadata'] = {
        "campaign": lead_data.campaign,
        "ad_name": lead_data.ad_name,
        "form_name": lead_data.form_name,
        "additional_data": lead_data.additional_data
    }
    
    await db.leads.insert_one(lead_dict)
    
    # Send WhatsApp notification for new enquiry
    await send_whatsapp_notification(
        new_lead.number, 
        "enquiry_saved",
        {"name": new_lead.name, "course": program['name']}
    )
    
    logging.info(f"Webhook lead captured: {new_lead.name} for branch {branch['name']}")
    
    return {
        "success": True,
        "lead_id": new_lead.lead_id,
        "message": f"Lead captured successfully for {branch['name']}"
    }

@api_router.post("/admin/branches/{branch_id}/regenerate-webhook-key")
async def regenerate_webhook_key(branch_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Regenerate webhook key for a branch - Super Admin only"""
    import secrets
    new_key = secrets.token_urlsafe(32)
    
    result = await db.branches.update_one(
        {"id": branch_id},
        {"$set": {"webhook_key": new_key}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    return {"webhook_key": new_key, "message": "Webhook key regenerated successfully"}

@api_router.get("/admin/branches/{branch_id}/webhook-info")
async def get_branch_webhook_info(request: Request, branch_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Get webhook URL and key for a branch - Super Admin only"""
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    # Generate webhook key if not exists
    if not branch.get('webhook_key'):
        import secrets
        webhook_key = secrets.token_urlsafe(32)
        await db.branches.update_one({"id": branch_id}, {"$set": {"webhook_key": webhook_key}})
        branch['webhook_key'] = webhook_key
    
    base_url = _get_public_base_url(request)
    webhook_url = f"{base_url}/api/webhooks/leads/{branch['webhook_key']}"
    
    return {
        "branch_id": branch_id,
        "branch_name": branch['name'],
        "webhook_key": branch['webhook_key'],
        "webhook_url": webhook_url,
        "usage_instructions": {
            "google_ads": "Use this URL as your Google Ads Lead Form Webhook URL",
            "meta": "Use this URL as your Facebook Lead Ads Webhook URL",
            "method": "POST",
            "content_type": "application/json",
            "sample_payload": {
                "name": "John Doe",
                "phone": "9876543210",
                "email": "john@example.com",
                "source": "Google Ads",
                "campaign": "Summer Course 2024",
                "program_name": "Optional - Will match with existing programs"
            }
        }
    }

# ============================================
# BROWSER PUSH NOTIFICATION SUBSCRIPTIONS
# ============================================

@api_router.post("/push-subscriptions")
async def save_push_subscription(subscription: PushSubscriptionCreate, current_user: User = Depends(get_current_user)):
    """Save browser push subscription for a user"""
    # Remove existing subscription for this user/endpoint
    await db.push_subscriptions.delete_many({
        "$or": [
            {"user_id": current_user.id},
            {"endpoint": subscription.endpoint}
        ]
    })
    
    new_sub = PushSubscription(
        user_id=current_user.id,
        endpoint=subscription.endpoint,
        keys=subscription.keys
    )
    
    sub_dict = new_sub.model_dump()
    sub_dict['created_at'] = sub_dict['created_at'].isoformat()
    
    await db.push_subscriptions.insert_one(sub_dict)
    return {"message": "Push subscription saved successfully"}

@api_router.delete("/push-subscriptions")
async def remove_push_subscription(current_user: User = Depends(get_current_user)):
    """Remove browser push subscription for current user"""
    await db.push_subscriptions.delete_many({"user_id": current_user.id})
    return {"message": "Push subscription removed"}

@api_router.get("/push-subscriptions/vapid-public-key")
async def get_vapid_public_key():
    """Get VAPID public key for browser push notifications"""
    # In production, generate proper VAPID keys
    vapid_public_key = os.environ.get('VAPID_PUBLIC_KEY', 'BEl62iUYgUivxIkv69yViEuiBIa-Ib9-SkvMeAtA3LFgDzkrxZJjSgSnfckjBJuBkr3qBUYIHBQFLXYp5Nksh8U')
    return {"publicKey": vapid_public_key}

# ============================================
# FOLLOW-UP REMINDERS - Due Soon Notifications
# ============================================

@api_router.get("/followups/due-soon")
async def get_due_soon_followups(current_user: User = Depends(get_current_user)):
    """
    Get follow-ups due within the next 10 minutes.
    Used by frontend to show alarm notifications.
    """
    now = datetime.now(timezone.utc)
    ten_minutes_later = now + timedelta(minutes=10)
    
    query = {
        "status": FollowUpStatus.PENDING.value,
        "followup_date": {
            "$gte": now.isoformat(),
            "$lte": ten_minutes_later.isoformat()
        }
    }
    
    # Filter by user if not admin
    if current_user.role != UserRole.ADMIN:
        query["created_by"] = current_user.id
    
    followups = await db.followups.find(query, {"_id": 0}).sort("followup_date", 1).to_list(100)
    
    for fu in followups:
        if isinstance(fu.get('created_at'), str):
            fu['created_at'] = datetime.fromisoformat(fu['created_at'])
        if isinstance(fu.get('followup_date'), str):
            fu['followup_date'] = datetime.fromisoformat(fu['followup_date'])
    
    return followups

@api_router.get("/followups/overdue")
async def get_overdue_followups(current_user: User = Depends(get_current_user)):
    """Get overdue follow-ups that haven't been completed"""
    now = datetime.now(timezone.utc)
    
    query = {
        "status": FollowUpStatus.PENDING.value,
        "followup_date": {"$lt": now.isoformat()}
    }
    
    if current_user.role != UserRole.ADMIN:
        query["created_by"] = current_user.id
    
    followups = await db.followups.find(query, {"_id": 0}).sort("followup_date", 1).to_list(100)
    
    for fu in followups:
        if isinstance(fu.get('created_at'), str):
            fu['created_at'] = datetime.fromisoformat(fu['created_at'])
        if isinstance(fu.get('followup_date'), str):
            fu['followup_date'] = datetime.fromisoformat(fu['followup_date'])
    
    return followups

# ============================================
# TASK NOTIFICATIONS - Auto-notify on task assignment
# ============================================

# ============================================
# CERTIFICATE MANAGEMENT SYSTEM
# ============================================

async def generate_certificate_id():
    """Generate unique certificate ID: ETI-2025-00001"""
    year = datetime.now().year
    # Get the count of certificates this year
    count = await db.certificate_requests.count_documents({
        "created_at": {"$regex": f"^{year}"}
    })
    return f"ETI-{year}-{str(count + 1).zfill(5)}"

async def generate_verification_id():
    """Generate unique verification ID for QR code"""
    import secrets
    return secrets.token_urlsafe(16)

@api_router.get("/public/enrollment/{enrollment_number}")
async def get_enrollment_for_certificate(enrollment_number: str):
    """Public endpoint to fetch enrollment details and all courses for certificate request.
    
    Finds the enrollment by number, then returns ALL courses (enrollments) that belong
    to the SAME student (matched by phone + branch for safety, fallback to name).
    """
    # Find the enrollment by enrollment number
    enrollment = await db.enrollments.find_one(
        {"enrollment_id": enrollment_number},
        {"_id": 0}
    )
    
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    
    student_name = enrollment.get('student_name', '')
    student_phone = enrollment.get('phone') or enrollment.get('student_phone') or ''
    branch_id = enrollment.get('branch_id')
    
    # Build match filter: prefer (phone AND branch) – most reliable. Fall back to name+branch.
    match_filters = []
    if student_phone:
        match_filters.append({
            "$or": [{"phone": student_phone}, {"student_phone": student_phone}],
            "branch_id": branch_id
        })
    if student_name:
        match_filters.append({"student_name": student_name, "branch_id": branch_id})
    
    all_enrollments = []
    seen_ids = set()
    for f in match_filters:
        docs = await db.enrollments.find(f, {"_id": 0}).to_list(100)
        for d in docs:
            if d.get('id') not in seen_ids:
                all_enrollments.append(d)
                seen_ids.add(d.get('id'))
    
    # Fallback in case above didn't match anything (legacy records missing phone)
    if not all_enrollments:
        all_enrollments = [enrollment]
    
    # Get branch details
    branch = await db.branches.find_one({"id": enrollment['branch_id']}, {"_id": 0, "name": 1})
    
    # Build courses list from all enrollments
    courses = []
    for enroll in all_enrollments:
        program = await db.programs.find_one({"id": enroll.get('program_id')}, {"_id": 0, "name": 1, "duration": 1}) if enroll.get('program_id') else None
        
        # Check fee status for this enrollment
        final_fee = enroll.get('final_fee', 0) or 0
        total_paid = enroll.get('total_paid', 0) or 0
        pending_fee = final_fee - total_paid
        fee_cleared = pending_fee <= 0
        
        # Check if certificate already requested for this enrollment
        existing_cert = await db.certificate_requests.find_one({
            "enrollment_number": enroll['enrollment_id'],
            "status": {"$in": ["Pending", "Approved", "Ready"]}
        })
        
        # Course completion check (was the course marked complete by the trainer?)
        course_completion = await db.course_completions.find_one(
            {"enrollment_id": enroll['id']}, {"_id": 0}
        )
        
        courses.append({
            "enrollment_id": enroll['enrollment_id'],
            "enrollment_db_id": enroll['id'],
            "program_id": enroll.get('program_id'),
            "program_name": (program['name'] if program else enroll.get('program_name', '')),
            "program_duration": (program['duration'] if program else ''),
            "enrollment_date": enroll.get('enrollment_date', enroll.get('created_at', '')),
            "fee_cleared": fee_cleared,
            "pending_fee": pending_fee if pending_fee > 0 else 0,
            "course_completed": course_completion is not None,
            "enrollment_status": enroll.get('status', 'Active'),
            "certificate_requested": existing_cert is not None,
            "certificate_status": existing_cert['status'] if existing_cert else None
        })
    
    return {
        "student_name": enrollment['student_name'],
        "phone": enrollment.get('phone', '') or enrollment.get('student_phone', ''),
        "email": enrollment.get('email', ''),
        "branch_name": branch['name'] if branch else '',
        "branch_id": enrollment['branch_id'],
        "courses": courses
    }

@api_router.post("/public/certificate-requests")
async def create_certificate_request(request_data: CertificateRequestCreate):
    """Public endpoint for students to request certificates.
    
    Eligibility checks (closed-loop):
      1. Enrollment must exist
      2. All fees must be cleared for that enrollment
      3. The course/exam must be marked complete by a trainer (course_completions entry)
      4. The exam must have been passed (not failed)
      5. No duplicate active request for the same enrollment
    """
    # Fetch enrollment
    enrollment = await db.enrollments.find_one(
        {"enrollment_id": request_data.enrollment_number},
        {"_id": 0}
    )
    
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found with this ID")
    
    # Check if fees are fully paid
    final_fee = enrollment.get('final_fee', 0) or 0
    total_paid = enrollment.get('total_paid', 0) or 0
    pending_fee = final_fee - total_paid
    
    if pending_fee > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot request certificate. Please clear pending fee of ₹{pending_fee:,.0f} first."
        )
    
    # Check that the course / exam is completed (trainer has marked complete)
    course_completion = await db.course_completions.find_one(
        {"enrollment_id": enrollment['id']}, {"_id": 0}
    )
    if not course_completion:
        raise HTTPException(
            status_code=400,
            detail="Cannot request certificate. Course/exam is not yet marked complete by the trainer."
        )
    
    # Exam must not be failed
    exam_status = (course_completion.get('exam_status') or '').strip()
    if exam_status and exam_status.lower() in ['failed', 'fail']:
        raise HTTPException(
            status_code=400,
            detail="Cannot request certificate. The exam result is marked as Failed."
        )
    
    # NOTE: Duplicate certificate requests are ALLOWED — one student may hold multiple
    # certificates against the same enrollment (e.g. reissue, name correction, multiple
    # skill modules under the same enrollment). Each request produces its own unique
    # certificate_id, registration_number and verification_id, so downstream download +
    # public-verify flows remain independent per certificate.
    existing = None
    
    # Get branch and program details
    branch = await db.branches.find_one({"id": enrollment['branch_id']}, {"_id": 0, "name": 1})
    program = await db.programs.find_one({"id": enrollment.get('program_id')}, {"_id": 0, "name": 1, "duration": 1}) if enrollment.get('program_id') else None
    
    # Generate IDs — registration_number has been removed from the product.
    # We keep only certificate_id + enrollment_number as identifiers on the certificate.
    certificate_id = await generate_certificate_id()
    verification_id = await generate_verification_id()
    registration_number = ""
    
    # Create certificate request
    cert_request = CertificateRequest(
        certificate_id=certificate_id,
        enrollment_id=enrollment['id'],
        enrollment_number=request_data.enrollment_number,
        student_name=enrollment['student_name'],
        program_name=program['name'] if program else enrollment.get('program_name', ''),
        program_duration=program['duration'] if program else '',
        branch_name=branch['name'] if branch else '',
        branch_id=enrollment['branch_id'],
        email=request_data.email,
        phone=request_data.phone,
        program_start_date=request_data.program_start_date,
        program_end_date=request_data.program_end_date,
        training_mode=request_data.training_mode,
        training_hours=request_data.training_hours,
        registration_number=registration_number,
        verification_id=verification_id
    )
    
    cert_dict = cert_request.model_dump()
    cert_dict['created_at'] = cert_dict['created_at'].isoformat()
    
    await db.certificate_requests.insert_one(cert_dict)
    
    return {
        "message": "Certificate request submitted successfully",
        "certificate_id": certificate_id,
        "status": "Pending"
    }

# Roles allowed to VIEW/DOWNLOAD certificate requests
# Per requirement: ONLY Front Desk Executive and Certificate Manager (+ Admin as super-user fallback)
CERT_VIEW_ROLES = [
    UserRole.ADMIN,
    UserRole.CERTIFICATE_MANAGER,
    UserRole.FRONT_DESK,
]
# Roles that see ALL branches (no branch filter)
CERT_ALL_BRANCH_ROLES = [UserRole.ADMIN, UserRole.CERTIFICATE_MANAGER]


@api_router.get("/certificate-requests")
async def get_certificate_requests(
    status: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """List certificate requests.
    
    - Admin & Certificate Manager see all branches.
    - Branch Admin / Counsellor / FDE see only requests from their own branch.
    """
    if current_user.role not in CERT_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if status:
        query["status"] = status
    # Scope by branch for non-admin roles
    if current_user.role not in CERT_ALL_BRANCH_ROLES:
        query["branch_id"] = current_user.branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    requests = await db.certificate_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    # Refresh student / program / branch from their source collections so that
    # any edits made post-submission reflect on the issued certificate.
    requests = [await _enrich_certificate_with_live_data(r) for r in requests]
    return requests

@api_router.get("/certificate-requests/{request_id}")
async def get_certificate_request(request_id: str, current_user: User = Depends(get_current_user)):
    """Get single certificate request."""
    if current_user.role not in CERT_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Access denied")
    
    request = await db.certificate_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="Certificate request not found")
    
    # Branch-scoped roles can only view their branch
    if current_user.role not in CERT_ALL_BRANCH_ROLES and request.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied for this branch")
    
    request = await _enrich_certificate_with_live_data(request)
    return request


async def _enrich_certificate_with_live_data(cert: dict) -> dict:
    """Merge the latest student/program/branch values from their source collections
    onto a certificate request document so that edits to the underlying enrollment
    or program (e.g. typo fix in student name, course duration update) reflect on
    the issued/verified certificate.

    NOTE: When the certificate has been manually edited or manually created by an
    Admin / Certificate Manager (``is_manually_edited`` / ``is_manually_created``),
    the stored values are treated as the source of truth and are NOT overwritten
    from enrollments / programs / branches.
    """
    if not cert:
        return cert
    # Respect manual overrides — do not overwrite Cert Manager edits.
    if cert.get("is_manually_edited") or cert.get("is_manually_created"):
        return cert
    try:
        enr = None
        if cert.get("enrollment_id"):
            enr = await db.enrollments.find_one(
                {"$or": [{"id": cert["enrollment_id"]}, {"enrollment_id": cert["enrollment_id"]}]},
                {"_id": 0},
            )
        elif cert.get("enrollment_number"):
            enr = await db.enrollments.find_one({"enrollment_id": cert["enrollment_number"]}, {"_id": 0})
        if enr:
            cert["student_name"] = enr.get("student_name") or cert.get("student_name")
            if enr.get("email"):
                cert["email"] = enr["email"]
            if enr.get("phone"):
                cert["phone"] = enr["phone"]
            if enr.get("program_id"):
                program = await db.programs.find_one({"id": enr["program_id"]}, {"_id": 0})
                if program:
                    cert["program_name"] = program.get("name") or cert.get("program_name")
                    if program.get("duration"):
                        cert["program_duration"] = program["duration"]
            if enr.get("branch_id"):
                branch = await db.branches.find_one({"id": enr["branch_id"]}, {"_id": 0})
                if branch and branch.get("name"):
                    cert["branch_name"] = branch["name"]
    except Exception as exc:
        logger.warning(f"Certificate live-enrichment failed for cert {cert.get('id')}: {exc}")
    return cert

@api_router.post("/certificate-requests/manual")
async def create_manual_certificate(
    payload: CertificateManualCreate,
    current_user: User = Depends(get_current_user)
):
    """Manually create a certificate — Admin / Certificate Manager only.

    Bypasses all eligibility gates (fees cleared, course completion, exam pass,
    duplicate check). This lets a Cert Manager issue multiple certificates for
    the same student (e.g. multiple modules against one enrollment, reissues,
    honorary certificates, name/course corrections issued as a new certificate).

    Defaults to status = "Approved" so it is immediately downloadable. Pass
    ``initial_status = "Pending"`` if you want it to still go through review.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.CERTIFICATE_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")

    initial_status = (payload.initial_status or "Approved").strip()
    if initial_status not in [CertificateStatus.PENDING.value, CertificateStatus.APPROVED.value]:
        raise HTTPException(
            status_code=400,
            detail="initial_status must be 'Pending' or 'Approved'"
        )

    # Resolve branch — prefer explicit branch_id, else Cert Manager's own branch, else Admin's.
    branch_id = payload.branch_id or current_user.branch_id
    branch_name = payload.branch_name or ""
    if branch_id and not branch_name:
        branch = await db.branches.find_one({"id": branch_id}, {"_id": 0, "name": 1})
        if branch:
            branch_name = branch.get("name", "")

    # Auto-generate identifiers when not supplied. Registration number is no longer
    # part of the certificate — we only rely on certificate_id + enrollment_number.
    certificate_id = payload.certificate_id or await generate_certificate_id()
    verification_id = await generate_verification_id()
    registration_number = ""

    # Try to link an enrollment_id if the enrollment_number matches an existing one.
    enrollment_db_id = ""
    if payload.enrollment_number:
        enr = await db.enrollments.find_one(
            {"enrollment_id": payload.enrollment_number},
            {"_id": 0, "id": 1}
        )
        if enr:
            enrollment_db_id = enr.get("id", "")

    now_iso = datetime.now(timezone.utc).isoformat()

    cert = CertificateRequest(
        certificate_id=certificate_id,
        enrollment_id=enrollment_db_id or payload.enrollment_number or "",
        enrollment_number=payload.enrollment_number or "",
        student_name=payload.student_name,
        program_name=payload.program_name,
        program_duration=payload.program_duration or "",
        branch_name=branch_name,
        branch_id=branch_id or "",
        email=payload.email or "",
        phone=payload.phone or "",
        program_start_date=payload.program_start_date,
        program_end_date=payload.program_end_date,
        training_mode=payload.training_mode,
        training_hours=payload.training_hours,
        registration_number=registration_number,
        verification_id=verification_id,
        status=CertificateStatus(initial_status),
        is_manually_created=True,
        is_manually_edited=True,
    )

    cert_dict = cert.model_dump()
    cert_dict['created_at'] = cert_dict['created_at'].isoformat()

    # If created directly as Approved, stamp approval metadata.
    if initial_status == CertificateStatus.APPROVED.value:
        cert_dict['approved_at'] = now_iso
        cert_dict['approved_by'] = current_user.id
        cert_dict['approved_by_name'] = current_user.name

    await db.certificate_requests.insert_one(cert_dict)

    return {
        "message": "Certificate created successfully",
        "id": cert_dict['id'],
        "certificate_id": certificate_id,
        "verification_id": verification_id,
        "status": initial_status,
    }


@api_router.put("/certificate-requests/{request_id}")
async def update_certificate_request(
    request_id: str,
    update_data: CertificateRequestUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update ANY field on a certificate request. Admin / Certificate Manager only.

    Any successful edit flips ``is_manually_edited=True`` so downstream reads
    (list, get, verify) will NOT auto-revert the value from source collections.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.CERTIFICATE_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if update_dict.get('training_mode'):
        update_dict['training_mode'] = update_dict['training_mode'].value

    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")

    # Any manual edit locks the record from live-enrichment overwrites.
    update_dict['is_manually_edited'] = True
    update_dict['last_edited_by'] = current_user.id
    update_dict['last_edited_by_name'] = current_user.name
    update_dict['last_edited_at'] = datetime.now(timezone.utc).isoformat()

    result = await db.certificate_requests.update_one(
        {"id": request_id},
        {"$set": update_dict}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Certificate request not found")

    return {"message": "Certificate request updated successfully"}

@api_router.post("/certificate-requests/{request_id}/approve")
async def approve_certificate_request(request_id: str, current_user: User = Depends(get_current_user)):
    """Approve a certificate request and mark student's course as complete"""
    if current_user.role not in [UserRole.ADMIN, UserRole.CERTIFICATE_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get the certificate request first
    cert_request = await db.certificate_requests.find_one({"id": request_id}, {"_id": 0})
    if not cert_request:
        raise HTTPException(status_code=404, detail="Certificate request not found")
    
    if cert_request.get('status') != CertificateStatus.PENDING.value:
        raise HTTPException(status_code=400, detail="Certificate request is not pending")
    
    # Update certificate request status
    result = await db.certificate_requests.update_one(
        {"id": request_id, "status": CertificateStatus.PENDING.value},
        {"$set": {
            "status": CertificateStatus.APPROVED.value,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "approved_by": current_user.id,
            "approved_by_name": current_user.name
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Certificate request not found or not pending")
    
    # Auto-mark the student's enrollment as "Completed"
    enrollment_number = cert_request.get('enrollment_number')
    if enrollment_number:
        await db.enrollments.update_one(
            {"enrollment_id": enrollment_number},
            {"$set": {
                "status": "Completed",
                "completion_date": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return {"message": "Certificate request approved and student marked as course complete"}

@api_router.post("/certificate-requests/{request_id}/reject")
async def reject_certificate_request(
    request_id: str,
    reason: str = "",
    current_user: User = Depends(get_current_user)
):
    """Reject a certificate request"""
    if current_user.role not in [UserRole.ADMIN, UserRole.CERTIFICATE_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.certificate_requests.update_one(
        {"id": request_id, "status": CertificateStatus.PENDING.value},
        {"$set": {
            "status": CertificateStatus.REJECTED.value,
            "rejection_reason": reason,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "approved_by": current_user.id,
            "approved_by_name": current_user.name
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Certificate request not found or not pending")
    
    return {"message": "Certificate request rejected"}

@api_router.delete("/certificate-requests/{request_id}")
async def delete_certificate_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """Permanently delete a certificate request. After deletion the student can re-request it.
    Allowed for Admin and Certificate Manager only.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.CERTIFICATE_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    cert = await db.certificate_requests.find_one({"id": request_id}, {"_id": 0})
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate request not found")
    
    result = await db.certificate_requests.delete_one({"id": request_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Certificate request not found")
    
    return {"message": "Certificate request deleted. Student can now submit a new request."}


# POST alias for delete – some hosting providers (Cloudflare, certain CDNs / corporate proxies)
# strip the DELETE method. This alias lets the UI fall back to a plain POST that performs the
# same deletion when DELETE is blocked.
@api_router.post("/certificate-requests/{request_id}/delete")
async def delete_certificate_request_post(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    return await delete_certificate_request(request_id, current_user)


@api_router.post("/certificate-requests/{request_id}/download")
async def download_certificate(request_id: str, current_user: User = Depends(get_current_user)):
    """Generate and download certificate PDF (Admin, Cert Manager, Branch Admin, Counsellor, FDE).
    Only approved/ready certificates can be downloaded. Marks the request as Ready on first download."""
    if current_user.role not in CERT_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Access denied")
    
    cert_request = await db.certificate_requests.find_one({"id": request_id}, {"_id": 0})
    if not cert_request:
        raise HTTPException(status_code=404, detail="Certificate request not found")
    
    # Branch-scoped roles can only download certs of their branch
    if current_user.role not in CERT_ALL_BRANCH_ROLES and cert_request.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied for this branch")
    
    if cert_request['status'] not in [CertificateStatus.APPROVED.value, CertificateStatus.READY.value, CertificateStatus.PRINTED.value]:
        raise HTTPException(status_code=400, detail="Certificate must be approved before download")
    
    # Mark as ready on first download — only Admin / Cert Manager transition the status
    if cert_request['status'] == CertificateStatus.APPROVED.value and current_user.role in CERT_ALL_BRANCH_ROLES:
        await db.certificate_requests.update_one(
            {"id": request_id},
            {"$set": {
                "status": CertificateStatus.READY.value,
                "issued_at": datetime.now(timezone.utc).isoformat(),
                "issued_by": current_user.id
            }}
        )
        # NOTE: WhatsApp "certificate_ready" notification is fired from the
        # /mark-printed endpoint (not here on download). This way the message
        # is only sent once the physical certificate is actually printed and
        # ready to be handed over — preventing premature notifications.
    
    # Return certificate data for PDF generation on frontend.
    # Use .get() for every field so older records missing newer columns don't
    # crash the endpoint with a KeyError.
    return {
        "certificate_id": cert_request.get('certificate_id', ''),
        "enrollment_number": cert_request.get('enrollment_number', ''),
        "student_name": cert_request.get('student_name', ''),
        "program_name": cert_request.get('program_name', ''),
        "program_duration": cert_request.get('program_duration', ''),
        "branch_name": cert_request.get('branch_name', ''),
        "training_mode": cert_request.get('training_mode', 'Offline'),
        "training_hours": cert_request.get('training_hours', None),
        "program_start_date": cert_request.get('program_start_date', ''),
        "program_end_date": cert_request.get('program_end_date', ''),
        "verification_id": cert_request.get('verification_id', ''),
        "issued_date": datetime.now().strftime("%d-%m-%Y")
    }

@api_router.post("/certificate-requests/{request_id}/mark-printed")
async def mark_certificate_printed(request_id: str, current_user: User = Depends(get_current_user)):
    """Mark a downloaded (Ready) certificate as physically Printed and handed over.
    Allowed for Admin, Certificate Manager, Branch Admin and Front Desk Executive
    (i.e. the same roles that can download it)."""
    if current_user.role not in CERT_VIEW_ROLES + [UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")

    cert_request = await db.certificate_requests.find_one({"id": request_id}, {"_id": 0})
    if not cert_request:
        raise HTTPException(status_code=404, detail="Certificate request not found")

    # Branch-scoped roles can only act on certs of their branch
    if current_user.role not in CERT_ALL_BRANCH_ROLES and cert_request.get('branch_id') != current_user.branch_id:
        raise HTTPException(status_code=403, detail="Access denied for this branch")

    if cert_request['status'] == CertificateStatus.PRINTED.value:
        return {
            "message": "Certificate was already marked as printed.",
            "already_printed": True,
            "status": CertificateStatus.PRINTED.value,
        }

    if cert_request['status'] not in [CertificateStatus.READY.value, CertificateStatus.APPROVED.value]:
        raise HTTPException(
            status_code=400,
            detail="Only Approved or Ready certificates can be marked as printed."
        )

    await db.certificate_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": CertificateStatus.PRINTED.value,
            "printed_at": datetime.now(timezone.utc).isoformat(),
            "printed_by": current_user.id,
            "printed_by_name": current_user.name,
        }}
    )

    # Fire the WhatsApp "certificate_ready" notification to the student now that
    # the physical certificate is printed and ready to be collected/couriered.
    # We deliberately DO NOT send this on download — only when actually printed.
    try:
        if cert_request.get('phone'):
            await send_whatsapp_notification(
                cert_request['phone'],
                "certificate_ready",
                {
                    "name": cert_request.get('student_name', ''),
                    "certificate_id": cert_request.get('certificate_id', ''),
                    "course": cert_request.get('program_name', ''),
                }
            )
    except Exception as e:
        logging.error(f"certificate_ready WhatsApp send failed for {request_id}: {e}")

    return {
        "message": f"Certificate {cert_request.get('certificate_id')} marked as printed.",
        "status": CertificateStatus.PRINTED.value,
    }


@api_router.get("/public/verify/{verification_id}")
async def verify_certificate(verification_id: str):
    """Public endpoint to verify certificate authenticity via QR code"""
    cert = await db.certificate_requests.find_one(
        {"verification_id": verification_id},
        {"_id": 0}
    )
    
    if not cert:
        return {
            "verified": False,
            "message": "Certificate not found or invalid verification code"
        }
    
    if cert['status'] not in (CertificateStatus.READY.value, CertificateStatus.PRINTED.value):
        return {
            "verified": False,
            "message": f"Certificate status: {cert['status']} (not issued yet)"
        }
    
    # Pull latest student/program/branch from source-of-truth collections so that
    # corrections (e.g. fixing a typo in the student's name) reflect on the
    # verified certificate.
    cert = await _enrich_certificate_with_live_data(cert)
    
    return {
        "verified": True,
        "message": "Certificate is authentic and verified",
        "certificate_details": {
            "certificate_id": cert['certificate_id'],
            "enrollment_number": cert.get('enrollment_number', ''),
            "student_name": cert['student_name'],
            "program_name": cert['program_name'],
            "program_duration": cert['program_duration'],
            "branch_name": cert['branch_name'],
            "training_mode": cert['training_mode'],
            "issued_date": cert.get('issued_at', ''),
        }
    }

# ========== FEE REMINDER AUTOMATION ==========
async def send_fee_reminders():
    """
    Send WhatsApp fee reminders for pending payments.
    Sends reminders 7 days, 5 days, 3 days, 1 day before due date, and on the due date.
    """
    try:
        logger.info("Running fee reminder job...")
        today = datetime.now(timezone.utc).date()
        reminder_days = [7, 5, 3, 1, 0]  # Days before due date
        
        for days_before in reminder_days:
            target_date = today + timedelta(days=days_before)
            target_date_str = target_date.isoformat()
            
            # Find installments due on the target date
            installments = await db.installment_schedule.find(
                {"due_date": target_date_str, "status": {"$ne": "Paid"}},
                {"_id": 0}
            ).to_list(1000)
            
            for inst in installments:
                payment_plan = await db.payment_plans.find_one(
                    {"id": inst.get("payment_plan_id")},
                    {"_id": 0}
                )
                if not payment_plan:
                    continue
                
                enrollment = await db.enrollments.find_one(
                    {"id": payment_plan.get("enrollment_id")},
                    {"_id": 0}
                )
                if not enrollment:
                    continue
                
                # Send WhatsApp reminder
                reminder_text = f"on {target_date_str}" if days_before == 0 else f"in {days_before} day(s)"
                await send_whatsapp_notification(
                    enrollment.get('phone', ''),
                    "fee_reminder",
                    {
                        "name": enrollment.get('student_name', ''),
                        "amount_due": str(inst.get('amount', 0)),
                        "due_date": target_date_str
                    }
                )
                logger.info(f"Fee reminder sent to {enrollment.get('student_name')} for ₹{inst.get('amount')} due {reminder_text}")
        
        # Also check for one-time payments that are overdue
        # For one-time payments, send reminder if enrollment date was 23, 25, 27, 29, or 30 days ago
        # (Equivalent to 7, 5, 3, 1, 0 days before 30-day mark)
        for days_since in [23, 25, 27, 29, 30]:
            check_date = today - timedelta(days=days_since)
            check_date_str = check_date.isoformat()
            
            # Find enrollments from that date with pending one-time payments
            enrollments = await db.enrollments.find(
                {"enrollment_date": check_date_str},
                {"_id": 0}
            ).to_list(1000)
            
            for enrollment in enrollments:
                enrollment_id = enrollment.get('id')
                final_fee = enrollment.get('final_fee', 0)
                
                # Check if it's a one-time payment plan
                payment_plan = await db.payment_plans.find_one(
                    {"enrollment_id": enrollment_id},
                    {"_id": 0}
                )
                if payment_plan and payment_plan.get('plan_type') == 'Installments':
                    continue  # Skip installment plans
                
                # Get total paid
                payments = await db.payments.find(
                    {"enrollment_id": enrollment_id},
                    {"_id": 0, "amount": 1}
                ).to_list(100)
                total_paid = sum(p.get('amount', 0) for p in payments)
                pending_amount = final_fee - total_paid
                
                if pending_amount > 0:
                    days_left = 30 - days_since
                    due_date_str = (today + timedelta(days=days_left)).isoformat() if days_left > 0 else today.isoformat()
                    await send_whatsapp_notification(
                        enrollment.get('phone', ''),
                        "fee_reminder",
                        {
                            "name": enrollment.get('student_name', ''),
                            "amount_due": str(pending_amount),
                            "due_date": due_date_str
                        }
                    )
                    logger.info(f"One-time payment reminder sent to {enrollment.get('student_name')} for ₹{pending_amount}")
        
        logger.info("Fee reminder job completed successfully")
    except Exception as e:
        logger.error(f"Error in fee reminder job: {str(e)}")


async def send_fee_late_reminders():
    """One-shot WhatsApp 'fee_late' message the day AFTER an installment becomes
    overdue (and remains unpaid). Idempotent via late_reminder_sent flag on each
    installment / enrollment so a student never gets spammed."""
    try:
        logger.info("Running fee LATE reminder job...")
        today = datetime.now(timezone.utc).date()

        # Track A: overdue installments (any installment whose due_date is strictly
        # before today, status != Paid, late_reminder_sent != True)
        overdue_installments = await db.installment_schedule.find(
            {
                "due_date": {"$lt": today.isoformat()},
                "status": {"$ne": "Paid"},
                "late_reminder_sent": {"$ne": True},
            },
            {"_id": 0},
        ).to_list(2000)

        for inst in overdue_installments:
            payment_plan = await db.payment_plans.find_one({"id": inst.get("payment_plan_id")}, {"_id": 0})
            if not payment_plan:
                continue
            enrollment = await db.enrollments.find_one({"id": payment_plan.get("enrollment_id")}, {"_id": 0})
            if not enrollment:
                continue
            try:
                await send_whatsapp_notification(
                    enrollment.get("phone", ""),
                    "fee_late",
                    {"name": enrollment.get("student_name", "")},
                )
                await db.installment_schedule.update_one(
                    {"id": inst.get("id")},
                    {"$set": {"late_reminder_sent": True, "late_reminder_sent_at": datetime.now(timezone.utc).isoformat()}},
                )
                logger.info(f"fee_late sent (installment) to {enrollment.get('student_name')} for inst {inst.get('id')}")
            except Exception as e:
                logger.warning(f"fee_late failed for inst {inst.get('id')}: {e}")

        # Track B: one-time payment enrollments past the 30-day implicit due date
        for enrollment in await db.enrollments.find(
            {"late_reminder_sent": {"$ne": True}}, {"_id": 0}
        ).to_list(5000):
            enrollment_date_str = enrollment.get("enrollment_date")
            if not enrollment_date_str:
                continue
            try:
                enrollment_date = datetime.fromisoformat(enrollment_date_str).date() if "T" in enrollment_date_str else datetime.strptime(enrollment_date_str, "%Y-%m-%d").date()
            except Exception:
                continue
            if (today - enrollment_date).days <= 30:
                continue  # still within 30-day window
            # Skip if this enrollment has an installment plan (Track A handles it)
            payment_plan = await db.payment_plans.find_one({"enrollment_id": enrollment.get("id")}, {"_id": 0})
            if payment_plan and payment_plan.get("plan_type") == "Installments":
                continue
            # Compute pending
            payments = await db.payments.find({"enrollment_id": enrollment.get("id")}, {"_id": 0, "amount": 1}).to_list(100)
            total_paid = sum(p.get("amount", 0) for p in payments)
            pending = (enrollment.get("final_fee") or 0) - total_paid
            if pending <= 0:
                continue
            try:
                await send_whatsapp_notification(
                    enrollment.get("phone", ""),
                    "fee_late",
                    {"name": enrollment.get("student_name", "")},
                )
                await db.enrollments.update_one(
                    {"id": enrollment.get("id")},
                    {"$set": {"late_reminder_sent": True, "late_reminder_sent_at": datetime.now(timezone.utc).isoformat()}},
                )
                logger.info(f"fee_late sent (one-time) to {enrollment.get('student_name')}")
            except Exception as e:
                logger.warning(f"fee_late failed for enrollment {enrollment.get('id')}: {e}")

        logger.info("Fee LATE reminder job completed successfully")
    except Exception as e:
        logger.error(f"Error in fee_late job: {str(e)}")

async def send_birthday_wishes():
    """
    Send WhatsApp birthday wishes to students on their birthday.
    """
    try:
        logger.info("Running birthday wishes job...")
        today = datetime.now(timezone.utc)
        today_month_day = today.strftime("%m-%d")  # MM-DD format
        
        # Find students with birthday today
        # date_of_birth is stored as YYYY-MM-DD string
        enrollments = await db.enrollments.find(
            {"status": "Active"},
            {"_id": 0}
        ).to_list(10000)
        
        for enrollment in enrollments:
            dob = enrollment.get('date_of_birth', '')
            if not dob:
                continue
            
            # Check if birthday matches today (MM-DD)
            try:
                dob_month_day = dob[5:10]  # Extract MM-DD from YYYY-MM-DD
                if dob_month_day == today_month_day:
                    await send_whatsapp_notification(
                        enrollment.get('phone', ''),
                        "birthday_wishes",
                        {"name": enrollment.get('student_name', '')}
                    )
                    logger.info(f"Birthday wish sent to {enrollment.get('student_name')}")
            except Exception as e:
                logger.warning(f"Error parsing DOB for {enrollment.get('student_name')}: {e}")
        
        logger.info("Birthday wishes job completed successfully")
    except Exception as e:
        logger.error(f"Error in birthday wishes job: {str(e)}")

# Initialize scheduler
scheduler = AsyncIOScheduler()

# ========== WIZBANG BRAND MANAGER MODULE ==========
class BrandPlanDay(BaseModel):
    model_config = ConfigDict(extra="ignore")
    day: int  # 1..31
    content_type: Optional[str] = None  # "Reel", "Story", "Post", "Carousel", etc.
    caption: Optional[str] = None
    notes: Optional[str] = None
    image_url: Optional[str] = None
    platform: Optional[str] = None  # "Instagram", "Facebook", "LinkedIn", ...

class BrandPlanCreate(BaseModel):
    client_id: str
    title: str
    month: str  # YYYY-MM
    days: List[BrandPlanDay] = []

class BrandPlanUpdate(BaseModel):
    title: Optional[str] = None
    month: Optional[str] = None
    days: Optional[List[BrandPlanDay]] = None

class BrandSocialProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    platform: str  # Instagram / Facebook / LinkedIn / Twitter / YouTube
    url: str
    handle: Optional[str] = None
    initial_followers: int = 0
    initial_followers_date: Optional[str] = None  # YYYY-MM-DD

class BrandProfilesUpdate(BaseModel):
    profiles: List[BrandSocialProfile] = []

class BrandReportPost(BaseModel):
    model_config = ConfigDict(extra="ignore")
    url: Optional[str] = None
    caption: Optional[str] = None
    platform: Optional[str] = None
    reach: Optional[int] = 0
    likes: Optional[int] = 0
    comments: Optional[int] = 0
    shares: Optional[int] = 0

class BrandReportProfileStat(BaseModel):
    model_config = ConfigDict(extra="ignore")
    platform: str
    url: Optional[str] = None
    followers_initial: Optional[int] = 0
    followers_now: int = 0

class BrandReportCreate(BaseModel):
    client_id: str
    month: str  # YYYY-MM
    title: Optional[str] = None
    profile_stats: List[BrandReportProfileStat] = []
    total_reach: int = 0
    total_posts: int = 0
    total_engagement: Optional[int] = 0
    posts: List[BrandReportPost] = []
    notes: Optional[str] = None

# ----- BRAND CLIENTS (read-only mirror of Wizbang clients for Brand Manager) -----
@api_router.get("/wizbang/brand/clients")
async def brand_list_clients(current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    docs = await db.wizbang_clients.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    # Decorate with quick counts so the dashboard can render badges
    out = []
    for c in docs:
        plans = await db.brand_plans.count_documents({"client_id": c["id"]})
        reports = await db.brand_reports.count_documents({"client_id": c["id"]})
        out.append({**c, "plan_count": plans, "report_count": reports})
    return out

@api_router.get("/wizbang/brand/clients/{client_id}")
async def brand_get_client(client_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    c = await db.wizbang_clients.find_one({"id": client_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    profiles = await db.brand_client_profiles.find_one({"client_id": client_id}, {"_id": 0})
    return {"client": c, "profiles": (profiles or {}).get("profiles", [])}

# ----- SOCIAL PROFILES (per client) -----
@api_router.put("/wizbang/brand/clients/{client_id}/profiles")
async def brand_set_profiles(client_id: str, payload: BrandProfilesUpdate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    c = await db.wizbang_clients.find_one({"id": client_id}, {"_id": 0, "id": 1})
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    profiles_data = [p.model_dump() for p in payload.profiles]
    await db.brand_client_profiles.update_one(
        {"client_id": client_id},
        {"$set": {
            "client_id": client_id,
            "profiles": profiles_data,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.id,
        }},
        upsert=True,
    )
    return {"ok": True, "profiles": profiles_data}

# ----- CONTENT PLANS -----
@api_router.post("/wizbang/brand/plans")
async def brand_create_plan(payload: BrandPlanCreate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    c = await db.wizbang_clients.find_one({"id": payload.client_id}, {"_id": 0, "id": 1, "name": 1})
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    doc = {
        "id": str(uuid.uuid4()),
        "client_id": payload.client_id,
        "client_name": c.get("name"),
        "title": payload.title,
        "month": payload.month,
        "days": [d.model_dump() for d in payload.days],
        "share_token": str(uuid.uuid4()),
        "status": "Draft",
        "client_remarks": None,
        "accepted_at": None,
        "accepted_by_name": None,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.brand_plans.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/wizbang/brand/plans")
async def brand_list_plans(client_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    q = {}
    if client_id:
        q["client_id"] = client_id
    return await db.brand_plans.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)

@api_router.get("/wizbang/brand/plans/{plan_id}")
async def brand_get_plan(plan_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    p = await db.brand_plans.find_one({"id": plan_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Plan not found")
    return p

@api_router.put("/wizbang/brand/plans/{plan_id}")
async def brand_update_plan(plan_id: str, payload: BrandPlanUpdate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    update = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if "days" in update:
        # Pydantic models are already dumped by model_dump above? Need to check — they may be dicts
        days = update["days"]
        if days and isinstance(days[0], BrandPlanDay):
            update["days"] = [d.model_dump() for d in days]
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    r = await db.brand_plans.update_one({"id": plan_id}, {"$set": update})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Plan not found")
    return {"ok": True}

@api_router.delete("/wizbang/brand/plans/{plan_id}")
async def brand_delete_plan(plan_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    r = await db.brand_plans.delete_one({"id": plan_id})
    if not r.deleted_count:
        raise HTTPException(status_code=404, detail="Plan not found")
    return {"ok": True}

@api_router.post("/wizbang/brand/plans/{plan_id}/share")
async def brand_share_plan(plan_id: str, current_user: User = Depends(get_current_user)):
    """Flips a plan to 'Shared' status and returns a public URL fragment the FE can
    turn into a fully-qualified link to send the client."""
    _ensure_wizbang_brand_or_admin(current_user)
    p = await db.brand_plans.find_one({"id": plan_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Plan not found")
    token = p.get("share_token") or str(uuid.uuid4())
    await db.brand_plans.update_one(
        {"id": plan_id},
        {"$set": {"share_token": token, "status": "Shared", "shared_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"ok": True, "share_token": token, "public_path": f"/public/brand-plan/{token}"}

# ----- PUBLIC plan endpoints (no auth) -----
@api_router.get("/public/brand-plan/{token}")
async def public_get_brand_plan(token: str):
    p = await db.brand_plans.find_one({"share_token": token}, {"_id": 0, "created_by": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Plan not found or link expired")
    return p

@api_router.post("/public/brand-plan/{token}/respond")
async def public_respond_brand_plan(token: str, payload: dict = Body(...)):
    """Client opens public link and either adds remarks or accepts the plan."""
    p = await db.brand_plans.find_one({"share_token": token}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Plan not found")
    action = (payload.get("action") or "remark").lower()
    update = {
        "client_remarks": payload.get("remarks") or p.get("client_remarks"),
        "accepted_by_name": payload.get("accepted_by_name") or p.get("accepted_by_name"),
        "responded_at": datetime.now(timezone.utc).isoformat(),
    }
    if action == "accept":
        update["status"] = "Accepted"
        update["accepted_at"] = datetime.now(timezone.utc).isoformat()
    else:
        update["status"] = "Acknowledged"
    await db.brand_plans.update_one({"share_token": token}, {"$set": update})
    return {"ok": True, "status": update.get("status")}

# ----- MONTHLY REPORTS -----
def _render_brand_report_pdf(report: dict, client: dict, profiles: List[dict]) -> str:
    """Render the report as a PDF and return base64 string."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.colors import HexColor
    import base64

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=22, leading=26, textColor=HexColor("#0f172a"), spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, leading=13, textColor=HexColor("#475569"), spaceAfter=10)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, leading=16, textColor=HexColor("#0f172a"), spaceBefore=12, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=14, textColor=HexColor("#1f2937"))

    flow = []
    flow.append(Paragraph(f"{client.get('name','Client')} — Monthly Report", h1))
    flow.append(Paragraph(f"Month: {report.get('month','')} &nbsp;•&nbsp; Prepared by ETI Educom Brand Team", sub))

    if report.get("ai_summary"):
        flow.append(Paragraph("EXECUTIVE SUMMARY", h2))
        flow.append(Paragraph(report["ai_summary"].replace("\n", "<br/>"), body))

    # KPI table
    flow.append(Paragraph("HEADLINE NUMBERS", h2))
    kpi_data = [
        ["Total Reach", "Total Posts", "Total Engagement"],
        [f"{report.get('total_reach', 0):,}", f"{report.get('total_posts', 0):,}", f"{report.get('total_engagement', 0):,}"],
    ]
    t = Table(kpi_data, colWidths=[55*mm, 55*mm, 55*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.4, HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    flow.append(t)

    # Profile stats table
    if report.get("profile_stats"):
        flow.append(Paragraph("AUDIENCE GROWTH BY PLATFORM", h2))
        rows = [["Platform", "Start", "Now", "Δ", "Growth %"]]
        for s in report["profile_stats"]:
            ini = int(s.get("followers_initial") or 0)
            now = int(s.get("followers_now") or 0)
            delta = now - ini
            pct = f"{(delta / ini * 100):.1f}%" if ini else ("∞" if delta else "0%")
            rows.append([s.get("platform", ""), f"{ini:,}", f"{now:,}", f"{delta:+,}", pct])
        pt = Table(rows, colWidths=[35*mm, 30*mm, 30*mm, 30*mm, 30*mm])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#f8fafc"), HexColor("#ffffff")]),
            ("GRID", (0, 0), (-1, -1), 0.3, HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        flow.append(pt)

    # Posts table
    if report.get("posts"):
        flow.append(Paragraph("TOP POSTS THIS MONTH", h2))
        rows = [["#", "Platform", "Reach", "Likes", "Comments", "Link"]]
        for i, p in enumerate(report["posts"], start=1):
            link = p.get("url") or ""
            short = link if len(link) <= 38 else link[:36] + "…"
            rows.append([
                str(i),
                p.get("platform") or "",
                f"{int(p.get('reach') or 0):,}",
                f"{int(p.get('likes') or 0):,}",
                f"{int(p.get('comments') or 0):,}",
                short,
            ])
        pt = Table(rows, colWidths=[10*mm, 25*mm, 25*mm, 22*mm, 27*mm, 55*mm])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 1), (-2, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#f8fafc"), HexColor("#ffffff")]),
            ("GRID", (0, 0), (-1, -1), 0.3, HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        flow.append(pt)

    if report.get("notes"):
        flow.append(Paragraph("NOTES", h2))
        flow.append(Paragraph(report["notes"].replace("\n", "<br/>"), body))

    flow.append(Spacer(1, 16))
    flow.append(Paragraph(f"<font color='#94a3b8' size='8'>Generated on {datetime.now(timezone.utc).strftime('%d %b %Y')} · ETI Educom · This report is confidential.</font>", body))

    doc.build(flow)
    return base64.b64encode(buf.getvalue()).decode("ascii")

async def _generate_ai_summary(client_name: str, report_input: dict, profiles: List[dict]) -> str:
    """Use Claude Sonnet 4.5 (via Emergent LLM key) to write a 1-page exec summary."""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=os.environ["EMERGENT_LLM_KEY"],
            session_id=f"brandreport-{client_name}-{report_input.get('month','')}",
            system_message=(
                "You are a senior social-media strategist. Write a concise but data-rich monthly executive summary "
                "(plain text, 4–6 short paragraphs, no markdown) covering: (1) audience growth highlights, "
                "(2) content performance, (3) what worked / what didn't, (4) recommendations for next month. "
                "Tone: confident, client-facing."
            ),
        ).with_model("anthropic", "claude-sonnet-4-5")
        prompt = (
            f"CLIENT: {client_name}\n"
            f"MONTH: {report_input.get('month','')}\n"
            f"INITIAL PROFILES: {json.dumps(profiles)}\n"
            f"REPORT DATA: {json.dumps(report_input)}\n\n"
            "Write the summary now."
        )
        out = await chat.send_message(UserMessage(text=prompt))
        return (out or "").strip()
    except Exception as ex:
        logger.warning(f"Brand report AI summary failed, using template: {ex}")
        # Fallback template
        total_reach = report_input.get("total_reach", 0)
        total_posts = report_input.get("total_posts", 0)
        growth_lines = []
        for s in report_input.get("profile_stats", []):
            ini = int(s.get("followers_initial") or 0)
            now = int(s.get("followers_now") or 0)
            d = now - ini
            growth_lines.append(f"{s.get('platform','?')}: {ini:,} → {now:,} ({d:+,})")
        growth = "; ".join(growth_lines) if growth_lines else "No platform data captured."
        return (
            f"This month {client_name} published {total_posts} pieces of content, reaching an estimated {total_reach:,} accounts. "
            f"Platform growth — {growth}. "
            "Content mix and consistency were the primary drivers. Next month focus on doubling down on the top-performing "
            "formats and increasing posting frequency on the strongest channel."
        )

@api_router.post("/wizbang/brand/reports")
async def brand_create_report(payload: BrandReportCreate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    client = await db.wizbang_clients.find_one({"id": payload.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    profiles_doc = await db.brand_client_profiles.find_one({"client_id": payload.client_id}, {"_id": 0})
    profiles = (profiles_doc or {}).get("profiles", [])

    # Backfill followers_initial from saved profile config when not provided
    enriched_stats = []
    for s in [x.model_dump() for x in payload.profile_stats]:
        if not s.get("followers_initial"):
            match = next((p for p in profiles if p.get("platform") == s.get("platform")), None)
            if match:
                s["followers_initial"] = match.get("initial_followers", 0)
                s["url"] = s.get("url") or match.get("url")
        enriched_stats.append(s)

    report_input = {
        "month": payload.month,
        "total_reach": payload.total_reach,
        "total_posts": payload.total_posts,
        "total_engagement": payload.total_engagement or 0,
        "profile_stats": enriched_stats,
        "posts": [p.model_dump() for p in payload.posts],
        "notes": payload.notes or "",
    }

    ai_summary = await _generate_ai_summary(client.get("name", "Client"), report_input, profiles)
    pdf_b64 = _render_brand_report_pdf({**report_input, "ai_summary": ai_summary}, client, profiles)

    doc = {
        "id": str(uuid.uuid4()),
        "client_id": payload.client_id,
        "client_name": client.get("name"),
        "title": payload.title or f"{client.get('name')} — {payload.month} Monthly Report",
        "month": payload.month,
        "profile_stats": enriched_stats,
        "total_reach": payload.total_reach,
        "total_posts": payload.total_posts,
        "total_engagement": payload.total_engagement or 0,
        "posts": [p.model_dump() for p in payload.posts],
        "notes": payload.notes,
        "ai_summary": ai_summary,
        "pdf_b64": pdf_b64,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.brand_reports.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/wizbang/brand/reports")
async def brand_list_reports(client_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    q = {}
    if client_id:
        q["client_id"] = client_id
    # Strip heavy pdf_b64 from list
    return await db.brand_reports.find(q, {"_id": 0, "pdf_b64": 0}).sort("created_at", -1).to_list(500)

@api_router.get("/wizbang/brand/reports/{report_id}")
async def brand_get_report(report_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    r = await db.brand_reports.find_one({"id": report_id}, {"_id": 0})
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")
    return r

@api_router.delete("/wizbang/brand/reports/{report_id}")
async def brand_delete_report(report_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_brand_or_admin(current_user)
    await db.brand_reports.delete_one({"id": report_id})
    return {"ok": True}

# ========== GOOGLE SHEETS LEAD SYNC (META ADS) ==========
import csv as _csv
import io as _io
import re as _re
from typing import Tuple as _Tuple

def _sheet_to_csv_url(sheet_url: str) -> Optional[str]:
    """Convert a Google Sheets share URL to its CSV-export URL.
    Returns None when the URL doesn't look like a Google Sheet.
    """
    if not sheet_url:
        return None
    m = _re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
    if not m:
        return None
    sheet_id = m.group(1)
    gid_m = _re.search(r"[?#&]gid=(\d+)", sheet_url)
    gid = gid_m.group(1) if gid_m else "0"
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

def _normalise_header(h: str) -> str:
    return _re.sub(r"[^a-z0-9]+", "", (h or "").strip().lower())

# Map of normalised CSV header → canonical field
_SHEET_HEADER_MAP = {
    # Name
    "name": "name", "fullname": "name", "leadname": "name", "studentname": "name",
    "firstname": "name",
    # Email
    "email": "email", "emailid": "email", "emailaddress": "email",
    # Phone
    "phone": "phone", "phonenumber": "phone", "mobile": "phone", "mobilenumber": "phone",
    "contact": "phone", "contactnumber": "phone", "whatsappnumber": "phone",
    # Program / course
    "program": "program", "programname": "program", "course": "program",
    "coursename": "program", "interestedin": "program", "courseinterested": "program",
}

async def _resolve_program_id(branch_id: str, course_name: str) -> Optional[str]:
    """Best-effort match of a free-text course name to a Program. Falls back to the
    branch's first program if no match is found."""
    if course_name:
        # exact (case-insensitive) match
        prog = await db.programs.find_one(
            {"name": {"$regex": f"^{_re.escape(course_name.strip())}$", "$options": "i"}},
            {"_id": 0, "id": 1, "name": 1},
        )
        if prog:
            return prog["id"]
        # partial match
        prog = await db.programs.find_one(
            {"name": {"$regex": _re.escape(course_name.strip()), "$options": "i"}},
            {"_id": 0, "id": 1, "name": 1},
        )
        if prog:
            return prog["id"]
    # Fallback: any program (we just need a valid id so the lead is creatable)
    any_prog = await db.programs.find_one({}, {"_id": 0, "id": 1, "name": 1})
    return any_prog["id"] if any_prog else None

async def _fetch_sheet_rows(sheet_url: str) -> dict:
    """Fetch the CSV for a Google Sheet and return parsed rows plus a clean error.
    Returns: {ok, rows, error_code, error_message}
    error_code values: invalid_url | not_public | http_<code> | fetch_error | empty
    """
    csv_url = _sheet_to_csv_url(sheet_url)
    if not csv_url:
        return {"ok": False, "rows": [], "error_code": "invalid_url",
                "error_message": "This doesn't look like a Google Sheets link. Paste the full URL that contains /spreadsheets/d/..."}
    try:
        async with httpx.AsyncClient(timeout=20.0, follow_redirects=True) as client:
            resp = await client.get(csv_url)
            if resp.status_code != 200:
                # Google returns 400/404 when the sheet is private or the gid is wrong
                if resp.status_code in (400, 401, 403, 404):
                    return {"ok": False, "rows": [], "error_code": "not_public",
                            "error_message": "We can't open this sheet. In Google Sheets click Share → General access → 'Anyone with the link · Viewer', then try again."}
                return {"ok": False, "rows": [], "error_code": f"http_{resp.status_code}",
                        "error_message": f"Google Sheets returned HTTP {resp.status_code}. Please try again in a moment."}
            csv_text = resp.text
    except Exception as e:
        return {"ok": False, "rows": [], "error_code": "fetch_error",
                "error_message": f"Could not reach Google Sheets: {e}"}
    reader = _csv.reader(_io.StringIO(csv_text))
    rows = list(reader)
    if not rows:
        return {"ok": False, "rows": [], "error_code": "empty",
                "error_message": "The sheet is empty."}
    return {"ok": True, "rows": rows, "error_code": None, "error_message": None}


async def _sync_one_sheet(branch_id: str, sheet_url: str, last_row: int, column_mapping: Optional[dict] = None) -> dict:
    """Pull rows from a single Google Sheet, dedupe, create leads + fire enquiry_saved.
    `column_mapping` (optional): dict mapping column index (str) → canonical field
    ("name"|"email"|"phone"|"program"|"skip"). When provided, overrides auto-detection.
    Returns: {ok, imported, skipped, new_last_row, reason}."""
    fetch = await _fetch_sheet_rows(sheet_url)
    if not fetch["ok"]:
        return {"ok": False, "reason": fetch["error_code"], "imported": 0,
                "skipped": 0, "new_last_row": last_row}
    rows = fetch["rows"]
    if len(rows) < 2:
        return {"ok": True, "imported": 0, "skipped": 0, "new_last_row": last_row, "reason": "empty"}

    header = rows[0]
    col_idx: dict = {}
    if column_mapping:
        # User-defined mapping takes priority. Keys are column index as string.
        for k, v in column_mapping.items():
            try:
                idx = int(k)
            except (TypeError, ValueError):
                continue
            if v in ("name", "email", "phone", "program") and v not in col_idx and 0 <= idx < len(header):
                col_idx[v] = idx
    # Fill any remaining canonical fields via auto-detection (so partial mappings still work).
    for i, h in enumerate(header):
        key = _SHEET_HEADER_MAP.get(_normalise_header(h))
        if key and key not in col_idx:
            col_idx[key] = i

    if "phone" not in col_idx and "email" not in col_idx:
        return {"ok": False, "reason": "missing_required_columns", "imported": 0, "skipped": 0, "new_last_row": last_row}

    imported = 0
    skipped = 0
    new_last_row = last_row

    for r_idx, row in enumerate(rows[1:], start=1):
        if r_idx <= last_row:
            continue
        new_last_row = r_idx

        def _get(field):
            i = col_idx.get(field)
            if i is None or i >= len(row):
                return ""
            return (row[i] or "").strip()

        name = _get("name") or "Meta Lead"
        email = _get("email")
        phone = _re.sub(r"[^\d+]", "", _get("phone"))
        course = _get("program")

        if not (phone or email):
            skipped += 1
            continue

        # Dedup: skip if phone OR email matches any existing lead in this branch
        dedupe_terms = []
        if phone:
            dedupe_terms.append({"number": phone})
        if email:
            dedupe_terms.append({"email": email})
        if dedupe_terms:
            dedupe_q = {"branch_id": branch_id, "is_deleted": {"$ne": True}, "$or": dedupe_terms}
            existing = await db.leads.find_one(dedupe_q, {"_id": 0, "id": 1})
            if existing:
                skipped += 1
                continue

        program_id = await _resolve_program_id(branch_id, course)
        if not program_id:
            skipped += 1
            continue
        program_doc = await db.programs.find_one({"id": program_id}, {"_id": 0, "name": 1})

        lead_id = await generate_custom_id(branch_id, "L")
        new_lead = Lead(
            lead_id=lead_id,
            name=name,
            email=email or f"{phone}@meta.lead",
            number=phone or "",
            program_id=program_id,
            program_name=(program_doc or {}).get("name", course or ""),
            lead_source="Ads",
            status=LeadStatus.NEW,
            branch_id=branch_id,
            counsellor_id="system",
        )
        lead_dict = new_lead.model_dump()
        lead_dict["created_at"] = lead_dict["created_at"].isoformat()
        lead_dict["updated_at"] = lead_dict["updated_at"].isoformat()
        await db.leads.insert_one(lead_dict)
        imported += 1

        try:
            await send_whatsapp_notification(
                new_lead.number,
                "enquiry_saved",
                {"name": new_lead.name, "course": new_lead.program_name or ""},
            )
        except Exception as e:
            logger.warning(f"WhatsApp enquiry_saved failed for sheet lead {new_lead.id}: {e}")

    return {"ok": True, "imported": imported, "skipped": skipped, "new_last_row": new_last_row}


async def sync_meta_sheet_for_branch(branch: dict) -> dict:
    """Sync every Google Sheet configured for the branch (multi-sheet aware).
    Supports both legacy single `meta_ads_sheet_url` and new `meta_ads_sheets` list."""
    branch_id = branch.get("id")
    now_iso = datetime.now(timezone.utc).isoformat()
    totals = {"branch_id": branch_id, "ok": True, "imported": 0, "skipped": 0, "sheets": []}

    # Legacy single-sheet field (backwards compat)
    legacy_url = (branch.get("meta_ads_sheet_url") or "").strip()
    if legacy_url:
        res = await _sync_one_sheet(branch_id, legacy_url, int(branch.get("meta_ads_sheet_last_row") or 0))
        await db.branches.update_one(
            {"id": branch_id},
            {"$set": {
                "meta_ads_sheet_last_sync": now_iso,
                "meta_ads_sheet_last_row": res.get("new_last_row", 0),
            }},
        )
        totals["imported"] += res.get("imported", 0)
        totals["skipped"] += res.get("skipped", 0)
        totals["sheets"].append({"label": "legacy", "url": legacy_url, **res})

    # New multi-sheet list
    sheets = branch.get("meta_ads_sheets") or []
    for sh in sheets:
        sh_id = sh.get("id")
        sh_url = (sh.get("url") or "").strip()
        if not sh_url:
            continue
        res = await _sync_one_sheet(branch_id, sh_url, int(sh.get("last_row") or 0), sh.get("column_mapping") or None)
        # Persist this sheet's pointer
        await db.branches.update_one(
            {"id": branch_id, "meta_ads_sheets.id": sh_id},
            {"$set": {
                "meta_ads_sheets.$.last_row": res.get("new_last_row", 0),
                "meta_ads_sheets.$.last_sync": now_iso,
                "meta_ads_sheets.$.last_result": {"imported": res.get("imported", 0), "skipped": res.get("skipped", 0), "reason": res.get("reason")},
            }},
        )
        totals["imported"] += res.get("imported", 0)
        totals["skipped"] += res.get("skipped", 0)
        totals["sheets"].append({"id": sh_id, "label": sh.get("label"), "url": sh_url, **res})

    if not totals["sheets"]:
        return {"branch_id": branch_id, "ok": False, "reason": "no_sheets_configured", "imported": 0}

    return totals


async def sync_all_meta_sheets():
    """Cron entry-point — sync every branch that has any configured sheet (legacy or multi)."""
    try:
        branches = await db.branches.find(
            {"$or": [
                {"meta_ads_sheet_url": {"$exists": True, "$ne": None, "$nin": [""]}},
                {"meta_ads_sheets.0": {"$exists": True}},
            ]},
            {"_id": 0},
        ).to_list(500)
        for b in branches:
            try:
                res = await sync_meta_sheet_for_branch(b)
                if res.get("imported", 0):
                    logger.info(f"Meta sheet sync ({b.get('code')}): imported={res['imported']}")
            except Exception as e:
                logger.error(f"Meta sheet sync failed for branch {b.get('id')}: {e}")
    except Exception as e:
        logger.error(f"sync_all_meta_sheets error: {e}")


def _ensure_branch_access(current_user: User, branch_id: str):
    if current_user.role not in [UserRole.ADMIN, UserRole.BRANCH_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id != branch_id:
        raise HTTPException(status_code=403, detail="You can only configure your own branch")


# ---------- Legacy single-sheet endpoints (kept for backwards compat) ----------
@api_router.put("/branches/{branch_id}/meta-sheet")
async def update_branch_meta_sheet(
    branch_id: str,
    payload: dict = Body(...),
    current_user: User = Depends(get_current_user),
):
    """LEGACY: Set / clear the single Meta Ads Google Sheet URL."""
    _ensure_branch_access(current_user, branch_id)
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    new_url = (payload.get("meta_ads_sheet_url") or "").strip() or None
    update = {"meta_ads_sheet_url": new_url}
    if new_url != branch.get("meta_ads_sheet_url"):
        update["meta_ads_sheet_last_row"] = 0
        update["meta_ads_sheet_last_sync"] = None
    await db.branches.update_one({"id": branch_id}, {"$set": update})
    return {"ok": True, "meta_ads_sheet_url": new_url}


@api_router.post("/branches/{branch_id}/meta-sheet/sync")
async def trigger_meta_sheet_sync(branch_id: str, current_user: User = Depends(get_current_user)):
    """Manually trigger a sync from every configured Google Sheet (legacy + multi)."""
    _ensure_branch_access(current_user, branch_id)
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    if not branch.get("meta_ads_sheet_url") and not branch.get("meta_ads_sheets"):
        raise HTTPException(status_code=400, detail="No Google Sheets configured for this branch")
    return await sync_meta_sheet_for_branch(branch)


# ---------- Multi-sheet CRUD ----------
@api_router.get("/branches/{branch_id}/meta-sheets")
async def list_meta_sheets(branch_id: str, current_user: User = Depends(get_current_user)):
    _ensure_branch_access(current_user, branch_id)
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    return {"sheets": branch.get("meta_ads_sheets") or []}


@api_router.post("/branches/{branch_id}/meta-sheets/preview")
async def preview_meta_sheet(
    branch_id: str,
    payload: dict = Body(...),
    current_user: User = Depends(get_current_user),
):
    """Fetch the first few rows of a Google Sheet so the user can map columns
    before saving. Returns headers, a couple of sample rows, and an
    auto-detected mapping suggestion (column index → canonical field).
    Surfaces a clean error message when the sheet is private or invalid."""
    _ensure_branch_access(current_user, branch_id)
    url = (payload.get("url") or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="url is required")
    if not _sheet_to_csv_url(url):
        raise HTTPException(status_code=400, detail="Not a valid Google Sheets URL. Paste the full URL that contains /spreadsheets/d/...")
    fetch = await _fetch_sheet_rows(url)
    if not fetch["ok"]:
        raise HTTPException(status_code=400, detail=fetch["error_message"])
    rows = fetch["rows"]
    headers = rows[0] if rows else []
    sample_rows = rows[1:4]  # up to 3 preview rows
    # Auto-detect canonical mapping (column index as string → field)
    auto_mapping: dict = {}
    for i, h in enumerate(headers):
        key = _SHEET_HEADER_MAP.get(_normalise_header(h))
        if key and key not in auto_mapping.values():
            auto_mapping[str(i)] = key
    return {
        "ok": True,
        "headers": headers,
        "sample_rows": sample_rows,
        "row_count": max(0, len(rows) - 1),
        "auto_mapping": auto_mapping,
    }


@api_router.post("/branches/{branch_id}/meta-sheets")
async def add_meta_sheet(
    branch_id: str,
    payload: dict = Body(...),
    current_user: User = Depends(get_current_user),
):
    _ensure_branch_access(current_user, branch_id)
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    url = (payload.get("url") or "").strip()
    label = (payload.get("label") or "").strip() or "Sheet"
    column_mapping = payload.get("column_mapping") or {}
    # Normalise: keep only valid canonical targets
    valid_fields = {"name", "email", "phone", "program"}
    column_mapping = {str(k): v for k, v in column_mapping.items() if v in valid_fields}
    if not url:
        raise HTTPException(status_code=400, detail="url is required")
    if not _sheet_to_csv_url(url):
        raise HTTPException(status_code=400, detail="Not a valid Google Sheets URL")
    # Validate the sheet is reachable BEFORE saving, so the user gets a clear error.
    fetch = await _fetch_sheet_rows(url)
    if not fetch["ok"]:
        raise HTTPException(status_code=400, detail=fetch["error_message"])
    new_sheet = {
        "id": str(uuid.uuid4()),
        "url": url,
        "label": label,
        "last_row": 0,
        "last_sync": None,
        "column_mapping": column_mapping,
    }
    await db.branches.update_one({"id": branch_id}, {"$push": {"meta_ads_sheets": new_sheet}})
    return {"ok": True, "sheet": new_sheet}


@api_router.put("/branches/{branch_id}/meta-sheets/{sheet_id}")
async def update_meta_sheet_mapping(
    branch_id: str,
    sheet_id: str,
    payload: dict = Body(...),
    current_user: User = Depends(get_current_user),
):
    """Update column mapping (and optionally label) for an existing sheet."""
    _ensure_branch_access(current_user, branch_id)
    update_doc: dict = {}
    if "label" in payload:
        update_doc["meta_ads_sheets.$.label"] = (payload.get("label") or "Sheet").strip()
    if "column_mapping" in payload:
        valid_fields = {"name", "email", "phone", "program"}
        cm = payload.get("column_mapping") or {}
        update_doc["meta_ads_sheets.$.column_mapping"] = {
            str(k): v for k, v in cm.items() if v in valid_fields
        }
    if not update_doc:
        return {"ok": True}
    res = await db.branches.update_one(
        {"id": branch_id, "meta_ads_sheets.id": sheet_id},
        {"$set": update_doc},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sheet not found")
    return {"ok": True}


@api_router.delete("/branches/{branch_id}/meta-sheets/{sheet_id}")
async def delete_meta_sheet(branch_id: str, sheet_id: str, current_user: User = Depends(get_current_user)):
    _ensure_branch_access(current_user, branch_id)
    await db.branches.update_one(
        {"id": branch_id},
        {"$pull": {"meta_ads_sheets": {"id": sheet_id}}},
    )
    return {"ok": True}


@api_router.post("/branches/{branch_id}/meta-sheets/{sheet_id}/sync")
async def sync_meta_sheet_one(branch_id: str, sheet_id: str, current_user: User = Depends(get_current_user)):
    _ensure_branch_access(current_user, branch_id)
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    sheets = branch.get("meta_ads_sheets") or []
    sh = next((s for s in sheets if s.get("id") == sheet_id), None)
    if not sh:
        raise HTTPException(status_code=404, detail="Sheet not found")
    res = await _sync_one_sheet(branch_id, sh.get("url", ""), int(sh.get("last_row") or 0), sh.get("column_mapping") or None)
    await db.branches.update_one(
        {"id": branch_id, "meta_ads_sheets.id": sheet_id},
        {"$set": {
            "meta_ads_sheets.$.last_row": res.get("new_last_row", 0),
            "meta_ads_sheets.$.last_sync": datetime.now(timezone.utc).isoformat(),
            "meta_ads_sheets.$.last_result": {"imported": res.get("imported", 0), "skipped": res.get("skipped", 0), "reason": res.get("reason")},
        }},
    )
    return {"ok": True, "result": res}

# ========== DEMO REMINDER (30 MIN BEFORE) ==========
async def send_demo_reminders():
    """Look ~30 min ahead and send a 'reach 5 min early' WhatsApp to upcoming demos.
    Runs every 5 min — uses a wider 25..35 min window to be tolerant of the schedule.
    Marks each lead with demo_reminder_sent=True so we never double-fire."""
    try:
        now = datetime.now(timezone.utc)
        # Use IST date/time strings because demo_date/demo_time are entered by humans
        # in local time, not UTC. The server.py codebase stores them as plain
        # 'YYYY-MM-DD' / 'HH:MM' strings without TZ, so we compare in the server's
        # local clock (assumed IST). Convert UTC → IST (+5:30).
        ist = now + timedelta(hours=5, minutes=30)
        today_str = ist.strftime("%Y-%m-%d")

        # Window: [now+25min, now+35min] in IST
        win_start = ist + timedelta(minutes=25)
        win_end = ist + timedelta(minutes=35)

        leads = await db.leads.find(
            {
                "status": LeadStatus.DEMO_BOOKED.value,
                "demo_date": today_str,
                "demo_reminder_sent": {"$ne": True},
                "is_deleted": {"$ne": True},
            },
            {"_id": 0},
        ).to_list(2000)

        sent = 0
        for lead in leads:
            t = (lead.get("demo_time") or "").strip()
            if not t:
                continue
            # Accept HH:MM and HH:MM:SS
            try:
                hh, mm = t.split(":")[0:2]
                demo_dt = ist.replace(hour=int(hh), minute=int(mm), second=0, microsecond=0)
            except Exception:
                continue
            if win_start <= demo_dt <= win_end:
                try:
                    await send_whatsapp_notification(
                        lead.get("number", ""),
                        "demo_reminder",
                        {
                            "name": lead.get("name", ""),
                            "trainer": lead.get("trainer_name") or "your trainer",
                            "demo_time": t,
                        },
                    )
                    await db.leads.update_one({"id": lead["id"]}, {"$set": {"demo_reminder_sent": True}})
                    sent += 1
                except Exception as e:
                    logger.warning(f"demo_reminder failed for lead {lead.get('id')}: {e}")
        if sent:
            logger.info(f"Demo reminders sent: {sent}")
    except Exception as e:
        logger.error(f"send_demo_reminders error: {e}")

# ========== PLACEMENT MANAGER MODULE ==========
class InterviewStatus(str, Enum):
    SCHEDULED = "Scheduled"
    COMPLETED = "Completed"
    CANCELLED = "Cancelled"
    SELECTED = "Selected"
    REJECTED = "Rejected"

class PlacementInterview(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str  # enrollment.id
    student_name: str
    program_name: Optional[str] = None
    company_name: str
    role: Optional[str] = None
    interview_date: str  # YYYY-MM-DD
    interview_time: Optional[str] = None  # HH:MM
    mode: Optional[str] = "Online"  # Online / Walk-in / Telephonic
    location_or_link: Optional[str] = None
    status: InterviewStatus = InterviewStatus.SCHEDULED
    notes: Optional[str] = None
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PlacementRemark(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str
    remark: str
    created_by: Optional[str] = None
    created_by_name: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PlacementRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str
    student_name: str
    program_name: Optional[str] = None
    company_name: str
    designation: Optional[str] = None
    salary_lpa: Optional[float] = None
    city: Optional[str] = None
    joining_date: Optional[str] = None
    offer_letter_url: Optional[str] = None
    remarks: Optional[str] = None
    placed_by: Optional[str] = None
    placed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

def _require_placement_role(current_user: User):
    if current_user.role not in [UserRole.ADMIN, UserRole.PLACEMENT_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied — Placement Manager only")

@api_router.get("/placement/students")
async def placement_list_students(
    branch_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    """List students whose enrollment status is 'Completed' — i.e. course finished
    and ready for placement."""
    _require_placement_role(current_user)
    q = {"status": EnrollmentStatus.COMPLETED.value}
    if branch_id:
        q["branch_id"] = branch_id
    enrollments = await db.enrollments.find(q, {"_id": 0}).sort("updated_at", -1).to_list(2000)

    out = []
    placed_ids = set(
        r["student_id"]
        for r in await db.placements.find({}, {"_id": 0, "student_id": 1}).to_list(5000)
    )
    for e in enrollments:
        sid = e.get("id")
        out.append({
            "id": sid,
            "student_id": e.get("student_id") or sid,
            "name": e.get("student_name"),
            "email": e.get("email"),
            "phone": e.get("phone"),
            "program_name": e.get("program_name"),
            "branch_id": e.get("branch_id"),
            "highest_qualification": e.get("highest_qualification"),
            "date_of_birth": e.get("date_of_birth"),
            "address": e.get("address"),
            "city": e.get("city"),
            "state": e.get("state"),
            "completed_at": e.get("updated_at"),
            "is_placed": sid in placed_ids,
        })
    return {"count": len(out), "students": out}

@api_router.get("/placement/students/{student_id}")
async def placement_student_detail(student_id: str, current_user: User = Depends(get_current_user)):
    _require_placement_role(current_user)
    e = await db.enrollments.find_one({"id": student_id}, {"_id": 0})
    if not e:
        raise HTTPException(status_code=404, detail="Student not found")
    interviews = await db.placement_interviews.find({"student_id": student_id}, {"_id": 0}).sort("interview_date", -1).to_list(200)
    remarks = await db.placement_remarks.find({"student_id": student_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    placement = await db.placements.find_one({"student_id": student_id}, {"_id": 0})
    return {"student": e, "interviews": interviews, "remarks": remarks, "placement": placement}

@api_router.post("/placement/resume/generate")
async def placement_generate_resume(payload: dict = Body(...), current_user: User = Depends(get_current_user)):
    """Generate an ATS-friendly resume PDF via Claude Sonnet 4.5. Returns a base64
    PDF the frontend can download immediately."""
    _require_placement_role(current_user)
    student_id = payload.get("student_id")
    if not student_id:
        raise HTTPException(status_code=400, detail="student_id required")
    e = await db.enrollments.find_one({"id": student_id}, {"_id": 0})
    if not e:
        raise HTTPException(status_code=404, detail="Student not found")

    # Caller-supplied overrides (skills / projects / experience / objective)
    extras = {
        "objective": payload.get("objective") or "",
        "skills": payload.get("skills") or "",  # comma-separated
        "projects": payload.get("projects") or "",  # free-form text
        "experience": payload.get("experience") or "",
        "certifications": payload.get("certifications") or "",
        "languages": payload.get("languages") or "English, Hindi",
    }

    student_ctx = {
        "name": e.get("student_name"),
        "email": e.get("email"),
        "phone": e.get("phone"),
        "address": e.get("address"),
        "city": e.get("city"),
        "state": e.get("state"),
        "date_of_birth": e.get("date_of_birth"),
        "highest_qualification": e.get("highest_qualification"),
        "program": e.get("program_name"),
    }

    # Ask Claude for a strict JSON resume payload, then render to PDF locally.
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=os.environ["EMERGENT_LLM_KEY"],
            session_id=f"resume-{student_id}",
            system_message=(
                "You are an expert career coach. Produce an ATS-ready resume as STRICT JSON only "
                "(no markdown, no preamble). Schema: {"
                "\"summary\": string, "
                "\"skills\": [string], "
                "\"education\": [{\"degree\": string, \"institution\": string, \"year\": string}], "
                "\"experience\": [{\"role\": string, \"company\": string, \"duration\": string, \"bullets\": [string]}], "
                "\"projects\": [{\"title\": string, \"description\": string, \"tech\": [string]}], "
                "\"certifications\": [string], "
                "\"languages\": [string]}"
            ),
        ).with_model("anthropic", "claude-sonnet-4-5")
        prompt = (
            "Build an ATS-optimised resume from this candidate data. Use strong action verbs, "
            "quantify achievements when possible, and tailor the summary toward the candidate's "
            "training program.\n\n"
            f"CANDIDATE: {json.dumps(student_ctx)}\n"
            f"USER-PROVIDED EXTRAS: {json.dumps(extras)}\n\n"
            "Return JSON only."
        )
        ai_resp = await chat.send_message(UserMessage(text=prompt))
        raw = ai_resp.strip()
        if raw.startswith("```"):
            raw = _re.sub(r"^```(?:json)?", "", raw).strip()
            raw = _re.sub(r"```$", "", raw).strip()
        resume_data = json.loads(raw)
    except Exception as ex:
        logger.warning(f"Resume LLM generation failed, falling back to template: {ex}")

        def _to_list(v):
            """Accept either a list (frontend-shape) or a comma-separated string."""
            if isinstance(v, list):
                return [str(x).strip() for x in v if x and str(x).strip()]
            return [s.strip() for s in (v or "").split(",") if s.strip()]

        resume_data = {
            "summary": extras.get("objective") or f"Recent {student_ctx.get('program') or 'graduate'} with strong fundamentals, ready to contribute.",
            "skills": _to_list(extras.get("skills")),
            "education": [{
                "degree": student_ctx.get("highest_qualification") or "",
                "institution": "",
                "year": "",
            }] if student_ctx.get("highest_qualification") else [],
            "experience": [],
            "projects": [],
            "certifications": _to_list(extras.get("certifications")),
            "languages": _to_list(extras.get("languages")),
        }

    # Render PDF using reportlab
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.colors import HexColor
    import base64

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=20, leading=24,
                        textColor=HexColor("#0f172a"), spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, leading=13,
                         textColor=HexColor("#475569"), spaceAfter=10)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, leading=14,
                        textColor=HexColor("#0f172a"), spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=13,
                          textColor=HexColor("#1f2937"))
    bullet = ParagraphStyle("bullet", parent=body, leftIndent=10, bulletIndent=0)

    flow = []
    flow.append(Paragraph(student_ctx.get("name") or "Candidate", h1))
    contact_bits = [b for b in [student_ctx.get("email"), student_ctx.get("phone"),
                                ", ".join([x for x in [student_ctx.get("city"), student_ctx.get("state")] if x])] if b]
    flow.append(Paragraph(" &nbsp;•&nbsp; ".join(contact_bits), sub))

    if resume_data.get("summary"):
        flow.append(Paragraph("PROFESSIONAL SUMMARY", h2))
        flow.append(Paragraph(resume_data["summary"], body))

    if resume_data.get("skills"):
        flow.append(Paragraph("CORE SKILLS", h2))
        flow.append(Paragraph(" • ".join(resume_data["skills"]), body))

    if resume_data.get("experience"):
        flow.append(Paragraph("EXPERIENCE", h2))
        for x in resume_data["experience"]:
            head = f"<b>{x.get('role','')}</b> — {x.get('company','')} <font color='#64748b'>({x.get('duration','')})</font>"
            flow.append(Paragraph(head, body))
            for b in x.get("bullets", []) or []:
                flow.append(Paragraph(f"• {b}", bullet))
            flow.append(Spacer(1, 4))

    if resume_data.get("projects"):
        flow.append(Paragraph("PROJECTS", h2))
        for p in resume_data["projects"]:
            tech = ", ".join(p.get("tech", []) or [])
            head = f"<b>{p.get('title','')}</b>" + (f" <font color='#64748b'>[{tech}]</font>" if tech else "")
            flow.append(Paragraph(head, body))
            if p.get("description"):
                flow.append(Paragraph(p["description"], body))
            flow.append(Spacer(1, 4))

    if resume_data.get("education"):
        flow.append(Paragraph("EDUCATION", h2))
        for ed in resume_data["education"]:
            flow.append(Paragraph(f"<b>{ed.get('degree','')}</b> — {ed.get('institution','')} <font color='#64748b'>({ed.get('year','')})</font>", body))

    if student_ctx.get("program"):
        flow.append(Paragraph("TRAINING", h2))
        flow.append(Paragraph(f"<b>{student_ctx['program']}</b> — ETI Educom", body))

    if resume_data.get("certifications"):
        flow.append(Paragraph("CERTIFICATIONS", h2))
        for c in resume_data["certifications"]:
            flow.append(Paragraph(f"• {c}", bullet))

    if resume_data.get("languages"):
        flow.append(Paragraph("LANGUAGES", h2))
        flow.append(Paragraph(", ".join(resume_data["languages"]), body))

    doc.build(flow)
    pdf_b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    # Persist the latest generation on the student record
    await db.placement_resumes.update_one(
        {"student_id": student_id},
        {"$set": {
            "student_id": student_id,
            "resume_json": resume_data,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "generated_by": current_user.id,
        }},
        upsert=True,
    )

    return {
        "ok": True,
        "filename": f"{(student_ctx.get('name') or 'resume').replace(' ', '_')}_resume.pdf",
        "pdf_base64": pdf_b64,
        "resume": resume_data,
    }

@api_router.post("/placement/interview")
async def placement_create_interview(payload: PlacementInterview = Body(...), current_user: User = Depends(get_current_user)):
    _require_placement_role(current_user)
    e = await db.enrollments.find_one({"id": payload.student_id}, {"_id": 0})
    if not e:
        raise HTTPException(status_code=404, detail="Student not found")
    d = payload.model_dump()
    d["created_by"] = current_user.id
    d["student_name"] = e.get("student_name")
    d["program_name"] = e.get("program_name")
    d["created_at"] = d["created_at"].isoformat()
    await db.placement_interviews.insert_one(d)
    d.pop("_id", None)
    return {"ok": True, "interview": d}

@api_router.get("/placement/interviews")
async def placement_list_interviews(
    student_id: Optional[str] = None,
    upcoming: Optional[bool] = None,
    current_user: User = Depends(get_current_user),
):
    _require_placement_role(current_user)
    q = {}
    if student_id:
        q["student_id"] = student_id
    if upcoming:
        q["interview_date"] = {"$gte": datetime.now(timezone.utc).strftime("%Y-%m-%d")}
        q["status"] = InterviewStatus.SCHEDULED.value
    return {"interviews": await db.placement_interviews.find(q, {"_id": 0}).sort("interview_date", 1).to_list(500)}

@api_router.put("/placement/interview/{interview_id}")
async def placement_update_interview(interview_id: str, payload: dict = Body(...), current_user: User = Depends(get_current_user)):
    _require_placement_role(current_user)
    allowed = {"status", "notes", "interview_date", "interview_time", "mode", "location_or_link", "role", "company_name"}
    patch = {k: v for k, v in payload.items() if k in allowed}
    if not patch:
        raise HTTPException(status_code=400, detail="Nothing to update")
    r = await db.placement_interviews.update_one({"id": interview_id}, {"$set": patch})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Interview not found")
    return {"ok": True}

@api_router.post("/placement/remark")
async def placement_add_remark(payload: PlacementRemark = Body(...), current_user: User = Depends(get_current_user)):
    _require_placement_role(current_user)
    d = payload.model_dump()
    d["created_by"] = current_user.id
    d["created_by_name"] = current_user.name
    d["created_at"] = d["created_at"].isoformat()
    await db.placement_remarks.insert_one(d)
    d.pop("_id", None)
    return {"ok": True, "remark": d}

@api_router.post("/placement/mark-placed")
async def placement_mark_placed(payload: PlacementRecord = Body(...), current_user: User = Depends(get_current_user)):
    _require_placement_role(current_user)
    e = await db.enrollments.find_one({"id": payload.student_id}, {"_id": 0})
    if not e:
        raise HTTPException(status_code=404, detail="Student not found")
    d = payload.model_dump()
    d["placed_by"] = current_user.id
    d["student_name"] = e.get("student_name")
    d["program_name"] = e.get("program_name")
    d["placed_at"] = d["placed_at"].isoformat()

    # Upsert so re-marking updates the record instead of duplicating it
    await db.placements.update_one(
        {"student_id": payload.student_id},
        {"$set": d},
        upsert=True,
    )

    # Fire a WhatsApp congrats message (uses existing certificate_ready template
    # as a generic congratulatory channel; safe fallback if a placement-specific
    # template isn't configured yet).
    try:
        await send_whatsapp_notification(
            e.get("phone", ""),
            "placement_confirmed",
            {
                "name": e.get("student_name", ""),
                "company": payload.company_name,
                "designation": payload.designation or "",
                "salary": str(payload.salary_lpa or ""),
                "city": payload.city or "",
            },
        )
    except Exception as ex:
        logger.warning(f"placement_confirmed WhatsApp failed: {ex}")

    return {"ok": True, "placement": d}

@api_router.get("/placement/placements")
async def placement_list_placements(current_user: User = Depends(get_current_user)):
    _require_placement_role(current_user)
    return {"placements": await db.placements.find({}, {"_id": 0}).sort("placed_at", -1).to_list(2000)}

@api_router.get("/placement/dashboard-stats")
async def placement_dashboard_stats(current_user: User = Depends(get_current_user)):
    _require_placement_role(current_user)
    completed = await db.enrollments.count_documents({"status": EnrollmentStatus.COMPLETED.value})
    placed = await db.placements.count_documents({})
    upcoming_interviews = await db.placement_interviews.count_documents({
        "interview_date": {"$gte": datetime.now(timezone.utc).strftime("%Y-%m-%d")},
        "status": InterviewStatus.SCHEDULED.value,
    })
    return {
        "completed_students": completed,
        "placed_students": placed,
        "pending_for_placement": max(0, completed - placed),
        "upcoming_interviews": upcoming_interviews,
    }

# ========== META/FACEBOOK INTEGRATION ==========

META_GRAPH_API_BASE = "https://graph.facebook.com/v20.0"

@api_router.get("/meta/config/{branch_id}")
async def get_meta_config(branch_id: str, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Get Meta configuration for a branch - Super Admin only"""
    config = await db.meta_configs.find_one({"branch_id": branch_id}, {"_id": 0})
    if not config:
        return {"configured": False, "message": "Meta not configured for this branch"}
    # Hide sensitive data
    config['app_secret'] = "***hidden***"
    config['access_token'] = "***hidden***" if config.get('access_token') else None
    return {"configured": True, "config": config}

@api_router.post("/meta/config")
async def create_meta_config(config: MetaConfigCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Create Meta configuration for a branch - Super Admin only"""
    existing = await db.meta_configs.find_one({"branch_id": config.branch_id})
    if existing:
        raise HTTPException(status_code=400, detail="Meta config already exists for this branch. Use PUT to update.")
    
    # Auto-fix Ad Account ID - add act_ prefix if missing
    config_data = config.model_dump()
    if config_data.get('ad_account_id') and not config_data['ad_account_id'].startswith('act_'):
        config_data['ad_account_id'] = f"act_{config_data['ad_account_id']}"
    
    meta_config = MetaConfig(**config_data)
    config_dict = meta_config.model_dump()
    config_dict['created_at'] = config_dict['created_at'].isoformat()
    config_dict['updated_at'] = config_dict['updated_at'].isoformat()
    
    await db.meta_configs.insert_one(config_dict)
    
    logger.info(f"Meta config created for branch {config.branch_id} by {current_user.email}")
    return {"message": "Meta configuration saved", "webhook_verify_token": meta_config.webhook_verify_token}

@api_router.put("/meta/config/{branch_id}")
async def update_meta_config(branch_id: str, update: MetaConfigUpdate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """Update Meta configuration for a branch - Super Admin only"""
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    # Auto-fix Ad Account ID - add act_ prefix if missing
    if update_data.get('ad_account_id') and not update_data['ad_account_id'].startswith('act_'):
        update_data['ad_account_id'] = f"act_{update_data['ad_account_id']}"
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.meta_configs.update_one(
        {"branch_id": branch_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Meta config not found for this branch")
    
    return {"message": "Meta configuration updated"}

@api_router.get("/meta/configs")
async def list_meta_configs(current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """List all Meta configurations - Super Admin only"""
    configs = await db.meta_configs.find({}, {"_id": 0, "app_secret": 0, "access_token": 0}).to_list(100)
    return configs

# Webhook for Facebook Lead Ads
@api_router.get("/webhooks/facebook-leads")
async def verify_facebook_webhook(request: Request):
    """Verify Facebook webhook - called by Meta during setup"""
    mode = request.query_params.get("hub.mode")
    verify_token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")
    
    logger.info(f"Facebook webhook verification: mode={mode}, token={verify_token}")
    
    if mode == "subscribe":
        # Find config with matching verify token
        config = await db.meta_configs.find_one({"webhook_verify_token": verify_token})
        if config:
            logger.info("Facebook webhook verification successful")
            return PlainTextResponse(challenge)
    
    logger.warning("Facebook webhook verification failed")
    raise HTTPException(status_code=403, detail="Verification failed")

@api_router.post("/webhooks/facebook-leads")
async def handle_facebook_lead_webhook(request: Request):
    """Handle incoming lead notifications from Facebook"""
    body = await request.body()
    
    try:
        payload = json.loads(body)
        logger.info(f"Facebook webhook received: {json.dumps(payload)[:500]}")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    
    # Process lead notifications
    if payload.get("object") == "page":
        for entry in payload.get("entry", []):
            page_id = entry.get("id")
            
            # Find branch config for this page
            config = await db.meta_configs.find_one({
                "$or": [{"page_id": page_id}, {"page_ids": page_id}],
                "is_active": True
            })
            
            if not config:
                logger.warning(f"No config found for page {page_id}")
                continue
            
            branch_id = config['branch_id']
            
            for change in entry.get("changes", []):
                if change.get("field") == "leadgen":
                    lead_data = change.get("value", {})
                    leadgen_id = lead_data.get("leadgen_id")
                    
                    if leadgen_id:
                        # Store the lead notification
                        meta_lead = MetaLead(
                            branch_id=branch_id,
                            leadgen_id=leadgen_id,
                            page_id=page_id,
                            form_id=lead_data.get("form_id"),
                            ad_id=lead_data.get("ad_id"),
                            fb_created_time=str(lead_data.get("created_time"))
                        )
                        
                        lead_dict = meta_lead.model_dump()
                        lead_dict['created_at'] = lead_dict['created_at'].isoformat()
                        await db.meta_leads.insert_one(lead_dict)
                        
                        # Try to fetch full lead data
                        try:
                            await fetch_and_sync_lead(config, meta_lead)
                        except Exception as e:
                            logger.error(f"Error fetching lead {leadgen_id}: {e}")
    
    return {"status": "ok"}

async def fetch_and_sync_lead(config: dict, meta_lead: MetaLead):
    """Fetch full lead data from Facebook and create CRM lead"""
    access_token = config.get('access_token')
    if not access_token:
        logger.warning("No access token for fetching lead data")
        return
    
    # Fetch lead data from Graph API
    url = f"{META_GRAPH_API_BASE}/{meta_lead.leadgen_id}"
    params = {
        "access_token": access_token,
        "fields": "id,created_time,field_data,ad_id,form_id,campaign_id,ad_name,campaign_name"
    }
    
    async with httpx.AsyncClient() as client:
        response = await client.get(url, params=params)
        if response.status_code != 200:
            logger.error(f"Error fetching lead: {response.text}")
            return
        
        lead_data = response.json()
    
    # Parse field data
    field_data = {}
    name = None
    email = None
    phone = None
    
    for field in lead_data.get("field_data", []):
        field_name = field.get("name", "").lower()
        values = field.get("values", [])
        value = values[0] if values else None
        field_data[field_name] = value
        
        if "name" in field_name or "full_name" in field_name:
            name = value
        elif "email" in field_name:
            email = value
        elif "phone" in field_name or "mobile" in field_name:
            phone = value
    
    # Update meta lead with parsed data
    await db.meta_leads.update_one(
        {"id": meta_lead.id},
        {"$set": {
            "name": name,
            "email": email,
            "phone": phone,
            "field_data": field_data,
            "campaign_id": lead_data.get("campaign_id"),
            "campaign_name": lead_data.get("campaign_name"),
            "ad_name": lead_data.get("ad_name")
        }}
    )
    
    # Create CRM lead if we have enough data
    if name and (email or phone):
        # Get default program for the branch
        branch = await db.branches.find_one({"id": config['branch_id']}, {"_id": 0})
        programs = await db.programs.find({}, {"_id": 0, "id": 1}).to_list(1)
        default_program_id = programs[0]['id'] if programs else None
        
        if default_program_id:
            # Generate custom lead ID
            lead_id = await generate_custom_id(config['branch_id'], 'L')
            
            crm_lead = Lead(
                lead_id=lead_id,
                name=name,
                email=email or f"{phone}@facebook.lead",
                number=phone or "",
                program_id=default_program_id,
                lead_source="Facebook Lead Ad",
                status=LeadStatus.NEW,
                branch_id=config['branch_id'],
                counsellor_id="system",  # Auto-imported lead
                meta_lead_id=meta_lead.id,
                meta_campaign=lead_data.get("campaign_name"),
                meta_ad=lead_data.get("ad_name")
            )
            
            lead_dict = crm_lead.model_dump()
            lead_dict['created_at'] = lead_dict['created_at'].isoformat()
            await db.leads.insert_one(lead_dict)
            
            # Update meta lead with CRM link
            await db.meta_leads.update_one(
                {"id": meta_lead.id},
                {"$set": {"crm_lead_id": crm_lead.id, "is_synced_to_crm": True}}
            )
            
            logger.info(f"Created CRM lead {crm_lead.id} from Facebook lead {meta_lead.leadgen_id}")

@api_router.get("/meta/leads")
async def get_meta_leads(
    branch_id: Optional[str] = None,
    synced: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """Get Facebook leads - Branch Admin sees their branch only"""
    query = {}
    
    if current_user.role == UserRole.BRANCH_ADMIN:
        query["branch_id"] = current_user.branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    if synced is not None:
        query["is_synced_to_crm"] = synced
    
    leads = await db.meta_leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return leads

@api_router.post("/meta/sync-ads/{branch_id}")
async def sync_ads_data(branch_id: str, current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.BRANCH_ADMIN]))):
    """Manually sync ads data from Facebook - fetches last 30 days"""
    config = await db.meta_configs.find_one({"branch_id": branch_id, "is_active": True})
    if not config:
        raise HTTPException(status_code=404, detail="Meta not configured for this branch")
    
    if not config.get('access_token') or not config.get('ad_account_id'):
        raise HTTPException(status_code=400, detail="Access token and Ad Account ID required for ads sync")
    
    access_token = config['access_token']
    ad_account_id = config['ad_account_id']
    
    # First check if there are any campaigns
    campaigns_url = f"{META_GRAPH_API_BASE}/{ad_account_id}/campaigns"
    campaigns_params = {
        "access_token": access_token,
        "fields": "id,name,status",
        "limit": 10
    }
    
    async with httpx.AsyncClient() as client:
        campaigns_response = await client.get(campaigns_url, params=campaigns_params, timeout=30.0)
        if campaigns_response.status_code != 200:
            error_data = campaigns_response.json() if campaigns_response.text else {}
            error_msg = error_data.get('error', {}).get('message', campaigns_response.text[:200])
            logger.error(f"Error checking campaigns: {error_msg}")
            raise HTTPException(status_code=400, detail=f"Facebook API error: {error_msg}")
        
        campaigns_data = campaigns_response.json()
        campaigns_list = campaigns_data.get("data", [])
        
        if not campaigns_list:
            # No campaigns exist - update last sync time and return
            await db.meta_configs.update_one(
                {"branch_id": branch_id},
                {"$set": {"last_sync_at": datetime.now(timezone.utc).isoformat()}}
            )
            return {
                "message": "No campaigns found in this Ad Account. Create campaigns in Meta Ads Manager first.",
                "campaigns_count": 0,
                "insights_count": 0,
                "period": "N/A"
            }
    
    # Fetch last 30 days of data
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=30)
    
    url = f"{META_GRAPH_API_BASE}/{ad_account_id}/insights"
    params = {
        "access_token": access_token,
        "fields": "account_name,campaign_id,campaign_name,impressions,reach,clicks,spend,cpc,cpm,ctr,actions",
        "time_range": json.dumps({
            "since": start_date.strftime('%Y-%m-%d'),
            "until": end_date.strftime('%Y-%m-%d')
        }),
        "level": "campaign",
        "time_increment": 1  # Daily breakdown
    }
    
    async with httpx.AsyncClient() as client:
        response = await client.get(url, params=params, timeout=30.0)
        if response.status_code != 200:
            error_data = response.json() if response.text else {}
            error_msg = error_data.get('error', {}).get('message', response.text[:200])
            logger.error(f"Error fetching ads insights: {error_msg}")
            
            # If insights not available but campaigns exist, return campaign count
            if "nonexisting field" in error_msg.lower() or "insights" in error_msg.lower():
                await db.meta_configs.update_one(
                    {"branch_id": branch_id},
                    {"$set": {"last_sync_at": datetime.now(timezone.utc).isoformat()}}
                )
                return {
                    "message": f"Found {len(campaigns_list)} campaigns but no performance data yet. Run your ads to see insights.",
                    "campaigns_count": len(campaigns_list),
                    "insights_count": 0,
                    "period": f"{start_date.date()} to {end_date.date()}"
                }
            
            raise HTTPException(status_code=400, detail=f"Facebook API error: {error_msg}")
        
        data = response.json()
    
    # Store insights
    insights_count = 0
    for insight in data.get("data", []):
        ad_insight = MetaAdInsight(
            branch_id=branch_id,
            date=insight.get("date_start", ""),
            account_id=ad_account_id,
            campaign_id=insight.get("campaign_id"),
            campaign_name=insight.get("campaign_name"),
            impressions=int(insight.get("impressions", 0)),
            reach=int(insight.get("reach", 0)),
            clicks=int(insight.get("clicks", 0)),
            spend=float(insight.get("spend", 0)),
            cpc=float(insight.get("cpc", 0)),
            cpm=float(insight.get("cpm", 0)),
            ctr=float(insight.get("ctr", 0)),
            level="campaign"
        )
        
        # Upsert - update if exists, insert if not
        await db.meta_ad_insights.update_one(
            {"branch_id": branch_id, "date": ad_insight.date, "campaign_id": ad_insight.campaign_id},
            {"$set": ad_insight.model_dump()},
            upsert=True
        )
        insights_count += 1
    
    # Update last sync time
    await db.meta_configs.update_one(
        {"branch_id": branch_id},
        {"$set": {"last_sync_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": f"Synced {insights_count} insights records", "period": f"{start_date.date()} to {end_date.date()}"}

@api_router.get("/meta/campaigns/{branch_id}")
async def get_meta_campaigns(
    branch_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get list of campaigns from Meta Ad Account - fetches directly from Facebook API"""
    # Check access
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id != branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    config = await db.meta_configs.find_one({"branch_id": branch_id, "is_active": True})
    if not config:
        raise HTTPException(status_code=404, detail="Meta not configured for this branch")
    
    if not config.get('access_token') or not config.get('ad_account_id'):
        raise HTTPException(status_code=400, detail="Access token and Ad Account ID required")
    
    access_token = config['access_token']
    ad_account_id = config['ad_account_id']
    
    # Fetch campaigns from Facebook
    url = f"{META_GRAPH_API_BASE}/{ad_account_id}/campaigns"
    params = {
        "access_token": access_token,
        "fields": "id,name,status,objective,daily_budget,lifetime_budget,created_time,start_time,stop_time,effective_status",
        "limit": 100
    }
    
    async with httpx.AsyncClient() as client:
        response = await client.get(url, params=params, timeout=30.0)
        if response.status_code != 200:
            logger.error(f"Error fetching campaigns: {response.text}")
            error_data = response.json() if response.text else {}
            error_msg = error_data.get('error', {}).get('message', response.text[:200])
            raise HTTPException(status_code=400, detail=f"Facebook API error: {error_msg}")
        
        data = response.json()
    
    campaigns = []
    for campaign in data.get("data", []):
        campaigns.append({
            "id": campaign.get("id"),
            "name": campaign.get("name"),
            "status": campaign.get("status"),
            "effective_status": campaign.get("effective_status"),
            "objective": campaign.get("objective"),
            "daily_budget": float(campaign.get("daily_budget", 0)) / 100 if campaign.get("daily_budget") else None,  # Convert from cents
            "lifetime_budget": float(campaign.get("lifetime_budget", 0)) / 100 if campaign.get("lifetime_budget") else None,
            "created_time": campaign.get("created_time"),
            "start_time": campaign.get("start_time"),
            "stop_time": campaign.get("stop_time")
        })
    
    return {
        "campaigns": campaigns,
        "total": len(campaigns),
        "ad_account_id": ad_account_id
    }

@api_router.get("/meta/analytics/{branch_id}")
async def get_meta_analytics(
    branch_id: str,
    days: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Get Meta ads analytics for branch - with AI analysis"""
    # Check access
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id != branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=days)
    
    # Get insights data
    insights = await db.meta_ad_insights.find({
        "branch_id": branch_id,
        "date": {"$gte": start_date.strftime('%Y-%m-%d'), "$lte": end_date.strftime('%Y-%m-%d')}
    }, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Get meta leads for this branch in the same period
    meta_leads = await db.meta_leads.find({
        "branch_id": branch_id,
        "created_at": {"$gte": start_date.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    # Get CRM leads from Facebook source
    crm_leads = await db.leads.find({
        "branch_id": branch_id,
        "lead_source": "Facebook Lead Ad",
        "created_at": {"$gte": start_date.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    # Calculate aggregates
    total_spend = sum(i.get('spend', 0) for i in insights)
    total_impressions = sum(i.get('impressions', 0) for i in insights)
    total_reach = sum(i.get('reach', 0) for i in insights)
    total_clicks = sum(i.get('clicks', 0) for i in insights)
    total_leads = len(meta_leads)
    converted_leads = len([l for l in crm_leads if l.get('status') == 'Converted'])
    
    # Cost per lead
    cost_per_lead = total_spend / total_leads if total_leads > 0 else 0
    
    # Campaign breakdown
    campaign_stats = {}
    for insight in insights:
        campaign = insight.get('campaign_name', 'Unknown')
        if campaign not in campaign_stats:
            campaign_stats[campaign] = {
                'spend': 0, 'impressions': 0, 'clicks': 0, 'leads': 0
            }
        campaign_stats[campaign]['spend'] += insight.get('spend', 0)
        campaign_stats[campaign]['impressions'] += insight.get('impressions', 0)
        campaign_stats[campaign]['clicks'] += insight.get('clicks', 0)
    
    # Count leads per campaign
    for lead in meta_leads:
        campaign = lead.get('campaign_name', 'Unknown')
        if campaign in campaign_stats:
            campaign_stats[campaign]['leads'] += 1
    
    # Daily trend
    daily_data = {}
    for insight in insights:
        date = insight.get('date')
        if date not in daily_data:
            daily_data[date] = {'spend': 0, 'impressions': 0, 'clicks': 0, 'leads': 0}
        daily_data[date]['spend'] += insight.get('spend', 0)
        daily_data[date]['impressions'] += insight.get('impressions', 0)
        daily_data[date]['clicks'] += insight.get('clicks', 0)
    
    for lead in meta_leads:
        date = lead.get('created_at', '')[:10]
        if date in daily_data:
            daily_data[date]['leads'] += 1
    
    # AI Analysis
    ai_analysis = None
    if LLM_AVAILABLE and os.environ.get('EMERGENT_LLM_KEY'):
        try:
            analytics_summary = f"""
Facebook Ads Analytics for Branch (Last {days} days):

OVERALL METRICS:
- Total Spend: ${total_spend:.2f}
- Total Impressions: {total_impressions:,}
- Total Reach: {total_reach:,}
- Total Clicks: {total_clicks:,}
- CTR: {(total_clicks/total_impressions*100) if total_impressions > 0 else 0:.2f}%
- Total Leads from Facebook: {total_leads}
- Converted Leads: {converted_leads}
- Cost Per Lead: ${cost_per_lead:.2f}
- Conversion Rate: {(converted_leads/total_leads*100) if total_leads > 0 else 0:.1f}%

CAMPAIGN BREAKDOWN:
{json.dumps(campaign_stats, indent=2)}

DAILY TREND (Last 7 days):
{json.dumps(dict(list(daily_data.items())[:7]), indent=2)}
"""
            
            chat = LlmChat(
                api_key=os.environ.get('EMERGENT_LLM_KEY'),
                session_id=f"meta-analytics-{branch_id}-{datetime.now().strftime('%Y%m%d')}",
                system_message="""You are a digital marketing analyst specializing in Facebook/Meta ads. Analyze the provided data and give actionable insights.

Your response MUST be valid JSON:
{
  "performance_summary": "2-3 sentence summary of overall performance",
  "top_campaign": {"name": "campaign name", "reason": "why it's best"},
  "underperforming": {"name": "campaign name", "issue": "what's wrong"},
  "spend_efficiency": "excellent|good|average|poor",
  "recommendations": [
    "specific recommendation 1",
    "specific recommendation 2",
    "specific recommendation 3"
  ],
  "trend_insight": "observation about daily/weekly trend",
  "roi_assessment": "assessment of return on ad spend"
}

Be specific, data-driven, and actionable."""
            ).with_model("openai", "gpt-4o")
            
            user_message = UserMessage(text=f"Analyze this Facebook Ads data:\n\n{analytics_summary}")
            ai_response = await chat.send_message(user_message)
            
            try:
                ai_text = ai_response.strip()
                if ai_text.startswith("```"):
                    ai_text = ai_text.split("```")[1]
                    if ai_text.startswith("json"):
                        ai_text = ai_text[4:]
                ai_analysis = json.loads(ai_text)
            except json.JSONDecodeError:
                ai_analysis = {"raw_response": ai_response}
                
        except Exception as e:
            logger.error(f"Meta analytics AI analysis failed: {e}")
    
    return {
        "period": {"start": start_date.strftime('%Y-%m-%d'), "end": end_date.strftime('%Y-%m-%d')},
        "summary": {
            "total_spend": round(total_spend, 2),
            "total_impressions": total_impressions,
            "total_reach": total_reach,
            "total_clicks": total_clicks,
            "ctr": round((total_clicks/total_impressions*100) if total_impressions > 0 else 0, 2),
            "total_leads": total_leads,
            "converted_leads": converted_leads,
            "cost_per_lead": round(cost_per_lead, 2),
            "conversion_rate": round((converted_leads/total_leads*100) if total_leads > 0 else 0, 1)
        },
        "campaigns": campaign_stats,
        "daily_trend": daily_data,
        "ai_analysis": ai_analysis,
        "ai_powered": ai_analysis is not None
    }

# ========== ADMIN RESET ENDPOINT ==========
@api_router.post("/admin/reset-system")
async def reset_system(current_user: User = Depends(require_role([UserRole.ADMIN]))):
    """
    Reset all system data - SUPER ADMIN ONLY.
    This will delete ALL data from the system except the super admin account.
    USE WITH EXTREME CAUTION!
    """
    try:
        # Keep track of collections to clear
        collections_to_clear = [
            "leads", "followups", "enrollments", "payments", "payment_plans",
            "installment_schedule", "expenses", "tasks", "notifications",
            "certificate_requests", "quiz_attempts", "exam_bookings",
            "push_subscriptions"
        ]
        
        deleted_counts = {}
        for collection_name in collections_to_clear:
            result = await db[collection_name].delete_many({})
            deleted_counts[collection_name] = result.deleted_count
        
        # Reset branch counters (don't delete branches)
        await db.branches.update_many(
            {},
            {"$set": {"lead_counter": 0, "enrollment_counter": 0, "receipt_counter": 0}}
        )
        
        # Delete all users except the current super admin
        user_delete_result = await db.users.delete_many({
            "id": {"$ne": current_user.id}
        })
        deleted_counts["users"] = user_delete_result.deleted_count
        
        logger.warning(f"SYSTEM RESET performed by {current_user.email}. Data deleted: {deleted_counts}")
        
        return {
            "success": True,
            "message": "System data has been reset successfully",
            "deleted_records": deleted_counts,
            "preserved": {
                "branches": "All branches preserved with counters reset",
                "programs": "All programs preserved",
                "expense_categories": "All expense categories preserved",
                "lead_sources": "All lead sources preserved",
                "whatsapp_settings": "WhatsApp settings preserved",
                "current_admin": current_user.email
            }
        }
    except Exception as e:
        logger.error(f"Error during system reset: {str(e)}")
        raise HTTPException(status_code=500, detail=f"System reset failed: {str(e)}")

# ========== ROYALTY MANAGEMENT ==========

@api_router.get("/royalty/branch/{branch_id}")
async def get_branch_royalty(
    branch_id: str,
    month: int = None,
    year: int = None,
    current_user: User = Depends(get_current_user)
):
    """Get royalty calculation for a specific branch for a given month"""
    # Check access
    if current_user.role == UserRole.BRANCH_ADMIN and current_user.branch_id != branch_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Default to last month
    today = datetime.now(timezone.utc)
    if month is None or year is None:
        # Get last month
        if today.month == 1:
            month = 12
            year = today.year - 1
        else:
            month = today.month - 1
            year = today.year
    
    # Get branch info
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    royalty_percentage = branch.get('royalty_percentage', 0)
    
    # Calculate date range for the month
    start_date = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    
    # Get all enrollment payments for the branch in this period
    # Convert to ISO strings for comparison
    start_str = start_date.isoformat()
    end_str = end_date.isoformat()
    
    payments = await db.payments.find({
        "branch_id": branch_id,
        "payment_date": {"$gte": start_str, "$lt": end_str},
        "payment_type": {"$ne": "Certificate Fee"}  # Exclude certificate fees
    }, {"_id": 0}).to_list(10000)
    
    total_collection = sum(p.get('amount', 0) for p in payments)
    royalty_amount = round(total_collection * (royalty_percentage / 100), 2)
    
    # Due date is 5th of next month
    if month == 12:
        due_date = datetime(year + 1, 1, 5)
    else:
        due_date = datetime(year, month + 1, 5)
    
    # Check if already paid (stored in royalty_payments collection)
    royalty_payment = await db.royalty_payments.find_one({
        "branch_id": branch_id,
        "month": month,
        "year": year
    }, {"_id": 0})
    
    return {
        "branch_id": branch_id,
        "branch_name": branch.get('name'),
        "month": month,
        "year": year,
        "month_name": datetime(year, month, 1).strftime('%B %Y'),
        "royalty_percentage": royalty_percentage,
        "total_collection": total_collection,
        "royalty_amount": royalty_amount,
        "due_date": due_date.strftime('%Y-%m-%d'),
        "due_date_display": f"5th {datetime(year, month + 1 if month < 12 else 1, 1).strftime('%B %Y') if month < 12 else datetime(year + 1, 1, 1).strftime('%B %Y')}",
        "is_paid": royalty_payment.get('is_paid', False) if royalty_payment else False,
        "paid_date": royalty_payment.get('paid_date') if royalty_payment else None,
        "payment_count": len(payments)
    }

@api_router.get("/royalty/all")
async def get_all_branches_royalty(
    month: int = None,
    year: int = None,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Get royalty for all branches - Super Admin only"""
    # Default to last month
    today = datetime.now(timezone.utc)
    if month is None or year is None:
        if today.month == 1:
            month = 12
            year = today.year - 1
        else:
            month = today.month - 1
            year = today.year
    
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    
    royalty_data = []
    total_royalty = 0
    total_paid = 0
    total_pending = 0
    
    for branch in branches:
        branch_id = branch['id']
        royalty_percentage = branch.get('royalty_percentage', 0)
        
        # Calculate date range
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        # Get payments
        payments = await db.payments.find({
            "branch_id": branch_id,
            "payment_date": {"$gte": start_str, "$lt": end_str},
            "payment_type": {"$ne": "Certificate Fee"}
        }, {"_id": 0}).to_list(10000)
        
        total_collection = sum(p.get('amount', 0) for p in payments)
        royalty_amount = round(total_collection * (royalty_percentage / 100), 2)
        
        # Check if paid
        royalty_payment = await db.royalty_payments.find_one({
            "branch_id": branch_id,
            "month": month,
            "year": year
        }, {"_id": 0})
        
        is_paid = royalty_payment.get('is_paid', False) if royalty_payment else False
        
        royalty_data.append({
            "branch_id": branch_id,
            "branch_name": branch.get('name'),
            "city": branch.get('city'),
            "royalty_percentage": royalty_percentage,
            "total_collection": total_collection,
            "royalty_amount": royalty_amount,
            "is_paid": is_paid,
            "paid_date": royalty_payment.get('paid_date') if royalty_payment else None
        })
        
        total_royalty += royalty_amount
        if is_paid:
            total_paid += royalty_amount
        else:
            total_pending += royalty_amount
    
    # Due date
    if month == 12:
        due_date = f"5th January {year + 1}"
    else:
        due_date = f"5th {datetime(year, month + 1, 1).strftime('%B %Y')}"
    
    return {
        "month": month,
        "year": year,
        "month_name": datetime(year, month, 1).strftime('%B %Y'),
        "due_date": due_date,
        "summary": {
            "total_branches": len(branches),
            "total_royalty": round(total_royalty, 2),
            "total_paid": round(total_paid, 2),
            "total_pending": round(total_pending, 2)
        },
        "branches": sorted(royalty_data, key=lambda x: x['royalty_amount'], reverse=True)
    }

@api_router.post("/royalty/mark-paid/{branch_id}")
async def mark_royalty_paid(
    branch_id: str,
    month: int,
    year: int,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Mark royalty as paid for a branch - Super Admin only"""
    await db.royalty_payments.update_one(
        {"branch_id": branch_id, "month": month, "year": year},
        {"$set": {
            "branch_id": branch_id,
            "month": month,
            "year": year,
            "is_paid": True,
            "paid_date": datetime.now(timezone.utc).isoformat(),
            "marked_by": current_user.email
        }},
        upsert=True
    )
    
    return {"message": "Royalty marked as paid"}

# ========== AUDIT LOGS ==========

@api_router.get("/audit-logs")
async def get_audit_logs(
    page: int = 1,
    limit: int = 50,
    user_id: str = None,
    entity_type: str = None,
    action: str = None,
    start_date: str = None,
    end_date: str = None,
    current_user: User = Depends(get_current_user)
):
    """Get audit logs - Branch Admin sees their branch users, Super Admin sees Branch Admins"""
    query = {}
    
    if current_user.role == UserRole.BRANCH_ADMIN:
        # Branch Admin sees logs of users in their branch (excluding other Branch Admins)
        query["branch_id"] = current_user.branch_id
        query["user_role"] = {"$in": ["trainer", "counsellor", "fde", "front_desk_executive"]}
    elif current_user.role == UserRole.ADMIN:
        # Super Admin sees Branch Admin logs
        query["user_role"] = "branch_admin"
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Apply filters
    if user_id:
        query["user_id"] = user_id
    if entity_type:
        query["entity_type"] = entity_type
    if action:
        query["action"] = action
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    skip = (page - 1) * limit
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.audit_logs.count_documents(query)
    
    return {
        "logs": logs,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit
    }

@api_router.get("/audit-logs/summary")
async def get_audit_logs_summary(
    days: int = 7,
    current_user: User = Depends(get_current_user)
):
    """Get audit log summary for dashboard"""
    query = {}
    
    if current_user.role == UserRole.BRANCH_ADMIN:
        query["branch_id"] = current_user.branch_id
        query["user_role"] = {"$in": ["trainer", "counsellor", "fde", "front_desk_executive"]}
    elif current_user.role == UserRole.ADMIN:
        query["user_role"] = "branch_admin"
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get logs from last N days
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    query["created_at"] = {"$gte": start_date}
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Summarize by user
    user_activity = {}
    action_counts = {}
    entity_counts = {}
    
    for log in logs:
        # By user
        user_key = f"{log['user_name']} ({log['user_email']})"
        if user_key not in user_activity:
            user_activity[user_key] = {"count": 0, "last_action": None}
        user_activity[user_key]["count"] += 1
        if not user_activity[user_key]["last_action"]:
            user_activity[user_key]["last_action"] = log['created_at']
        
        # By action
        action_counts[log['action']] = action_counts.get(log['action'], 0) + 1
        
        # By entity
        entity_counts[log['entity_type']] = entity_counts.get(log['entity_type'], 0) + 1
    
    return {
        "period_days": days,
        "total_actions": len(logs),
        "user_activity": [{"user": k, "count": v["count"], "last_action": v["last_action"]} for k, v in sorted(user_activity.items(), key=lambda x: x[1]["count"], reverse=True)],
        "action_breakdown": action_counts,
        "entity_breakdown": entity_counts,
        "recent_logs": logs[:10]
    }

# ======================
# RESPONSIBILITIES ENDPOINTS
# ======================

@api_router.get("/responsibilities")
async def get_responsibilities(current_user: User = Depends(get_current_user)):
    """Get responsibilities for the current user based on their role"""
    query = {
        "is_active": True,
        "$or": [
            {"user_id": current_user.id},  # Assigned specifically to this user
            {"user_id": None, "role": current_user.role}  # Assigned to their role
        ]
    }
    
    # If user has a branch, also include responsibilities for their branch
    if current_user.branch_id:
        query["$or"].append({"branch_id": current_user.branch_id, "role": current_user.role})
    
    responsibilities = await db.responsibilities.find(query, {"_id": 0}).to_list(100)
    
    # Sort by priority
    priority_order = {"high": 0, "medium": 1, "low": 2}
    responsibilities.sort(key=lambda x: priority_order.get(x.get("priority", "medium"), 1))
    
    return responsibilities

@api_router.get("/responsibilities/all")
async def get_all_responsibilities(
    role: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Get all responsibilities (Super Admin only)"""
    query = {}
    if role:
        query["role"] = role
    if branch_id:
        query["branch_id"] = branch_id
    
    responsibilities = await db.responsibilities.find(query, {"_id": 0}).to_list(500)
    return responsibilities

@api_router.post("/responsibilities")
async def create_responsibility(
    data: ResponsibilityCreate,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Create a new responsibility (Super Admin only)"""
    responsibility = Responsibility(
        **data.model_dump(),
        created_by=current_user.id
    )
    
    resp_dict = responsibility.model_dump()
    resp_dict['created_at'] = resp_dict['created_at'].isoformat()
    
    await db.responsibilities.insert_one(resp_dict)
    return responsibility

@api_router.put("/responsibilities/{resp_id}")
async def update_responsibility(
    resp_id: str,
    data: ResponsibilityUpdate,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Update a responsibility (Super Admin only)"""
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = await db.responsibilities.update_one(
        {"id": resp_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Responsibility not found")
    
    updated = await db.responsibilities.find_one({"id": resp_id}, {"_id": 0})
    return updated

@api_router.delete("/responsibilities/{resp_id}")
async def delete_responsibility(
    resp_id: str,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Delete a responsibility (Super Admin only)"""
    result = await db.responsibilities.delete_one({"id": resp_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Responsibility not found")
    
    return {"message": "Responsibility deleted successfully"}

# ============================================================================
# WIZBANG (Finance) ROLE — Global income / expense ledger with locked opening
# bank balance. Only one Wizbang user is allowed system-wide (enforced at
# registration). Admin can also view all Wizbang data.
# ============================================================================

class WizbangTransactionType(str, Enum):
    INCOME = "income"
    EXPENSE = "expense"

class WizbangTransactionCreate(BaseModel):
    type: WizbangTransactionType
    date: str  # YYYY-MM-DD
    amount: float
    category: str
    vendor_name: str
    payment_mode: PaymentMode
    description: Optional[str] = None
    receipt_url: Optional[str] = None

def _ensure_wizbang_or_admin(user: User):
    if user.role not in (UserRole.WIZBANG, UserRole.ADMIN):
        raise HTTPException(status_code=403, detail="Wizbang access required")

def _ensure_wizbang_brand_or_admin(user: User):
    """Allow Wizbang Brand Manager + Wizbang Admin (legacy enum value 'Wizbang') + Admin
    to access brand-side resources. Brand Manager gets read-only access to clients."""
    if user.role not in (UserRole.WIZBANG_BRAND_MANAGER, UserRole.WIZBANG, UserRole.ADMIN):
        raise HTTPException(status_code=403, detail="Brand Manager access required")

@api_router.get("/wizbang/account")
async def get_wizbang_account(current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    account = await db.wizbang_account.find_one({}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Wizbang account not configured. Ask the Super Admin to create a Wizbang user with an opening balance.")
    # Compute current balance
    total_income = 0.0
    total_expense = 0.0
    async for tx in db.wizbang_transactions.find({}, {"_id": 0, "type": 1, "amount": 1}):
        if tx["type"] == WizbangTransactionType.INCOME.value:
            total_income += float(tx.get("amount", 0))
        else:
            total_expense += float(tx.get("amount", 0))
    return {
        **account,
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "current_balance": round(account["opening_balance"] + total_income - total_expense, 2),
    }

@api_router.post("/wizbang/transactions")
async def create_wizbang_transaction(payload: WizbangTransactionCreate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    account = await db.wizbang_account.find_one({}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=400, detail="Wizbang account not configured")
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    
    record = {
        "id": str(uuid.uuid4()),
        "type": payload.type.value,
        "date": payload.date,
        "amount": float(payload.amount),
        "category": payload.category.strip(),
        "vendor_name": payload.vendor_name.strip(),
        "payment_mode": payload.payment_mode.value,
        "description": (payload.description or "").strip(),
        "receipt_url": payload.receipt_url,
        "created_by": current_user.id,
        "created_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wizbang_transactions.insert_one(record)
    return {k: v for k, v in record.items() if k != "_id"}

@api_router.get("/wizbang/transactions")
async def list_wizbang_transactions(type: Optional[str] = None, limit: int = 200, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    query = {}
    if type in (WizbangTransactionType.INCOME.value, WizbangTransactionType.EXPENSE.value):
        query["type"] = type
    docs = await db.wizbang_transactions.find(query, {"_id": 0}).sort("date", -1).to_list(length=limit)
    return docs

@api_router.delete("/wizbang/transactions/{tx_id}")
async def delete_wizbang_transaction(tx_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    result = await db.wizbang_transactions.delete_one({"id": tx_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {"message": "Transaction deleted"}

@api_router.get("/wizbang/dashboard")
async def wizbang_dashboard(current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    account = await db.wizbang_account.find_one({}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Wizbang account not configured")
    
    # Aggregate all transactions
    total_income = 0.0
    total_expense = 0.0
    regular_income = 0.0      # excludes credit repayments
    regular_expense = 0.0     # excludes credit-out entries
    credit_in_total = 0.0     # money received back from credits
    credit_out_total = 0.0    # money given out as credit
    monthly = {}  # {"YYYY-MM": {"income": x, "expense": y}}
    recent = []
    
    today = datetime.now(timezone.utc).date()
    current_month_key = today.strftime("%Y-%m")
    month_income = 0.0
    month_expense = 0.0
    
    async for tx in db.wizbang_transactions.find({}, {"_id": 0}).sort("date", -1):
        amt = float(tx.get("amount", 0))
        tx_type = tx.get("type")
        tx_date = tx.get("date", "")[:10]
        is_credit_linked = bool(tx.get("linked_credit_id"))
        month_key = tx_date[:7] if len(tx_date) >= 7 else ""
        if month_key:
            if month_key not in monthly:
                monthly[month_key] = {"income": 0.0, "expense": 0.0}
            if tx_type == WizbangTransactionType.INCOME.value:
                monthly[month_key]["income"] += amt
                total_income += amt
                if is_credit_linked:
                    credit_in_total += amt
                else:
                    regular_income += amt
                if month_key == current_month_key:
                    month_income += amt
            else:
                monthly[month_key]["expense"] += amt
                total_expense += amt
                if is_credit_linked:
                    credit_out_total += amt
                else:
                    regular_expense += amt
                if month_key == current_month_key:
                    month_expense += amt
        if len(recent) < 10:
            recent.append(tx)
    
    # Build last 12 months series (oldest -> newest)
    chart = []
    for i in range(11, -1, -1):
        ref_year = today.year
        ref_month = today.month - i
        while ref_month <= 0:
            ref_month += 12
            ref_year -= 1
        key = f"{ref_year:04d}-{ref_month:02d}"
        label = datetime(ref_year, ref_month, 1).strftime("%b %Y")
        bucket = monthly.get(key, {"income": 0.0, "expense": 0.0})
        chart.append({
            "month": key,
            "label": label,
            "income": round(bucket["income"], 2),
            "expense": round(bucket["expense"], 2),
            "net": round(bucket["income"] - bucket["expense"], 2),
        })
    
    current_balance = round(account["opening_balance"] + total_income - total_expense, 2)
    
    # Outstanding credit (money given out that hasn't been repaid yet)
    outstanding_credit = 0.0
    async for c in db.wizbang_credits.find({"is_repaid": False}, {"_id": 0, "amount": 1}):
        outstanding_credit += float(c.get("amount", 0))
    
    return {
        "account": {
            "opening_balance": account["opening_balance"],
            "opening_balance_set_at": account.get("opening_balance_set_at"),
            "owner_email": account.get("owner_email"),
        },
        "current_balance": current_balance,
        "outstanding_credit": round(outstanding_credit, 2),
        "breakdown": {
            "opening_balance": round(account["opening_balance"], 2),
            "regular_income": round(regular_income, 2),
            "credit_repaid_in": round(credit_in_total, 2),
            "regular_expense": round(regular_expense, 2),
            "credit_given_out": round(credit_out_total, 2),
            "current_balance": current_balance,
        },
        "totals": {
            "total_income": round(total_income, 2),
            "total_expense": round(total_expense, 2),
            "net": round(total_income - total_expense, 2),
        },
        "this_month": {
            "month": current_month_key,
            "income": round(month_income, 2),
            "expense": round(month_expense, 2),
            "net": round(month_income - month_expense, 2),
        },
        "chart_12_months": chart,
        "recent_transactions": recent,
    }


# ============================================================================
# WIZBANG — Clients / Credits / Invoices / Service Agreements
# ============================================================================

WIZBANG_BANK_DETAILS = {
    "bank_name": "Bank of Baroda",
    "branch": "Dhangu Road",
    "account_name": "Wizbang",
    "account_number": "65190200001543",
    "ifsc": "BARB0VJPAKO",
}

WIZBANG_INVOICE_TERMS = [
    "Payment due within 7 days of invoice date.",
    "Ad budgets and third-party costs are billed separately.",
    "Delays in approvals may affect delivery timelines.",
    "Additional revisions/scope changes may incur extra charges.",
    "Services may be paused for overdue payments.",
    "All deliverables become client property upon full payment.",
]

WIZBANG_LOGO_URL = "https://customer-assets.emergentagent.com/job_management-platform-1/artifacts/jcy4ud3n_image.png"


class WizbangClientCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    gstin: Optional[str] = None
    notes: Optional[str] = None


class WizbangClientUpdate(BaseModel):
    name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    gstin: Optional[str] = None
    notes: Optional[str] = None


class WizbangCreditCreate(BaseModel):
    recipient_name: str
    amount: float
    purpose: str
    given_date: str  # YYYY-MM-DD
    notes: Optional[str] = None


class WizbangInvoiceItem(BaseModel):
    description: str
    quantity: float = 1
    rate: float
    amount: float  # quantity * rate (sent by client to avoid float drift)


class WizbangInvoiceCreate(BaseModel):
    client_id: str
    issue_date: str  # YYYY-MM-DD
    due_date: str  # YYYY-MM-DD
    items: List[WizbangInvoiceItem]
    tax_enabled: bool = False
    tax_rate: float = 18.0  # percentage if tax_enabled
    notes: Optional[str] = None


class WizbangInvoiceUpdate(BaseModel):
    status: Optional[str] = None  # Draft / Sent / Paid / Overdue / Cancelled


class WizbangAgreementCreate(BaseModel):
    client_id: str
    title: str
    scope_brief: str  # short paragraph describing scope; the system expands it
    effective_date: str  # YYYY-MM-DD
    term_months: int = 12  # validity in months


def _build_agreement_content(client: dict, payload: WizbangAgreementCreate) -> str:
    """Generate a complete, lawyer-friendly service agreement body from a short scope brief."""
    today_str = datetime.now(timezone.utc).strftime("%d %B %Y")
    return f"""SERVICE AGREEMENT

This Service Agreement ("Agreement") is entered into on {today_str}
between:

WIZBANG ("Service Provider"), having its principal place of business in India,
hereinafter referred to as the "Company",
and

{client.get("name", "")} ("Client"),
{client.get("address") or "address on record"},
{("GSTIN: " + client["gstin"]) if client.get("gstin") else ""}
hereinafter referred to as the "Client".

1. SCOPE OF SERVICES
The Company agrees to provide the following services to the Client:

{payload.scope_brief.strip()}

The above scope may be refined in writing (including email) by mutual consent.

2. EFFECTIVE DATE & TERM
This Agreement shall be effective from {payload.effective_date} and shall remain
in force for a period of {payload.term_months} months unless terminated earlier
in accordance with this Agreement.

3. FEES & PAYMENT
3.1 Fees for the services shall be invoiced as per the Company's standard
invoicing terms and as detailed in invoices raised by the Company against this
engagement.
3.2 Payment is due within 7 (seven) days of invoice date.
3.3 Ad budgets and third-party costs are billed separately and are payable in
advance.
3.4 Services may be paused at the Company's sole discretion for overdue
payments.

4. CLIENT RESPONSIBILITIES
The Client agrees to (a) provide timely access to the assets, accounts and
approvals required to deliver the services; (b) designate a single point of
contact for approvals; and (c) acknowledge that delays in approvals may affect
delivery timelines.

5. CHANGES & REVISIONS
Additional revisions or scope changes beyond what is described in Section 1
may incur additional charges, which shall be communicated to and approved by
the Client in writing before being undertaken.

6. INTELLECTUAL PROPERTY
All deliverables created under this Agreement shall become the property of the
Client upon full payment of all applicable invoices. The Company retains the
right to display the work in its portfolio unless otherwise agreed in writing.

7. CONFIDENTIALITY
Both parties shall keep all non-public information shared during the
engagement strictly confidential and shall not disclose it to any third party
without the other party's prior written consent, except where required by law.

8. LIMITATION OF LIABILITY
The Company's total liability under this Agreement shall not exceed the fees
paid by the Client for the immediately preceding three (3) months of service.

9. TERMINATION
Either party may terminate this Agreement with 15 (fifteen) days' written
notice. All fees for work completed up to the date of termination shall remain
payable.

10. GOVERNING LAW & JURISDICTION
This Agreement shall be governed by and construed in accordance with the laws
of India. The courts at Pathankot, Punjab shall have exclusive jurisdiction
over any disputes arising under this Agreement.

11. ENTIRE AGREEMENT
This Agreement constitutes the entire understanding between the parties on the
subject matter and supersedes all prior discussions and writings.

12. ACCEPTANCE
By electronically signing below, the Client confirms that they have read,
understood and agreed to the terms of this Agreement. Electronic acceptance
shall be deemed equivalent to a handwritten signature under the Information
Technology Act, 2000.
"""


def _next_serial(prefix: str, collection_name: str) -> str:
    """Year-scoped serial like WIZ-2026-0001."""
    # Note: this is async-safe enough for our scale (single Wizbang user).
    return prefix  # placeholder — real version is async below


async def _next_invoice_number():
    year = datetime.now(timezone.utc).year
    count = await db.wizbang_invoices.count_documents({"invoice_number": {"$regex": f"^WIZ-{year}-"}})
    return f"WIZ-{year}-{count + 1:04d}"


async def _next_agreement_number():
    year = datetime.now(timezone.utc).year
    count = await db.wizbang_agreements.count_documents({"agreement_number": {"$regex": f"^AGR-{year}-"}})
    return f"AGR-{year}-{count + 1:04d}"


# ----- CLIENTS -----
@api_router.post("/wizbang/clients")
async def wizbang_create_client(payload: WizbangClientCreate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_by"] = current_user.id
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.wizbang_clients.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api_router.get("/wizbang/clients")
async def wizbang_list_clients(current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    docs = await db.wizbang_clients.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs


@api_router.put("/wizbang/clients/{client_id}")
async def wizbang_update_client(client_id: str, payload: WizbangClientUpdate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.wizbang_clients.update_one({"id": client_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    return {"message": "Client updated"}


@api_router.delete("/wizbang/clients/{client_id}")
async def wizbang_delete_client(client_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    used = await db.wizbang_invoices.count_documents({"client_id": client_id})
    if used:
        raise HTTPException(status_code=400, detail=f"Cannot delete: client has {used} invoice(s)")
    result = await db.wizbang_clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    return {"message": "Client deleted"}


# ----- CREDITS (money given on credit / repayments) -----
@api_router.post("/wizbang/credits")
async def wizbang_create_credit(payload: WizbangCreditCreate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    credit_id = str(uuid.uuid4())
    # Backing expense transaction so current_balance reflects the outflow
    tx_id = str(uuid.uuid4())
    tx_doc = {
        "id": tx_id,
        "type": WizbangTransactionType.EXPENSE.value,
        "date": payload.given_date,
        "amount": float(payload.amount),
        "category": "Credit Out",
        "vendor_name": payload.recipient_name,
        "payment_mode": PaymentMode.CASH.value,
        "description": f"Credit given: {payload.purpose}",
        "receipt_url": None,
        "linked_credit_id": credit_id,
        "created_by": current_user.id,
        "created_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wizbang_transactions.insert_one(tx_doc)
    credit_doc = {
        "id": credit_id,
        "recipient_name": payload.recipient_name,
        "amount": float(payload.amount),
        "purpose": payload.purpose,
        "given_date": payload.given_date,
        "notes": payload.notes,
        "is_repaid": False,
        "repaid_at": None,
        "repaid_amount": 0.0,
        "expense_tx_id": tx_id,
        "income_tx_id": None,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wizbang_credits.insert_one(credit_doc)
    return {k: v for k, v in credit_doc.items() if k != "_id"}


@api_router.get("/wizbang/credits")
async def wizbang_list_credits(status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    query = {}
    if status == "outstanding":
        query["is_repaid"] = False
    elif status == "repaid":
        query["is_repaid"] = True
    docs = await db.wizbang_credits.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs


@api_router.post("/wizbang/credits/{credit_id}/repay")
async def wizbang_repay_credit(credit_id: str, repaid_amount: Optional[float] = None, repaid_date: Optional[str] = None,
                                current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    credit = await db.wizbang_credits.find_one({"id": credit_id}, {"_id": 0})
    if not credit:
        raise HTTPException(status_code=404, detail="Credit not found")
    if credit.get("is_repaid"):
        raise HTTPException(status_code=400, detail="Credit already marked as repaid")
    amt = float(repaid_amount) if repaid_amount is not None else float(credit["amount"])
    date_str = repaid_date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Income transaction
    tx_id = str(uuid.uuid4())
    tx_doc = {
        "id": tx_id,
        "type": WizbangTransactionType.INCOME.value,
        "date": date_str,
        "amount": amt,
        "category": "Credit Repaid",
        "vendor_name": credit["recipient_name"],
        "payment_mode": PaymentMode.CASH.value,
        "description": f"Repayment of credit: {credit['purpose']}",
        "receipt_url": None,
        "linked_credit_id": credit_id,
        "created_by": current_user.id,
        "created_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wizbang_transactions.insert_one(tx_doc)
    await db.wizbang_credits.update_one(
        {"id": credit_id},
        {"$set": {"is_repaid": True, "repaid_at": date_str, "repaid_amount": amt, "income_tx_id": tx_id}}
    )
    return {"message": "Credit marked as repaid", "income_tx_id": tx_id}


@api_router.delete("/wizbang/credits/{credit_id}")
async def wizbang_delete_credit(credit_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    credit = await db.wizbang_credits.find_one({"id": credit_id}, {"_id": 0})
    if not credit:
        raise HTTPException(status_code=404, detail="Credit not found")
    # Also delete backing transaction(s) so balance is reverted
    if credit.get("expense_tx_id"):
        await db.wizbang_transactions.delete_one({"id": credit["expense_tx_id"]})
    if credit.get("income_tx_id"):
        await db.wizbang_transactions.delete_one({"id": credit["income_tx_id"]})
    await db.wizbang_credits.delete_one({"id": credit_id})
    return {"message": "Credit deleted"}


# ----- INVOICES -----
@api_router.post("/wizbang/invoices")
async def wizbang_create_invoice(payload: WizbangInvoiceCreate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    client = await db.wizbang_clients.find_one({"id": payload.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if not payload.items:
        raise HTTPException(status_code=400, detail="At least one line item is required")
    
    subtotal = sum(float(i.amount) for i in payload.items)
    tax_amount = round(subtotal * (payload.tax_rate / 100), 2) if payload.tax_enabled else 0.0
    total = round(subtotal + tax_amount, 2)
    
    invoice_number = await _next_invoice_number()
    invoice_doc = {
        "id": str(uuid.uuid4()),
        "invoice_number": invoice_number,
        "client_id": payload.client_id,
        "client_snapshot": {
            "name": client.get("name"),
            "address": client.get("address"),
            "email": client.get("email"),
            "phone": client.get("phone"),
            "gstin": client.get("gstin"),
        },
        "items": [i.model_dump() for i in payload.items],
        "subtotal": round(subtotal, 2),
        "tax_enabled": payload.tax_enabled,
        "tax_rate": payload.tax_rate if payload.tax_enabled else 0,
        "tax_amount": tax_amount,
        "total": total,
        "issue_date": payload.issue_date,
        "due_date": payload.due_date,
        "status": "Sent",
        "notes": payload.notes,
        "terms": WIZBANG_INVOICE_TERMS,
        "bank_details": WIZBANG_BANK_DETAILS,
        "logo_url": WIZBANG_LOGO_URL,
        "public_token": uuid.uuid4().hex,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wizbang_invoices.insert_one(invoice_doc)
    return {k: v for k, v in invoice_doc.items() if k != "_id"}


@api_router.get("/wizbang/invoices")
async def wizbang_list_invoices(current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    docs = await db.wizbang_invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs


@api_router.get("/wizbang/invoices/{invoice_id}")
async def wizbang_get_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    inv = await db.wizbang_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return inv


@api_router.put("/wizbang/invoices/{invoice_id}")
async def wizbang_update_invoice(invoice_id: str, payload: WizbangInvoiceUpdate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.wizbang_invoices.update_one({"id": invoice_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": "Invoice updated"}


@api_router.delete("/wizbang/invoices/{invoice_id}")
async def wizbang_delete_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    inv = await db.wizbang_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    # If the invoice has a linked income transaction (i.e. payment was recorded),
    # also delete that so the balance is restored.
    if inv.get("income_tx_id"):
        await db.wizbang_transactions.delete_one({"id": inv["income_tx_id"]})
    await db.wizbang_invoices.delete_one({"id": invoice_id})
    return {"message": "Invoice deleted"}


class InvoicePaymentPayload(BaseModel):
    payment_date: str  # YYYY-MM-DD
    payment_mode: PaymentMode
    amount: Optional[float] = None  # defaults to invoice.total
    notes: Optional[str] = None


@api_router.post("/wizbang/invoices/{invoice_id}/record-payment")
async def wizbang_record_invoice_payment(invoice_id: str, payload: InvoicePaymentPayload,
                                          current_user: User = Depends(get_current_user)):
    """Mark an invoice as Paid AND post a matching income transaction in one shot."""
    _ensure_wizbang_or_admin(current_user)
    inv = await db.wizbang_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.get("status") == "Paid":
        raise HTTPException(status_code=400, detail="Invoice is already marked Paid")
    if inv.get("status") == "Cancelled":
        raise HTTPException(status_code=400, detail="Cancelled invoices cannot be paid")
    amount = float(payload.amount) if payload.amount is not None else float(inv["total"])
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    tx_id = str(uuid.uuid4())
    description = f"Payment for invoice {inv['invoice_number']}"
    if payload.notes:
        description += f" — {payload.notes}"
    tx_doc = {
        "id": tx_id,
        "type": WizbangTransactionType.INCOME.value,
        "date": payload.payment_date,
        "amount": amount,
        "category": "Invoice Payment",
        "vendor_name": (inv.get("client_snapshot") or {}).get("name", "Client"),
        "payment_mode": payload.payment_mode.value,
        "description": description,
        "receipt_url": None,
        "linked_invoice_id": invoice_id,
        "linked_invoice_number": inv["invoice_number"],
        "created_by": current_user.id,
        "created_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wizbang_transactions.insert_one(tx_doc)
    await db.wizbang_invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "Paid",
            "paid_date": payload.payment_date,
            "paid_amount": amount,
            "payment_mode": payload.payment_mode.value,
            "income_tx_id": tx_id,
        }}
    )
    return {"message": "Payment recorded and invoice marked Paid", "income_tx_id": tx_id, "amount": amount}


@api_router.get("/public/wizbang/invoice/{token}")
async def public_view_invoice(token: str):
    """Public, no-auth view of an invoice via its share token."""
    inv = await db.wizbang_invoices.find_one({"public_token": token}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return inv


# ----- SERVICE AGREEMENTS -----
@api_router.post("/wizbang/agreements")
async def wizbang_create_agreement(payload: WizbangAgreementCreate, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    client = await db.wizbang_clients.find_one({"id": payload.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    agreement_number = await _next_agreement_number()
    content = _build_agreement_content(client, payload)
    doc = {
        "id": str(uuid.uuid4()),
        "agreement_number": agreement_number,
        "client_id": payload.client_id,
        "client_snapshot": {
            "name": client.get("name"),
            "email": client.get("email"),
            "phone": client.get("phone"),
            "address": client.get("address"),
            "gstin": client.get("gstin"),
        },
        "title": payload.title,
        "scope_brief": payload.scope_brief,
        "effective_date": payload.effective_date,
        "term_months": payload.term_months,
        "content": content,
        "status": "Sent",  # Sent -> Signed
        "public_token": uuid.uuid4().hex,
        "signed_by_name": None,
        "signed_by_email": None,
        "signed_at": None,
        "signed_ip": None,
        "logo_url": WIZBANG_LOGO_URL,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wizbang_agreements.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api_router.get("/wizbang/agreements")
async def wizbang_list_agreements(current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    docs = await db.wizbang_agreements.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs


@api_router.get("/wizbang/agreements/{agreement_id}")
async def wizbang_get_agreement(agreement_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    a = await db.wizbang_agreements.find_one({"id": agreement_id}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Agreement not found")
    return a


@api_router.delete("/wizbang/agreements/{agreement_id}")
async def wizbang_delete_agreement(agreement_id: str, current_user: User = Depends(get_current_user)):
    _ensure_wizbang_or_admin(current_user)
    a = await db.wizbang_agreements.find_one({"id": agreement_id}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Agreement not found")
    if a.get("status") == "Signed":
        raise HTTPException(status_code=400, detail="Signed agreements cannot be deleted")
    await db.wizbang_agreements.delete_one({"id": agreement_id})
    return {"message": "Agreement deleted"}


@api_router.get("/public/wizbang/agreement/{token}")
async def public_view_agreement(token: str):
    """Public no-auth view of an agreement."""
    a = await db.wizbang_agreements.find_one({"public_token": token}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Agreement not found")
    return a


class AgreementSignPayload(BaseModel):
    full_name: str
    email: Optional[EmailStr] = None
    agree: bool


@api_router.post("/public/wizbang/agreement/{token}/sign")
async def public_sign_agreement(token: str, payload: AgreementSignPayload, request: Request):
    """Public endpoint: client signs the agreement. One-time only; once signed it is locked."""
    a = await db.wizbang_agreements.find_one({"public_token": token}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Agreement not found")
    if a.get("status") == "Signed":
        raise HTTPException(status_code=400, detail="This agreement has already been signed and is locked")
    if not payload.agree:
        raise HTTPException(status_code=400, detail="You must agree to the terms to sign")
    if not payload.full_name or not payload.full_name.strip():
        raise HTTPException(status_code=400, detail="Full name is required")
    # Capture IP
    fwd = request.headers.get("x-forwarded-for", "")
    ip = fwd.split(",")[0].strip() if fwd else (request.client.host if request.client else "unknown")
    signed_at = datetime.now(timezone.utc).isoformat()
    await db.wizbang_agreements.update_one(
        {"public_token": token},
        {"$set": {
            "status": "Signed",
            "signed_by_name": payload.full_name.strip(),
            "signed_by_email": payload.email,
            "signed_at": signed_at,
            "signed_ip": ip,
        }}
    )
    return {"message": "Agreement signed successfully", "signed_at": signed_at}


app.include_router(api_router)

@app.middleware("http")
async def add_no_crawl_headers(request: Request, call_next):
    """Prevent search engine indexing across the entire app."""
    response = await call_next(request)
    response.headers["X-Robots-Tag"] = "noindex, nofollow, noarchive, nosnippet, noimageindex"
    return response

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

async def seed_default_users():
    """Idempotently create the canonical seed accounts on cold start.

    Skips any user whose email already exists. Passwords are read from env vars
    (SEED_*_PASSWORD) with sensible defaults that match /app/memory/test_credentials.md.
    Never overwrites an existing user's password or role.
    """
    seeds = [
        {
            "email": os.environ.get("SEED_ADMIN_EMAIL", "admin@etieducom.com"),
            "password": os.environ.get("SEED_ADMIN_PASSWORD", "admin@123"),
            "name": "Super Admin",
            "role": UserRole.ADMIN,
        },
        {
            "email": os.environ.get("SEED_PLACEMENT_EMAIL", "placement@etieducom.com"),
            "password": os.environ.get("SEED_PLACEMENT_PASSWORD", "placement@123"),
            "name": "Placement Manager",
            "role": UserRole.PLACEMENT_MANAGER,
        },
        {
            "email": os.environ.get("SEED_BRAND_MGR_EMAIL", "brandmgr@etieducom.com"),
            "password": os.environ.get("SEED_BRAND_MGR_PASSWORD", "brand@123"),
            "name": "Brand Manager",
            "role": UserRole.WIZBANG_BRAND_MANAGER,
        },
        {
            "email": os.environ.get("SEED_WIZBANG_EMAIL", "wizbang@etieducom.com"),
            "password": os.environ.get("SEED_WIZBANG_PASSWORD", "wizbang@123"),
            "name": "Wizbang Finance",
            "role": UserRole.WIZBANG,
        },
    ]

    created = []
    for seed in seeds:
        try:
            existing = await db.users.find_one({"email": seed["email"]}, {"_id": 0, "id": 1})
            if existing:
                continue
            new_user = User(
                email=seed["email"],
                name=seed["name"],
                role=seed["role"],
                hashed_password=get_password_hash(seed["password"]),
            )
            user_dict = new_user.model_dump()
            user_dict["created_at"] = user_dict["created_at"].isoformat()
            await db.users.insert_one(user_dict)
            created.append(seed["email"])
        except Exception as e:
            logger.error(f"Failed to seed user {seed['email']}: {e}")

    if created:
        logger.info(f"Seeded default users: {', '.join(created)}")
    else:
        logger.info("Seed users already present; nothing to insert")


@app.on_event("startup")
async def startup_event():
    """Initialize scheduler and create indexes on app startup"""
    # Create MongoDB indexes for performance
    try:
        await db.enrollments.create_index("id", unique=True)
        await db.enrollments.create_index("branch_id")
        await db.enrollments.create_index("enrollment_date")
        await db.enrollments.create_index("status")
        await db.enrollments.create_index("phone")
        await db.enrollments.create_index("email")
        await db.payments.create_index("enrollment_id")
        await db.payments.create_index("branch_id")
        await db.payments.create_index("payment_date")
        await db.payment_plans.create_index("enrollment_id", unique=True)
        await db.addon_courses.create_index("enrollment_id")
        await db.leads.create_index("id", unique=True)
        await db.leads.create_index("branch_id")
        await db.leads.create_index("counsellor_id")
        await db.leads.create_index("status")
        await db.leads.create_index("is_deleted")
        await db.users.create_index("email", unique=True)
        await db.users.create_index("branch_id")
        await db.followups.create_index("lead_id")
        await db.followups.create_index("branch_id")
        await db.followups.create_index("status")
        await db.followups.create_index("followup_date")
        await db.batches.create_index("branch_id")
        await db.batches.create_index("trainer_id")
        await db.attendance.create_index([("batch_id", 1), ("date", 1)])
        await db.attendance.create_index("enrollment_id")
        await db.audit_logs.create_index("created_at")
        await db.audit_logs.create_index("branch_id")
        await db.notifications.create_index("user_id")
        await db.installment_schedule.create_index("payment_plan_id")
        await db.installment_schedule.create_index("enrollment_id")
        await db.branches.create_index("id", unique=True)
        await db.programs.create_index("id", unique=True)
        await db.student_batch_assignments.create_index("enrollment_id")
        await db.student_batch_assignments.create_index("batch_id")
        logger.info("MongoDB indexes created successfully")
    except Exception as e:
        logger.error(f"Error creating indexes: {e}")

    # Idempotent seed of default role accounts (admin/placement/brand manager)
    try:
        await seed_default_users()
    except Exception as e:
        logger.error(f"Error seeding default users: {e}")

    # One-time cleanup: registration_number was removed from the certificate spec
    # (previously auto-generated from `count_documents + 1`, which collided across
    # concurrent creates and deletes, giving multiple students the same value).
    # Wipe any lingering legacy values so re-downloaded PDFs & verify page never
    # show them. Idempotent — a no-op once all rows are already blank.
    try:
        purge = await db.certificate_requests.update_many(
            {"registration_number": {"$nin": ["", None]}},
            {"$set": {"registration_number": ""}}
        )
        if purge.modified_count:
            logger.info(
                f"Cleared legacy registration_number on {purge.modified_count} certificate(s)"
            )
    except Exception as e:
        logger.error(f"Error purging legacy registration_number: {e}")
    
    # Schedule fee reminders to run daily at 9:00 AM
    scheduler.add_job(
        send_fee_reminders,
        CronTrigger(hour=9, minute=0),
        id="fee_reminders",
        replace_existing=True
    )
    
    # Schedule birthday wishes to run daily at 8:00 AM
    scheduler.add_job(
        send_birthday_wishes,
        CronTrigger(hour=8, minute=0),
        id="birthday_wishes",
        replace_existing=True
    )

    # Schedule demo reminders every 5 minutes (fires WhatsApp ~30 min before each demo)
    scheduler.add_job(
        send_demo_reminders,
        CronTrigger(minute="*/5"),
        id="demo_reminders",
        replace_existing=True,
    )

    # Schedule Meta Ads Google Sheet sync every minute (near real-time)
    scheduler.add_job(
        sync_all_meta_sheets,
        CronTrigger(minute="*"),
        id="meta_sheet_sync",
        replace_existing=True,
    )

    # Schedule fee LATE reminders daily at 9:30 AM (just after the regular fee reminders)
    scheduler.add_job(
        send_fee_late_reminders,
        CronTrigger(hour=9, minute=30),
        id="fee_late_reminders",
        replace_existing=True,
    )

    scheduler.start()
    logger.info("Scheduler started - Fee 9AM, FeeLate 9:30AM, Birthday 8AM, Demo reminders /5min, Meta sheet /1min")
    
    # Migration: Add branch_id to existing followups that don't have it
    try:
        followups_without_branch = await db.followups.find({"branch_id": {"$exists": False}}, {"_id": 0, "id": 1, "lead_id": 1}).to_list(10000)
        if followups_without_branch:
            logger.info(f"Migrating {len(followups_without_branch)} followups to add branch_id")
            for fu in followups_without_branch:
                lead = await db.leads.find_one({"id": fu.get('lead_id')}, {"_id": 0, "branch_id": 1})
                if lead and lead.get('branch_id'):
                    await db.followups.update_one(
                        {"id": fu['id']},
                        {"$set": {"branch_id": lead['branch_id']}}
                    )
            logger.info("Followup migration completed")
    except Exception as e:
        logger.error(f"Error during followup migration: {e}")

    # Migration: Update WhatsApp template names to ETI defaults if blank
    try:
        eti_templates = {
            "enquiry_saved": "eti_enquiry_confirmation",
            "demo_booked": "eti_demo_confirmation",
            "enrollment_confirmed": "eti_enrollment_confirmation",
            "payment_received": "eti_payment",
            "fee_reminder": "eti_fee_reminder",
            "birthday_wishes": "eti_birthday_wishes",
            "certificate_ready": "eti_certificate",
        }
        eti_namespace = "73fda5e9_77e9_445f_82ac_9c2e532b32f4"
        ws = await db.whatsapp_settings.find_one({}, {"_id": 0})
        if ws:
            events = ws.get("events", {}) or {}
            changed = False
            for event_key, default_template in eti_templates.items():
                cfg = events.get(event_key, {}) or {}
                if not cfg.get("template_name"):
                    cfg["template_name"] = default_template
                    changed = True
                if not cfg.get("namespace"):
                    cfg["namespace"] = eti_namespace
                    changed = True
                events[event_key] = cfg
            if changed:
                await db.whatsapp_settings.update_one({"id": ws.get("id")}, {"$set": {"events": events}})
                logger.info("WhatsApp template defaults migrated to ETI templates")
    except Exception as e:
        logger.error(f"Error during WhatsApp template migration: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()
    logger.info("Scheduler and database connection closed")
